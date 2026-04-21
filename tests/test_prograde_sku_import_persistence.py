import importlib
import csv
import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


class ProgradeSkuImportPersistenceTests(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self._db_path = Path(self._tmpdir.name) / "prograde_sku_import_persistence.db"
        self._previous_db_path = os.environ.get("PROGRADE_DB_PATH")
        self._previous_preserve_sku_edits = os.environ.get("PROGRADE_PRESERVE_SKU_EDITS_ON_START")
        os.environ["PROGRADE_DB_PATH"] = str(self._db_path)
        os.environ.pop("PROGRADE_PRESERVE_SKU_EDITS_ON_START", None)

        import blueprints.prograde.db as prograde_db

        self.db = importlib.reload(prograde_db)
        self.db.init_db()

    def tearDown(self):
        if self._previous_db_path is None:
            os.environ.pop("PROGRADE_DB_PATH", None)
        else:
            os.environ["PROGRADE_DB_PATH"] = self._previous_db_path
        if self._previous_preserve_sku_edits is None:
            os.environ.pop("PROGRADE_PRESERVE_SKU_EDITS_ON_START", None)
        else:
            os.environ["PROGRADE_PRESERVE_SKU_EDITS_ON_START"] = self._previous_preserve_sku_edits
        try:
            self._tmpdir.cleanup()
        except PermissionError:
            pass

    def _write_bigtex_workbook(self, workbook_path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["MCAT", "Model", "Item Number", "Bed Length", "Tongue"])
        ws.append(["DUMP", "14LP", "BT-UPDATE", 16, 5])
        ws.append(["UTILITY", "35SA", "BT-NEW", 12, 4])
        wb.save(workbook_path)

    def _first_seed_item(self, seed_filename: str) -> str:
        seed_path = Path(__file__).resolve().parents[1] / "data" / "seed" / seed_filename
        with seed_path.open("r", newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                item_number = str(row.get("item_number") or "").strip().upper()
                if item_number:
                    return item_number
        self.fail(f"No item_number rows found in {seed_filename}")

    def _write_pj_workbook(self, workbook_path: Path):
        wb = Workbook()
        toc = wb.active
        toc.title = "ToC"
        toc["A1"] = "Utility"
        toc["A1"].font = Font(bold=True)
        toc["A2"] = "Utility Trailers [UT]"

        ut_sheet = wb.create_sheet("UT")
        ut_sheet.append(["Code", "Description"])
        ut_sheet.append(["PJ-UPDATE", "16' Utility Trailer"])
        ut_sheet.append(["PJ-NEW", "14' Utility Trailer"])

        wb.save(workbook_path)

    def test_import_bigtex_workbook_upserts_without_deleting_existing_rows(self):
        now = datetime.utcnow().isoformat()
        with self.db.get_db() as conn:
            conn.execute("DELETE FROM bigtex_skus")
            conn.execute(
                """
                INSERT INTO bigtex_skus
                (item_number, mcat, model, bed_length, tongue, total_footprint, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("BT-KEEP", "GOOSENECK", "22GN", 20.0, 6.0, 26.0, now),
            )
            conn.execute(
                """
                INSERT INTO bigtex_skus
                (item_number, mcat, model, bed_length, tongue, total_footprint, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                ("BT-UPDATE", "DUMP", "14LP", 14.0, 5.0, 19.0, now),
            )

        workbook_path = Path(self._tmpdir.name) / "bigtex_import.xlsx"
        self._write_bigtex_workbook(workbook_path)
        result = self.db.import_bigtex_skus_from_workbook(workbook_path=workbook_path, sheet_name="Data")

        with self.db.get_db() as conn:
            total_rows = int(conn.execute("SELECT COUNT(*) FROM bigtex_skus").fetchone()[0] or 0)
            keep_row = conn.execute(
                "SELECT model, bed_length, tongue, total_footprint FROM bigtex_skus WHERE item_number=?",
                ("BT-KEEP",),
            ).fetchone()
            update_row = conn.execute(
                "SELECT model, bed_length, tongue, total_footprint FROM bigtex_skus WHERE item_number=?",
                ("BT-UPDATE",),
            ).fetchone()
            new_row = conn.execute(
                "SELECT model FROM bigtex_skus WHERE item_number=?",
                ("BT-NEW",),
            ).fetchone()

        self.assertEqual(total_rows, 3)
        self.assertIsNotNone(keep_row)
        self.assertEqual(str(keep_row["model"]), "22GN")
        self.assertEqual(float(keep_row["total_footprint"]), 26.0)
        self.assertIsNotNone(update_row)
        self.assertEqual(float(update_row["bed_length"]), 16.0)
        self.assertEqual(float(update_row["tongue"]), 5.0)
        self.assertEqual(float(update_row["total_footprint"]), 21.0)
        self.assertIsNotNone(new_row)
        self.assertEqual(result["row_count"], 2)
        self.assertEqual(result["created_count"], 1)
        self.assertEqual(result["updated_count"], 1)
        self.assertEqual(result["total_row_count"], 3)

    def test_import_pj_workbook_upserts_without_deleting_existing_rows(self):
        now = datetime.utcnow().isoformat()
        with self.db.get_db() as conn:
            conn.execute("DELETE FROM pj_skus")
            conn.execute(
                """
                INSERT INTO pj_skus
                (item_number, model, pj_category, description, gvwr, bed_length_stated, bed_length_measured,
                 tongue_group, tongue_feet, total_footprint, dump_side_height_ft, can_nest_inside_dump,
                 gn_axle_droppable, tongue_overlap_allowed, pairing_rule, notes, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "PJ-KEEP",
                    "UL",
                    "utility",
                    "Keep Existing",
                    None,
                    12.0,
                    12.0,
                    "c_channel",
                    2.0,
                    14.0,
                    None,
                    0,
                    0,
                    0,
                    None,
                    None,
                    now,
                ),
            )
            conn.execute(
                """
                INSERT INTO pj_skus
                (item_number, model, pj_category, description, gvwr, bed_length_stated, bed_length_measured,
                 tongue_group, tongue_feet, total_footprint, dump_side_height_ft, can_nest_inside_dump,
                 gn_axle_droppable, tongue_overlap_allowed, pairing_rule, notes, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "PJ-UPDATE",
                    "UT",
                    "utility",
                    "Old Description",
                    None,
                    12.0,
                    12.0,
                    "c_channel",
                    2.0,
                    14.0,
                    None,
                    0,
                    0,
                    0,
                    None,
                    None,
                    now,
                ),
            )

        workbook_path = Path(self._tmpdir.name) / "pj_import.xlsx"
        self._write_pj_workbook(workbook_path)
        result = self.db.import_pj_skus_from_workbook(workbook_path=workbook_path, toc_sheet_name="ToC")

        with self.db.get_db() as conn:
            total_rows = int(conn.execute("SELECT COUNT(*) FROM pj_skus").fetchone()[0] or 0)
            keep_row = conn.execute(
                "SELECT description FROM pj_skus WHERE item_number=?",
                ("PJ-KEEP",),
            ).fetchone()
            update_row = conn.execute(
                "SELECT model, description, bed_length_stated FROM pj_skus WHERE item_number=?",
                ("PJ-UPDATE",),
            ).fetchone()
            new_row = conn.execute(
                "SELECT model FROM pj_skus WHERE item_number=?",
                ("PJ-NEW",),
            ).fetchone()

        self.assertEqual(total_rows, 3)
        self.assertIsNotNone(keep_row)
        self.assertEqual(str(keep_row["description"]), "Keep Existing")
        self.assertIsNotNone(update_row)
        self.assertEqual(str(update_row["model"]), "UT")
        self.assertEqual(str(update_row["description"]), "16' Utility Trailer")
        self.assertEqual(float(update_row["bed_length_stated"]), 16.0)
        self.assertIsNotNone(new_row)
        self.assertEqual(result["row_count"], 2)
        self.assertEqual(result["created_count"], 1)
        self.assertEqual(result["updated_count"], 1)
        self.assertEqual(result["total_row_count"], 3)

    def test_init_db_backfills_missing_seed_rows_on_restart(self):
        bt_item = self._first_seed_item("bigtex_skus.csv")
        pj_item = self._first_seed_item("pj_skus.csv")
        with self.db.get_db() as conn:
            bt_before = int(conn.execute("SELECT COUNT(*) FROM bigtex_skus").fetchone()[0] or 0)
            pj_before = int(conn.execute("SELECT COUNT(*) FROM pj_skus").fetchone()[0] or 0)
            conn.execute("DELETE FROM bigtex_skus WHERE item_number=?", (bt_item,))
            conn.execute("DELETE FROM pj_skus WHERE item_number=?", (pj_item,))

        with self.db.get_db() as conn:
            bt_missing = int(conn.execute("SELECT COUNT(*) FROM bigtex_skus WHERE item_number=?", (bt_item,)).fetchone()[0] or 0)
            pj_missing = int(conn.execute("SELECT COUNT(*) FROM pj_skus WHERE item_number=?", (pj_item,)).fetchone()[0] or 0)
            self.assertEqual(bt_missing, 0)
            self.assertEqual(pj_missing, 0)

        self.db.init_db()

        with self.db.get_db() as conn:
            bt_after = int(conn.execute("SELECT COUNT(*) FROM bigtex_skus").fetchone()[0] or 0)
            pj_after = int(conn.execute("SELECT COUNT(*) FROM pj_skus").fetchone()[0] or 0)
            bt_restored = int(conn.execute("SELECT COUNT(*) FROM bigtex_skus WHERE item_number=?", (bt_item,)).fetchone()[0] or 0)
            pj_restored = int(conn.execute("SELECT COUNT(*) FROM pj_skus WHERE item_number=?", (pj_item,)).fetchone()[0] or 0)

        self.assertEqual(bt_after, bt_before)
        self.assertEqual(pj_after, pj_before)
        self.assertEqual(bt_restored, 1)
        self.assertEqual(pj_restored, 1)

    def test_init_db_overwrites_existing_sku_edits_on_restart(self):
        bt_item = self._first_seed_item("bigtex_skus.csv")
        pj_item = self._first_seed_item("pj_skus.csv")
        with self.db.get_db() as conn:
            bt_original = conn.execute(
                "SELECT model FROM bigtex_skus WHERE item_number=?",
                (bt_item,),
            ).fetchone()
            pj_original = conn.execute(
                "SELECT description FROM pj_skus WHERE item_number=?",
                (pj_item,),
            ).fetchone()
            conn.execute("UPDATE bigtex_skus SET model=? WHERE item_number=?", ("CUSTOM-BT-MODEL", bt_item))
            conn.execute("UPDATE pj_skus SET description=? WHERE item_number=?", ("Custom PJ Description", pj_item))

        self.db.init_db()

        with self.db.get_db() as conn:
            bt_row = conn.execute("SELECT model FROM bigtex_skus WHERE item_number=?", (bt_item,)).fetchone()
            pj_row = conn.execute("SELECT description FROM pj_skus WHERE item_number=?", (pj_item,)).fetchone()

        self.assertIsNotNone(bt_original)
        self.assertIsNotNone(pj_original)
        self.assertIsNotNone(bt_row)
        self.assertIsNotNone(pj_row)
        self.assertEqual(str(bt_row["model"]), str(bt_original["model"]))
        self.assertEqual(str(pj_row["description"]), str(pj_original["description"]))

    def test_preserve_flag_keeps_existing_sku_edits_on_restart(self):
        bt_item = self._first_seed_item("bigtex_skus.csv")
        pj_item = self._first_seed_item("pj_skus.csv")
        with self.db.get_db() as conn:
            conn.execute("UPDATE bigtex_skus SET model=? WHERE item_number=?", ("CUSTOM-BT-MODEL", bt_item))
            conn.execute("UPDATE pj_skus SET description=? WHERE item_number=?", ("Custom PJ Description", pj_item))

        os.environ["PROGRADE_PRESERVE_SKU_EDITS_ON_START"] = "true"
        import blueprints.prograde.db as prograde_db
        self.db = importlib.reload(prograde_db)
        self.db.init_db()

        with self.db.get_db() as conn:
            bt_row = conn.execute("SELECT model FROM bigtex_skus WHERE item_number=?", (bt_item,)).fetchone()
            pj_row = conn.execute("SELECT description FROM pj_skus WHERE item_number=?", (pj_item,)).fetchone()

        self.assertIsNotNone(bt_row)
        self.assertIsNotNone(pj_row)
        self.assertEqual(str(bt_row["model"]), "CUSTOM-BT-MODEL")
        self.assertEqual(str(pj_row["description"]), "Custom PJ Description")


if __name__ == "__main__":
    unittest.main()
