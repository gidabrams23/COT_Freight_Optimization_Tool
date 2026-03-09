import json
import logging
import math
import os
import re
import secrets
import subprocess
import threading
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from flask import (
    Flask,
    redirect,
    render_template,
    request,
    url_for,
    Response,
    jsonify,
    session,
    abort,
)

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    from openpyxl.drawing.image import Image as OpenPyxlImage
except Exception:  # pragma: no cover - optional dependency path
    OpenPyxlImage = None

try:
    from PIL import Image as PILImage, ImageDraw
except Exception:  # pragma: no cover - optional dependency path
    PILImage = None
    ImageDraw = None

import db
from services import (
    load_builder,
    orders as order_service,
    order_categories,
    stack_calculator,
    geo_utils,
    tsp_solver,
    customer_rules,
    routing_service,
    replay_evaluator,
)
from services.cost_calculator import (
    FUEL_SURCHARGE_SETTING_KEY,
    DEFAULT_FUEL_SURCHARGE_PER_MILE,
    STOP_FEE_SETTING_KEY,
    DEFAULT_STOP_FEE,
    MIN_LOAD_COST_SETTING_KEY,
    DEFAULT_MIN_LOAD_COST,
    DEFAULT_RATE_PER_MILE,
)
from services.optimizer_engine import OptimizerEngine
from services.order_importer import OrderImporter

logger = logging.getLogger(__name__)
ROOT_DIR = Path(__file__).resolve().parent


def _coerce_iso_date(raw_value):
    text = (raw_value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return None


def _git_last_updated_date():
    try:
        output = subprocess.check_output(
            ["git", "log", "-1", "--format=%cs"],
            cwd=str(ROOT_DIR),
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
    except Exception:
        return None
    return _coerce_iso_date(output)


def _source_last_updated_date():
    latest_ts = 0.0
    tracked_paths = [
        ROOT_DIR / "app.py",
        ROOT_DIR / "db.py",
        ROOT_DIR / "templates" / "login.html",
        ROOT_DIR / "static" / "styles.css",
    ]
    for path in tracked_paths:
        try:
            if path.exists():
                latest_ts = max(latest_ts, path.stat().st_mtime)
        except OSError:
            continue
    if latest_ts <= 0:
        return None
    return datetime.fromtimestamp(latest_ts).date().isoformat()


def _resolve_app_updated_on():
    return (
        _coerce_iso_date(os.environ.get("APP_UPDATED_ON"))
        or _git_last_updated_date()
        or _source_last_updated_date()
        or date.today().isoformat()
    )


APP_UPDATED_ON = _resolve_app_updated_on()
APP_VERSION = (os.environ.get("APP_VERSION") or "").strip() or f"v{APP_UPDATED_ON.replace('-', '.')}"
APP_RELEASE_LABEL = f"{APP_VERSION} | Updated {APP_UPDATED_ON}"


class UploadValidationError(Exception):
    def __init__(self, message, summary=None):
        super().__init__(message)
        self.summary = summary or {}


def _is_local_dev_mode():
    env_hint = (os.environ.get("APP_ENV") or os.environ.get("FLASK_ENV") or "").strip().lower()
    if env_hint in {"dev", "development", "local"}:
        return True
    return os.environ.get("FLASK_DEBUG", "").strip() == "1"


def _env_bool(name, default=False):
    raw = os.environ.get(name)
    if raw is None:
        return bool(default)
    normalized = str(raw).strip().lower()
    if normalized in {"1", "true", "yes", "on", "y"}:
        return True
    if normalized in {"0", "false", "no", "off", "n"}:
        return False
    return bool(default)


app = Flask(__name__)
_configured_secret = (os.environ.get("FLASK_SECRET_KEY") or "").strip()
if not _configured_secret and not _is_local_dev_mode():
    raise RuntimeError(
        "FLASK_SECRET_KEY must be set for non-development environments."
    )
if not _configured_secret:
    _configured_secret = "dev-session-key"
    logger.warning("Using development session secret key.")
app.secret_key = _configured_secret
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=_env_bool(
        "SESSION_COOKIE_SECURE",
        default=not _is_local_dev_mode(),
    ),
)
_raw_web_concurrency = (os.environ.get("WEB_CONCURRENCY") or "").strip()
try:
    _configured_web_concurrency = int(_raw_web_concurrency) if _raw_web_concurrency else 1
except ValueError:
    _configured_web_concurrency = 1
if _configured_web_concurrency > 1:
    logger.warning(
        "WEB_CONCURRENCY=%s detected. Re-optimization job status is process-local; "
        "set WEB_CONCURRENCY=1 to avoid cross-worker status loss.",
        _configured_web_concurrency,
    )

@app.template_filter("short_date")
def short_date(value):
    if value is None:
        return "—"

    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        raw = str(value).strip()
        if not raw:
            return "—"

        parsed = None
        try:
            parsed = date.fromisoformat(raw)
        except ValueError:
            try:
                parsed = datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
            except ValueError:
                for fmt in ("%m/%d/%Y", "%Y/%m/%d", "%b %d %Y", "%b %d, %Y"):
                    try:
                        parsed = datetime.strptime(raw, fmt).date()
                        break
                    except ValueError:
                        continue

        if parsed is None:
            return raw

    return f"{parsed.strftime('%b')} {parsed.day}"


@app.template_filter("est_datetime")
def est_datetime(value):
    return _format_est_datetime_label(value)

PLANT_CODES = ["GA", "TX", "VA", "IA", "OR", "NV", "CL"]
RATE_MATRIX_PLANT_DISPLAY_ORDER = ["GA", "OR", "TX", "IA", "VA", "NV"]
PLANT_NAMES = {
    "GA": "Lavonia",
    "IA": "Missouri Valley",
    "TX": "Mexia",
    "VA": "Montross",
    "CL": "Callao",
    "OR": "Coburg",
    "NV": "Winnemucca",
}
PLANT_DEFAULT_TRAILER_TYPE_OVERRIDES = {
    "VA": "FLATBED_48",
    "NV": "STEP_DECK_48",
}
PLANT_DEFAULT_AUTO_HOTSHOT_OVERRIDES = {}
FIXED_CAPACITY_TRAILER_TYPES = {"STEP_DECK_48", "FLATBED_48", "HOTSHOT", "WEDGE"}
STATUS_PROPOSED = "PROPOSED"
STATUS_DRAFT = "DRAFT"
STATUS_APPROVED = "APPROVED"
LOAD_NUMBER_START_PATTERN = re.compile(r"^\d{4}$")
OPTIMIZER_V2_ENABLED = (
    os.environ.get("OPTIMIZER_V2_ENABLED", "1").strip().lower() in {"1", "true", "yes", "on"}
)
ROLE_ADMIN = "admin"
ROLE_PLANNER = "planner"
APP_TIMEZONE = ZoneInfo("America/New_York")
SESSION_PROFILE_ID_KEY = "profile_id"
SESSION_PROFILE_NAME_KEY = "profile_name"
SESSION_PROFILE_DEFAULT_PLANTS_KEY = "profile_default_plants"
SESSION_PROFILE_SANDBOX_KEY = "profile_is_sandbox"
SESSION_ACTIVE_PLANNING_ID_KEY = "active_planning_session_id"
ORDER_REMOVAL_REASONS = [
    "Customer mixing conflict",
    "Capacity exceeded",
    "Geographic infeasibility",
    "Delivery date conflict",
    "Other",
]
LOAD_REJECTION_REASONS = [
    "Customer mixing conflict",
    "Capacity exceeded",
    "Geographic infeasibility",
    "Route inefficiency",
    "Delivery date conflicts",
    "Other",
]
LOAD_CARRIER_SELECTION_OPTIONS = [
    {"key": "fls", "label": "FLS"},
    {"key": "lst", "label": "LST"},
    {"key": "ryder", "label": "Ryder Dedicated"},
    {"key": "alternate", "label": "Alternate Trailer"},
]
ADMIN_ONLY_PATH_PREFIXES = (
    "/access/manage",
    "/planning-sessions/replay",
)
OPTIMIZER_DEFAULTS_SETTING_KEY = "optimizer_defaults"
UTILIZATION_GRADE_THRESHOLDS_SETTING_KEY = "utilization_grade_thresholds"
REPLAY_EVAL_PRESET_SETTING_KEY = "replay_eval_preset"
STOP_COLOR_PALETTE_SETTING_KEY = "stop_color_palette"
TRAILER_ASSIGNMENT_RULES_SETTING_KEY = "trailer_assignment_rules"
STRATEGIC_CUSTOMERS_SETTING_KEY = "strategic_customers"
PLANNER_SETTING_OVERRIDE_SETTING_PREFIX = "planner_setting_override::"
PLANNER_TRAILER_RULES_OVERRIDE_SETTING_PREFIX = "planner_trailer_assignment_rules_override::"
RATE_TABLE_CONTEXTS_SETTING_KEY = "rate_table_contexts"
RYDER_DEDICATED_RATE_TABLE_SETTING_KEY = "ryder_dedicated_rate_table"
LST_RATE_TABLE_SETTING_KEY = "lst_rate_matrix"
ALTERNATE_TRAILER_RATES_SETTING_KEY = "alternate_trailer_rates"
DEFAULT_RATE_CHANGE_METADATA_SETTING_KEY = "default_rate_change_metadata"
DEFAULT_UTILIZATION_GRADE_THRESHOLDS = {"A": 85, "B": 70, "C": 55, "D": 40}
DEFAULT_TRAILER_ASSIGNMENT_RULES = {
    "livestock_wedge_enabled": True,
    "livestock_category_tokens": ["LIVESTOCK"],
}
DEFAULT_TRACTOR_SUPPLY_CARGO_WEDGE_MIN_ITEM_LENGTH_FT = 20.0
DEFAULT_TRACTOR_SUPPLY_UTA_WEDGE_MIN_ITEM_LENGTH_FT = 22.0
DEFAULT_RATE_TABLE_CONTEXTS = {
    "default_rate_table_key": "FLS",
    "carrier_dedicated_ryder_rate_table_key": "LST",
    "trailer_hotshot_rate_table_key": "ALTERNATE_TRAILERS",
}
RATE_TABLE_KEY_OPTIONS = [
    {"key": "FLS", "label": "FLS (DEFAULT MATRIX)"},
    {"key": "RYDER_DEDICATED", "label": "RYDER DEDICATED"},
    {"key": "LST", "label": "LST (CARRIER MATRIX)"},
    {"key": "ALTERNATE_TRAILERS", "label": "ALTERNATE TRAILERS"},
]
ALTERNATE_TRAILER_SECTION_DEFINITIONS = [
    {"code": "HOT_SHOT", "label": "Hot Shot"},
    {"code": "WEDGE", "label": "Wedge"},
    {"code": "FLAT_BED", "label": "Flat Bed"},
]
DEFAULT_RYDER_DEDICATED_RATE_TABLE = {
    "name": "Ryder Dedicated",
    "applies_to": "ALL_STATES",
    "notes": "Round-trip rates include fuel surcharge and stop-off.",
    "plants": list(RATE_MATRIX_PLANT_DISPLAY_ORDER),
    "rates_by_plant": {
        "GA": 3.68,
        "OR": None,
        "TX": 4.18,
        "IA": 4.60,
        "VA": 5.00,
        "NV": None,
    },
    "fuel_surcharge": 0.0,
    "per_stop": 0.0,
    "load_minimum": 0.0,
}
DEFAULT_LST_RATE_TABLE_SETTINGS = {
    "carrier": "LST",
    "plants": list(RATE_MATRIX_PLANT_DISPLAY_ORDER),
    "states": [],
    "state_lane_labels": {},
    "matrix": {},
    "fuel_surcharge": 0.0,
    "per_stop": 0.0,
    "load_minimum": 0.0,
}
DEFAULT_ALTERNATE_TRAILER_RATES = {
    "sections": [
        {
            "code": section["code"],
            "label": section["label"],
            "plants": list(RATE_MATRIX_PLANT_DISPLAY_ORDER),
            "rates_by_plant": {plant: None for plant in RATE_MATRIX_PLANT_DISPLAY_ORDER},
            "placeholders_by_plant": {
                plant: {
                    "per_stop": 0.0,
                    "load_minimum": 0.0,
                    "apply_fuel_surcharge": True,
                    "requires_return_miles": True,
                }
                for plant in RATE_MATRIX_PLANT_DISPLAY_ORDER
            },
        }
        for section in ALTERNATE_TRAILER_SECTION_DEFINITIONS
    ],
}
DEFAULT_STACK_OVERFLOW_MAX_HEIGHT = 5
DEFAULT_MAX_BACK_OVERHANG_FT = 4.0
DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT = 7.0
DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT = 16.0
DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT = 6.0
DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES = ["USA", "UTA"]
DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED = True
DEFAULT_STOP_COLOR_PALETTE = [
    "#6FAD47",
    "#01B0F0",
    "#EC7D31",
    "#FFFFFF",
    "#FE0000",
    "#A56CD2",
    "#A5A6A6",
    "#FED966",
    "#FE6699",
    "#6CF8FB",
    "#5F87CC",
    "#FEFF00",
]
FALLBACK_STOP_COLOR = "#64748B"
HEX_COLOR_PATTERN = re.compile(r"^#[0-9A-F]{6}$")
TRAILER_PROFILE_OPTIONS = stack_calculator.trailer_profile_options()
TUTORIAL_MANIFEST_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "docs",
    "tutorial",
    "tutorial_manifest.json",
)
TUTORIAL_ALLOWED_MEDIA_TYPES = {"image", "video"}
TUTORIAL_ALLOWED_AUDIENCE = {"all", ROLE_ADMIN, ROLE_PLANNER}
TUTORIAL_NOTE_LABELS = {
    "tip": "Tip",
    "warning": "Warning",
    "required": "Required",
}
TUTORIAL_NAV_ENABLED = _env_bool("TUTORIAL_NAV_ENABLED", default=False)
REOPT_JOB_RETENTION_SEC = 60 * 60
REOPT_JOB_MAX_ENTRIES = 200
_REOPT_JOB_LOCK = threading.Lock()
_REOPT_JOBS = {}
ACCESS_PROFILES_SEED_PATH = Path(
    os.environ.get("ACCESS_PROFILES_SEED_PATH", str(ROOT_DIR / "data" / "seed" / "access_profiles.csv"))
)
ACCESS_PROFILES_SEED_COLUMNS = ["name", "is_admin", "is_sandbox", "allowed_plants", "default_plants", "created_at"]


def _sync_access_profiles_seed_snapshot():
    try:
        profiles = db.list_access_profiles()
        ACCESS_PROFILES_SEED_PATH.parent.mkdir(parents=True, exist_ok=True)
        with ACCESS_PROFILES_SEED_PATH.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=ACCESS_PROFILES_SEED_COLUMNS)
            writer.writeheader()
            for profile in profiles:
                writer.writerow(
                    {
                        "name": (profile.get("name") or "").strip(),
                        "is_admin": 1 if profile.get("is_admin") else 0,
                        "is_sandbox": 1 if profile.get("is_sandbox") else 0,
                        "allowed_plants": profile.get("allowed_plants") or "ALL",
                        "default_plants": profile.get("default_plants") or "ALL",
                        "created_at": profile.get("created_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    }
                )
    except Exception:
        logger.warning("Unable to sync access profile seed snapshot at %s", ACCESS_PROFILES_SEED_PATH)


db.init_db()
db.ensure_default_access_profiles(
    [
        {
            "name": "Admin",
            "is_admin": True,
            "is_sandbox": False,
            "allowed_plants": "ALL",
            "default_plants": "ALL",
        },
        {
            "name": "Chris",
            "is_admin": False,
            "is_sandbox": False,
            "allowed_plants": "ALL",
            "default_plants": "OR",
        },
        {
            "name": "Basil",
            "is_admin": False,
            "is_sandbox": False,
            "allowed_plants": "ALL",
            "default_plants": "IA",
        },
        {
            "name": "Mario",
            "is_admin": False,
            "is_sandbox": False,
            "allowed_plants": "ALL",
            "default_plants": "TX",
        },
        {
            "name": "Ed",
            "is_admin": False,
            "is_sandbox": False,
            "allowed_plants": "ALL",
            "default_plants": "GA,VA,NV",
        },
        {
            "name": "Kissaryn",
            "is_admin": False,
            "is_sandbox": False,
            "allowed_plants": "ALL",
            "default_plants": "GA,VA,NV",
        },
        {
            "name": "Judy",
            "is_admin": False,
            "is_sandbox": False,
            "allowed_plants": "ALL",
            "default_plants": "GA,VA,NV",
        },
    ]
)
_sync_access_profiles_seed_snapshot()
db.ensure_default_planning_settings(
    {
        "strategic_customers": json.dumps(
            [
                {
                    "label": "Lowe's",
                    "patterns": ["LOWE'S", "LOWES"],
                    "include_in_optimizer_workbench": True,
                },
                {
                    "label": "Tractor Supply",
                    "patterns": ["TRACTOR SUPPLY", "TRACTORSUPPLY"],
                    "wedge_min_item_length_ft": 20,
                    "include_in_optimizer_workbench": True,
                },
                {
                    "label": "Ace",
                    "patterns": ["ACE HARDWARE CORPORATION", "ACE HARDWARE"],
                    "include_in_optimizer_workbench": True,
                },
                {
                    "label": "TrailersPlus",
                    "patterns": ["TRAILERSPLUS", "TRAILER'S PLUS"],
                    "include_in_optimizer_workbench": True,
                },
                {
                    "label": "COT Sample",
                    "patterns": ["COT SAMPLE", "COTSAMPLE"],
                    "ignore_for_optimization": True,
                    "include_in_optimizer_workbench": False,
                },
                {
                    "label": "Carolina Equipment",
                    "patterns": ["CAROLINA EQUIPMENT", "CAROLINAEQUIPMENT"],
                    "ignore_for_optimization": True,
                    "include_in_optimizer_workbench": False,
                },
            ],
            separators=(",", ":"),
        ),
        STOP_FEE_SETTING_KEY: DEFAULT_STOP_FEE,
        MIN_LOAD_COST_SETTING_KEY: DEFAULT_MIN_LOAD_COST,
        FUEL_SURCHARGE_SETTING_KEY: DEFAULT_FUEL_SURCHARGE_PER_MILE,
        OPTIMIZER_DEFAULTS_SETTING_KEY: json.dumps(
            {
                "trailer_type": load_builder.DEFAULT_BUILD_PARAMS.get("trailer_type", "STEP_DECK"),
                "capacity_feet": float(load_builder.DEFAULT_BUILD_PARAMS.get("capacity_feet", 53) or 53),
                "time_window_days": int(load_builder.DEFAULT_BUILD_PARAMS.get("time_window_days", 7) or 7),
                "geo_radius": float(load_builder.DEFAULT_BUILD_PARAMS.get("geo_radius", 100) or 100),
                "max_detour_pct": float(load_builder.DEFAULT_BUILD_PARAMS.get("max_detour_pct", 15) or 15),
                "stack_overflow_max_height": int(
                    load_builder.DEFAULT_BUILD_PARAMS.get(
                        "stack_overflow_max_height",
                        DEFAULT_STACK_OVERFLOW_MAX_HEIGHT,
                    )
                    or DEFAULT_STACK_OVERFLOW_MAX_HEIGHT
                ),
                "max_back_overhang_ft": float(
                    load_builder.DEFAULT_BUILD_PARAMS.get(
                        "max_back_overhang_ft",
                        DEFAULT_MAX_BACK_OVERHANG_FT,
                    )
                    or DEFAULT_MAX_BACK_OVERHANG_FT
                ),
                "upper_two_across_max_length_ft": float(
                    load_builder.DEFAULT_BUILD_PARAMS.get(
                        "upper_two_across_max_length_ft",
                        DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT,
                    )
                    or DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT
                ),
                "upper_deck_exception_max_length_ft": float(
                    load_builder.DEFAULT_BUILD_PARAMS.get(
                        "upper_deck_exception_max_length_ft",
                        DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
                    )
                    or DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT
                ),
                "upper_deck_exception_overhang_allowance_ft": float(
                    load_builder.DEFAULT_BUILD_PARAMS.get(
                        "upper_deck_exception_overhang_allowance_ft",
                        DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
                    )
                    or DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT
                ),
                "upper_deck_exception_categories": stack_calculator.normalize_upper_deck_exception_categories(
                    load_builder.DEFAULT_BUILD_PARAMS.get("upper_deck_exception_categories"),
                    default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
                ),
                "equal_length_deck_length_order_enabled": DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED,
            }
        ),
        REPLAY_EVAL_PRESET_SETTING_KEY: json.dumps(replay_evaluator.DEFAULT_REPLAY_PRESET),
        UTILIZATION_GRADE_THRESHOLDS_SETTING_KEY: json.dumps(DEFAULT_UTILIZATION_GRADE_THRESHOLDS),
        STOP_COLOR_PALETTE_SETTING_KEY: json.dumps(DEFAULT_STOP_COLOR_PALETTE),
        TRAILER_ASSIGNMENT_RULES_SETTING_KEY: json.dumps(DEFAULT_TRAILER_ASSIGNMENT_RULES),
        RATE_TABLE_CONTEXTS_SETTING_KEY: json.dumps(DEFAULT_RATE_TABLE_CONTEXTS),
        RYDER_DEDICATED_RATE_TABLE_SETTING_KEY: json.dumps(DEFAULT_RYDER_DEDICATED_RATE_TABLE),
        LST_RATE_TABLE_SETTING_KEY: json.dumps(DEFAULT_LST_RATE_TABLE_SETTINGS),
        ALTERNATE_TRAILER_RATES_SETTING_KEY: json.dumps(DEFAULT_ALTERNATE_TRAILER_RATES),
        DEFAULT_RATE_CHANGE_METADATA_SETTING_KEY: json.dumps({}),
    }
)


def _backfill_legacy_sessions():
    orphan_loads = db.list_loads_without_session()
    if not orphan_loads:
        return
    grouped = {}
    for load in orphan_loads:
        plant = load.get("origin_plant") or "UNKNOWN"
        grouped.setdefault(plant, []).append(load)

    for plant, loads in grouped.items():
        session_code = f"LEGACY_{plant}"
        existing = db.get_planning_session_by_code(session_code)
        if existing:
            session_id = existing.get("id")
        else:
            created_at = min(
                (entry.get("created_at") for entry in loads if entry.get("created_at")), default=None
            )
            session_id = db.create_planning_session(
                session_code=session_code,
                plant_code=plant,
                created_by="Legacy",
                config_json=json.dumps({}),
                horizon_end=None,
                status="ARCHIVED",
                created_at=created_at,
            )
        load_ids = [entry.get("id") for entry in loads if entry.get("id")]
        db.assign_loads_to_session(session_id, load_ids)


_backfill_legacy_sessions()


@app.route("/session", methods=["GET", "POST"])
def session_setup():
    # Kept for backwards-compatibility with older links/bookmarks.
    # Access is now handled via /login.
    next_url = request.values.get("next") or request.args.get("next")
    return redirect(url_for("login", next=next_url))


@app.route("/session/reset")
def session_reset():
    session.clear()
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    next_url = _safe_next_url(request.values.get("next")) or ""
    profiles = db.list_access_profiles()
    login_profiles = []
    for profile in profiles:
        is_admin = bool(profile.get("is_admin"))
        is_sandbox = bool(profile.get("is_sandbox"))
        role_label = "Administrator Account" if is_admin else "Planner Account"
        if is_sandbox and not is_admin:
            role_label = "Planner Sandbox Account"
        login_profiles.append(
            {
                "id": profile.get("id"),
                "name": (profile.get("name") or "Unnamed").strip() or "Unnamed",
                "is_admin": is_admin,
                "is_sandbox": is_sandbox,
                "role_label": role_label,
                "focus_plants": _profile_focus_plants(profile),
            }
        )
    selected_profile_id = None
    error = None

    if request.method == "POST":
        profile_id = request.form.get("profile_id")
        try:
            profile_id = int(profile_id)
        except (TypeError, ValueError):
            profile_id = None
        selected_profile_id = profile_id

        profile = db.get_access_profile(profile_id) if profile_id else None
        if not profile:
            error = "Select a valid account."
        elif profile.get("is_admin"):
            password = request.form.get("password") or ""
            expected = (os.environ.get("ADMIN_PASSWORD") or "").strip()
            if not expected:
                if _is_local_dev_mode():
                    expected = "admin"
                elif password == "admin":
                    # Local fallback: allow one known default when env config is missing.
                    expected = "admin"
                else:
                    error = "Admin password is not configured."
            if not error and password != expected:
                error = "Invalid admin password."

        if not error and profile:
            _apply_profile_to_session(profile, reset_filters=True)
            return redirect(_safe_next_url_for_profile(profile, next_url))

    if selected_profile_id is None and login_profiles:
        selected_profile_id = login_profiles[0]["id"]
    selected_profile = next(
        (profile for profile in login_profiles if profile["id"] == selected_profile_id),
        None,
    )

    return render_template(
        "login.html",
        profiles=login_profiles,
        selected_profile=selected_profile,
        selected_profile_id=selected_profile_id,
        plant_names=PLANT_NAMES,
        error=error,
        next_url=next_url,
    )


def _safe_next_url(value):
    value = (value or "").strip()
    if not value:
        return None
    if value.startswith("/"):
        return value
    return None


def _is_admin_only_path(path_value):
    path = str(path_value or "").strip()
    if not path:
        return False
    return any(
        path == prefix or path.startswith(f"{prefix}/")
        for prefix in ADMIN_ONLY_PATH_PREFIXES
    )


def _safe_next_url_for_profile(profile, requested_next):
    next_url = _safe_next_url(requested_next)
    if not next_url:
        return url_for("orders")
    if not bool((profile or {}).get("is_admin")) and _is_admin_only_path(next_url):
        return url_for("orders")
    return next_url


def _json_session_expired_response():
    next_url = request.full_path if request else ""
    return jsonify(
        {
            "error": "Session expired",
            "redirect_url": url_for("login", next=next_url),
        }
    ), 401


def _normalize_order_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _line_signature(lines):
    signature = []
    for line in lines or []:
        item = _normalize_order_value(line.get("item"))
        sku = _normalize_order_value(line.get("sku"))
        try:
            qty_value = int(float(line.get("qty") or 0))
        except (TypeError, ValueError):
            qty_value = 0
        signature.append((item, sku, str(qty_value)))
    signature.sort()
    return signature


def _diff_order(existing, incoming, existing_lines=None, incoming_lines=None):
    fields = [
        "cust_name",
        "address1",
        "address2",
        "city",
        "state",
        "zip",
        "line_count",
    ]
    changes = {}
    for field in fields:
        old = _normalize_order_value(existing.get(field))
        new = _normalize_order_value(incoming.get(field))
        if old != new:
            changes[field] = {"from": old, "to": new}
    if _line_signature(existing_lines) != _line_signature(incoming_lines):
        changes["line_items"] = {"from": "changed", "to": "changed"}
    return changes


_DIMENSION_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)")


def _category_from_bin(bin_code):
    if not bin_code:
        return None
    normalized = str(bin_code).strip().upper()
    return normalized or None


def _extract_dimensions(value):
    if not value:
        return None
    match = _DIMENSION_RE.search(str(value))
    if not match:
        return None
    try:
        first = float(match.group(1))
        second = float(match.group(2))
    except (TypeError, ValueError):
        return None
    return (first, second)


def _suggest_sku_for_item(item, desc, specs, category_hint=None):
    item_dim = _extract_dimensions(item) or _extract_dimensions(desc)
    if not item_dim:
        return None
    if category_hint:
        candidates = [
            spec for spec in specs if (spec.get("category") or "").upper() == category_hint
        ]
    else:
        candidates = specs
    if not candidates:
        return None
    best = None
    best_score = -1e9
    for spec in candidates:
        spec_dim = spec.get("_dim")
        if not spec_dim:
            continue
        diff = abs(spec_dim[0] - item_dim[0]) + abs(spec_dim[1] - item_dim[1])
        score = 100 - (diff * 10)
        if spec_dim == item_dim:
            score += 40
        if spec.get("category") and spec.get("category") != "UNKNOWN":
            score += 2
        if score > best_score:
            best = spec
            best_score = score
    return best


def _build_unmapped_suggestions(unmapped_items):
    if not unmapped_items:
        return []
    specs = db.list_sku_specs()
    prepared_specs = []
    for spec in specs:
        spec_copy = dict(spec)
        spec_copy["_dim"] = _extract_dimensions(spec.get("sku")) or _extract_dimensions(
            spec.get("description") or spec.get("notes")
        )
        prepared_specs.append(spec_copy)

    grouped = {}
    for entry in unmapped_items:
        item = (entry.get("item") or "").strip().upper()
        if not item:
            continue
        if item not in grouped:
            grouped[item] = {
                "item": item,
                "desc": (entry.get("desc") or "").strip(),
                "bin": (entry.get("bin") or "").strip(),
                "bin_counts": {},
                "count": 0,
            }
        grouped[item]["count"] += 1
        bin_code = (entry.get("bin") or "").strip()
        if bin_code:
            grouped[item]["bin_counts"][bin_code] = (
                grouped[item]["bin_counts"].get(bin_code, 0) + 1
            )

    suggestions = []
    for payload in grouped.values():
        if payload.get("bin_counts"):
            payload["bin"] = max(
                payload["bin_counts"].items(), key=lambda entry: entry[1]
            )[0]
        category_hint = _category_from_bin(payload.get("bin"))
        payload["category_hint"] = category_hint or ""
        suggestion = _suggest_sku_for_item(
            payload["item"],
            payload.get("desc"),
            prepared_specs,
            category_hint=category_hint,
        )
        if suggestion:
            payload["suggested"] = {
                "sku": suggestion.get("sku"),
                "description": suggestion.get("description") or suggestion.get("notes") or "",
                "category": category_hint or suggestion.get("category") or "UNKNOWN",
                "length_with_tongue_ft": suggestion.get("length_with_tongue_ft") or 0,
                "max_stack_step_deck": suggestion.get("max_stack_step_deck") or 1,
                "max_stack_flat_bed": suggestion.get("max_stack_flat_bed") or 1,
            }
        suggestions.append(payload)

    suggestions.sort(key=lambda entry: entry.get("item") or "")
    return suggestions


def _handle_order_upload(file):
    importer = OrderImporter()
    summary = importer.parse_csv(file)
    unmapped_items = summary.get("unmapped_items") or []
    if unmapped_items:
        raise UploadValidationError(
            (
                "Upload blocked: some SKUs are unmapped or missing required dimensions/stack limits. "
                "Add SKU specs, then re-upload."
            ),
            summary=summary,
        )

    orders = summary.get("orders") or []
    so_nums = []
    for order in orders:
        so_num = (order.get("so_num") or "").strip()
        if so_num:
            so_nums.append(so_num)

    existing_map = {}
    if so_nums:
        for entry in db.list_orders_by_so_nums_any(so_nums):
            key = (entry.get("so_num") or "").strip()
            if key:
                existing_map[key] = entry

    existing_lines_map = {}
    if existing_map:
        existing_lines = db.list_order_lines_by_so_nums(list(existing_map.keys()))
        for line in existing_lines:
            key = (line.get("so_num") or "").strip()
            existing_lines_map.setdefault(key, []).append(line)

    incoming_lines_map = {}
    for line in summary.get("order_lines") or []:
        key = (line.get("so_num") or "").strip()
        if key:
            incoming_lines_map.setdefault(key, []).append(line)

    new_orders = 0
    existing_orders = 0
    changed_orders = 0
    unchanged_orders = 0
    reopened_orders = 0
    changed_keys = set()
    changes_payload = []

    seen_at = datetime.utcnow().isoformat(timespec="seconds")
    dropped = []
    for order in orders:
        so_num = (order.get("so_num") or "").strip()
        if not so_num:
            continue
        existing = existing_map.get(so_num)
        if existing:
            existing_orders += 1
            was_closed = (existing.get("status") or "").strip().upper() == "CLOSED"
            if was_closed:
                reopened_orders += 1
            else:
                diff = _diff_order(
                    existing,
                    order,
                    existing_lines=existing_lines_map.get(so_num, []),
                    incoming_lines=incoming_lines_map.get(so_num, []),
                )
                if diff:
                    changed_orders += 1
                    changed_keys.add(so_num)
                    changes_payload.append(
                        {
                            "so_num": so_num,
                            "plant": (order.get("plant") or "").strip(),
                            "changes_json": json.dumps(diff),
                        }
                    )
                else:
                    unchanged_orders += 1
        else:
            new_orders += 1
        order["status"] = "OPEN"
        order["last_seen_at"] = seen_at
        order["closed_at"] = None

    db.upsert_order_lines(summary.get("order_lines") or [])
    db.upsert_orders(orders)

    if so_nums:
        db.mark_orders_seen(so_nums, seen_at=seen_at)
        prior_open = set(db.list_open_order_so_nums())
        current_set = set(so_nums)
        dropped = sorted(prior_open - current_set)
        if dropped:
            db.mark_orders_closed(dropped, closed_at=seen_at)
        db.purge_closed_orders(retention_days=30)

    upload_id = db.add_upload_history(
        {
            "filename": getattr(file, "filename", ""),
            "total_rows": summary.get("total_rows"),
            "total_orders": len(orders),
            "new_orders": new_orders,
            "duplicate_orders": existing_orders,
            "changed_orders": changed_orders,
            "unchanged_orders": unchanged_orders,
            "reopened_orders": reopened_orders,
            "dropped_orders": len(dropped) if so_nums else 0,
            "mapping_rate": summary.get("mapping_rate"),
            "unmapped_count": len(summary.get("unmapped_items") or []),
        }
    )
    if changes_payload:
        db.add_upload_order_changes(upload_id, changes_payload)
    if so_nums:
        db.update_orders_upload_meta(so_nums, upload_id, changed_keys)
    db.add_upload_unmapped_items(upload_id, summary.get("unmapped_items") or [])

    summary["upload_id"] = upload_id
    summary["duplicate_orders"] = existing_orders
    summary["new_orders"] = new_orders
    summary["changed_orders"] = changed_orders
    summary["unchanged_orders"] = unchanged_orders
    summary["reopened_orders"] = reopened_orders
    summary["dropped_orders"] = len(dropped) if so_nums else 0
    return summary


def _normalize_so_num_for_load_report_match(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def _normalize_load_report_column(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _parse_load_report_rows(file):
    filename = str(getattr(file, "filename", "") or "").strip()
    suffix = os.path.splitext(filename)[1].lower()

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        parsed = replay_evaluator.parse_report(file)
        return parsed.get("rows") or []

    stream = file
    if hasattr(file, "stream"):
        stream = file.stream
    if hasattr(stream, "seek"):
        stream.seek(0)
    raw_bytes = stream.read()
    if hasattr(stream, "seek"):
        stream.seek(0)
    if not raw_bytes:
        raise ValueError("Upload is empty.")

    text = raw_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("Load report is missing a header row.")

    aliases = {
        "load_number": {
            "load_number",
            "load_no",
            "load",
            "load_id",
        },
        "order_number": {
            "order_number",
            "order_no",
            "order",
            "so_num",
            "sonum",
            "sales_order",
            "name",
        },
    }
    field_lookup = {}
    for raw_name in reader.fieldnames:
        normalized = _normalize_load_report_column(raw_name)
        for canonical, alias_set in aliases.items():
            if canonical in field_lookup:
                continue
            if normalized in alias_set:
                field_lookup[canonical] = raw_name

    missing = [field for field in ("load_number", "order_number") if field not in field_lookup]
    if missing:
        raise ValueError(
            "Missing required columns: "
            + ", ".join(missing)
            + ". Expected Load Number and Name (SO #)."
        )

    rows = []
    for raw in reader:
        load_number = str(raw.get(field_lookup["load_number"]) or "").strip().strip('"')
        order_number = _normalize_so_num_for_load_report_match(raw.get(field_lookup["order_number"]))
        if not load_number or not order_number:
            continue
        rows.append(
            {
                "load_number": load_number,
                "order_number": order_number,
            }
        )
    return rows


def _handle_load_report_upload(file):
    report_rows = _parse_load_report_rows(file)
    if not report_rows:
        raise ValueError("Load report has no valid rows.")

    assignments_by_so = {}
    duplicate_rows = 0
    conflicting_orders = 0
    for row in report_rows:
        so_num = _normalize_so_num_for_load_report_match(row.get("order_number"))
        load_number = str(row.get("load_number") or "").strip()
        if not so_num or not load_number:
            continue
        existing_load = assignments_by_so.get(so_num)
        if existing_load:
            duplicate_rows += 1
            if existing_load != load_number:
                conflicting_orders += 1
            continue
        assignments_by_so[so_num] = load_number

    if not assignments_by_so:
        raise ValueError("Load report has no usable SO # and Load Number pairs.")

    assignments = [
        {"so_num": so_num, "load_number": load_number}
        for so_num, load_number in assignments_by_so.items()
    ]
    open_orders = db.list_orders_by_so_nums_any(list(assignments_by_so.keys()), include_closed=False)
    open_so_nums = {
        _normalize_so_num_for_load_report_match(row.get("so_num"))
        for row in open_orders
        if _normalize_so_num_for_load_report_match(row.get("so_num"))
    }
    matched_open_orders = sum(1 for so_num in assignments_by_so if so_num in open_so_nums)

    summary = {
        "filename": getattr(file, "filename", ""),
        "total_rows": len(report_rows),
        "valid_rows": len(report_rows),
        "unique_orders": len(assignments_by_so),
        "unique_loads": len({value for value in assignments_by_so.values() if value}),
        "duplicate_rows": duplicate_rows,
        "conflicting_orders": conflicting_orders,
        "matched_open_orders": matched_open_orders,
        "unmatched_open_orders": max(len(assignments_by_so) - matched_open_orders, 0),
    }
    upload_id = db.add_load_report_upload(summary)
    db.replace_latest_load_report_assignments(upload_id, assignments)
    summary["upload_id"] = upload_id
    return summary


@app.route("/access/switch", methods=["POST"])
def access_switch():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    profile_id = request.form.get("profile_id")
    try:
        profile_id = int(profile_id)
    except (TypeError, ValueError):
        profile_id = None

    profile = db.get_access_profile(profile_id) if profile_id else None
    if not profile:
        return redirect(url_for("orders"))

    _apply_profile_to_session(profile, reset_filters=True)
    next_url = _safe_next_url_for_profile(profile, request.form.get("next"))
    return redirect(next_url)


@app.route("/access/manage", methods=["GET", "POST"])
def access_manage():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    if _get_session_role() != ROLE_ADMIN:
        return redirect(url_for("orders"))

    error = None
    edit_profile = None
    edit_allowed = []
    edit_defaults = []
    edit_id = request.args.get("edit_id")
    if edit_id:
        try:
            edit_id = int(edit_id)
        except (TypeError, ValueError):
            edit_id = None
    if edit_id:
        edit_profile = db.get_access_profile(edit_id)
        if edit_profile:
            allowed_parsed = _parse_plant_filters(edit_profile.get("allowed_plants"))
            edit_allowed = allowed_parsed if allowed_parsed else list(PLANT_CODES)
            defaults_parsed = _parse_plant_filters(edit_profile.get("default_plants"))
            # Empty means "no focus (all plants)"
            edit_defaults = defaults_parsed if defaults_parsed else []

    if request.method == "POST":
        action = (request.form.get("action") or "").strip().lower() or "create"
        profile_id = request.form.get("profile_id")
        try:
            profile_id = int(profile_id) if profile_id else None
        except (TypeError, ValueError):
            profile_id = None

        name = (request.form.get("name") or "").strip()
        is_admin_flag = (request.form.get("is_admin") or "").strip().lower() in {
            "1",
            "true",
            "on",
            "yes",
        }
        is_sandbox_flag = (request.form.get("is_sandbox") or "").strip().lower() in {
            "1",
            "true",
            "on",
            "yes",
        }
        allowed = [_normalize_plant_code(code) for code in request.form.getlist("allowed_plants")]
        allowed = [code for code in allowed if code in PLANT_CODES]
        default_plants = [_normalize_plant_code(code) for code in request.form.getlist("default_plants")]
        default_plants = [code for code in default_plants if code in PLANT_CODES]

        if not name:
            error = "Profile name is required."
        elif action == "update" and not profile_id:
            error = "Missing profile to update."
        elif action not in {"create", "update"}:
            error = "Unsupported action."
        elif is_admin_flag:
            try:
                if action == "create":
                    db.create_access_profile(name, True, "ALL", "ALL", is_sandbox=False)
                else:
                    db.update_access_profile(profile_id, name, True, "ALL", "ALL", is_sandbox=False)
                _sync_access_profiles_seed_snapshot()
                return redirect(url_for("access_manage"))
            except Exception:
                error = "Unable to save profile. Profile names must be unique."
        else:
            if not allowed:
                error = "Select at least one allowed plant."
            else:
                default_plants = [code for code in default_plants if code in allowed]
                allowed_csv = ",".join(allowed) if len(allowed) < len(PLANT_CODES) else "ALL"
                # Empty defaults => no focus (treat as ALL)
                default_csv = ",".join(default_plants) if default_plants else "ALL"
                try:
                    if action == "create":
                        db.create_access_profile(
                            name,
                            False,
                            allowed_csv,
                            default_csv,
                            is_sandbox=is_sandbox_flag,
                        )
                    else:
                        db.update_access_profile(
                            profile_id,
                            name,
                            False,
                            allowed_csv,
                            default_csv,
                            is_sandbox=is_sandbox_flag,
                        )
                    _sync_access_profiles_seed_snapshot()
                    return redirect(url_for("access_manage"))
                except Exception:
                    error = "Unable to save profile. Profile names must be unique."

    profiles = db.list_access_profiles()
    return render_template(
        "access_manage.html",
        profiles=profiles,
        plants=PLANT_CODES,
        plant_names=PLANT_NAMES,
        error=error,
        edit_profile=edit_profile,
        edit_allowed=edit_allowed,
        edit_defaults=edit_defaults,
    )


@app.route("/access/delete", methods=["POST"])
def access_delete():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()

    profile_id = request.form.get("profile_id")
    try:
        profile_id = int(profile_id)
    except (TypeError, ValueError):
        return redirect(url_for("access_manage"))

    profiles = db.list_access_profiles()
    if len(profiles) <= 1:
        return redirect(url_for("access_manage"))

    admin_count = sum(1 for profile in profiles if profile.get("is_admin"))
    target = next((p for p in profiles if p.get("id") == profile_id), None)
    if not target:
        return redirect(url_for("access_manage"))

    if target.get("is_admin") and admin_count <= 1:
        return redirect(url_for("access_manage"))

    # If deleting the active session profile, switch to Admin first.
    if session.get(SESSION_PROFILE_ID_KEY) == profile_id:
        fallback = db.get_access_profile_by_name("Admin")
        if fallback and fallback.get("id") != profile_id:
            _apply_profile_to_session(fallback, reset_filters=True)
        else:
            other = next((p for p in profiles if p.get("id") != profile_id), None)
            if other:
                _apply_profile_to_session(other, reset_filters=True)

    db.delete_access_profile(profile_id)
    _sync_access_profiles_seed_snapshot()
    return redirect(url_for("access_manage"))


def _default_optimize_form(plant_code=None):
    form_data = dict(load_builder.DEFAULT_BUILD_PARAMS)
    optimizer_defaults = _get_optimizer_default_settings()
    form_data["trailer_type"] = optimizer_defaults["trailer_type"]
    form_data["capacity_feet"] = str(optimizer_defaults["capacity_feet"])
    form_data["max_detour_pct"] = str(optimizer_defaults["max_detour_pct"])
    form_data["time_window_days"] = str(optimizer_defaults["time_window_days"])
    form_data["geo_radius"] = str(optimizer_defaults["geo_radius"])
    form_data["stack_overflow_max_height"] = str(
        optimizer_defaults["stack_overflow_max_height"]
    )
    form_data["max_back_overhang_ft"] = str(
        optimizer_defaults["max_back_overhang_ft"]
    )
    form_data["upper_two_across_max_length_ft"] = str(
        optimizer_defaults.get("upper_two_across_max_length_ft", DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT)
    )
    form_data["upper_deck_exception_max_length_ft"] = str(
        optimizer_defaults.get(
            "upper_deck_exception_max_length_ft",
            DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
        )
    )
    form_data["upper_deck_exception_overhang_allowance_ft"] = str(
        optimizer_defaults.get(
            "upper_deck_exception_overhang_allowance_ft",
            DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
        )
    )
    form_data["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
        optimizer_defaults.get("upper_deck_exception_categories"),
        default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
    )
    form_data["order_category_scope"] = order_categories.normalize_order_category_scope(
        form_data.get("order_category_scope"),
        default=order_categories.ORDER_CATEGORY_SCOPE_ALL,
    )
    form_data["ignore_due_date"] = _coerce_bool_value(form_data.get("ignore_due_date"))
    if not form_data.get("orders_start_date"):
        form_data["orders_start_date"] = date.today().strftime("%Y-%m-%d")
    resolved_plant = _normalize_plant_code(plant_code) or _normalize_plant_code(form_data.get("origin_plant"))
    if not resolved_plant and PLANT_CODES:
        resolved_plant = PLANT_CODES[0]
    if resolved_plant:
        plant_defaults = _get_optimizer_settings_for_plant(
            resolved_plant,
            optimizer_defaults=optimizer_defaults,
        )
        form_data["origin_plant"] = resolved_plant
        form_data["trailer_type"] = plant_defaults["trailer_type"]
        form_data["capacity_feet"] = str(plant_defaults["capacity_feet"])
    return form_data


def _distinct(values):
    return sorted({value for value in values if value})


_HIDDEN_RATE_MATRIX_ROWS = {"DETENTION", "MIN", "NY ZIP 100", "TONU", "STOP"}


def _is_rate_matrix_display_state(state):
    normalized = (state or "").strip().upper()
    return bool(normalized) and normalized not in _HIDDEN_RATE_MATRIX_ROWS


def _order_rate_matrix_plants(plants):
    normalized = [str(plant or "").strip().upper() for plant in plants if str(plant or "").strip()]
    unique = list(dict.fromkeys(normalized))
    preferred = [plant for plant in RATE_MATRIX_PLANT_DISPLAY_ORDER if plant in unique]
    remaining = sorted([plant for plant in unique if plant not in RATE_MATRIX_PLANT_DISPLAY_ORDER])
    return preferred + remaining


def _context_plants():
    return list(RATE_MATRIX_PLANT_DISPLAY_ORDER)


def _build_rate_matrix(rates):
    plants = _order_rate_matrix_plants(
        {rate["origin_plant"] for rate in rates if rate.get("origin_plant")}
    )
    states = sorted(
        {
            rate["destination_state"]
            for rate in rates
            if _is_rate_matrix_display_state(rate.get("destination_state"))
        }
    )
    matrix = {state: {plant: None for plant in plants} for state in states}
    for rate in rates:
        state = rate.get("destination_state")
        plant = rate.get("origin_plant")
        if state and plant and state in matrix and matrix[state][plant] is None:
            matrix[state][plant] = rate.get("rate_per_mile")
    return plants, states, matrix


def _build_rate_matrix_records(rates):
    plants = _order_rate_matrix_plants(
        {rate["origin_plant"] for rate in rates if rate.get("origin_plant")}
    )
    states = sorted(
        {
            rate["destination_state"]
            for rate in rates
            if _is_rate_matrix_display_state(rate.get("destination_state"))
        }
    )
    matrix = {state: {plant: None for plant in plants} for state in states}
    for rate in rates:
        state = rate.get("destination_state")
        plant = rate.get("origin_plant")
        if not state or not plant or state not in matrix:
            continue
        current = matrix[state][plant]
        if not current or (rate.get("effective_year") or 0) > (current.get("effective_year") or 0):
            matrix[state][plant] = rate
    return plants, states, matrix


def _serialize_rate_record(record):
    if not isinstance(record, dict):
        return None
    rate_per_mile = _coerce_optional_non_negative_float(record.get("rate_per_mile"))
    if rate_per_mile is None:
        return None
    return {
        "id": record.get("id"),
        "rate_per_mile": round(rate_per_mile, 2),
        "effective_year": int(record.get("effective_year") or datetime.now().year),
    }


def _build_serialized_rate_matrix(plants, states, matrix_records):
    serialized = {}
    for state in states:
        row = {}
        row_records = matrix_records.get(state) or {}
        for plant in plants:
            row[plant] = _serialize_rate_record(row_records.get(plant))
        serialized[state] = row
    return serialized


def _coerce_non_negative_float(value, default):
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        parsed = float(default)
    if parsed < 0:
        return 0.0
    return parsed


def _coerce_non_negative_int(value, default):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = int(default)
    return max(parsed, 0)


def _coerce_utilization_grade_thresholds(raw_value):
    defaults = dict(DEFAULT_UTILIZATION_GRADE_THRESHOLDS)
    if not isinstance(raw_value, dict):
        return defaults
    a = _coerce_non_negative_int(raw_value.get("A"), defaults["A"])
    b = _coerce_non_negative_int(raw_value.get("B"), defaults["B"])
    c = _coerce_non_negative_int(raw_value.get("C"), defaults["C"])
    d = _coerce_non_negative_int(raw_value.get("D"), defaults["D"])

    a = min(a, 100)
    b = min(b, 99)
    c = min(c, 99)
    d = min(d, 99)

    if b >= a:
        b = max(a - 1, 0)
    if c >= b:
        c = max(b - 1, 0)
    if d >= c:
        d = max(c - 1, 0)
    return {"A": a, "B": b, "C": c, "D": d}


def _get_planning_float_setting(setting_key, default_value):
    setting = _get_effective_planning_setting(setting_key)
    return round(_coerce_non_negative_float(setting.get("value_text"), default_value), 2)


def _get_stop_fee_amount():
    return _get_planning_float_setting(STOP_FEE_SETTING_KEY, DEFAULT_STOP_FEE)


def _get_load_minimum_amount():
    return _get_planning_float_setting(MIN_LOAD_COST_SETTING_KEY, DEFAULT_MIN_LOAD_COST)


def _get_fuel_surcharge_per_mile():
    return _get_planning_float_setting(
        FUEL_SURCHARGE_SETTING_KEY,
        DEFAULT_FUEL_SURCHARGE_PER_MILE,
    )


def _get_rates_overview_metrics():
    return {
        "stop_fee": _get_stop_fee_amount(),
        "load_minimum": _get_load_minimum_amount(),
        "fuel_surcharge": _get_fuel_surcharge_per_mile(),
    }


def _normalize_rate_table_key(value, default="FLS"):
    key = str(value or "").strip().upper()
    valid_keys = {option["key"] for option in RATE_TABLE_KEY_OPTIONS}
    if key in valid_keys:
        return key
    return default if default in valid_keys else "FLS"


def _get_rate_table_contexts():
    defaults = dict(DEFAULT_RATE_TABLE_CONTEXTS)
    setting = _get_effective_planning_setting(RATE_TABLE_CONTEXTS_SETTING_KEY)
    raw_text = (setting.get("value_text") or "").strip()
    parsed = None
    if raw_text:
        try:
            parsed = json.loads(raw_text)
        except json.JSONDecodeError:
            parsed = None
    if isinstance(parsed, dict):
        defaults["default_rate_table_key"] = _normalize_rate_table_key(
            parsed.get("default_rate_table_key"),
            defaults["default_rate_table_key"],
        )
        defaults["carrier_dedicated_ryder_rate_table_key"] = _normalize_rate_table_key(
            parsed.get("carrier_dedicated_ryder_rate_table_key"),
            defaults["carrier_dedicated_ryder_rate_table_key"],
        )
        defaults["trailer_hotshot_rate_table_key"] = _normalize_rate_table_key(
            parsed.get("trailer_hotshot_rate_table_key"),
            defaults["trailer_hotshot_rate_table_key"],
        )
    return defaults


def _get_json_planning_setting(setting_key):
    setting = _get_effective_planning_setting(setting_key)
    raw_text = (setting.get("value_text") or "").strip()
    if not raw_text:
        return None
    try:
        return json.loads(raw_text)
    except json.JSONDecodeError:
        return None


def _coerce_optional_non_negative_float(value):
    if value in (None, ""):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if parsed < 0:
        return 0.0
    return parsed


def _normalize_alternate_trailer_code(value):
    text = str(value or "").strip().upper().replace("-", " ").replace("_", " ")
    if not text:
        return ""
    if "STEP" in text and "DECK" in text:
        return "STEP_DECK"
    if "HOT" in text:
        return "HOT_SHOT"
    if "WEDGE" in text:
        return "WEDGE"
    if "FLAT" in text:
        return "FLAT_BED"
    return ""


def _alternate_trailer_label_for_code(code):
    normalized = _normalize_alternate_trailer_code(code)
    for definition in ALTERNATE_TRAILER_SECTION_DEFINITIONS:
        if definition["code"] == normalized:
            return definition["label"]
    return normalized.replace("_", " ").title()


def _copy_default_alternate_trailer_rates():
    return json.loads(json.dumps(DEFAULT_ALTERNATE_TRAILER_RATES))


def _coerce_alternate_apply_fuel_surcharge(value, *, default=True):
    if isinstance(value, dict):
        if "apply_fuel_surcharge" in value:
            return bool(_coerce_bool_value(value.get("apply_fuel_surcharge")))
        if "fuel_surcharge_per_mile" in value:
            legacy = _coerce_optional_non_negative_float(value.get("fuel_surcharge_per_mile"))
            if legacy is not None:
                return legacy > 0
        return bool(default)
    if value is None:
        return bool(default)
    return bool(_coerce_bool_value(value))


def _default_alternate_trailer_placeholder():
    return {
        "per_stop": 0.0,
        "load_minimum": 0.0,
        "apply_fuel_surcharge": True,
        "requires_return_miles": True,
    }


def _get_ryder_dedicated_rate_table():
    defaults = json.loads(json.dumps(DEFAULT_RYDER_DEDICATED_RATE_TABLE))
    payload = _get_json_planning_setting(RYDER_DEDICATED_RATE_TABLE_SETTING_KEY)
    plants = _context_plants()
    rates_by_plant = {plant: defaults.get("rates_by_plant", {}).get(plant) for plant in plants}
    if not isinstance(payload, dict):
        defaults["plants"] = plants
        defaults["rates_by_plant"] = rates_by_plant
        return defaults

    raw_rates = payload.get("rates_by_plant")
    if isinstance(raw_rates, dict):
        for plant in plants:
            if plant not in raw_rates:
                continue
            parsed = _coerce_optional_non_negative_float(raw_rates.get(plant))
            rates_by_plant[plant] = round(parsed, 2) if parsed is not None else None

    # Backward compatibility: previously saved as row objects.
    payload_rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    for row in payload_rows:
        if not isinstance(row, dict):
            continue
        plant = str(row.get("plant") or "").strip().upper()
        if plant not in plants:
            continue
        parsed_rate = _coerce_optional_non_negative_float(row.get("rate_per_mile"))
        if parsed_rate is not None:
            rates_by_plant[plant] = round(parsed_rate, 2)

    fuel = _coerce_optional_non_negative_float(payload.get("fuel_surcharge", defaults.get("fuel_surcharge")))
    stop = _coerce_optional_non_negative_float(payload.get("per_stop", defaults.get("per_stop")))
    minimum = _coerce_optional_non_negative_float(payload.get("load_minimum", defaults.get("load_minimum")))

    defaults["name"] = str(payload.get("name") or defaults.get("name") or "").strip() or "Ryder Dedicated"
    defaults["applies_to"] = str(payload.get("applies_to") or defaults.get("applies_to") or "").strip().upper() or "ALL_STATES"
    defaults["notes"] = str(payload.get("notes") or defaults.get("notes") or "").strip()
    defaults["plants"] = plants
    defaults["rates_by_plant"] = rates_by_plant
    defaults["fuel_surcharge"] = round(fuel, 2) if fuel is not None else 0.0
    defaults["per_stop"] = round(stop, 2) if stop is not None else 0.0
    defaults["load_minimum"] = round(minimum, 2) if minimum is not None else 0.0
    return defaults


def _get_default_rate_change_metadata():
    payload = _get_json_planning_setting(DEFAULT_RATE_CHANGE_METADATA_SETTING_KEY)
    if not isinstance(payload, dict):
        return {"changed_index": {}, "changed_count": 0}
    changed_index = {}
    for cell in payload.get("changed_cells") or []:
        if not isinstance(cell, dict):
            continue
        origin = (cell.get("origin_plant") or "").strip().upper()
        destination = (cell.get("destination_state") or "").strip().upper()
        if len(destination) != 2 or not origin:
            continue
        changed_index[f"{origin}|{destination}"] = cell
    payload["changed_index"] = changed_index
    payload["changed_count"] = len(changed_index)
    return payload


def _get_lst_rate_matrix():
    defaults = json.loads(json.dumps(DEFAULT_LST_RATE_TABLE_SETTINGS))
    payload = _get_json_planning_setting(LST_RATE_TABLE_SETTING_KEY)
    if not isinstance(payload, dict):
        return defaults
    plants = _context_plants()
    states = [str(value).strip().upper() for value in payload.get("states") or [] if str(value or "").strip()]
    matrix_payload = payload.get("matrix")
    matrix = {}
    if isinstance(matrix_payload, dict):
        for state, row in matrix_payload.items():
            normalized_state = str(state or "").strip().upper()
            if not normalized_state:
                continue
            normalized_row = {}
            if isinstance(row, dict):
                for plant, rate in row.items():
                    normalized_plant = str(plant or "").strip().upper()
                    if normalized_plant not in plants:
                        continue
                    try:
                        normalized_row[normalized_plant] = float(rate) if rate not in (None, "") else None
                    except (TypeError, ValueError):
                        normalized_row[normalized_plant] = None
            matrix[normalized_state] = normalized_row
    if not states:
        states = sorted(matrix.keys())
    fuel = _coerce_optional_non_negative_float(payload.get("fuel_surcharge", defaults.get("fuel_surcharge")))
    stop = _coerce_optional_non_negative_float(payload.get("per_stop", defaults.get("per_stop")))
    minimum = _coerce_optional_non_negative_float(payload.get("load_minimum", defaults.get("load_minimum")))
    defaults["carrier"] = str(payload.get("carrier") or defaults.get("carrier") or "").strip() or "LST"
    payload["plants"] = plants
    payload["states"] = states
    payload["matrix"] = matrix
    payload["fuel_surcharge"] = round(fuel, 2) if fuel is not None else 0.0
    payload["per_stop"] = round(stop, 2) if stop is not None else 0.0
    payload["load_minimum"] = round(minimum, 2) if minimum is not None else 0.0
    payload["carrier"] = defaults["carrier"]
    payload["state_lane_labels"] = payload.get("state_lane_labels") if isinstance(payload.get("state_lane_labels"), dict) else {}
    return payload


def _get_alternate_trailer_rates():
    defaults = _copy_default_alternate_trailer_rates()
    payload = _get_json_planning_setting(ALTERNATE_TRAILER_RATES_SETTING_KEY)
    if not isinstance(payload, dict):
        return defaults
    plants = _context_plants()
    sections_by_code = {
        section["code"]: section for section in defaults.get("sections") or [] if section.get("code")
    }
    for section in payload.get("sections") or []:
        if not isinstance(section, dict):
            continue
        code = _normalize_alternate_trailer_code(section.get("code") or section.get("trailer_type"))
        if not code or code not in sections_by_code:
            continue
        normalized = sections_by_code[code]
        normalized["label"] = _alternate_trailer_label_for_code(code)
        normalized["plants"] = plants
        rates = normalized.get("rates_by_plant") or {}
        placeholders = normalized.get("placeholders_by_plant") or {}

        raw_rates = section.get("rates_by_plant")
        if isinstance(raw_rates, dict):
            for plant in plants:
                if plant not in raw_rates:
                    continue
                parsed = _coerce_optional_non_negative_float(raw_rates.get(plant))
                rates[plant] = round(parsed, 2) if parsed is not None else None

        raw_placeholders = section.get("placeholders_by_plant")
        if isinstance(raw_placeholders, dict):
            for plant in plants:
                plant_payload = raw_placeholders.get(plant)
                if not isinstance(plant_payload, dict):
                    continue
                current = placeholders.get(plant) or {}
                for key in ("per_stop", "load_minimum"):
                    if key not in plant_payload:
                        continue
                    parsed = _coerce_optional_non_negative_float(plant_payload.get(key))
                    current[key] = round(parsed, 2) if parsed is not None else None
                if "apply_fuel_surcharge" in plant_payload or "fuel_surcharge_per_mile" in plant_payload:
                    current["apply_fuel_surcharge"] = _coerce_alternate_apply_fuel_surcharge(
                        plant_payload,
                        default=True,
                    )
                if "requires_return_miles" in plant_payload:
                    current["requires_return_miles"] = _coerce_bool_value(
                        plant_payload.get("requires_return_miles")
                    )
                elif "dedicated_round_trip_miles" in plant_payload:
                    current["requires_return_miles"] = _coerce_bool_value(
                        plant_payload.get("dedicated_round_trip_miles")
                    )
                placeholders[plant] = current

        # Backward compatibility for older carrier-row payload format.
        rows = section.get("rows") if isinstance(section.get("rows"), list) else []
        for row in rows:
            if not isinstance(row, dict):
                continue
            plant = str(row.get("plant") or "").strip().upper()
            if plant not in plants:
                continue
            parsed_rate = _coerce_optional_non_negative_float(row.get("rate_per_mile"))
            if parsed_rate is not None and rates.get(plant) is None:
                rates[plant] = round(parsed_rate, 2)

        for plant in plants:
            rates.setdefault(plant, None)
            plant_placeholders = placeholders.get(plant) or {}
            for key in ("per_stop", "load_minimum"):
                plant_placeholders.setdefault(key, 0.0)
            plant_placeholders["apply_fuel_surcharge"] = _coerce_alternate_apply_fuel_surcharge(
                plant_placeholders,
                default=True,
            )
            plant_placeholders.pop("fuel_surcharge_per_mile", None)
            plant_placeholders.setdefault("requires_return_miles", True)
            placeholders[plant] = plant_placeholders
        normalized["rates_by_plant"] = rates
        normalized["placeholders_by_plant"] = placeholders

    defaults["sections"] = [sections_by_code[definition["code"]] for definition in ALTERNATE_TRAILER_SECTION_DEFINITIONS]
    return defaults


def _serialize_optional_money(value):
    parsed = _coerce_optional_non_negative_float(value)
    return round(parsed, 2) if parsed is not None else None


def _plant_color_map():
    return {
        "GA": "#22c55e",
        "TX": "#3b82f6",
        "VA": "#06b6d4",
        "IA": "#a855f7",
        "OR": "#f97316",
        "NV": "#ef4444",
    }


def _build_rates_v2_payload(
    rate_plants,
    rate_states,
    rate_matrix_serialized,
    global_rate_metrics,
    fuel_surcharge_per_mile,
    ryder_dedicated_rate_table,
    lst_rate_matrix,
    alternate_trailer_rates,
):
    plants = _context_plants()
    plant_names = {plant: PLANT_NAMES.get(plant, plant) for plant in plants}
    plant_colors = _plant_color_map()

    fls_lanes = {}
    for state in rate_states:
        row = {}
        row_source = rate_matrix_serialized.get(state) or {}
        for plant in plants:
            record = row_source.get(plant)
            if isinstance(record, dict) and record.get("rate_per_mile") is not None:
                row[plant] = {
                    "rate_per_mile": round(float(record.get("rate_per_mile")), 2),
                    "rate_id": record.get("id"),
                    "effective_year": int(record.get("effective_year") or datetime.now().year),
                }
            else:
                row[plant] = None
        fls_lanes[state] = row

    ryder_plants = ryder_dedicated_rate_table.get("plants") or plants
    ryder_rates = ryder_dedicated_rate_table.get("rates_by_plant") or {}
    ryder_lanes = {
        "ALL_STATES": {
            plant: (
                {"rate_per_mile": round(float(ryder_rates.get(plant)), 2)}
                if _coerce_optional_non_negative_float(ryder_rates.get(plant)) is not None
                else None
            )
            for plant in plants
        }
    }

    lst_states = [str(state).strip().upper() for state in (lst_rate_matrix.get("states") or []) if str(state or "").strip()]
    lst_matrix = lst_rate_matrix.get("matrix") or {}
    lst_labels = lst_rate_matrix.get("state_lane_labels") if isinstance(lst_rate_matrix.get("state_lane_labels"), dict) else {}
    lst_lanes = {}
    for state in lst_states:
        source_row = lst_matrix.get(state) if isinstance(lst_matrix.get(state), dict) else {}
        row = {}
        for plant in plants:
            parsed = _coerce_optional_non_negative_float(source_row.get(plant))
            row[plant] = {"rate_per_mile": round(parsed, 2)} if parsed is not None else None
        lst_lanes[state] = row

    trailer_sections = []
    alternate_sections = alternate_trailer_rates.get("sections") if isinstance(alternate_trailer_rates, dict) else []
    alternate_rows = []
    alternate_lanes = {}
    for section in alternate_sections or []:
        if not isinstance(section, dict):
            continue
        code = _normalize_alternate_trailer_code(section.get("code") or section.get("trailer_type"))
        if not code:
            continue
        label = section.get("label") or _alternate_trailer_label_for_code(code)
        rates_by_plant = section.get("rates_by_plant") if isinstance(section.get("rates_by_plant"), dict) else {}
        placeholders = section.get("placeholders_by_plant") if isinstance(section.get("placeholders_by_plant"), dict) else {}
        section_rates_row = {}
        for plant in plants:
            parsed_rate = _coerce_optional_non_negative_float(rates_by_plant.get(plant))
            section_rates_row[plant] = {"rate_per_mile": round(parsed_rate, 2)} if parsed_rate is not None else None
        alternate_rows.append({"key": code, "label": label})
        alternate_lanes[code] = section_rates_row

        trailer_section = {
            "code": code,
            "label": label,
            "icon": {
                "HOT_SHOT": "bolt",
                "WEDGE": "change_history",
                "FLAT_BED": "view_stream",
            }.get(code, "local_shipping"),
            "plants": list(plants),
            "rows": [
                {"field": "rate_per_mile", "label": "All States Rate/Mile", "values": {}},
                {"field": "per_stop", "label": "Per Stop Fee", "values": {}},
                {"field": "apply_fuel_surcharge", "label": "Fuel Surcharge (Use Top Box)", "values": {}},
                {"field": "load_minimum", "label": "Load Minimum", "values": {}},
                {"field": "requires_return_miles", "label": "Roundtrip Miles Required", "values": {}},
            ],
        }
        for plant in plants:
            placeholders_by_plant = placeholders.get(plant) if isinstance(placeholders.get(plant), dict) else {}
            trailer_section["rows"][0]["values"][plant] = _serialize_optional_money(rates_by_plant.get(plant))
            trailer_section["rows"][1]["values"][plant] = _serialize_optional_money(placeholders_by_plant.get("per_stop"))
            trailer_section["rows"][2]["values"][plant] = _coerce_alternate_apply_fuel_surcharge(
                placeholders_by_plant,
                default=True,
            )
            trailer_section["rows"][3]["values"][plant] = _serialize_optional_money(placeholders_by_plant.get("load_minimum"))
            trailer_section["rows"][4]["values"][plant] = bool(
                placeholders_by_plant.get("requires_return_miles", True)
            )
        trailer_sections.append(trailer_section)

    return {
        "plants": plants,
        "plant_names": plant_names,
        "plant_colors": plant_colors,
        "carrier_order": ["fls", "ryder", "lst", "alternate"],
        "carriers": {
            "fls": {
                "key": "fls",
                "label": "FLS",
                "rows": [{"key": state, "label": state} for state in rate_states],
                "lanes": fls_lanes,
                "accessorial": {
                    "per_stop": round(float(global_rate_metrics.get("stop_fee") or 0.0), 2),
                    "load_minimum": round(float(global_rate_metrics.get("load_minimum") or 0.0), 2),
                    "fuel_surcharge": round(float(fuel_surcharge_per_mile or 0.0), 2),
                },
            },
            "ryder": {
                "key": "ryder",
                "label": "Ryder Dedicated",
                "rows": [{"key": "ALL_STATES", "label": "ALL STATES"}],
                "lanes": ryder_lanes,
                "accessorial": {
                    "per_stop": _serialize_optional_money(ryder_dedicated_rate_table.get("per_stop")) or 0.0,
                    "load_minimum": _serialize_optional_money(ryder_dedicated_rate_table.get("load_minimum")) or 0.0,
                    "fuel_surcharge": _serialize_optional_money(ryder_dedicated_rate_table.get("fuel_surcharge")) or 0.0,
                },
            },
            "lst": {
                "key": "lst",
                "label": "LST",
                "rows": [
                    {"key": state, "label": lst_labels.get(state) or state}
                    for state in lst_states
                ],
                "lanes": lst_lanes,
                "accessorial": {
                    "per_stop": _serialize_optional_money(lst_rate_matrix.get("per_stop")) or 0.0,
                    "load_minimum": _serialize_optional_money(lst_rate_matrix.get("load_minimum")) or 0.0,
                    "fuel_surcharge": _serialize_optional_money(lst_rate_matrix.get("fuel_surcharge")) or 0.0,
                },
            },
            "alternate": {
                "key": "alternate",
                "label": "Alternate Trailer Types",
                "rows": alternate_rows,
                "lanes": alternate_lanes,
                "accessorial": {
                    "per_stop": 0.0,
                    "load_minimum": 0.0,
                    "fuel_surcharge": 0.0,
                },
            },
        },
        "trailer_sections": trailer_sections,
    }


def _active_planner_profile_name():
    try:
        if _get_session_role() != ROLE_PLANNER:
            return ""
        return (_get_session_profile_name() or _get_session_role() or "").strip()
    except RuntimeError:
        return ""


def _planner_setting_override_setting_key(setting_key, profile_name=None):
    profile = str(profile_name or "").strip().upper()
    key = str(setting_key or "").strip()
    if not profile or not key:
        return ""
    return f"{PLANNER_SETTING_OVERRIDE_SETTING_PREFIX}{profile}::{key}"


def _get_effective_planning_setting(setting_key, profile_name=None):
    key = str(setting_key or "").strip()
    if not key:
        return {}
    resolved_profile = str(profile_name or "").strip()
    if not resolved_profile:
        resolved_profile = _active_planner_profile_name()
    if resolved_profile:
        override_key = _planner_setting_override_setting_key(key, resolved_profile)
        if override_key:
            override_row = db.get_planning_setting(override_key) or {}
            if (override_row.get("value_text") or "").strip():
                return override_row
    return db.get_planning_setting(key) or {}


def _upsert_scoped_planning_setting(setting_key, value_text, profile_name=None):
    key = str(setting_key or "").strip()
    if not key:
        return
    resolved_profile = str(profile_name or "").strip()
    if not resolved_profile:
        resolved_profile = _active_planner_profile_name()
    if resolved_profile:
        db.upsert_planning_setting(
            _planner_setting_override_setting_key(key, resolved_profile),
            value_text,
        )
        return
    db.upsert_planning_setting(key, value_text)


def _merge_trailer_assignment_rules(base_rules, parsed):
    merged = dict(base_rules or DEFAULT_TRAILER_ASSIGNMENT_RULES)
    if not isinstance(parsed, dict):
        return merged
    if "livestock_wedge_enabled" in parsed:
        merged["livestock_wedge_enabled"] = _coerce_bool_value(
            parsed.get("livestock_wedge_enabled")
        )
    tokens = parsed.get("livestock_category_tokens")
    if isinstance(tokens, str):
        tokens = [tokens]
    if isinstance(tokens, (list, tuple, set)):
        normalized_tokens = []
        for token in tokens:
            text = str(token or "").strip().upper()
            if text and text not in normalized_tokens:
                normalized_tokens.append(text)
        if normalized_tokens:
            merged["livestock_category_tokens"] = normalized_tokens
    return merged


def _planner_trailer_rules_override_setting_key(profile_name=None):
    profile = str(profile_name or "").strip().upper()
    if not profile:
        return ""
    return f"{PLANNER_TRAILER_RULES_OVERRIDE_SETTING_PREFIX}{profile}"


def _get_planner_trailer_rule_overrides(profile_name=None):
    resolved_profile = str(profile_name or "").strip()
    if not resolved_profile:
        try:
            resolved_profile = _get_session_profile_name()
        except RuntimeError:
            resolved_profile = ""
    key = _planner_trailer_rules_override_setting_key(resolved_profile)
    if not key:
        return {}
    setting = db.get_planning_setting(key) or {}
    raw_text = (setting.get("value_text") or "").strip()
    if not raw_text:
        return {}
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        return {}
    if not isinstance(parsed, dict):
        return {}
    sanitized = {}
    if "livestock_wedge_enabled" in parsed:
        sanitized["livestock_wedge_enabled"] = _coerce_bool_value(
            parsed.get("livestock_wedge_enabled")
        )
    return sanitized


def _get_trailer_assignment_rules():
    defaults = dict(DEFAULT_TRAILER_ASSIGNMENT_RULES)
    setting = db.get_planning_setting(TRAILER_ASSIGNMENT_RULES_SETTING_KEY) or {}
    raw_text = (setting.get("value_text") or "").strip()
    if not raw_text:
        return defaults
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        parsed = None
    return _merge_trailer_assignment_rules(defaults, parsed)


def _get_effective_trailer_assignment_rules(profile_name=None):
    base_rules = _get_trailer_assignment_rules()
    overrides = _get_planner_trailer_rule_overrides(profile_name)
    if not overrides:
        return base_rules
    return _merge_trailer_assignment_rules(base_rules, overrides)


def _carrier_section_code_for_trailer_type(trailer_type):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    if trailer_key == "HOTSHOT":
        return "HOT_SHOT"
    if trailer_key.startswith("FLATBED"):
        return "FLAT_BED"
    if trailer_key.startswith("STEP_DECK"):
        return "STEP_DECK"
    if trailer_key == "WEDGE":
        return "WEDGE"
    return ""


def _build_load_carrier_pricing_context():
    fls_lookup = {}
    for row in db.list_rate_matrix() or []:
        origin = str((row or {}).get("origin_plant") or "").strip().upper()
        destination = str((row or {}).get("destination_state") or "").strip().upper()
        rate = _coerce_optional_non_negative_float((row or {}).get("rate_per_mile"))
        if not origin or not destination or rate is None:
            continue
        key = (origin, destination)
        if key not in fls_lookup:
            fls_lookup[key] = round(float(rate), 4)

    lst_payload = _get_lst_rate_matrix()
    lst_lookup = {}
    for state, state_row in (lst_payload.get("matrix") or {}).items():
        destination = str(state or "").strip().upper()
        if not destination or not isinstance(state_row, dict):
            continue
        for plant, raw_rate in state_row.items():
            origin = str(plant or "").strip().upper()
            rate = _coerce_optional_non_negative_float(raw_rate)
            if not origin or rate is None:
                continue
            lst_lookup[(origin, destination)] = round(float(rate), 4)

    alternate_payload = _get_alternate_trailer_rates()
    alternate_sections = {}
    for section in (alternate_payload.get("sections") or []):
        if not isinstance(section, dict):
            continue
        code = _normalize_alternate_trailer_code(section.get("code") or section.get("trailer_type"))
        if not code:
            continue
        alternate_sections[code] = section

    return {
        "fls_lookup": fls_lookup,
        "fls_accessorial": {
            "per_stop": _get_stop_fee_amount(),
            "fuel_surcharge": _get_fuel_surcharge_per_mile(),
            "load_minimum": _get_load_minimum_amount(),
        },
        "lst_lookup": lst_lookup,
        "lst_accessorial": {
            "per_stop": _coerce_optional_non_negative_float(lst_payload.get("per_stop")) or 0.0,
            "fuel_surcharge": _coerce_optional_non_negative_float(lst_payload.get("fuel_surcharge")) or 0.0,
            "load_minimum": _coerce_optional_non_negative_float(lst_payload.get("load_minimum")) or 0.0,
        },
        "ryder_table": _get_ryder_dedicated_rate_table(),
        "alternate_sections": alternate_sections,
    }


def _normalize_route_leg_miles(leg_miles, total_miles, expected_count):
    expected = max(int(expected_count or 0), 0)
    if expected <= 0:
        return []
    normalized = []
    for value in list(leg_miles or [])[:expected]:
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            parsed = 0.0
        normalized.append(max(parsed, 0.0))
    while len(normalized) < expected:
        normalized.append(0.0)

    total_value = max(float(total_miles or 0.0), 0.0)
    current_sum = sum(normalized)
    if total_value > 0 and current_sum <= 0:
        even = total_value / expected
        return [even for _ in range(expected)]
    if total_value > 0 and current_sum > 0:
        scale = total_value / current_sum
        return [value * scale for value in normalized]
    return normalized


def _matrix_carrier_candidate(
    *,
    carrier_key,
    carrier_label,
    source_label,
    origin_plant,
    stop_states,
    leg_miles,
    lookup,
    stop_count,
    per_stop_fee,
    fuel_surcharge_per_mile,
    load_minimum,
    requires_return_miles=False,
    fallback_rate=None,
    require_all_lanes=False,
):
    if not stop_states:
        return None
    linehaul_cost = 0.0
    for index, state in enumerate(stop_states):
        rate = lookup.get((origin_plant, state))
        if rate is None:
            if require_all_lanes:
                return None
            rate = fallback_rate if fallback_rate is not None else DEFAULT_RATE_PER_MILE
        miles = leg_miles[index] if index < len(leg_miles) else 0.0
        linehaul_cost += max(float(miles or 0.0), 0.0) * float(rate)

    total_miles = max(sum(leg_miles), 0.0)
    fuel_cost = total_miles * max(float(fuel_surcharge_per_mile or 0.0), 0.0)
    stop_cost = max(int(stop_count or 0), 0) * max(float(per_stop_fee or 0.0), 0.0)
    subtotal = linehaul_cost + fuel_cost + stop_cost
    minimum = max(float(load_minimum or 0.0), 0.0)
    total_cost = max(subtotal, minimum)
    effective_rate = (linehaul_cost / total_miles) if total_miles > 0 else float(fallback_rate or DEFAULT_RATE_PER_MILE)
    return {
        "carrier_key": carrier_key,
        "carrier_label": carrier_label,
        "rate_source_label": source_label,
        "pricing_note": "Lane matrix by origin plant and destination state.",
        "requires_return_miles": bool(requires_return_miles),
        "includes_fuel_and_stop": False,
        "per_stop_fee": round(float(per_stop_fee or 0.0), 4),
        "fuel_surcharge_per_mile": round(float(fuel_surcharge_per_mile or 0.0), 4),
        "load_minimum": round(minimum, 2),
        "rate_per_mile": round(effective_rate + float(fuel_surcharge_per_mile or 0.0), 4),
        "linehaul_rate_per_mile": round(effective_rate, 4),
        "linehaul_cost": round(linehaul_cost, 2),
        "fuel_cost": round(fuel_cost, 2),
        "stop_cost": round(stop_cost, 2),
        "subtotal_before_floor": round(subtotal, 2),
        "total_cost": round(total_cost, 2),
    }


def _flat_carrier_candidate(
    *,
    carrier_key,
    carrier_label,
    source_label,
    rate_per_mile,
    total_miles,
    stop_count,
    per_stop_fee,
    fuel_surcharge_per_mile,
    load_minimum,
    includes_fuel_and_stop=False,
    requires_return_miles=False,
    pricing_note="",
):
    base_rate = _coerce_optional_non_negative_float(rate_per_mile)
    if base_rate is None:
        return None
    miles = max(float(total_miles or 0.0), 0.0)
    base_rate = float(base_rate)
    linehaul_cost = miles * base_rate
    fuel_cost = miles * max(float(fuel_surcharge_per_mile or 0.0), 0.0)
    stop_cost = max(int(stop_count or 0), 0) * max(float(per_stop_fee or 0.0), 0.0)
    subtotal = linehaul_cost + fuel_cost + stop_cost
    minimum = max(float(load_minimum or 0.0), 0.0)
    total_cost = max(subtotal, minimum)
    return {
        "carrier_key": carrier_key,
        "carrier_label": carrier_label,
        "rate_source_label": source_label,
        "pricing_note": pricing_note,
        "requires_return_miles": bool(requires_return_miles),
        "includes_fuel_and_stop": bool(includes_fuel_and_stop),
        "per_stop_fee": round(float(per_stop_fee or 0.0), 4),
        "fuel_surcharge_per_mile": round(float(fuel_surcharge_per_mile or 0.0), 4),
        "load_minimum": round(minimum, 2),
        "rate_per_mile": round(base_rate + float(fuel_surcharge_per_mile or 0.0), 4),
        "linehaul_rate_per_mile": round(base_rate, 4),
        "linehaul_cost": round(linehaul_cost, 2),
        "fuel_cost": round(fuel_cost, 2),
        "stop_cost": round(stop_cost, 2),
        "subtotal_before_floor": round(subtotal, 2),
        "total_cost": round(total_cost, 2),
    }


def _load_has_lowes_order(lines):
    return any(
        customer_rules.is_lowes_customer((line or {}).get("cust_name") or "")
        for line in (lines or [])
    )


def _alternate_requires_return_hint(lines, trailer_type, origin_plant, carrier_context):
    trailer_code = _carrier_section_code_for_trailer_type(trailer_type)
    if not trailer_code:
        return False
    section = (carrier_context.get("alternate_sections") or {}).get(trailer_code)
    if not isinstance(section, dict):
        return False
    rates = section.get("rates_by_plant") if isinstance(section.get("rates_by_plant"), dict) else {}
    if _coerce_optional_non_negative_float(rates.get(origin_plant)) is None:
        return False
    placeholders = section.get("placeholders_by_plant") if isinstance(section.get("placeholders_by_plant"), dict) else {}
    plant_row = placeholders.get(origin_plant) if isinstance(placeholders.get(origin_plant), dict) else {}
    return bool(_coerce_bool_value(plant_row.get("requires_return_miles", True)))


def _resolve_load_carrier_pricing(
    *,
    lines,
    trailer_type,
    origin_plant,
    ordered_stops,
    route_legs,
    total_miles,
    stop_count,
    requires_return_to_origin,
    carrier_context,
    forced_carrier_key=None,
):
    origin = str(origin_plant or "").strip().upper()
    stops = list(ordered_stops or [])
    stop_states = [str((stop or {}).get("state") or "").strip().upper() for stop in stops]
    if requires_return_to_origin:
        stop_states = stop_states + [origin]
    leg_miles = _normalize_route_leg_miles(route_legs, total_miles, len(stop_states))

    fls_accessorial = carrier_context.get("fls_accessorial") or {}
    fls_candidate = _matrix_carrier_candidate(
        carrier_key="fls",
        carrier_label="FLS",
        source_label="FLS Default Matrix",
        origin_plant=origin,
        stop_states=stop_states,
        leg_miles=leg_miles,
        lookup=carrier_context.get("fls_lookup") or {},
        stop_count=stop_count,
        per_stop_fee=fls_accessorial.get("per_stop", DEFAULT_STOP_FEE),
        fuel_surcharge_per_mile=fls_accessorial.get("fuel_surcharge", DEFAULT_FUEL_SURCHARGE_PER_MILE),
        load_minimum=fls_accessorial.get("load_minimum", DEFAULT_MIN_LOAD_COST),
        requires_return_miles=requires_return_to_origin,
        fallback_rate=DEFAULT_RATE_PER_MILE,
        require_all_lanes=False,
    )

    lst_accessorial = carrier_context.get("lst_accessorial") or {}
    lst_candidate = _matrix_carrier_candidate(
        carrier_key="lst",
        carrier_label="LST",
        source_label="LST Matrix",
        origin_plant=origin,
        stop_states=stop_states,
        leg_miles=leg_miles,
        lookup=carrier_context.get("lst_lookup") or {},
        stop_count=stop_count,
        per_stop_fee=lst_accessorial.get("per_stop", 0.0),
        fuel_surcharge_per_mile=lst_accessorial.get("fuel_surcharge", 0.0),
        load_minimum=lst_accessorial.get("load_minimum", 0.0),
        requires_return_miles=requires_return_to_origin,
        fallback_rate=None,
        require_all_lanes=True,
    )

    ryder_table = carrier_context.get("ryder_table") or {}
    ryder_rate = ((ryder_table.get("rates_by_plant") or {}).get(origin))
    ryder_candidate = _flat_carrier_candidate(
        carrier_key="ryder",
        carrier_label="Ryder Dedicated",
        source_label="Ryder Dedicated (All States)",
        rate_per_mile=ryder_rate,
        total_miles=total_miles,
        stop_count=stop_count,
        per_stop_fee=ryder_table.get("per_stop", 0.0),
        fuel_surcharge_per_mile=ryder_table.get("fuel_surcharge", 0.0),
        load_minimum=ryder_table.get("load_minimum", 0.0),
        includes_fuel_and_stop=True,
        requires_return_miles=True,
        pricing_note="Includes dedicated fleet return miles pricing.",
    )

    trailer_code = _carrier_section_code_for_trailer_type(trailer_type)
    alternate_candidate = None
    if trailer_code:
        section = (carrier_context.get("alternate_sections") or {}).get(trailer_code)
        if isinstance(section, dict):
            rates_by_plant = section.get("rates_by_plant") if isinstance(section.get("rates_by_plant"), dict) else {}
            placeholders_by_plant = section.get("placeholders_by_plant") if isinstance(section.get("placeholders_by_plant"), dict) else {}
            plant_placeholder = placeholders_by_plant.get(origin) if isinstance(placeholders_by_plant.get(origin), dict) else {}
            default_fuel_surcharge = _coerce_optional_non_negative_float(
                fls_accessorial.get("fuel_surcharge")
            )
            if default_fuel_surcharge is None:
                default_fuel_surcharge = DEFAULT_FUEL_SURCHARGE_PER_MILE
            apply_fuel_surcharge = _coerce_alternate_apply_fuel_surcharge(
                plant_placeholder,
                default=True,
            )
            alternate_candidate = _flat_carrier_candidate(
                carrier_key="alternate",
                carrier_label=_alternate_trailer_label_for_code(trailer_code),
                source_label=f"{_alternate_trailer_label_for_code(trailer_code)} Alternate Trailer",
                rate_per_mile=rates_by_plant.get(origin),
                total_miles=total_miles,
                stop_count=stop_count,
                per_stop_fee=plant_placeholder.get("per_stop", 0.0),
                fuel_surcharge_per_mile=default_fuel_surcharge if apply_fuel_surcharge else 0.0,
                load_minimum=plant_placeholder.get("load_minimum", 0.0),
                includes_fuel_and_stop=False,
                requires_return_miles=bool(_coerce_bool_value(plant_placeholder.get("requires_return_miles", True))),
                pricing_note="Plant-specific alternate trailer pricing.",
            )

    candidate_by_key = {
        "fls": fls_candidate,
        "lst": lst_candidate,
        "ryder": ryder_candidate,
        "alternate": alternate_candidate,
    }
    forced_key = str(forced_carrier_key or "").strip().lower()
    if forced_key:
        forced_candidate = candidate_by_key.get(forced_key)
        if forced_candidate:
            forced_payload = dict(forced_candidate)
            forced_payload["selection_reason"] = "Carrier override selected by planner."
            forced_payload["override_applied"] = True
            forced_payload["requested_carrier_key"] = forced_key
            return forced_payload

    has_lowes = _load_has_lowes_order(lines)

    if has_lowes and ryder_candidate is not None:
        ryder_candidate["selection_reason"] = "Lowe's order detected; using dedicated fleet carrier."
        return ryder_candidate

    if alternate_candidate is not None:
        alternate_candidate["selection_reason"] = "Alternate trailer pricing applied."
        return alternate_candidate

    if lst_candidate and fls_candidate:
        if lst_candidate["total_cost"] <= fls_candidate["total_cost"]:
            lst_candidate["selection_reason"] = "Selected the lower total cost between LST and FLS."
            return lst_candidate
        fls_candidate["selection_reason"] = "Selected the lower total cost between FLS and LST."
        return fls_candidate

    if fls_candidate:
        fls_candidate["selection_reason"] = "Default FLS matrix applied."
        return fls_candidate

    fallback_miles = max(float(total_miles or 0.0), 0.0)
    fallback_stop_fee = float(fls_accessorial.get("per_stop", DEFAULT_STOP_FEE))
    fallback_fuel = float(fls_accessorial.get("fuel_surcharge", DEFAULT_FUEL_SURCHARGE_PER_MILE))
    fallback_minimum = float(fls_accessorial.get("load_minimum", DEFAULT_MIN_LOAD_COST))
    fallback_linehaul = fallback_miles * DEFAULT_RATE_PER_MILE
    fallback_fuel_cost = fallback_miles * fallback_fuel
    fallback_stop_cost = max(int(stop_count or 0), 0) * fallback_stop_fee
    fallback_subtotal = fallback_linehaul + fallback_fuel_cost + fallback_stop_cost
    fallback_total = max(fallback_subtotal, fallback_minimum)

    return {
        "carrier_key": "fls",
        "carrier_label": "FLS",
        "rate_source_label": "FLS Default Matrix",
        "pricing_note": "Fallback default carrier rate applied.",
        "selection_reason": "Fallback default carrier rate applied.",
        "requires_return_miles": bool(requires_return_to_origin),
        "includes_fuel_and_stop": False,
        "per_stop_fee": float(fls_accessorial.get("per_stop", DEFAULT_STOP_FEE)),
        "fuel_surcharge_per_mile": float(fls_accessorial.get("fuel_surcharge", DEFAULT_FUEL_SURCHARGE_PER_MILE)),
        "load_minimum": float(fls_accessorial.get("load_minimum", DEFAULT_MIN_LOAD_COST)),
        "rate_per_mile": float(DEFAULT_RATE_PER_MILE + fls_accessorial.get("fuel_surcharge", DEFAULT_FUEL_SURCHARGE_PER_MILE)),
        "linehaul_rate_per_mile": float(DEFAULT_RATE_PER_MILE),
        "linehaul_cost": round(fallback_linehaul, 2),
        "fuel_cost": round(fallback_fuel_cost, 2),
        "stop_cost": round(fallback_stop_cost, 2),
        "subtotal_before_floor": round(fallback_subtotal, 2),
        "total_cost": round(fallback_total, 2),
    }


def _build_freight_breakdown(load, stop_fee_amount, fuel_surcharge_per_mile, load_minimum_amount):
    def _as_float(value, default=0.0):
        try:
            return float(value)
        except (TypeError, ValueError):
            return float(default)

    carrier_pricing = load.get("carrier_pricing") if isinstance(load.get("carrier_pricing"), dict) else {}
    estimated_cost = max(_as_float(load.get("estimated_cost"), 0.0), 0.0)
    estimated_miles = max(
        _as_float(load.get("estimated_miles"), _as_float(load.get("route_distance"), 0.0)),
        0.0,
    )
    stop_count = max(int(_as_float(load.get("stop_count"), 0.0)), 0)
    stop_fee = max(
        _as_float(carrier_pricing.get("per_stop_fee"), _as_float(stop_fee_amount, DEFAULT_STOP_FEE)),
        0.0,
    )
    fuel_surcharge = max(
        _as_float(
            carrier_pricing.get("fuel_surcharge_per_mile"),
            _as_float(fuel_surcharge_per_mile, DEFAULT_FUEL_SURCHARGE_PER_MILE),
        ),
        0.0,
    )
    load_minimum = max(
        _as_float(carrier_pricing.get("load_minimum"), _as_float(load_minimum_amount, DEFAULT_MIN_LOAD_COST)),
        0.0,
    )

    stop_cost = stop_count * stop_fee
    avg_rate_per_mile = _as_float(load.get("rate_per_mile"), 0.0)
    if avg_rate_per_mile <= 0 and estimated_miles > 0:
        avg_rate_per_mile = max((estimated_cost - stop_cost) / estimated_miles, 0.0)

    linehaul_rate_per_mile = max(avg_rate_per_mile - fuel_surcharge, 0.0)
    linehaul_cost = estimated_miles * linehaul_rate_per_mile
    fuel_cost = estimated_miles * fuel_surcharge
    subtotal_before_floor = linehaul_cost + fuel_cost + stop_cost

    adjustment_cost = estimated_cost - subtotal_before_floor
    if abs(adjustment_cost) < 0.01:
        adjustment_cost = 0.0

    min_floor_applied = (
        load_minimum > 0
        and estimated_cost + 0.01 >= load_minimum
        and subtotal_before_floor < load_minimum
    )
    adjustment_label = "Minimum load floor adjustment" if min_floor_applied else "Adjustment"

    return {
        "estimated_miles": round(estimated_miles, 1),
        "stop_count": stop_count,
        "stop_fee": round(stop_fee, 2),
        "stop_cost": round(stop_cost, 2),
        "avg_rate_per_mile": round(avg_rate_per_mile, 4),
        "linehaul_rate_per_mile": round(linehaul_rate_per_mile, 4),
        "linehaul_cost": round(linehaul_cost, 2),
        "fuel_surcharge_per_mile": round(fuel_surcharge, 4),
        "fuel_cost": round(fuel_cost, 2),
        "subtotal_before_floor": round(subtotal_before_floor, 2),
        "adjustment_cost": round(adjustment_cost, 2),
        "adjustment_label": adjustment_label,
        "min_load_cost": round(load_minimum, 2),
        "total_cost": round(estimated_cost, 2),
        "carrier_key": str(carrier_pricing.get("carrier_key") or load.get("carrier_key") or "fls"),
        "carrier_label": str(carrier_pricing.get("carrier_label") or load.get("carrier_label") or "FLS"),
        "rate_source_label": str(carrier_pricing.get("rate_source_label") or ""),
        "selection_reason": str(carrier_pricing.get("selection_reason") or ""),
        "pricing_note": str(carrier_pricing.get("pricing_note") or ""),
        "requires_return_miles": bool(
            _coerce_bool_value(
                carrier_pricing.get(
                    "requires_return_miles",
                    load.get("return_to_origin"),
                )
            )
        ),
        "includes_fuel_and_stop": bool(
            _coerce_bool_value(carrier_pricing.get("includes_fuel_and_stop", False))
        ),
    }


def _profile_capacity_for_trailer(trailer_type, default_capacity=53.0):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    config = stack_calculator.TRAILER_CONFIGS.get(
        trailer_key,
        stack_calculator.TRAILER_CONFIGS["STEP_DECK"],
    )
    return _coerce_non_negative_float(config.get("capacity"), default_capacity)


def _capacity_for_trailer_setting(trailer_type, requested_capacity=None, default_capacity=53.0):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    profile_capacity = _profile_capacity_for_trailer(trailer_key, default_capacity)
    if profile_capacity > 0:
        return round(profile_capacity, 2)
    return round(_coerce_non_negative_float(default_capacity, 53.0), 2)


def _default_trailer_type_for_plant(plant_code, fallback="STEP_DECK"):
    normalized = _normalize_plant_code(plant_code)
    preferred = PLANT_DEFAULT_TRAILER_TYPE_OVERRIDES.get(normalized, fallback)
    return stack_calculator.normalize_trailer_type(preferred, default=fallback)


def _resolve_auto_hotshot_enabled_for_plant(plant_code, trailer_rules=None, settings_row=None):
    normalized = _normalize_plant_code(plant_code)
    fallback = True
    if normalized in PLANT_DEFAULT_AUTO_HOTSHOT_OVERRIDES:
        fallback = bool(PLANT_DEFAULT_AUTO_HOTSHOT_OVERRIDES[normalized])
    row = settings_row
    if row is None and normalized:
        row = db.get_optimizer_settings(normalized) or {}
    override_value = (row or {}).get("auto_hotshot_enabled")
    if override_value in (None, ""):
        return bool(fallback)
    return _coerce_bool_value(override_value)


def _get_optimizer_settings_for_plant(plant_code, optimizer_defaults=None, trailer_rules=None, settings_row=None):
    normalized = _normalize_plant_code(plant_code)
    base_defaults = dict(optimizer_defaults or _get_optimizer_default_settings())
    fallback_trailer = stack_calculator.normalize_trailer_type(
        base_defaults.get("trailer_type"),
        default="STEP_DECK",
    )
    trailer_type = _default_trailer_type_for_plant(
        normalized,
        fallback=fallback_trailer,
    )
    row = settings_row
    if row is None and normalized:
        row = db.get_optimizer_settings(normalized) or {}
    if (row or {}).get("trailer_type"):
        trailer_type = stack_calculator.normalize_trailer_type(
            row.get("trailer_type"),
            default=trailer_type,
        )
    capacity_feet = _capacity_for_trailer_setting(
        trailer_type,
        requested_capacity=(row or {}).get("capacity_feet", base_defaults.get("capacity_feet")),
        default_capacity=base_defaults.get("capacity_feet", 53.0),
    )
    return {
        "plant_code": normalized,
        "trailer_type": trailer_type,
        "capacity_feet": capacity_feet,
        "auto_hotshot_enabled": _resolve_auto_hotshot_enabled_for_plant(
            normalized,
            trailer_rules=trailer_rules,
            settings_row=row,
        ),
    }


def _build_optimizer_plant_defaults(optimizer_defaults=None, trailer_rules=None, plant_rows=None):
    defaults = dict(optimizer_defaults or _get_optimizer_default_settings())
    rules = trailer_rules or _get_effective_trailer_assignment_rules()
    source_rows = list(plant_rows or [])
    if not source_rows:
        source_rows = [{"plant_code": code, "name": PLANT_NAMES.get(code, code)} for code in PLANT_CODES]
    indexed_rows = {
        _normalize_plant_code(row.get("plant_code")): row
        for row in source_rows
        if _normalize_plant_code(row.get("plant_code"))
    }

    rows = []
    for plant_code in sorted(
        indexed_rows.keys(),
        key=lambda code: (PLANT_CODES.index(code) if code in PLANT_CODES else 10**6, code),
    ):
        row = indexed_rows.get(plant_code) or {}
        plant_settings_row = db.get_optimizer_settings(plant_code) or {}
        settings = _get_optimizer_settings_for_plant(
            plant_code,
            optimizer_defaults=defaults,
            trailer_rules=rules,
            settings_row=plant_settings_row,
        )
        rows.append(
            {
                "plant_code": plant_code,
                "plant_name": row.get("name") or PLANT_NAMES.get(plant_code, plant_code),
                "trailer_type": settings["trailer_type"],
                "capacity_feet": settings["capacity_feet"],
                "auto_hotshot_enabled": bool(settings["auto_hotshot_enabled"]),
            }
        )
    return rows


def _build_optimizer_workbench_trailer_defaults(plants, optimizer_defaults=None):
    options = stack_calculator.trailer_profile_options()
    option_map = {
        str(option.get("value") or "").strip().upper(): option
        for option in options
        if str(option.get("value") or "").strip()
    }
    defaults = {}
    for plant_code in (plants or []):
        normalized = _normalize_plant_code(plant_code)
        if not normalized:
            continue
        settings = _get_optimizer_settings_for_plant(
            normalized,
            optimizer_defaults=optimizer_defaults,
        )
        trailer_key = stack_calculator.normalize_trailer_type(
            settings.get("trailer_type"),
            default="STEP_DECK",
        )
        option = option_map.get(trailer_key, {})
        defaults[normalized] = {
            "trailer_type": trailer_key,
            "label": option.get("label") or trailer_key.replace("_", " ").title(),
            "capacity_feet": _coerce_non_negative_float(settings.get("capacity_feet"), 53.0),
        }
    return {
        "options": options,
        "defaults_by_plant": defaults,
    }


def _get_optimizer_default_settings():
    defaults = {
        "trailer_type": stack_calculator.normalize_trailer_type(
            load_builder.DEFAULT_BUILD_PARAMS.get("trailer_type"),
            default="STEP_DECK",
        ),
        "capacity_feet": _coerce_non_negative_float(load_builder.DEFAULT_BUILD_PARAMS.get("capacity_feet"), 53),
        "max_detour_pct": _coerce_non_negative_float(load_builder.DEFAULT_BUILD_PARAMS.get("max_detour_pct"), 15),
        "time_window_days": _coerce_non_negative_int(load_builder.DEFAULT_BUILD_PARAMS.get("time_window_days"), 7),
        "geo_radius": _coerce_non_negative_float(load_builder.DEFAULT_BUILD_PARAMS.get("geo_radius"), 100),
        "stack_overflow_max_height": _coerce_non_negative_int(
            load_builder.DEFAULT_BUILD_PARAMS.get("stack_overflow_max_height"),
            DEFAULT_STACK_OVERFLOW_MAX_HEIGHT,
        ),
        "max_back_overhang_ft": _coerce_non_negative_float(
            load_builder.DEFAULT_BUILD_PARAMS.get("max_back_overhang_ft"),
            DEFAULT_MAX_BACK_OVERHANG_FT,
        ),
        "upper_two_across_max_length_ft": _coerce_non_negative_float(
            load_builder.DEFAULT_BUILD_PARAMS.get("upper_two_across_max_length_ft"),
            DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT,
        ),
        "upper_deck_exception_max_length_ft": _coerce_non_negative_float(
            load_builder.DEFAULT_BUILD_PARAMS.get("upper_deck_exception_max_length_ft"),
            DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
        ),
        "upper_deck_exception_overhang_allowance_ft": _coerce_non_negative_float(
            load_builder.DEFAULT_BUILD_PARAMS.get("upper_deck_exception_overhang_allowance_ft"),
            DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
        ),
        "upper_deck_exception_categories": stack_calculator.normalize_upper_deck_exception_categories(
            load_builder.DEFAULT_BUILD_PARAMS.get("upper_deck_exception_categories"),
            default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
        ),
        "equal_length_deck_length_order_enabled": (
            _coerce_bool_value(load_builder.DEFAULT_BUILD_PARAMS.get("equal_length_deck_length_order_enabled"))
            if load_builder.DEFAULT_BUILD_PARAMS.get("equal_length_deck_length_order_enabled") is not None
            else DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED
        ),
    }
    setting = _get_effective_planning_setting(OPTIMIZER_DEFAULTS_SETTING_KEY)
    raw_text = (setting.get("value_text") or "").strip()
    if raw_text:
        try:
            parsed = json.loads(raw_text)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, dict):
            defaults["trailer_type"] = stack_calculator.normalize_trailer_type(
                parsed.get("trailer_type"),
                default=defaults["trailer_type"],
            )
            defaults["capacity_feet"] = _coerce_non_negative_float(parsed.get("capacity_feet"), defaults["capacity_feet"])
            defaults["max_detour_pct"] = _coerce_non_negative_float(parsed.get("max_detour_pct"), defaults["max_detour_pct"])
            defaults["time_window_days"] = _coerce_non_negative_int(
                parsed.get("time_window_days"),
                defaults["time_window_days"],
            )
            defaults["geo_radius"] = _coerce_non_negative_float(parsed.get("geo_radius"), defaults["geo_radius"])
            defaults["stack_overflow_max_height"] = _coerce_non_negative_int(
                parsed.get("stack_overflow_max_height"),
                defaults["stack_overflow_max_height"],
            )
            defaults["max_back_overhang_ft"] = _coerce_non_negative_float(
                parsed.get("max_back_overhang_ft"),
                defaults["max_back_overhang_ft"],
            )
            defaults["upper_two_across_max_length_ft"] = _coerce_non_negative_float(
                parsed.get("upper_two_across_max_length_ft"),
                defaults["upper_two_across_max_length_ft"],
            )
            defaults["upper_deck_exception_max_length_ft"] = _coerce_non_negative_float(
                parsed.get("upper_deck_exception_max_length_ft"),
                defaults["upper_deck_exception_max_length_ft"],
            )
            defaults["upper_deck_exception_overhang_allowance_ft"] = _coerce_non_negative_float(
                parsed.get("upper_deck_exception_overhang_allowance_ft"),
                defaults["upper_deck_exception_overhang_allowance_ft"],
            )
            defaults["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
                parsed.get("upper_deck_exception_categories"),
                default=defaults["upper_deck_exception_categories"],
            )
            defaults["equal_length_deck_length_order_enabled"] = (
                _coerce_bool_value(parsed.get("equal_length_deck_length_order_enabled"))
                if parsed.get("equal_length_deck_length_order_enabled") is not None
                else bool(defaults["equal_length_deck_length_order_enabled"])
            )
    defaults["max_back_overhang_ft"] = round(defaults["max_back_overhang_ft"], 2)
    defaults["upper_two_across_max_length_ft"] = round(defaults["upper_two_across_max_length_ft"], 2)
    defaults["upper_deck_exception_max_length_ft"] = round(
        defaults["upper_deck_exception_max_length_ft"],
        2,
    )
    defaults["upper_deck_exception_overhang_allowance_ft"] = round(
        defaults["upper_deck_exception_overhang_allowance_ft"],
        2,
    )
    defaults["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
        defaults.get("upper_deck_exception_categories"),
        default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
    )
    defaults["equal_length_deck_length_order_enabled"] = bool(
        defaults.get(
            "equal_length_deck_length_order_enabled",
            DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED,
        )
    )
    defaults["capacity_feet"] = _capacity_for_trailer_setting(
        defaults.get("trailer_type"),
        requested_capacity=defaults.get("capacity_feet"),
        default_capacity=53.0,
    )
    return defaults


def _get_stack_capacity_assumptions():
    defaults = _get_optimizer_default_settings()
    return {
        "stack_overflow_max_height": _coerce_non_negative_int(
            defaults.get("stack_overflow_max_height"),
            DEFAULT_STACK_OVERFLOW_MAX_HEIGHT,
        ),
        "max_back_overhang_ft": round(
            _coerce_non_negative_float(
                defaults.get("max_back_overhang_ft"),
                DEFAULT_MAX_BACK_OVERHANG_FT,
            ),
            2,
        ),
        "upper_two_across_max_length_ft": round(
            _coerce_non_negative_float(
                defaults.get("upper_two_across_max_length_ft"),
                DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT,
            ),
            2,
        ),
        "upper_deck_exception_max_length_ft": round(
            _coerce_non_negative_float(
                defaults.get("upper_deck_exception_max_length_ft"),
                DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
            ),
            2,
        ),
        "upper_deck_exception_overhang_allowance_ft": round(
            _coerce_non_negative_float(
                defaults.get("upper_deck_exception_overhang_allowance_ft"),
                DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
            ),
            2,
        ),
        "upper_deck_exception_categories": stack_calculator.normalize_upper_deck_exception_categories(
            defaults.get("upper_deck_exception_categories"),
            default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
        ),
        "equal_length_deck_length_order_enabled": bool(
            defaults.get(
                "equal_length_deck_length_order_enabled",
                DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED,
            )
        ),
    }


def _get_replay_eval_preset():
    defaults = dict(replay_evaluator.DEFAULT_REPLAY_PRESET)
    optimizer_defaults = _get_optimizer_default_settings()
    defaults["trailer_type"] = optimizer_defaults.get("trailer_type") or defaults.get("trailer_type")
    defaults["capacity_feet"] = float(optimizer_defaults.get("capacity_feet") or defaults.get("capacity_feet") or 53.0)
    defaults["max_detour_pct"] = float(optimizer_defaults.get("max_detour_pct") or defaults.get("max_detour_pct") or 15.0)
    defaults["time_window_days"] = int(optimizer_defaults.get("time_window_days") or defaults.get("time_window_days") or 7)
    defaults["geo_radius"] = float(optimizer_defaults.get("geo_radius") or defaults.get("geo_radius") or 100.0)
    defaults["stack_overflow_max_height"] = int(
        optimizer_defaults.get("stack_overflow_max_height")
        or defaults.get("stack_overflow_max_height")
        or DEFAULT_STACK_OVERFLOW_MAX_HEIGHT
    )
    defaults["max_back_overhang_ft"] = float(
        optimizer_defaults.get("max_back_overhang_ft")
        or defaults.get("max_back_overhang_ft")
        or DEFAULT_MAX_BACK_OVERHANG_FT
    )
    defaults["upper_two_across_max_length_ft"] = float(
        optimizer_defaults.get("upper_two_across_max_length_ft")
        or defaults.get("upper_two_across_max_length_ft")
        or DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT
    )
    defaults["upper_deck_exception_max_length_ft"] = float(
        optimizer_defaults.get("upper_deck_exception_max_length_ft")
        or defaults.get("upper_deck_exception_max_length_ft")
        or DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT
    )
    defaults["upper_deck_exception_overhang_allowance_ft"] = float(
        optimizer_defaults.get("upper_deck_exception_overhang_allowance_ft")
        or defaults.get("upper_deck_exception_overhang_allowance_ft")
        or DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT
    )
    defaults["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
        optimizer_defaults.get("upper_deck_exception_categories")
        or defaults.get("upper_deck_exception_categories")
        or DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
        default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
    )
    defaults["equal_length_deck_length_order_enabled"] = (
        _coerce_bool_value(optimizer_defaults.get("equal_length_deck_length_order_enabled"))
        if optimizer_defaults.get("equal_length_deck_length_order_enabled") is not None
        else bool(
            defaults.get(
                "equal_length_deck_length_order_enabled",
                DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED,
            )
        )
    )

    setting = db.get_planning_setting(REPLAY_EVAL_PRESET_SETTING_KEY) or {}
    raw_text = (setting.get("value_text") or "").strip()
    parsed = None
    if raw_text:
        try:
            parsed = json.loads(raw_text)
        except json.JSONDecodeError:
            parsed = None
    if isinstance(parsed, dict):
        defaults.update(parsed)

    defaults["trailer_type"] = stack_calculator.normalize_trailer_type(
        defaults.get("trailer_type"),
        default="STEP_DECK",
    )
    defaults["capacity_feet"] = _coerce_non_negative_float(defaults.get("capacity_feet"), 53)
    defaults["max_detour_pct"] = _coerce_non_negative_float(defaults.get("max_detour_pct"), 15)
    defaults["time_window_days"] = _coerce_non_negative_int(defaults.get("time_window_days"), 7)
    defaults["geo_radius"] = _coerce_non_negative_float(defaults.get("geo_radius"), 100)
    defaults["stack_overflow_max_height"] = _coerce_non_negative_int(
        defaults.get("stack_overflow_max_height"),
        DEFAULT_STACK_OVERFLOW_MAX_HEIGHT,
    )
    defaults["max_back_overhang_ft"] = round(
        _coerce_non_negative_float(
            defaults.get("max_back_overhang_ft"),
            DEFAULT_MAX_BACK_OVERHANG_FT,
        ),
        2,
    )
    defaults["upper_two_across_max_length_ft"] = round(
        _coerce_non_negative_float(
            defaults.get("upper_two_across_max_length_ft"),
            DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT,
        ),
        2,
    )
    defaults["upper_deck_exception_max_length_ft"] = round(
        _coerce_non_negative_float(
            defaults.get("upper_deck_exception_max_length_ft"),
            DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
        ),
        2,
    )
    defaults["upper_deck_exception_overhang_allowance_ft"] = round(
        _coerce_non_negative_float(
            defaults.get("upper_deck_exception_overhang_allowance_ft"),
            DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
        ),
        2,
    )
    defaults["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
        defaults.get("upper_deck_exception_categories"),
        default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
    )
    defaults["equal_length_deck_length_order_enabled"] = (
        _coerce_bool_value(defaults.get("equal_length_deck_length_order_enabled"))
        if defaults.get("equal_length_deck_length_order_enabled") is not None
        else DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED
    )
    defaults["ops_parity_enabled"] = _coerce_bool_value(defaults.get("ops_parity_enabled"))
    defaults["ops_parity_max_utilization_pct"] = _coerce_non_negative_float(
        defaults.get("ops_parity_max_utilization_pct"),
        replay_evaluator.DEFAULT_REPLAY_PRESET.get("ops_parity_max_utilization_pct", 120.0),
    )
    defaults["algorithm_version"] = "v2"
    defaults["enforce_time_window"] = bool(defaults.get("enforce_time_window", True))
    return defaults


def _parse_replay_summary(raw_json):
    if not raw_json:
        return {}
    if isinstance(raw_json, dict):
        return raw_json
    try:
        parsed = json.loads(raw_json)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _get_utilization_grade_thresholds():
    setting = _get_effective_planning_setting(UTILIZATION_GRADE_THRESHOLDS_SETTING_KEY)
    raw_text = (setting.get("value_text") or "").strip()
    if not raw_text:
        return dict(DEFAULT_UTILIZATION_GRADE_THRESHOLDS)
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        return dict(DEFAULT_UTILIZATION_GRADE_THRESHOLDS)
    return _coerce_utilization_grade_thresholds(parsed)


def _build_utilization_grade_rows(thresholds):
    a = thresholds["A"]
    b = thresholds["B"]
    c = thresholds["C"]
    d = thresholds["D"]
    f_max = max(d - 1, 0)
    return [
        {"grade": "A", "min_pct": a, "label": f">= {a}%"},
        {"grade": "B", "min_pct": b, "label": f">= {b}% and < {a}%"},
        {"grade": "C", "min_pct": c, "label": f">= {c}% and < {b}%"},
        {"grade": "D", "min_pct": d, "label": f">= {d}% and < {c}%"},
        {"grade": "F", "min_pct": None, "f_max": f_max, "label": f"< {d}%"},
    ]


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(value).date()
    except (TypeError, ValueError):
        pass
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except (TypeError, ValueError):
            continue
    return None


def _parse_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except (TypeError, ValueError):
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except (TypeError, ValueError):
            continue
    return None


def _to_est_datetime(value):
    parsed = _parse_datetime(value)
    if not parsed:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(APP_TIMEZONE)


def _format_est_datetime_label(value):
    parsed = _to_est_datetime(value)
    if not parsed:
        return ""
    return f"{parsed.strftime('%b')} {parsed.day}, {parsed.year} {parsed.strftime('%I:%M %p')} ET"


def _format_datetime_label(value):
    return _format_est_datetime_label(value)


def _resolve_today_override(value):
    token = (value or "").strip().lower()
    if token == "clear":
        session.pop("today_override", None)
        return None
    if value:
        parsed = _parse_date(value)
        if parsed:
            session["today_override"] = parsed.strftime("%Y-%m-%d")
        return parsed
    stored = session.get("today_override")
    return _parse_date(stored)


def _compute_load_progress_snapshot(plant_scope=None, all_loads=None, allowed_plants=None):
    allowed_plants = allowed_plants or _get_allowed_plants()
    plant_scope = plant_scope or allowed_plants
    if all_loads is None:
        all_loads = load_builder.list_loads(None, include_stack_metrics=False)
    all_loads = [load for load in all_loads if load.get("origin_plant") in allowed_plants]
    loads_for_progress = [load for load in all_loads if load.get("origin_plant") in plant_scope]

    optimized_loads = []
    for load in loads_for_progress:
        build_source = (load.get("build_source") or "OPTIMIZED").upper()
        if build_source == "MANUAL":
            continue
        optimized_loads.append(load)

    optimized_order_ids = {
        line.get("so_num")
        for load in optimized_loads
        for line in load.get("lines", [])
        if line.get("so_num")
    }
    order_status_map = {so_num: "UNASSIGNED" for so_num in optimized_order_ids if so_num}

    status_priority = {
        "UNASSIGNED": 0,
        STATUS_PROPOSED: 1,
        STATUS_DRAFT: 2,
        STATUS_APPROVED: 3,
    }

    for load in loads_for_progress:
        load_status = (load.get("status") or STATUS_PROPOSED).upper()
        for line in load.get("lines", []):
            so_num = line.get("so_num")
            if so_num in order_status_map:
                current = order_status_map.get(so_num, "UNASSIGNED")
                if status_priority.get(load_status, 0) > status_priority.get(current, 0):
                    order_status_map[so_num] = load_status

    order_status_counts = {
        "unassigned": 0,
        "proposed": 0,
        "draft": 0,
        "approved": 0,
    }
    for status in order_status_map.values():
        if status == STATUS_PROPOSED:
            order_status_counts["proposed"] += 1
        elif status == STATUS_DRAFT:
            order_status_counts["draft"] += 1
        elif status == STATUS_APPROVED:
            order_status_counts["approved"] += 1
        else:
            order_status_counts["unassigned"] += 1

    load_status_counts = {"proposed": 0, "draft": 0, "approved": 0}
    for load in loads_for_progress:
        status = (load.get("status") or STATUS_PROPOSED).upper()
        if status == STATUS_DRAFT:
            load_status_counts["draft"] += 1
        elif status == STATUS_APPROVED:
            load_status_counts["approved"] += 1
        else:
            load_status_counts["proposed"] += 1

    total_orders = len(order_status_map)
    approved_orders = order_status_counts["approved"]
    progress_pct = round((approved_orders / total_orders) * 100, 1) if total_orders else 0.0

    return {
        "order_status_counts": order_status_counts,
        "load_status_counts": load_status_counts,
        "approved_orders": approved_orders,
        "total_orders": total_orders,
        "progress_pct": progress_pct,
        "draft_tab_count": load_status_counts["draft"] + load_status_counts["proposed"],
        "final_tab_count": load_status_counts["approved"],
    }


def _due_status(due_date_value, today=None, due_soon_days=14):
    """Return a simple bucket for UI badges."""
    today = today or date.today()
    due_date = _parse_date(due_date_value)
    if not due_date:
        return ""
    if due_date < today:
        return "PAST_DUE"
    if due_date <= today + timedelta(days=due_soon_days):
        return "DUE_SOON"
    return ""


def _annotate_orders_due_status(orders, today=None):
    today = today or date.today()
    for order in orders or []:
        order["due_status"] = _due_status(order.get("due_date"), today=today)
    return orders


def _city_abbr(value):
    if not value:
        return ""
    parts = [part for part in str(value).replace("-", " ").split() if part]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0][:3].upper()
    abbr = "".join(part[0] for part in parts)
    return abbr[:3].upper()


def _utilization_grade(utilization_pct):
    thresholds = _get_utilization_grade_thresholds()
    if utilization_pct >= thresholds["A"]:
        return "A"
    if utilization_pct >= thresholds["B"]:
        return "B"
    if utilization_pct >= thresholds["C"]:
        return "C"
    if utilization_pct >= thresholds["D"]:
        return "D"
    return "F"


def _normalize_plant_code(value):
    return (value or "").strip().upper()


def _parse_plant_filters(value):
    if value is None:
        return None
    if isinstance(value, (list, tuple)):
        raw_values = value
    else:
        text = str(value).strip()
        if not text:
            return []
        if text.upper() == "ALL":
            return []
        raw_values = text.replace(",", " ").split()

    normalized = []
    for entry in raw_values:
        code = _normalize_plant_code(entry)
        if not code:
            continue
        if code == "ALL":
            return []
        if code not in normalized:
            normalized.append(code)
    return normalized


def _parse_strategic_customers(value_text):
    return customer_rules.parse_strategic_customers(value_text)


def _line_category_for_trailer_rules(line, sku_specs):
    sku = (line or {}).get("sku")
    spec = (sku_specs or {}).get(sku) if sku else None
    for value in (
        (spec or {}).get("category"),
        (line or {}).get("category"),
        (line or {}).get("bin"),
    ):
        text = str(value or "").strip().upper()
        if text:
            return text
    return ""


def _line_max_unit_length_for_trailer_rules(line, sku_specs):
    sku = (line or {}).get("sku")
    spec = (sku_specs or {}).get(sku) if sku else None
    lengths = [
        (line or {}).get("unit_length_ft"),
        (line or {}).get("length_with_tongue_ft"),
        (spec or {}).get("length_with_tongue_ft"),
    ]
    max_length = 0.0
    for value in lengths:
        parsed = _coerce_non_negative_float(value, 0.0)
        max_length = max(max_length, parsed)
    return max_length


def _auto_trailer_rule_annotation(
    load,
    lines,
    trailer_type,
    schematic,
    sku_specs,
    stop_sequence_map=None,
    assumptions=None,
    trailer_assignment_rules=None,
    strategic_customers=None,
):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    rules = trailer_assignment_rules or _get_trailer_assignment_rules()
    strategic_entries = strategic_customers if strategic_customers is not None else []
    exceeds_capacity = bool((schematic or {}).get("exceeds_capacity"))
    build_source = str((load or {}).get("build_source") or "").strip().upper()
    load_status = str((load or {}).get("status") or "").strip().upper()
    is_manual_build = build_source == "MANUAL"
    is_active_draft = load_status in {STATUS_PROPOSED, STATUS_DRAFT, ""}

    flatbed_label = "48' Flatbed" if trailer_key == "FLATBED_48" else "53' Flatbed"
    if trailer_key.startswith("FLATBED"):
        step_schematic, _, _ = _calculate_load_schematic(
            lines,
            sku_specs,
            "STEP_DECK",
            stop_sequence_map=stop_sequence_map,
            assumptions=assumptions,
        )
        if step_schematic.get("exceeds_capacity"):
            return (
                "Auto Trailer Rule",
                f"Changed from 53' Step Deck to {flatbed_label} because the load does not fit on the 43' / 10' split.",
            )

    if trailer_key == "HOTSHOT" and not exceeds_capacity and bool(rules.get("auto_assign_hotshot_enabled", True)):
        return (
            "Auto Trailer Rule",
            "Changed from 53' Step Deck to 40' Hotshot because the load fits on a hotshot trailer.",
        )

    if (
        (trailer_key.startswith("FLATBED") or trailer_key == "HOTSHOT")
        and is_active_draft
        and not is_manual_build
    ):
        return (
            "Auto Trailer Rule",
            "Trailer was auto-selected by draft fit/rule logic for this load.",
        )

    if trailer_key != "WEDGE":
        return "", ""

    categories = {
        _line_category_for_trailer_rules(line, sku_specs)
        for line in (lines or [])
    }
    categories.discard("")
    max_unit_length_ft = max(
        (_line_max_unit_length_for_trailer_rules(line, sku_specs) for line in (lines or [])),
        default=0.0,
    )
    max_cargo_unit_length_ft = max(
        (
            _line_max_unit_length_for_trailer_rules(line, sku_specs)
            for line in (lines or [])
            if (
                _line_category_for_trailer_rules(line, sku_specs) == "CARGO"
                or _line_category_for_trailer_rules(line, sku_specs).startswith("CARGO-")
            )
        ),
        default=0.0,
    )
    max_uta_unit_length_ft = max(
        (
            _line_max_unit_length_for_trailer_rules(line, sku_specs)
            for line in (lines or [])
            if (
                _line_category_for_trailer_rules(line, sku_specs) == "UTA"
                or _line_category_for_trailer_rules(line, sku_specs).startswith("UTA-")
            )
        ),
        default=0.0,
    )
    customer_names = {
        str((line or {}).get("cust_name") or "").strip()
        for line in (lines or [])
        if str((line or {}).get("cust_name") or "").strip()
    }
    matching_rules = [
        customer_rules.find_matching_strategic_customer(name, strategic_entries)
        for name in customer_names
    ]
    matching_rules = [entry for entry in matching_rules if entry]

    if any(bool((entry or {}).get("default_wedge_51")) for entry in matching_rules):
        return (
            "Auto Trailer Rule",
            "Changed from 53' Step Deck to 51' Wedge due to strategic customer wedge default.",
        )

    for entry in matching_rules:
        threshold_ft = _coerce_non_negative_float((entry or {}).get("wedge_min_item_length_ft"), 0.0)
        if threshold_ft > 0 and max_unit_length_ft >= threshold_ft:
            return (
                "Auto Trailer Rule",
                f"Changed from 53' Step Deck to 51' Wedge due to customer length rule ({threshold_ft:.0f} ft+).",
            )

    livestock_tokens = {
        str(token or "").strip().upper()
        for token in (rules.get("livestock_category_tokens") or [])
        if str(token or "").strip()
    }
    if bool(rules.get("livestock_wedge_enabled")) and categories.intersection(livestock_tokens):
        return (
            "Auto Trailer Rule",
            "Changed from 53' Step Deck to 51' Wedge because load includes LIVESTOCK-category items.",
        )

    has_tractor_supply = any(customer_rules.is_tractor_supply_customer(name) for name in customer_names)
    tractor_supply_cargo_trigger = (
        max_cargo_unit_length_ft >= DEFAULT_TRACTOR_SUPPLY_CARGO_WEDGE_MIN_ITEM_LENGTH_FT
    )
    tractor_supply_uta_trigger = (
        max_uta_unit_length_ft >= DEFAULT_TRACTOR_SUPPLY_UTA_WEDGE_MIN_ITEM_LENGTH_FT
    )
    if has_tractor_supply and (tractor_supply_cargo_trigger or tractor_supply_uta_trigger):
        applied_rules = []
        if tractor_supply_cargo_trigger:
            applied_rules.append(
                f"CARGO {DEFAULT_TRACTOR_SUPPLY_CARGO_WEDGE_MIN_ITEM_LENGTH_FT:.0f}+ ft"
            )
        if tractor_supply_uta_trigger:
            applied_rules.append(
                f"UTA {DEFAULT_TRACTOR_SUPPLY_UTA_WEDGE_MIN_ITEM_LENGTH_FT:.0f}+ ft"
            )
        return (
            "Auto Trailer Rule",
            "Changed from 53' Step Deck to 51' Wedge for Tractor Supply "
            + " / ".join(applied_rules)
            + " units.",
        )

    return "", ""


def _parse_manual_so_nums(raw_text):
    text = str(raw_text or "").strip()
    if not text:
        return []
    parsed = []
    seen = set()
    for token in re.split(r"[\s,;]+", text):
        value = token.strip().strip("\"'")
        if not value:
            continue
        normalized = value.upper()
        if normalized in {"SO_NUM", "ORDER", "ORDERS", "ORDER#", "SO#"}:
            continue
        if value in seen:
            continue
        seen.add(value)
        parsed.append(value)
    return parsed


def _sku_category_lookup():
    lookup = {}
    for spec in db.list_sku_specs():
        sku_key = str(spec.get("sku") or "").strip().upper()
        if sku_key and sku_key not in lookup:
            lookup[sku_key] = str(spec.get("category") or "").strip()
    return lookup


def _line_order_category_token(line, sku_category_by_sku=None):
    sku_category_by_sku = sku_category_by_sku or {}
    sku_key = str((line or {}).get("sku") or "").strip().upper()
    sku_category = sku_category_by_sku.get(sku_key, "")
    if sku_category:
        return sku_category
    for key in ("category", "bin"):
        value = str((line or {}).get(key) or "").strip()
        if value:
            return value
    return ""


def _group_order_category_scope(lines, sku_category_by_sku=None):
    tokens = [
        _line_order_category_token(line, sku_category_by_sku=sku_category_by_sku)
        for line in (lines or [])
    ]
    return order_categories.order_category_scope_from_tokens(tokens)


def _scope_snapshot_base(mode, selected_order_category_scope):
    return {
        "mode": mode,
        "order_category_scope": order_categories.normalize_order_category_scope(
            selected_order_category_scope,
            default=order_categories.ORDER_CATEGORY_SCOPE_ALL,
        ),
        "category_counts": order_categories.empty_category_counts(),
        "orders_in_scope": 0,
        "lines_in_scope": 0,
    }


def _manual_order_scope_snapshot(
    raw_text,
    origin_plant=None,
    allowed_plants=None,
    order_category_scope=order_categories.ORDER_CATEGORY_SCOPE_ALL,
):
    parsed_so_nums = _parse_manual_so_nums(raw_text)
    snapshot = {
        **_scope_snapshot_base("manual", order_category_scope),
        "pasted_count": len(parsed_so_nums),
        "matched_count": 0,
        "matched_in_scope_count": 0,
        "not_found_count": 0,
        "matched_so_nums": [],
        "not_found_so_nums": [],
        "matched_plant": "",
        "matched_plants": [],
    }
    if not parsed_so_nums:
        return snapshot

    allowed = {
        _normalize_plant_code(code)
        for code in (allowed_plants or _get_allowed_plants())
        if _normalize_plant_code(code)
    }
    origin = _normalize_plant_code(origin_plant)

    matched_so_nums = []
    matched_set = set()
    matched_plants = set()
    rows = db.list_orders_by_so_nums_any(parsed_so_nums, include_closed=False)
    for row in rows:
        so_num = str(row.get("so_num") or "").strip()
        plant_code = _normalize_plant_code(row.get("plant"))
        if (
            not so_num
            or so_num in matched_set
            or plant_code not in allowed
            or (origin and plant_code != origin)
            or _coerce_bool_value(row.get("is_excluded"))
        ):
            continue
        matched_so_nums.append(so_num)
        matched_set.add(so_num)
        matched_plants.add(plant_code)

    sku_category_by_sku = _sku_category_lookup()
    grouped_lines = {}
    if matched_set:
        matched_lines = db.list_order_lines_by_so_nums(list(matched_set))
        for line in matched_lines:
            plant_code = _normalize_plant_code(line.get("plant"))
            if (
                _coerce_bool_value(line.get("is_excluded"))
                or plant_code not in allowed
                or (origin and plant_code != origin)
            ):
                continue
            so_num = str(line.get("so_num") or "").strip()
            if not so_num:
                continue
            grouped_lines.setdefault(so_num, []).append(line)

    not_found = [so_num for so_num in parsed_so_nums if so_num not in matched_set]
    snapshot["matched_so_nums"] = matched_so_nums
    snapshot["matched_count"] = len(matched_so_nums)
    snapshot["not_found_so_nums"] = not_found
    snapshot["not_found_count"] = len(not_found)
    snapshot["matched_plants"] = sorted(matched_plants)
    snapshot["matched_plant"] = snapshot["matched_plants"][0] if len(snapshot["matched_plants"]) == 1 else ""

    selected_scope = snapshot["order_category_scope"]
    for so_num in matched_so_nums:
        lines = grouped_lines.get(so_num) or []
        if not lines:
            continue
        group_scope = _group_order_category_scope(lines, sku_category_by_sku=sku_category_by_sku)
        snapshot["category_counts"][group_scope] += 1
        if selected_scope != order_categories.ORDER_CATEGORY_SCOPE_ALL and group_scope != selected_scope:
            continue
        snapshot["orders_in_scope"] += 1
        snapshot["lines_in_scope"] += len(lines)

    snapshot["matched_in_scope_count"] = snapshot["orders_in_scope"]
    return snapshot


def _auto_order_scope_snapshot(
    origin_plant,
    state_filters=None,
    customer_filters=None,
    orders_start_date=None,
    batch_end_date=None,
    batch_horizon_enabled=False,
    order_category_scope=order_categories.ORDER_CATEGORY_SCOPE_ALL,
):
    origin = _normalize_plant_code(origin_plant)
    snapshot = _scope_snapshot_base("auto", order_category_scope)
    if not origin:
        return snapshot

    min_due_date = _parse_date(orders_start_date)
    min_due_iso = min_due_date.strftime("%Y-%m-%d") if min_due_date else None
    batch_due_date = _parse_date(batch_end_date) if batch_horizon_enabled else None

    state_set = {
        str(value or "").strip().upper()
        for value in (state_filters or [])
        if str(value or "").strip()
    }
    customer_set = {
        str(value or "").strip().casefold()
        for value in (customer_filters or [])
        if str(value or "").strip()
    }

    order_lines = db.list_order_lines_for_optimization(origin, min_due_date=min_due_iso)
    if not order_lines:
        return snapshot

    sku_category_by_sku = _sku_category_lookup()
    summary_rows = db.list_orders_for_optimization(origin)
    summary_by_so_num = {
        str(row.get("so_num") or "").strip(): row
        for row in summary_rows
        if str(row.get("so_num") or "").strip()
    }

    grouped_lines = {}
    for line in order_lines:
        so_num = str(line.get("so_num") or "").strip()
        if not so_num:
            continue
        grouped_lines.setdefault(so_num, []).append(line)

    selected_scope = snapshot["order_category_scope"]
    for so_num, lines in grouped_lines.items():
        if not lines:
            continue
        summary = summary_by_so_num.get(so_num) or {}
        state_value = str(summary.get("state") or lines[0].get("state") or "").strip().upper()
        if state_set and state_value not in state_set:
            continue
        customer_value = str(summary.get("cust_name") or lines[0].get("cust_name") or "").strip().casefold()
        if customer_set and customer_value not in customer_set:
            continue
        if batch_due_date:
            due_date_value = _parse_date(summary.get("due_date") or lines[0].get("due_date"))
            if due_date_value and due_date_value > batch_due_date:
                continue
        group_scope = _group_order_category_scope(lines, sku_category_by_sku=sku_category_by_sku)
        snapshot["category_counts"][group_scope] += 1
        if selected_scope != order_categories.ORDER_CATEGORY_SCOPE_ALL and group_scope != selected_scope:
            continue
        snapshot["orders_in_scope"] += 1
        snapshot["lines_in_scope"] += len(lines)
    return snapshot


def _sku_is_planner_input(spec):
    source = (spec.get("source") or "").strip().lower()
    if source == "planner":
        return True
    if source == "system":
        return False
    return bool((spec.get("added_at") or "").strip())


def _sku_source_label(spec):
    return "Planner Input" if _sku_is_planner_input(spec) else "Cheat Sheet"


def _source_led_source_text(spec):
    if not isinstance(spec, dict):
        return ""
    base = "Manually inputted on order entry" if _sku_is_planner_input(spec) else "Came from cheat sheet"
    updated_by = str(spec.get("updated_by") or "").strip()
    updated_at_raw = str(spec.get("updated_at") or "").strip()
    if not updated_at_raw:
        return base
    try:
        parsed = datetime.fromisoformat(updated_at_raw.replace("Z", "+00:00"))
    except ValueError:
        parsed_date = _parse_date(updated_at_raw)
        if not parsed_date:
            return base
        timestamp_label = parsed_date.strftime("%m/%d/%y")
    else:
        timestamp_label = parsed.strftime("%m/%d/%y")
    if updated_by:
        return f"{base} | Updated by {updated_by} on {timestamp_label}"
    return f"{base} | Updated on {timestamp_label}"


def _build_source_led_cheat_sheet_rows(specs):
    spec_by_sku = {}
    for spec in specs or []:
        sku_key = (spec.get("sku") or "").strip().upper()
        if sku_key and sku_key not in spec_by_sku:
            spec_by_sku[sku_key] = spec

    importer = OrderImporter()
    rows = []
    with db.get_connection() as connection:
        observed = connection.execute(
            """
            WITH totals AS (
                SELECT
                    UPPER(TRIM(COALESCE(plant, ''))) AS plant,
                    UPPER(TRIM(COALESCE(bin, ''))) AS bin_raw,
                    UPPER(TRIM(COALESCE(item, ''))) AS item_num,
                    COUNT(*) AS line_count,
                    COUNT(DISTINCT COALESCE(NULLIF(TRIM(so_num), ''), printf('ROW_%d', id))) AS order_count
                FROM order_lines
                WHERE TRIM(COALESCE(item, '')) <> ''
                  AND TRIM(COALESCE(plant, '')) <> ''
                GROUP BY 1, 2, 3
            ),
            ranked_desc AS (
                SELECT
                    UPPER(TRIM(COALESCE(plant, ''))) AS plant,
                    UPPER(TRIM(COALESCE(bin, ''))) AS bin_raw,
                    UPPER(TRIM(COALESCE(item, ''))) AS item_num,
                    TRIM(COALESCE(item_desc, '')) AS item_desc,
                    COUNT(*) AS desc_count,
                    ROW_NUMBER() OVER (
                        PARTITION BY
                            UPPER(TRIM(COALESCE(plant, ''))),
                            UPPER(TRIM(COALESCE(bin, ''))),
                            UPPER(TRIM(COALESCE(item, '')))
                        ORDER BY COUNT(*) DESC, TRIM(COALESCE(item_desc, '')) ASC
                    ) AS rn
                FROM order_lines
                WHERE TRIM(COALESCE(item, '')) <> ''
                  AND TRIM(COALESCE(plant, '')) <> ''
                GROUP BY 1, 2, 3, 4
            )
            SELECT
                t.plant,
                t.bin_raw,
                t.item_num,
                COALESCE(d.item_desc, '') AS item_desc,
                t.line_count,
                t.order_count
            FROM totals t
            LEFT JOIN ranked_desc d
                ON d.plant = t.plant
               AND d.bin_raw = t.bin_raw
               AND d.item_num = t.item_num
               AND d.rn = 1
            ORDER BY t.item_num ASC, t.plant ASC, t.bin_raw ASC
            """
        ).fetchall()

    for row in observed:
        plant = (row["plant"] or "").strip().upper()
        bin_raw = (row["bin_raw"] or "").strip().upper()
        item_num = (row["item_num"] or "").strip().upper()
        item_desc = (row["item_desc"] or "").strip()
        bin_code = importer._extract_code(bin_raw)
        mapped_sku = importer.lookup_sku(
            item_num,
            plant=plant,
            bin_code=bin_code,
            bin_raw=bin_raw,
        )

        spec = None
        if mapped_sku:
            spec = spec_by_sku.get(str(mapped_sku).strip().upper())

        category = (spec.get("category") or "").strip() if spec else ""
        mapped_description = ""
        length_with_tongue_ft = None
        max_stack_step_deck = None
        max_stack_flat_bed = None
        mapped_source = ""

        if spec:
            mapped_description = (spec.get("description") or spec.get("notes") or "").strip()
            length_with_tongue_ft = spec.get("length_with_tongue_ft")
            max_stack_step_deck = spec.get("max_stack_step_deck")
            max_stack_flat_bed = spec.get("max_stack_flat_bed")
            mapped_source = _source_led_source_text(spec)
            mapping_status = "Mapped"
        elif mapped_sku:
            cargo_length = importer._cargo_length_from_item(item_num, mapped_sku)
            if cargo_length:
                mapped_description = "Derived from cargo length rule"
                category = bin_code or bin_raw or "CARGO"
                length_with_tongue_ft = cargo_length
                max_stack_step_deck = 1
                max_stack_flat_bed = 1
                mapped_source = "Derived from cargo rule"
                mapping_status = "Mapped (Cargo Rule)"
            else:
                mapped_source = "Mapped SKU has no cheat sheet spec"
                mapping_status = "Mapped SKU Missing Spec"
        else:
            mapped_source = "Unmapped item"
            mapping_status = "Unmapped"

        rows.append(
            {
                "plant": plant,
                "bin": bin_raw,
                "item_num": item_num,
                "item_desc": item_desc,
                "mapped_sku": mapped_sku or "",
                "mapped_description": mapped_description,
                "category": category,
                "length_with_tongue_ft": length_with_tongue_ft,
                "max_stack_step_deck": max_stack_step_deck,
                "max_stack_flat_bed": max_stack_flat_bed,
                "mapped_source": mapped_source,
                "mapping_status": mapping_status,
                "line_count": int(row["line_count"] or 0),
                "order_count": int(row["order_count"] or 0),
            }
        )

    collapsed = {}
    for row in rows:
        key = (row.get("item_num") or "", row.get("bin") or "")
        group = collapsed.setdefault(
            key,
            {
                "item_num": row.get("item_num") or "",
                "bin": row.get("bin") or "",
                "line_count": 0,
                "order_count": 0,
                "desc_weights": {},
                "mapped_variants": {},
            },
        )
        group["line_count"] += int(row.get("line_count") or 0)
        group["order_count"] += int(row.get("order_count") or 0)

        item_desc = (row.get("item_desc") or "").strip()
        if item_desc:
            group["desc_weights"][item_desc] = group["desc_weights"].get(item_desc, 0) + int(
                row.get("line_count") or 0
            )

        mapped_sku = (row.get("mapped_sku") or "").strip()
        variant_key = mapped_sku or "__UNMAPPED__"
        variant = group["mapped_variants"].get(variant_key)
        if not variant:
            variant = {
                "mapped_sku": mapped_sku,
                "mapped_description": row.get("mapped_description") or "",
                "category": row.get("category") or "",
                "length_with_tongue_ft": row.get("length_with_tongue_ft"),
                "max_stack_step_deck": row.get("max_stack_step_deck"),
                "max_stack_flat_bed": row.get("max_stack_flat_bed"),
                "mapped_source": row.get("mapped_source") or "",
                "mapping_status": row.get("mapping_status") or "Unmapped",
                "weight": 0,
            }
            group["mapped_variants"][variant_key] = variant
        variant["weight"] += int(row.get("line_count") or 0)

    deduped_rows = []
    for group in collapsed.values():
        item_desc = ""
        if group["desc_weights"]:
            item_desc = max(
                group["desc_weights"].items(),
                key=lambda pair: (pair[1], pair[0]),
            )[0]

        mapped_variants = [
            payload
            for key, payload in group["mapped_variants"].items()
            if key != "__UNMAPPED__"
        ]

        if not mapped_variants:
            row = {
                "item_num": group["item_num"],
                "item_desc": item_desc,
                "bin": group["bin"],
                "mapped_sku": "",
                "mapped_sku_list": [],
                "mapped_description": "",
                "category": "",
                "length_with_tongue_ft": None,
                "max_stack_step_deck": None,
                "max_stack_flat_bed": None,
                "mapped_source": "",
                "mapping_status": "Unmapped",
                "mapping_conflict": False,
                "can_edit_specs": False,
                "line_count": group["line_count"],
                "order_count": group["order_count"],
            }
            deduped_rows.append(row)
            continue

        if len(mapped_variants) == 1:
            winner = mapped_variants[0]
            row = {
                "item_num": group["item_num"],
                "item_desc": item_desc,
                "bin": group["bin"],
                "mapped_sku": winner.get("mapped_sku") or "",
                "mapped_sku_list": [winner.get("mapped_sku")] if winner.get("mapped_sku") else [],
                "mapped_description": winner.get("mapped_description") or "",
                "category": winner.get("category") or "",
                "length_with_tongue_ft": winner.get("length_with_tongue_ft"),
                "max_stack_step_deck": winner.get("max_stack_step_deck"),
                "max_stack_flat_bed": winner.get("max_stack_flat_bed"),
                "mapped_source": winner.get("mapped_source") or "",
                "mapping_status": winner.get("mapping_status") or "Mapped",
                "mapping_conflict": False,
                "can_edit_specs": (
                    bool((winner.get("mapped_sku") or "").strip())
                    and (winner.get("mapping_status") or "").strip() == "Mapped"
                ),
                "line_count": group["line_count"],
                "order_count": group["order_count"],
            }
            deduped_rows.append(row)
            continue

        mapped_skus = sorted(
            [
                variant.get("mapped_sku")
                for variant in mapped_variants
                if (variant.get("mapped_sku") or "").strip()
            ]
        )
        conflict_label = ", ".join(mapped_skus[:3])
        if len(mapped_skus) > 3:
            conflict_label += f" +{len(mapped_skus) - 3} more"
        deduped_rows.append(
            {
                "item_num": group["item_num"],
                "item_desc": item_desc,
                "bin": group["bin"],
                "mapped_sku": conflict_label,
                "mapped_sku_list": mapped_skus,
                "mapped_description": "Multiple mapped SKUs across plants. Review lookup scope rules.",
                "category": "",
                "length_with_tongue_ft": None,
                "max_stack_step_deck": None,
                "max_stack_flat_bed": None,
                "mapped_source": "",
                "mapping_status": "Plant Mapping Conflict",
                "mapping_conflict": True,
                "can_edit_specs": False,
                "line_count": group["line_count"],
                "order_count": group["order_count"],
            }
        )

    deduped_rows.sort(key=lambda row: (row.get("item_num") or "", row.get("bin") or ""))
    return deduped_rows


def _sync_legacy_plant_filter(selected, allowed):
    if not selected or len(selected) >= len(allowed):
        session["plant_filter"] = "ALL"
    elif len(selected) == 1:
        session["plant_filter"] = selected[0]
    else:
        session["plant_filter"] = ",".join(selected)


def _get_session_profile_name():
    name = (session.get(SESSION_PROFILE_NAME_KEY) or "").strip()
    return name or None


def _profile_allowed_plants(profile):
    parsed = _parse_plant_filters(profile.get("allowed_plants"))
    # In profile storage, empty/ALL means "all plants".
    if not parsed:
        return list(PLANT_CODES)
    return [code for code in parsed if code in PLANT_CODES]


def _profile_default_plants(profile, allowed):
    parsed = _parse_plant_filters(profile.get("default_plants"))
    # Empty/ALL means "default to all allowed plants". Represent as empty list.
    if not parsed:
        return []
    filtered = [code for code in parsed if code in allowed]
    if not filtered or len(filtered) >= len(allowed):
        return []
    return filtered


def _profile_focus_plants(profile):
    allowed = _profile_allowed_plants(profile)
    defaults = _profile_default_plants(profile, allowed)
    return defaults or allowed or list(PLANT_CODES)


def _apply_profile_to_session(profile, *, reset_filters=False):
    allowed = _profile_allowed_plants(profile)
    default_plants = _profile_default_plants(profile, allowed)
    is_sandbox_profile = bool(profile.get("is_sandbox"))

    session[SESSION_PROFILE_ID_KEY] = profile["id"]
    session[SESSION_PROFILE_NAME_KEY] = profile["name"]
    session[SESSION_PROFILE_DEFAULT_PLANTS_KEY] = default_plants
    session[SESSION_PROFILE_SANDBOX_KEY] = is_sandbox_profile
    session.pop(SESSION_ACTIVE_PLANNING_ID_KEY, None)

    session["role"] = ROLE_ADMIN if profile.get("is_admin") else ROLE_PLANNER
    session["allowed_plants"] = allowed

    if reset_filters or session.get("plant_filters") is None:
        session["plant_filters"] = list(default_plants)
        _sync_legacy_plant_filter(session["plant_filters"], allowed)


def _ensure_active_profile():
    profile_id = session.get(SESSION_PROFILE_ID_KEY)
    profile = db.get_access_profile(profile_id) if profile_id else None
    if not profile:
        return None

    # Apply if missing/mismatched or if allowed plants were cleared.
    if (
        session.get(SESSION_PROFILE_ID_KEY) != profile["id"]
        or not _get_allowed_plants()
        or SESSION_PROFILE_SANDBOX_KEY not in session
        or bool(session.get(SESSION_PROFILE_SANDBOX_KEY)) != bool(profile.get("is_sandbox"))
    ):
        _apply_profile_to_session(profile, reset_filters=True)
    return profile


def _get_session_role():
    role = session.get("role")
    return role if role in {ROLE_ADMIN, ROLE_PLANNER} else None


def _is_session_sandbox():
    return bool(session.get(SESSION_PROFILE_SANDBOX_KEY))


def _get_active_planning_session_id():
    value = session.get(SESSION_ACTIVE_PLANNING_ID_KEY)
    try:
        return int(value) if value else None
    except (TypeError, ValueError):
        return None


def _set_active_planning_session_id(session_id):
    if session_id:
        session[SESSION_ACTIVE_PLANNING_ID_KEY] = int(session_id)
    else:
        session.pop(SESSION_ACTIVE_PLANNING_ID_KEY, None)


def _get_allowed_plants():
    allowed = session.get("allowed_plants") or []
    return [code for code in allowed if code in PLANT_CODES]


def _resolve_plant_filters(selected):
    role = _get_session_role()
    allowed = _get_allowed_plants()
    if not role or not allowed:
        return []

    parsed = _parse_plant_filters(selected)
    if parsed is not None:
        filtered = [code for code in parsed if code in allowed]
        if not filtered or len(filtered) >= len(allowed):
            session["plant_filters"] = []
        else:
            session["plant_filters"] = filtered

    current = session.get("plant_filters")
    if current is None:
        legacy = session.get("plant_filter")
        if legacy and legacy != "ALL":
            current = [legacy] if legacy in allowed else []
        else:
            current = []
        session["plant_filters"] = current

    current = [code for code in current if code in allowed]
    if not current or len(current) >= len(allowed):
        session["plant_filters"] = []
        _sync_legacy_plant_filter([], allowed)
        return []

    session["plant_filters"] = current
    _sync_legacy_plant_filter(current, allowed)
    return current


def _get_current_plant_filters():
    allowed = _get_allowed_plants()
    current = session.get("plant_filters")
    if current is None:
        legacy = session.get("plant_filter")
        if legacy and legacy != "ALL":
            current = [legacy] if legacy in allowed else []
        else:
            current = []
    current = [code for code in current if code in allowed]
    if not current or len(current) >= len(allowed):
        return []
    return current


def _format_plant_filter_label(selected, allowed):
    if not selected or len(selected) >= len(allowed):
        return "All Plants"
    return ", ".join(selected)


def _build_plant_filter_cards(allowed, selected):
    if not allowed:
        return []
    selected_set = set(selected or [])
    util_by_plant = {}
    with db.get_connection() as connection:
        rows = connection.execute(
            """
            SELECT origin_plant, AVG(utilization_pct) AS avg_util
            FROM loads
            WHERE status = 'APPROVED'
              AND DATE(created_at) >= DATE('now', '-6 days')
            GROUP BY origin_plant
            """
        ).fetchall()
        for row in rows:
            util_by_plant[row["origin_plant"]] = row["avg_util"]

    cards = []
    for plant in allowed:
        avg_util = util_by_plant.get(plant)
        util_value = round(avg_util, 0) if avg_util is not None else None
        util_width = min(max(int(util_value), 0), 100) if util_value is not None else 0
        status = "neutral"
        if util_value is not None:
            if util_value >= 80:
                status = "success"
            elif util_value >= 70:
                status = "warning"
            else:
                status = "error"
        cards.append(
            {
                "code": plant,
                "name": PLANT_NAMES.get(plant, plant),
                "utilization": util_value,
                "util_width": util_width,
                "util_display": f"{int(util_value)}%" if util_value is not None else "--",
                "status": status,
                "selected": plant in selected_set,
            }
        )
    return cards


def _week_bounds(today):
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    next_start = end + timedelta(days=1)
    next_end = next_start + timedelta(days=6)
    return start, end, next_start, next_end


def _default_batch_end_date(today=None):
    """Default planning horizon: Friday of the week that is two weeks from today."""
    today = today or date.today()
    anchor = today + timedelta(days=14)
    # Friday = 4 (Mon=0 ... Sun=6)
    return anchor + timedelta(days=(4 - anchor.weekday()))


def _sync_planning_session_status(session_id, loads=None):
    if not session_id:
        return None
    planning_session = db.get_planning_session(session_id)
    if not planning_session:
        return None
    current_status = (planning_session.get("status") or "").upper()
    if current_status in {"ARCHIVED", "LEGACY"}:
        return current_status
    if loads is None:
        next_status = db.compute_planning_session_status(session_id) or "DRAFT"
    else:
        total = len(loads)
        approved = sum(
            1 for load in loads if (load.get("status") or "").upper() == STATUS_APPROVED
        )
        next_status = "COMPLETED" if total > 0 and total == approved else "DRAFT"
    if next_status and next_status != current_status:
        db.update_planning_session_status(session_id, next_status)
        planning_session["status"] = next_status
    return planning_session.get("status") or next_status


def _normalize_session_status(value):
    status = (value or "").upper()
    if not status or status == "ACTIVE":
        return "DRAFT"
    return status


def _slug_session_label(value):
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").strip())
    return cleaned.strip("_").upper() or "PLANNER"


def _serialize_session_config(form_data, params):
    return json.dumps(
        {
            "origin_plant": params.get("origin_plant"),
            "capacity_feet": params.get("capacity_feet"),
            "trailer_type": params.get("trailer_type"),
            "max_detour_pct": params.get("max_detour_pct"),
            "time_window_days": params.get("time_window_days"),
            "geo_radius": params.get("geo_radius"),
            "stack_overflow_max_height": params.get("stack_overflow_max_height"),
            "max_back_overhang_ft": params.get("max_back_overhang_ft"),
            "upper_two_across_max_length_ft": params.get("upper_two_across_max_length_ft"),
            "upper_deck_exception_max_length_ft": params.get("upper_deck_exception_max_length_ft"),
            "upper_deck_exception_overhang_allowance_ft": params.get("upper_deck_exception_overhang_allowance_ft"),
            "upper_deck_exception_categories": params.get("upper_deck_exception_categories") or [],
            "enforce_time_window": params.get("enforce_time_window"),
            "batch_horizon_enabled": params.get("batch_horizon_enabled"),
            "batch_end_date": form_data.get("batch_end_date") or "",
            "orders_start_date": form_data.get("orders_start_date") or "",
            "ignore_due_date": bool(params.get("ignore_due_date")),
            "state_filters": params.get("state_filters") or [],
            "customer_filters": params.get("customer_filters") or [],
            "ignore_past_due": params.get("ignore_past_due"),
            "algorithm_version": params.get("algorithm_version") or "v2",
            "compare_algorithms": bool(params.get("compare_algorithms")),
            "optimize_mode": params.get("optimize_mode") or "auto",
            "manual_order_input": form_data.get("manual_order_input") or "",
            "selected_so_nums": params.get("selected_so_nums") or [],
            "order_category_scope": order_categories.normalize_order_category_scope(
                params.get("order_category_scope"),
                default=order_categories.ORDER_CATEGORY_SCOPE_ALL,
            ),
        }
    )


def _create_planning_session(created_by, plant_code, form_data, params):
    today_value = date.today()
    date_stamp = today_value.strftime("%m%d%Y")
    created_by = (created_by or "").strip() or "Planner"
    planner_slug = _slug_session_label(created_by)
    plant_code = (plant_code or "").strip().upper() or "ALL"
    sequence = db.count_planning_sessions_for_day(created_by, plant_code, today_value.strftime("%Y-%m-%d")) + 1
    session_code = f"{planner_slug}_{date_stamp}_{plant_code}_S{sequence}"
    while db.get_planning_session_by_code(session_code):
        sequence += 1
        session_code = f"{planner_slug}_{date_stamp}_{plant_code}_S{sequence}"
    horizon_end = form_data.get("batch_end_date") if form_data.get("batch_horizon_enabled") else None
    config_json = _serialize_session_config(form_data, params)
    return db.create_planning_session(
        session_code=session_code,
        plant_code=plant_code,
        created_by=created_by,
        config_json=config_json,
        horizon_end=horizon_end or None,
        status="DRAFT",
        is_sandbox=_is_session_sandbox(),
    )


def _find_resumable_planning_session(plant_code):
    if not plant_code:
        return None
    sessions = db.list_planning_sessions({"plant_code": plant_code})
    for entry in sessions:
        if not _can_access_planning_session(entry):
            continue
        status = _normalize_session_status(entry.get("status"))
        if status == "DRAFT" and (entry.get("load_count") or 0) > 0:
            return entry
    return None


def _is_full_truckload(load):
    utilization = load.get("utilization_pct") or 0
    order_numbers = {line.get("so_num") for line in load.get("lines", []) if line.get("so_num")}
    is_single_order = len(order_numbers) <= 1
    return is_single_order and (load.get("over_capacity") or utilization > 90)


def _period_range(period, today):
    key = (period or "").strip().lower()
    if key == "today":
        return today, today
    if key in {"this_month", "month_to_date"}:
        start = today.replace(day=1)
        return start, today
    if key in {"last_30_days", "last30", "last_30"}:
        return today - timedelta(days=29), today
    if key in {"last_7_days", "last7"}:
        return today - timedelta(days=6), today
    if key in {"ytd", "year_to_date"}:
        return date(today.year, 1, 1), today
    if key == "last_quarter":
        quarter = ((today.month - 1) // 3) + 1
        if quarter == 1:
            target_year = today.year - 1
            target_quarter = 4
        else:
            target_year = today.year
            target_quarter = quarter - 1

        start_month = ((target_quarter - 1) * 3) + 1
        start = date(target_year, start_month, 1)
        if start_month == 10:
            end = date(target_year, 12, 31)
        else:
            end = date(target_year, start_month + 3, 1) - timedelta(days=1)
        return start, end
    # default to this_week
    start, end, _, _ = _week_bounds(today)
    return start, end


def _build_load_thumbnail(load, sku_specs, stop_color_palette=None, max_blocks=4):
    lines = db.list_load_lines(load["id"])
    if not lines:
        return []

    trailer_type = stack_calculator.normalize_trailer_type(load.get("trailer_type"), default="STEP_DECK")
    zip_coords = geo_utils.load_zip_coordinates()
    ordered_stops = _ordered_stops_for_lines(lines, load.get("origin_plant"), zip_coords)
    ordered_stops = _apply_load_route_direction(ordered_stops, load=load)
    stop_sequence_map = _stop_sequence_map_from_ordered_stops(ordered_stops)
    order_colors = _build_order_colors_for_lines(
        lines,
        stop_sequence_map=stop_sequence_map,
        stop_palette=stop_color_palette,
    )
    schematic, line_items, _ = _calculate_load_schematic(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
    )
    positions = []
    for pos in schematic.get("positions", []) or []:
        colors = []
        for item in pos.get("items", []) or []:
            color = order_colors.get(item.get("order_id"), "#30363D")
            units = item.get("units") or 0
            for _ in range(min(units, max_blocks)):
                colors.append(color)
                if len(colors) >= max_blocks:
                    break
            if len(colors) >= max_blocks:
                break
        positions.append(
            {
                "width_pct": pos.get("width_pct") or 0,
                "colors": colors,
            }
        )
    return positions


def _line_stop_key(state, zip_code):
    state_value = (state or "").strip().upper()
    raw_zip = (zip_code or "").strip()
    normalized_zip = geo_utils.normalize_zip(raw_zip)
    zip_value = normalized_zip or raw_zip
    return f"{state_value}|{zip_value}"


def _normalize_order_identifier(value, fallback=""):
    text = str(value or "").strip()
    return text if text else str(fallback or "").strip()


def _stop_sequence_map_from_ordered_stops(ordered_stops):
    sequence = {}
    for idx, stop in enumerate(ordered_stops or [], start=1):
        key = _line_stop_key(stop.get("state"), stop.get("zip"))
        if key and key not in sequence:
            sequence[key] = idx
    return sequence


def _normalize_hex_color(value, fallback):
    fallback_text = str(fallback or FALLBACK_STOP_COLOR).strip().upper()
    if not HEX_COLOR_PATTERN.fullmatch(fallback_text):
        fallback_text = FALLBACK_STOP_COLOR
    raw = str(value or "").strip().upper()
    if not raw:
        return fallback_text
    if not raw.startswith("#"):
        raw = f"#{raw}"
    if HEX_COLOR_PATTERN.fullmatch(raw):
        return raw
    return fallback_text


def _get_stop_color_palette():
    defaults = list(DEFAULT_STOP_COLOR_PALETTE)
    setting = _get_effective_planning_setting(STOP_COLOR_PALETTE_SETTING_KEY)
    raw_value = (setting.get("value_text") or "").strip()
    if not raw_value:
        return defaults
    try:
        parsed = json.loads(raw_value)
    except json.JSONDecodeError:
        return defaults

    if isinstance(parsed, list):
        normalized = []
        for idx, default_color in enumerate(defaults):
            source = parsed[idx] if idx < len(parsed) else default_color
            normalized.append(_normalize_hex_color(source, default_color))
        return normalized

    if isinstance(parsed, dict):
        normalized = []
        for idx, default_color in enumerate(defaults, start=1):
            source = parsed.get(str(idx))
            if source is None:
                source = parsed.get(idx)
            normalized.append(_normalize_hex_color(source, default_color))
        return normalized

    return defaults


def _color_for_stop_sequence(stop_sequence, stop_palette=None):
    palette = list(stop_palette or _get_stop_color_palette())
    if not palette:
        return FALLBACK_STOP_COLOR
    sequence = _coerce_int_value(stop_sequence, 0)
    if sequence <= 0:
        return FALLBACK_STOP_COLOR
    return palette[(sequence - 1) % len(palette)]


def _build_order_colors_for_lines(lines, stop_sequence_map=None, stop_palette=None):
    palette = list(stop_palette or _get_stop_color_palette())
    order_ids = []
    order_stop_map = {}
    for line in lines or []:
        order_id = _normalize_order_identifier(line.get("so_num"))
        if not order_id:
            continue
        if order_id not in order_ids:
            order_ids.append(order_id)
        stop_sequence = None
        if stop_sequence_map:
            stop_sequence = stop_sequence_map.get(
                _line_stop_key(line.get("state"), line.get("zip"))
            )
        stop_value = _coerce_int_value(stop_sequence, 0)
        if stop_value <= 0:
            continue
        current = order_stop_map.get(order_id)
        if current is None or stop_value < current:
            order_stop_map[order_id] = stop_value

    order_colors = {}
    fallback_idx = 0
    for order_id in order_ids:
        stop_sequence = order_stop_map.get(order_id)
        if stop_sequence:
            order_colors[order_id] = _color_for_stop_sequence(stop_sequence, palette)
            continue
        if palette:
            order_colors[order_id] = palette[fallback_idx % len(palette)]
            fallback_idx += 1
        else:
            order_colors[order_id] = FALLBACK_STOP_COLOR
    return order_colors


def _requires_return_to_origin(lines):
    if not lines:
        return False
    strategic_setting = _get_effective_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY)
    strategic_customers = _parse_strategic_customers(strategic_setting.get("value_text") or "")
    return any(
        bool(
            (
                customer_rules.find_matching_strategic_customer(
                    (line or {}).get("cust_name") or "",
                    strategic_customers,
                )
                or {}
            ).get("requires_return_to_origin")
        )
        for line in (lines or [])
    )


def _ordered_stops_for_lines(lines, origin_plant, zip_coords, return_to_origin=None):
    stop_map = {}
    for line in lines or []:
        state = (line.get("state") or "").strip().upper()
        zip_code = (line.get("zip") or "").strip()
        key = _line_stop_key(state, zip_code)
        if key in stop_map:
            continue
        normalized_zip = geo_utils.normalize_zip(zip_code)
        coords = zip_coords.get(normalized_zip) if normalized_zip else None
        stop_map[key] = {
            "state": state,
            "zip": zip_code,
            "coords": coords,
        }
    stops = list(stop_map.values())
    origin_coords = geo_utils.plant_coords_for_code(origin_plant)
    if return_to_origin is None:
        return_to_origin = _requires_return_to_origin(lines)
    if origin_coords:
        route_result = routing_service.get_routing_service().build_route(
            origin_coords,
            stops,
            return_to_origin=bool(return_to_origin),
            objective="distance",
            include_geometry=False,
        )
        ordered_stops = route_result.get("ordered_stops") or []
        if ordered_stops:
            return ordered_stops
        return tsp_solver.solve_route(
            origin_coords,
            stops,
            return_to_origin=bool(return_to_origin),
        )
    return stops


def _is_load_route_reversed(load):
    return bool(_coerce_int_value((load or {}).get("route_reversed"), 0))


def _apply_load_route_direction(ordered_stops, load=None, reverse_route=None):
    stops = list(ordered_stops or [])
    if reverse_route is None:
        reverse_route = _is_load_route_reversed(load)
    if reverse_route and len(stops) > 1:
        stops.reverse()
    return stops


def _load_route_display_metrics(load, route_nodes, use_cached_route=True):
    expected_leg_count = max(len(route_nodes) - 1, 0)
    stored_legs = load.get("route_legs") or []
    route_legs = []
    use_cached_route = bool(use_cached_route)

    if use_cached_route and isinstance(stored_legs, list) and len(stored_legs) == expected_leg_count:
        for value in stored_legs:
            try:
                route_legs.append(round(float(value)))
            except (TypeError, ValueError):
                route_legs.append(None)
    else:
        for idx in range(expected_leg_count):
            origin_leg = route_nodes[idx].get("coords")
            dest_leg = route_nodes[idx + 1].get("coords")
            if origin_leg and dest_leg:
                route_legs.append(round(geo_utils.haversine_distance_coords(origin_leg, dest_leg)))
            else:
                route_legs.append(None)

    stored_total = load.get("route_total_miles")
    total_route_distance = None
    if use_cached_route:
        if stored_total is None:
            stored_total = load.get("route_distance")
        try:
            total_route_distance = float(stored_total) if stored_total is not None else None
        except (TypeError, ValueError):
            total_route_distance = None
    if total_route_distance is None:
        total_route_distance = sum(leg for leg in route_legs if leg is not None)

    stored_geometry = load.get("route_geometry") or []
    route_geometry = []
    if use_cached_route and isinstance(stored_geometry, list):
        for point in stored_geometry:
            if (
                isinstance(point, (list, tuple))
                and len(point) >= 2
                and point[0] is not None
                and point[1] is not None
            ):
                try:
                    route_geometry.append([float(point[0]), float(point[1])])
                except (TypeError, ValueError):
                    continue
    if not route_geometry:
        for node in route_nodes:
            coords = node.get("coords")
            if coords:
                route_geometry.append([float(coords[0]), float(coords[1])])

    return {
        "route_legs": route_legs,
        "route_distance": round(total_route_distance) if total_route_distance else 0,
        "route_geometry": route_geometry,
    }


def _build_route_stops_for_lines(lines, zip_coords):
    stops = []
    stop_map = {}
    for line in lines or []:
        zip_code = geo_utils.normalize_zip(line.get("zip"))
        state = (line.get("state") or "").strip().upper()
        key = f"{zip_code}|{state}"
        if key not in stop_map:
            coords = zip_coords.get(zip_code) if zip_code else None
            stop_map[key] = {
                "zip": zip_code,
                "state": state,
                "lat": coords[0] if coords else None,
                "lng": coords[1] if coords else None,
                "coords": coords,
            }
        if line.get("cust_name"):
            stop_map[key].setdefault("customers", set()).add(line.get("cust_name"))
        if line.get("city") and not stop_map[key].get("city"):
            stop_map[key]["city"] = line.get("city")

    for stop in stop_map.values():
        stops.append(
            {
                "zip": stop.get("zip"),
                "state": stop.get("state"),
                "city": stop.get("city") or "",
                "customers": sorted(stop.get("customers") or []),
                "lat": stop.get("lat"),
                "lng": stop.get("lng"),
                "coords": stop.get("coords"),
            }
        )
    return stops


def _build_schematic_line_items(lines, sku_specs, trailer_type, stop_sequence_map=None):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    is_step_deck = trailer_key.startswith("STEP_DECK")
    line_items = []
    for line in lines or []:
        sku = line.get("sku")
        spec = sku_specs.get(sku) if sku else None
        if is_step_deck:
            max_stack = (spec or {}).get("max_stack_step_deck") or (spec or {}).get("max_stack_flat_bed") or 1
            upper_max_stack = (spec or {}).get("max_stack_flat_bed") or max_stack or 1
        else:
            max_stack = (spec or {}).get("max_stack_flat_bed") or 1
            upper_max_stack = max_stack
        stop_sequence = None
        if stop_sequence_map:
            stop_sequence = stop_sequence_map.get(_line_stop_key(line.get("state"), line.get("zip")))
        line_items.append(
            {
                "item": line.get("item"),
                "item_desc": line.get("item_desc"),
                "sku": sku,
                "qty": line.get("qty") or 0,
                "unit_length_ft": line.get("unit_length_ft") or 0,
                "max_stack_height": max_stack,
                "upper_deck_max_stack_height": upper_max_stack,
                "category": (spec or {}).get("category", ""),
                "order_id": _normalize_order_identifier(line.get("so_num")),
                "stop_sequence": stop_sequence,
            }
        )
    return line_items


def _calculate_load_schematic(
    lines,
    sku_specs,
    trailer_type,
    stop_sequence_map=None,
    assumptions=None,
):
    order_numbers = {
        _normalize_order_identifier(line.get("so_num"))
        for line in (lines or [])
        if _normalize_order_identifier(line.get("so_num"))
    }
    assumptions = assumptions or _get_stack_capacity_assumptions()
    line_items = _build_schematic_line_items(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
    )
    preserve_order_contiguity = len(order_numbers) <= 1
    schematic = stack_calculator.calculate_stack_configuration(
        line_items,
        trailer_type=trailer_type,
        preserve_order_contiguity=preserve_order_contiguity,
        stack_overflow_max_height=assumptions.get("stack_overflow_max_height"),
        max_back_overhang_ft=assumptions.get("max_back_overhang_ft"),
        upper_two_across_max_length_ft=assumptions.get("upper_two_across_max_length_ft"),
        upper_deck_exception_max_length_ft=assumptions.get("upper_deck_exception_max_length_ft"),
        upper_deck_exception_overhang_allowance_ft=assumptions.get("upper_deck_exception_overhang_allowance_ft"),
        upper_deck_exception_categories=assumptions.get("upper_deck_exception_categories"),
    )
    return schematic, line_items, order_numbers


def _coerce_float_value(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _coerce_int_value(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def _coerce_bool_value(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "on", "y"}


def _warning(code, message, position_id=None, deck=None):
    payload = {"code": code, "message": message, "severity": "warning"}
    if position_id:
        payload["position_id"] = position_id
    if deck:
        payload["deck"] = deck
    return payload


def _trailer_config_for_type(trailer_type):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    base = stack_calculator.TRAILER_CONFIGS.get(
        trailer_key,
        stack_calculator.TRAILER_CONFIGS["STEP_DECK"],
    )
    return {
        "type": trailer_key,
        "capacity": _coerce_float_value(base.get("capacity"), 53.0),
        "lower": _coerce_float_value(base.get("lower"), 53.0),
        "upper": _coerce_float_value(base.get("upper"), 0.0),
    }


def _build_schematic_units(lines, sku_specs, trailer_type, stop_sequence_map=None, order_colors=None):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    is_step_deck = trailer_key.startswith("STEP_DECK")
    order_colors = order_colors or {}
    units = []
    for line_idx, line in enumerate(lines or [], start=1):
        sku = line.get("sku")
        spec = sku_specs.get(sku) if sku else None
        if is_step_deck:
            max_stack = (
                (spec or {}).get("max_stack_step_deck")
                or (spec or {}).get("max_stack_flat_bed")
                or 1
            )
            upper_max_stack = (spec or {}).get("max_stack_flat_bed") or max_stack or 1
        else:
            max_stack = (spec or {}).get("max_stack_flat_bed") or 1
            upper_max_stack = max_stack
        max_stack = max(_coerce_int_value(max_stack, 1), 1)
        upper_max_stack = max(_coerce_int_value(upper_max_stack, max_stack), 1)
        qty = max(_coerce_int_value(line.get("qty"), 0), 0)
        if qty <= 0:
            continue
        unit_length_ft = _coerce_float_value(line.get("unit_length_ft"), 0.0)
        order_id = _normalize_order_identifier(line.get("so_num"))
        stop_sequence = None
        if stop_sequence_map:
            stop_sequence = stop_sequence_map.get(_line_stop_key(line.get("state"), line.get("zip")))
        order_line_id = line.get("order_line_id") or line.get("id") or line_idx
        fallback_color = _color_for_stop_sequence(stop_sequence)
        for unit_index in range(qty):
            unit_id = f"ol{order_line_id}-u{unit_index + 1}"
            units.append(
                {
                    "unit_id": unit_id,
                    "order_id": order_id,
                    "order_line_id": order_line_id,
                    "sku": sku,
                    "item": line.get("item"),
                    "item_desc": line.get("item_desc"),
                    "unit_length_ft": unit_length_ft,
                    "max_stack": max_stack,
                    "upper_max_stack": upper_max_stack,
                    "category": (spec or {}).get("category", ""),
                    "stop_sequence": stop_sequence,
                    "color": order_colors.get(order_id, fallback_color),
                }
            )
    return units


def _unit_item_key_from_unit(unit):
    return (
        unit.get("order_id") or "",
        unit.get("sku") or "",
        unit.get("item") or "",
        unit.get("item_desc") or "",
        round(_coerce_float_value(unit.get("unit_length_ft"), 0.0), 4),
        max(_coerce_int_value(unit.get("max_stack"), 1), 1),
        max(_coerce_int_value(unit.get("upper_max_stack"), unit.get("max_stack") or 1), 1),
        unit.get("category") or "",
    )


def _unit_item_key_from_schematic_item(item):
    return (
        item.get("order_id") or "",
        item.get("sku") or "",
        item.get("item") or "",
        item.get("item_desc") or "",
        round(_coerce_float_value(item.get("unit_length_ft"), 0.0), 4),
        max(_coerce_int_value(item.get("max_stack"), 1), 1),
        max(_coerce_int_value(item.get("upper_max_stack"), item.get("max_stack") or 1), 1),
        item.get("category") or "",
    )


def _layout_from_schematic(schematic, units):
    buckets = {}
    for unit in units:
        key = _unit_item_key_from_unit(unit)
        buckets.setdefault(key, []).append(unit["unit_id"])
    remaining_by_id = {unit["unit_id"]: unit for unit in units}
    positions = []

    def _pop_first_matching(order_id, sku):
        for unit_id, unit in list(remaining_by_id.items()):
            if (unit.get("order_id") or "") != (order_id or ""):
                continue
            if (unit.get("sku") or "") != (sku or ""):
                continue
            remaining_by_id.pop(unit_id, None)
            key = _unit_item_key_from_unit(unit)
            bucket = buckets.get(key) or []
            if unit_id in bucket:
                bucket.remove(unit_id)
            return unit_id
        return None

    for idx, pos in enumerate((schematic or {}).get("positions") or [], start=1):
        unit_ids = []
        for item in pos.get("items") or []:
            needed = max(_coerce_int_value(item.get("units"), 0), 0)
            if needed <= 0:
                continue
            item_key = _unit_item_key_from_schematic_item(item)
            bucket = buckets.get(item_key) or []
            while needed > 0 and bucket:
                unit_id = bucket.pop(0)
                if unit_id not in remaining_by_id:
                    continue
                remaining_by_id.pop(unit_id, None)
                unit_ids.append(unit_id)
                needed -= 1
            while needed > 0:
                fallback_id = _pop_first_matching(item.get("order_id"), item.get("sku"))
                if not fallback_id:
                    break
                unit_ids.append(fallback_id)
                needed -= 1
        if unit_ids:
            positions.append(
                {
                    "position_id": f"p{idx}",
                    "deck": (pos.get("deck") or "lower"),
                    "unit_ids": unit_ids,
                }
            )

    for unit_id in sorted(remaining_by_id.keys()):
        positions.append(
            {
                "position_id": f"p{len(positions) + 1}",
                "deck": "lower",
                "unit_ids": [unit_id],
            }
        )

    return {"positions": positions}


def _normalize_edit_layout(layout, units_by_id, trailer_type):
    config = _trailer_config_for_type(trailer_type)
    has_upper = config["upper"] > 0
    raw_positions = (layout or {}).get("positions") or []
    normalized_positions = []
    seen = set()

    for raw in raw_positions:
        deck = (raw.get("deck") or "lower").strip().lower()
        if deck not in {"lower", "upper"}:
            deck = "lower"
        if deck == "upper" and not has_upper:
            deck = "lower"
        raw_unit_ids = raw.get("unit_ids") or []
        unit_ids = []
        for value in raw_unit_ids:
            unit_id = str(value or "").strip()
            if not unit_id or unit_id not in units_by_id:
                continue
            if unit_id in seen:
                raise ValueError(f"Duplicate unit in layout: {unit_id}")
            seen.add(unit_id)
            unit_ids.append(unit_id)
        if unit_ids:
            normalized_positions.append(
                {
                    "position_id": "",
                    "deck": deck,
                    "unit_ids": unit_ids,
                }
            )

    expected_ids = set(units_by_id.keys())
    if expected_ids != seen:
        missing = sorted(expected_ids - seen)
        extra = sorted(seen - expected_ids)
        details = []
        if missing:
            details.append(f"missing units: {', '.join(missing[:5])}")
        if extra:
            details.append(f"unknown units: {', '.join(extra[:5])}")
        raise ValueError("Invalid layout payload (" + "; ".join(details) + ")")

    for idx, pos in enumerate(normalized_positions, start=1):
        pos["position_id"] = f"p{idx}"

    return {"positions": normalized_positions}


def _two_across_unit_group_key(unit):
    stop_sequence = _coerce_int_value((unit or {}).get("stop_sequence"), 0)
    if stop_sequence > 0:
        return f"stop:{stop_sequence}"
    order_id = str((unit or {}).get("order_id") or "").strip()
    if order_id:
        return f"order:{order_id}"
    sku = str((unit or {}).get("sku") or "").strip().upper()
    if sku:
        return f"sku:{sku}"
    return "misc"


def _upper_two_across_stack_sort_key(
    unit,
    unit_id="",
    equal_length_deck_length_order_enabled=True,
):
    length_ft = _coerce_float_value((unit or {}).get("unit_length_ft"), 0.0)
    deck_length_ft = stack_calculator.item_deck_length_ft(unit, fallback_length_ft=length_ft)
    stop_sequence = _coerce_int_value((unit or {}).get("stop_sequence"), 0)
    # Missing stop sequence is treated as latest/unknown so it remains lower in the stack.
    stop_rank = stop_sequence if stop_sequence > 0 else 999999
    order_id = str((unit or {}).get("order_id") or "").strip().upper()
    sku = str((unit or {}).get("sku") or "").strip().upper()
    return (
        -length_ft,
        (-deck_length_ft if equal_length_deck_length_order_enabled else 0.0),
        -stop_rank,
        order_id,
        sku,
        str(unit_id or ""),
    )


def _sort_upper_two_across_unit_ids(
    unit_ids,
    units_by_id,
    equal_length_deck_length_order_enabled=True,
):
    groups = []
    for raw_unit_id in unit_ids or []:
        unit_id = str(raw_unit_id or "").strip()
        if not unit_id or unit_id not in units_by_id:
            continue
        unit = units_by_id.get(unit_id) or {}
        group_key = _two_across_unit_group_key(unit)
        if groups and groups[-1]["group_key"] == group_key:
            groups[-1]["unit_ids"].append(unit_id)
            continue
        groups.append(
            {
                "group_key": group_key,
                "unit_ids": [unit_id],
            }
        )

    if not groups:
        return []

    for group in groups:
        group["unit_ids"].sort(
            key=lambda current_unit_id: _upper_two_across_stack_sort_key(
                units_by_id.get(current_unit_id) or {},
                current_unit_id,
                equal_length_deck_length_order_enabled=equal_length_deck_length_order_enabled,
            )
        )
        first_unit_id = (group.get("unit_ids") or [""])[0]
        group["sort_key"] = _upper_two_across_stack_sort_key(
            units_by_id.get(first_unit_id) or {},
            first_unit_id,
            equal_length_deck_length_order_enabled=equal_length_deck_length_order_enabled,
        )

    groups.sort(key=lambda group: group.get("sort_key") or ())
    ordered = []
    for group in groups:
        ordered.extend(group.get("unit_ids") or [])
    return ordered


def _auto_stack_upper_two_across_layout(layout, units_by_id, trailer_type, assumptions=None):
    assumptions = assumptions or _get_stack_capacity_assumptions()
    trailer_config = _trailer_config_for_type(trailer_type)
    if not trailer_config["type"].startswith("STEP_DECK"):
        return layout
    if _coerce_float_value(trailer_config.get("upper"), 0.0) <= 0.0:
        return layout

    source_positions = list((layout or {}).get("positions") or [])
    if not source_positions:
        return layout

    upper_two_across_max_length_ft = round(
        _coerce_non_negative_float(
            assumptions.get("upper_two_across_max_length_ft"),
            DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT,
        ),
        2,
    )
    equal_length_deck_length_order_enabled = (
        _coerce_bool_value(assumptions.get("equal_length_deck_length_order_enabled"))
        if assumptions.get("equal_length_deck_length_order_enabled") is not None
        else True
    )

    probe_positions = []
    for idx, raw_pos in enumerate(source_positions, start=1):
        deck = (raw_pos.get("deck") or "lower").strip().lower()
        if deck != "upper":
            continue
        position_id = raw_pos.get("position_id") or f"p{idx}"
        unit_ids = [
            str(value or "").strip()
            for value in (raw_pos.get("unit_ids") or [])
            if str(value or "").strip() in units_by_id
        ]
        if not unit_ids:
            continue
        units = [units_by_id[unit_id] for unit_id in unit_ids]
        length_ft = max(_coerce_float_value(unit.get("unit_length_ft"), 0.0) for unit in units)
        effective_units = []
        capacity_used = 0.0
        for unit in units:
            effective_max_stack = max(
                _coerce_int_value(
                    unit.get("upper_max_stack"),
                    unit.get("max_stack") or 1,
                ),
                1,
            )
            capacity_used += 1.0 / effective_max_stack
            adjusted = dict(unit)
            adjusted["max_stack"] = effective_max_stack
            effective_units.append(adjusted)

        probe_positions.append(
            {
                "position_id": position_id,
                "deck": "upper",
                "length_ft": round(length_ft, 2),
                "items": _aggregate_position_items(effective_units),
                "capacity_used": round(capacity_used, 4),
            }
        )

    if not probe_positions:
        return layout

    stack_calculator.apply_upper_usage_metadata(
        probe_positions,
        trailer_config,
        upper_two_across_max_length_ft,
    )
    two_across_position_ids = {
        str(position.get("position_id") or "").strip()
        for position in probe_positions
        if bool(position.get("two_across_applied"))
    }
    if not two_across_position_ids:
        return layout

    updated_positions = []
    changed = False
    for idx, raw_pos in enumerate(source_positions, start=1):
        pos = dict(raw_pos or {})
        position_id = str(pos.get("position_id") or f"p{idx}").strip()
        if position_id in two_across_position_ids:
            original_unit_ids = [
                str(value or "").strip()
                for value in (pos.get("unit_ids") or [])
                if str(value or "").strip() in units_by_id
            ]
            sorted_unit_ids = _sort_upper_two_across_unit_ids(
                original_unit_ids,
                units_by_id,
                equal_length_deck_length_order_enabled=equal_length_deck_length_order_enabled,
            )
            if sorted_unit_ids and sorted_unit_ids != original_unit_ids:
                pos["unit_ids"] = sorted_unit_ids
                changed = True
        updated_positions.append(pos)

    if not changed:
        return layout
    return {"positions": updated_positions}


def _aggregate_position_items(units):
    grouped = []
    for unit in units:
        unit_length_ft = _coerce_float_value(unit.get("unit_length_ft"), 0.0)
        deck_length_ft = stack_calculator.item_deck_length_ft(unit, fallback_length_ft=unit_length_ft)
        signature = (
            unit.get("item"),
            unit.get("sku"),
            unit.get("item_desc"),
            unit.get("category"),
            max(_coerce_int_value(unit.get("max_stack"), 1), 1),
            max(_coerce_int_value(unit.get("upper_max_stack"), unit.get("max_stack") or 1), 1),
            unit_length_ft,
            deck_length_ft,
            unit.get("order_id"),
            unit.get("stop_sequence"),
        )
        if grouped and grouped[-1]["signature"] == signature:
            grouped[-1]["units"] += 1
            continue
        grouped.append(
            {
                "signature": signature,
                "item": unit.get("item"),
                "sku": unit.get("sku"),
                "item_desc": unit.get("item_desc"),
                "category": unit.get("category"),
                "max_stack": max(_coerce_int_value(unit.get("max_stack"), 1), 1),
                "upper_max_stack": max(
                    _coerce_int_value(unit.get("upper_max_stack"), unit.get("max_stack") or 1),
                    1,
                ),
                "unit_length_ft": unit_length_ft,
                "deck_length_ft": deck_length_ft,
                "order_id": unit.get("order_id"),
                "stop_sequence": unit.get("stop_sequence"),
                "units": 1,
            }
        )
    for group in grouped:
        group.pop("signature", None)
    return grouped


def _unit_capacity_fraction(max_stack):
    return 1.0 / max(_coerce_int_value(max_stack, 1), 1)


def _position_allows_single_overflow(
    units,
    stack_overflow_max_height,
    max_stack_utilization_multiplier,
    capacity_used=None,
):
    threshold = _coerce_int_value(stack_overflow_max_height, 0)
    if threshold <= 0:
        return False
    if capacity_used is None:
        capacity_used = sum(
            _unit_capacity_fraction(unit.get("max_stack"))
            for unit in (units or [])
        )
    if capacity_used <= (1.0 + 1e-6):
        return False
    if capacity_used > (max_stack_utilization_multiplier + 1e-6):
        return False

    for idx, unit in enumerate(units or []):
        unit_max_stack = max(_coerce_int_value(unit.get("max_stack"), 1), 1)
        if unit_max_stack < threshold:
            continue
        if (capacity_used - _unit_capacity_fraction(unit_max_stack)) > (1.0 + 1e-6):
            continue
        base_heights = {
            max(_coerce_int_value(base_unit.get("max_stack"), 1), 1)
            for base_idx, base_unit in enumerate(units or [])
            if base_idx != idx
        }
        # Must already be a mixed-height base stack before adding overflow.
        if len(base_heights) >= 2:
            return True
    return False


def _build_schematic_from_layout(layout, units_by_id, trailer_type, assumptions=None):
    assumptions = assumptions or _get_stack_capacity_assumptions()
    stack_overflow_max_height = _coerce_non_negative_int(
        assumptions.get("stack_overflow_max_height"),
        DEFAULT_STACK_OVERFLOW_MAX_HEIGHT,
    )
    max_back_overhang_ft = round(
        _coerce_non_negative_float(
            assumptions.get("max_back_overhang_ft"),
            DEFAULT_MAX_BACK_OVERHANG_FT,
        ),
        2,
    )
    upper_two_across_max_length_ft = round(
        _coerce_non_negative_float(
            assumptions.get("upper_two_across_max_length_ft"),
            DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT,
        ),
        2,
    )
    upper_deck_exception_max_length_ft = round(
        _coerce_non_negative_float(
            assumptions.get("upper_deck_exception_max_length_ft"),
            DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
        ),
        2,
    )
    upper_deck_exception_overhang_allowance_ft = round(
        _coerce_non_negative_float(
            assumptions.get("upper_deck_exception_overhang_allowance_ft"),
            DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
        ),
        2,
    )
    upper_deck_exception_categories = stack_calculator.normalize_upper_deck_exception_categories(
        assumptions.get("upper_deck_exception_categories"),
        default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
    )
    equal_length_deck_length_order_enabled = (
        _coerce_bool_value(assumptions.get("equal_length_deck_length_order_enabled"))
        if assumptions.get("equal_length_deck_length_order_enabled") is not None
        else True
    )
    max_stack_utilization_multiplier = (
        1.0 + (1.0 / stack_overflow_max_height)
        if stack_overflow_max_height > 0
        else 1.0
    )
    trailer_config = _trailer_config_for_type(trailer_type)
    lower_length = trailer_config["lower"]
    upper_length = trailer_config["upper"]
    capacity = trailer_config["capacity"]
    has_upper = upper_length > 0

    positions = []
    warnings = []
    upper_deck_length_exceeds = []
    effective_units_by_position = {}

    def _effective_unit_max_stack(unit, deck_name):
        if deck_name == "upper" and trailer_config["type"].startswith("STEP_DECK"):
            return max(
                _coerce_int_value(
                    unit.get("upper_max_stack"),
                    unit.get("max_stack") or 1,
                ),
                1,
            )
        return max(_coerce_int_value(unit.get("max_stack"), 1), 1)

    for idx, raw_pos in enumerate((layout or {}).get("positions") or [], start=1):
        unit_ids = raw_pos.get("unit_ids") or []
        units = [units_by_id.get(unit_id) for unit_id in unit_ids if unit_id in units_by_id]
        units = [unit for unit in units if unit]
        if not units:
            continue
        length_ft = max(_coerce_float_value(unit.get("unit_length_ft"), 0.0) for unit in units)
        capacity_used = sum(
            1.0 / max(_coerce_int_value(unit.get("max_stack"), 1), 1)
            for unit in units
        )
        deck = (raw_pos.get("deck") or "lower").strip().lower()
        if deck not in {"lower", "upper"}:
            deck = "lower"
        if deck == "upper" and not has_upper:
            deck = "lower"
        position_id = raw_pos.get("position_id") or f"p{idx}"

        effective_units = []
        capacity_used = 0.0
        for unit in units:
            effective_max_stack = _effective_unit_max_stack(unit, deck)
            capacity_used += 1.0 / effective_max_stack
            adjusted = dict(unit)
            adjusted["max_stack"] = effective_max_stack
            effective_units.append(adjusted)
        effective_units_by_position[position_id] = effective_units

        if deck == "upper":
            position_probe = {
                "length_ft": length_ft,
                "items": [{"category": unit.get("category")} for unit in units],
            }
            upper_length_limit = stack_calculator.upper_deck_position_length_limit_ft(
                position_probe,
                trailer_config,
                upper_deck_exception_max_length_ft,
                upper_deck_exception_categories,
            )
            too_long_units = [
                unit
                for unit in units
                if _coerce_float_value(unit.get("unit_length_ft"), 0.0) > (upper_length_limit + 1e-6)
            ]
            if too_long_units:
                upper_deck_length_exceeds.append(
                    (position_id, upper_length_limit)
                )

        aggregated_items = _aggregate_position_items(units)
        if deck == "upper" and trailer_config["type"].startswith("STEP_DECK"):
            for item in aggregated_items:
                item["max_stack"] = max(
                    _coerce_int_value(
                        item.get("upper_max_stack"),
                        item.get("max_stack") or 1,
                    ),
                    1,
                )

        positions.append(
            {
                "position_id": position_id,
                "deck": deck,
                "length_ft": round(length_ft, 2),
                "items": aggregated_items,
                "capacity_used": round(capacity_used, 4),
                "overflow_applied": False,
                "overflow_note": None,
                "units_count": len(units),
                "top_stop_sequence": units[-1].get("stop_sequence"),
                "top_length_ft": _coerce_float_value(units[-1].get("unit_length_ft"), length_ft),
                "top_deck_length_ft": stack_calculator.item_deck_length_ft(
                    units[-1],
                    fallback_length_ft=_coerce_float_value(units[-1].get("unit_length_ft"), length_ft),
                ),
            }
        )

    upper_usage_meta = stack_calculator.apply_upper_usage_metadata(
        positions,
        trailer_config,
        upper_two_across_max_length_ft,
    )
    stack_index_by_position_id = stack_calculator.stack_display_index_map(
        positions,
        trailer_config=trailer_config,
    )

    for idx, (position_id, upper_length_limit) in enumerate(upper_deck_length_exceeds, start=1):
        stack_idx = int(stack_index_by_position_id.get(position_id, idx))
        warnings.append(
            _warning(
                "ITEM_TOO_BIG_FOR_UPPER_DECK",
                f"Stack {stack_idx}: one or more items exceed {upper_length_limit:.1f} ft upper deck limit.",
                position_id=position_id,
                deck="upper",
            )
        )

    for idx, pos in enumerate(positions, start=1):
        position_id = pos.get("position_id") or f"p{idx}"
        stack_idx = int(stack_index_by_position_id.get(position_id, idx))
        deck = (pos.get("deck") or "lower").strip().lower() or "lower"
        two_across_applied = bool(pos.get("two_across_applied")) and deck == "upper"
        capacity_used = _coerce_float_value(pos.get("capacity_used"), 0.0)
        effective_units = effective_units_by_position.get(position_id) or []
        overflow_applied = _position_allows_single_overflow(
            effective_units,
            stack_overflow_max_height=stack_overflow_max_height,
            max_stack_utilization_multiplier=max_stack_utilization_multiplier,
            capacity_used=capacity_used,
        )
        overflow_note = None
        if two_across_applied:
            pos["overflow_applied"] = False
            pos["overflow_note"] = None
            continue
        if capacity_used > (max_stack_utilization_multiplier + 1e-6):
            warnings.append(
                _warning(
                    "STACK_TOO_HIGH",
                    (
                        f"Stack {stack_idx} is {capacity_used * 100:.0f}% overfilled relative "
                        "to SKU-specific stacking maximums."
                    ),
                    position_id=position_id,
                    deck=deck,
                )
            )
        elif capacity_used > (1.0 + 1e-6):
            if overflow_applied:
                overflow_note = (
                    f"Stack {stack_idx} is {capacity_used * 100:.0f}% overfilled relative to "
                    "SKU-specific stacking maximums."
                )
                warnings.append(
                    _warning(
                        "STACK_OVERFLOW_ALLOWANCE_USED",
                        overflow_note,
                        position_id=position_id,
                        deck=deck,
                    )
                )
            else:
                warnings.append(
                    _warning(
                        "STACK_TOO_HIGH",
                        f"Stack {stack_idx} is {capacity_used * 100:.0f}% overfilled relative to SKU-specific stacking maximums.",
                        position_id=position_id,
                        deck=deck,
                    )
                )
        pos["overflow_applied"] = bool(overflow_applied)
        pos["overflow_note"] = overflow_note

    lower_total = sum(
        _coerce_float_value(pos.get("length_ft"), 0.0)
        for pos in positions
        if (pos.get("deck") or "lower") == "lower"
    )
    upper_total_raw = sum(
        _coerce_float_value(pos.get("length_ft"), 0.0)
        for pos in positions
        if (pos.get("deck") or "lower") == "upper"
    )
    upper_total_effective = sum(
        _coerce_float_value(pos.get("effective_length_ft"), _coerce_float_value(pos.get("length_ft"), 0.0))
        for pos in positions
        if (pos.get("deck") or "lower") == "upper"
    )

    def _append_overhang_warning(deck_name, overhang_ft, allowance_ft, deck_key):
        if overhang_ft <= 0.05:
            return
        if overhang_ft <= (allowance_ft + 1e-6):
            warnings.append(
                _warning(
                    "BACK_OVERHANG_IN_ALLOWANCE",
                    (
                        f"{deck_name} deck back overhang is {overhang_ft:.1f} ft "
                        f"(allowance {allowance_ft:.1f} ft)."
                    ),
                    deck=deck_key,
                )
            )
            return
        warnings.append(
            _warning(
                "ITEM_HANGS_OVER_DECK",
                (
                    f"{deck_name} deck back overhang is {overhang_ft:.1f} ft, "
                    f"exceeding allowance by {overhang_ft - allowance_ft:.1f} ft."
                ),
                deck=deck_key,
            )
        )

    _append_overhang_warning(
        "Lower",
        max(lower_total - lower_length, 0.0),
        max_back_overhang_ft,
        "lower",
    )
    if upper_length > 0:
        upper_eval = stack_calculator.evaluate_upper_deck_overhang(
            positions,
            trailer_config,
            max_back_overhang_ft=max_back_overhang_ft,
            upper_deck_exception_max_length_ft=upper_deck_exception_max_length_ft,
            upper_deck_exception_overhang_allowance_ft=upper_deck_exception_overhang_allowance_ft,
            upper_deck_exception_categories=upper_deck_exception_categories,
        )
        _append_overhang_warning(
            "Upper",
            upper_eval["upper_overhang_ft"],
            upper_eval["allowed_overhang_ft"],
            "upper",
        )

    for pos in positions:
        deck_length = upper_length if pos.get("deck") == "upper" else lower_length
        width_length = (
            _coerce_float_value(
                pos.get("effective_length_ft"),
                _coerce_float_value(pos.get("length_ft"), 0.0),
            )
            if (pos.get("deck") or "lower") == "upper"
            else _coerce_float_value(pos.get("length_ft"), 0.0)
        )
        if deck_length > 0:
            pos["width_pct"] = min(
                round((width_length / deck_length) * 100, 1),
                100,
            )
        else:
            pos["width_pct"] = 0

    total_linear_feet = lower_total + upper_total_effective
    lower_credit = 0.0
    upper_credit_raw = 0.0
    upper_length_used = 0.0
    for pos in positions:
        length_ft = _coerce_float_value(pos.get("length_ft"), 0.0)
        effective_length_ft = _coerce_float_value(pos.get("effective_length_ft"), length_ft)
        capacity_used = _coerce_float_value(pos.get("capacity_used"), 0.0)
        if pos.get("overflow_applied"):
            multiplier = min(capacity_used, max_stack_utilization_multiplier)
        else:
            multiplier = min(capacity_used, 1.0)
        multiplier = max(multiplier, 0.0)
        credit = (effective_length_ft if (pos.get("deck") or "lower") == "upper" else length_ft) * multiplier
        if (pos.get("deck") or "lower") == "upper":
            upper_credit_raw += credit
            upper_length_used += effective_length_ft
        else:
            lower_credit += credit

    upper_credit = upper_credit_raw
    if (
        trailer_config["type"].startswith("STEP_DECK")
        and upper_length > 0
        and upper_length_used > 0
        and upper_length_used < (upper_length - 1e-6)
    ):
        # Normalize occupied upper-deck stacks to the full 10' basis.
        upper_credit *= (upper_length / upper_length_used)

    total_credit_feet = lower_credit + upper_credit
    utilization_pct = (total_credit_feet / capacity) * 100 if capacity > 0 else 0.0
    max_stack_height = max((pos.get("units_count") or 0 for pos in positions), default=0)
    compatibility_issues = stack_calculator.check_stacking_compatibility(
        positions,
        trailer_config=trailer_config,
        equal_length_deck_length_order_enabled=equal_length_deck_length_order_enabled,
    )
    exceeds_capacity = stack_calculator.capacity_overflow_feet(
        {
            "positions": positions,
            "trailer_type": trailer_config["type"],
            "lower_deck_length": lower_length,
            "upper_deck_length": upper_length,
            "max_back_overhang_ft": max_back_overhang_ft,
            "upper_deck_exception_max_length_ft": upper_deck_exception_max_length_ft,
            "upper_deck_exception_overhang_allowance_ft": upper_deck_exception_overhang_allowance_ft,
            "upper_deck_exception_categories": upper_deck_exception_categories,
        }
    ) > 0.0
    utilization_grade = _utilization_grade(utilization_pct)

    for issue in compatibility_issues:
        warnings.append(_warning("COMPATIBILITY_ISSUE", issue))

    return (
        {
            "positions": positions,
            "total_linear_feet": round(total_linear_feet, 1),
            "utilization_pct": round(utilization_pct, 1),
            "max_stack_height": max_stack_height,
            "compatibility_issues": compatibility_issues,
            "exceeds_capacity": exceeds_capacity,
            "utilization_credit_ft": round(total_credit_feet, 1),
            "utilization_grade": utilization_grade,
            "trailer_type": trailer_config["type"],
            "capacity_feet": capacity,
            "lower_deck_length": lower_length,
            "upper_deck_length": upper_length,
            "lower_deck_used_length_ft": round(lower_total, 1),
            "upper_deck_raw_length_ft": round(upper_total_raw, 1),
            "upper_deck_effective_length_ft": round(upper_total_effective, 1),
            "upper_two_across_applied_count": int(upper_usage_meta.get("paired_positions") or 0),
            "upper_two_across_max_length_ft": round(
                _coerce_non_negative_float(
                    upper_usage_meta.get("threshold_ft"),
                    upper_two_across_max_length_ft,
                ),
                2,
            ),
            "stack_overflow_max_height": stack_overflow_max_height,
            "max_back_overhang_ft": max_back_overhang_ft,
            "upper_deck_exception_max_length_ft": upper_deck_exception_max_length_ft,
            "upper_deck_exception_overhang_allowance_ft": upper_deck_exception_overhang_allowance_ft,
            "upper_deck_exception_categories": list(upper_deck_exception_categories),
            "equal_length_deck_length_order_enabled": bool(equal_length_deck_length_order_enabled),
            "max_stack_utilization_multiplier": round(
                max_stack_utilization_multiplier,
                4,
            ),
        },
        warnings,
    )


def _remap_layout_for_trailer(layout, trailer_type):
    config = _trailer_config_for_type(trailer_type)
    has_upper = config["upper"] > 0
    remapped = {"positions": []}
    for pos in (layout or {}).get("positions") or []:
        deck = (pos.get("deck") or "lower").strip().lower()
        if deck not in {"lower", "upper"}:
            deck = "lower"
        if deck == "upper" and not has_upper:
            deck = "lower"
        remapped["positions"].append(
            {
                "position_id": pos.get("position_id"),
                "deck": deck,
                "unit_ids": list(pos.get("unit_ids") or []),
            }
        )
    return remapped


def _build_schematic_fragment_payload(load_data, status=None, tab=None):
    status = status or (load_data.get("status") or STATUS_PROPOSED).upper()
    tab = (tab or "").strip().lower()
    schematic_html = render_template(
        "partials/load_schematic_card.html",
        load=load_data,
        status=status,
        tab=tab,
        carrier_selection_options=LOAD_CARRIER_SELECTION_OPTIONS,
    )
    freight = load_data.get("freight_breakdown") or {}
    return {
        "schematic_html": schematic_html,
        "utilization_pct": round(load_data.get("utilization_pct") or 0),
        "utilization_grade": (load_data.get("schematic") or {}).get("utilization_grade") or "F",
        "over_capacity": bool(load_data.get("over_capacity")),
        "exceeds_capacity": bool((load_data.get("schematic") or {}).get("exceeds_capacity")),
        "warnings": list(load_data.get("schematic_warnings") or []),
        "warning_count": int(load_data.get("schematic_warning_count") or 0),
        "assumptions": dict(load_data.get("stack_assumptions") or {}),
        "carrier_key": str(load_data.get("carrier_key") or ""),
        "carrier_label": str(load_data.get("carrier_label") or ""),
        "carrier_source_label": str(load_data.get("carrier_source_label") or ""),
        "carrier_selection_reason": str(load_data.get("carrier_selection_reason") or ""),
        "rate_per_mile": float(load_data.get("rate_per_mile") or 0.0),
        "estimated_cost": float(load_data.get("estimated_cost") or 0.0),
        "freight_breakdown": {
            "stop_count": int(freight.get("stop_count") or 0),
            "stop_fee": float(freight.get("stop_fee") or 0.0),
            "stop_cost": float(freight.get("stop_cost") or 0.0),
            "estimated_miles": float(freight.get("estimated_miles") or 0.0),
            "linehaul_rate_per_mile": float(freight.get("linehaul_rate_per_mile") or 0.0),
            "linehaul_cost": float(freight.get("linehaul_cost") or 0.0),
            "fuel_surcharge_per_mile": float(freight.get("fuel_surcharge_per_mile") or 0.0),
            "fuel_cost": float(freight.get("fuel_cost") or 0.0),
            "adjustment_cost": float(freight.get("adjustment_cost") or 0.0),
            "adjustment_label": str(freight.get("adjustment_label") or "Adjustment"),
            "subtotal_before_floor": float(freight.get("subtotal_before_floor") or 0.0),
            "min_load_cost": float(freight.get("min_load_cost") or 0.0),
            "total_cost": float(freight.get("total_cost") or 0.0),
            "carrier_key": str(freight.get("carrier_key") or load_data.get("carrier_key") or ""),
            "carrier_label": str(freight.get("carrier_label") or load_data.get("carrier_label") or ""),
            "rate_source_label": str(freight.get("rate_source_label") or load_data.get("carrier_source_label") or ""),
            "selection_reason": str(freight.get("selection_reason") or load_data.get("carrier_selection_reason") or ""),
            "pricing_note": str(freight.get("pricing_note") or ""),
            "requires_return_miles": bool(_coerce_bool_value(freight.get("requires_return_miles"))),
        },
    }


def _build_load_schematic_edit_payload(load_id):
    load = db.get_load(load_id)
    if not load:
        return None

    trailer_type = stack_calculator.normalize_trailer_type(load.get("trailer_type"), default="STEP_DECK")
    assumptions = _get_stack_capacity_assumptions()
    lines = db.list_load_lines(load_id)
    sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
    stop_color_palette = _get_stop_color_palette()

    zip_coords = geo_utils.load_zip_coordinates()
    ordered_stops = _ordered_stops_for_lines(lines, load.get("origin_plant"), zip_coords)
    ordered_stops = _apply_load_route_direction(ordered_stops, load=load)
    stop_sequence_map = _stop_sequence_map_from_ordered_stops(ordered_stops)
    order_colors = _build_order_colors_for_lines(
        lines,
        stop_sequence_map=stop_sequence_map,
        stop_palette=stop_color_palette,
    )
    base_schematic, _, _ = _calculate_load_schematic(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
        assumptions=assumptions,
    )

    units = _build_schematic_units(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
        order_colors=order_colors,
    )
    units_by_id = {unit["unit_id"]: unit for unit in units}
    base_layout = _layout_from_schematic(base_schematic, units)
    layout = base_layout
    override = db.get_load_schematic_override(load_id)
    if override and (override.get("trailer_type") or "").strip().upper() == trailer_type:
        try:
            override_layout = json.loads(override.get("layout_json") or "{}")
            layout = _normalize_edit_layout(override_layout, units_by_id, trailer_type)
        except (json.JSONDecodeError, ValueError):
            layout = base_layout

    schematic, warnings = _build_schematic_from_layout(
        layout,
        units_by_id,
        trailer_type,
        assumptions=assumptions,
    )
    status = (load.get("status") or STATUS_PROPOSED).upper()

    return {
        "load_id": load_id,
        "status": status,
        "trailer_type": trailer_type,
        "can_edit": status != STATUS_APPROVED,
        "units": units,
        "layout": layout,
        "base_layout": base_layout,
        "metrics": {
            "utilization_pct": schematic.get("utilization_pct") or 0,
            "utilization_grade": schematic.get("utilization_grade") or "F",
            "total_linear_feet": schematic.get("total_linear_feet") or 0,
            "utilization_credit_ft": schematic.get("utilization_credit_ft") or 0,
            "exceeds_capacity": bool(schematic.get("exceeds_capacity")),
        },
        "warnings": warnings,
        "warning_count": len(warnings),
        "assumptions": assumptions,
        "schematic": schematic,
        "order_colors": order_colors,
        "load": load,
    }


def _build_orders_snapshot(orders, today=None, load_assignment_map=None):
    today = today or date.today()
    active_orders = [order for order in orders if not order.get("is_excluded")]
    load_assignment_map = load_assignment_map or {}
    next_14_end = today + timedelta(days=14)
    snapshot = {
        "total": len(active_orders),
        "on_load": 0,
        "past_due": 0,
        "due_next_14": 0,
        "due_14_plus": 0,
        "unassigned": 0,
    }

    normalized_assignment_map = {
        _normalize_so_num_for_load_report_match(key): value
        for key, value in (load_assignment_map or {}).items()
        if _normalize_so_num_for_load_report_match(key) and str(value or "").strip()
    }

    for order in active_orders:
        so_num = _normalize_so_num_for_load_report_match(order.get("so_num"))
        if so_num and normalized_assignment_map.get(so_num):
            snapshot["on_load"] += 1
            continue
        snapshot["unassigned"] += 1

        due_date = _parse_date(order.get("due_date"))
        if due_date:
            if due_date < today:
                snapshot["past_due"] += 1
            elif due_date <= next_14_end:
                snapshot["due_next_14"] += 1
            else:
                snapshot["due_14_plus"] += 1
        else:
            snapshot["due_14_plus"] += 1

    total_orders = snapshot["total"] or 0
    snapshot["on_load_pct"] = round((snapshot["on_load"] / total_orders) * 100, 1) if total_orders else 0.0
    snapshot["unassigned_pct"] = round((snapshot["unassigned"] / total_orders) * 100, 1) if total_orders else 0.0
    snapshot["timeline_total"] = snapshot["past_due"] + snapshot["due_next_14"] + snapshot["due_14_plus"]
    return snapshot


def _filter_out_past_due_orders(orders):
    filtered = []
    for order in orders or []:
        due_status = (order.get("due_status") or "").upper()
        if due_status == "PAST_DUE":
            continue
        filtered.append(order)
    return filtered


def _count_active_orders_by_plant_from_rows(orders, plants=None):
    counts = {plant: 0 for plant in (plants or [])}
    allowed = set(plants or [])
    for order in orders or []:
        if _coerce_bool_value(order.get("is_excluded")):
            continue
        plant_code = _normalize_plant_code(order.get("plant"))
        if not plant_code:
            continue
        if allowed and plant_code not in allowed:
            continue
        counts[plant_code] = counts.get(plant_code, 0) + 1
    return counts


def _clean_query_params(values):
    cleaned = {}
    for key, value in (values or {}).items():
        if value is None:
            continue
        normalized = str(value).strip()
        if not normalized:
            continue
        cleaned[key] = normalized
    return cleaned


def _require_session():
    profile = _ensure_active_profile()
    if not profile or not _get_allowed_plants():
        next_url = request.full_path if request else ""
        return redirect(url_for("login", next=next_url))
    return None


def _require_admin():
    if _get_session_role() != ROLE_ADMIN:
        abort(403)


def _require_settings_editor():
    if _get_session_role() not in {ROLE_ADMIN, ROLE_PLANNER}:
        abort(403)


def _can_access_planning_session(planning_session):
    if not planning_session:
        return False
    plant_code = _normalize_plant_code(planning_session.get("plant_code"))
    allowed_plants = set(_get_allowed_plants())
    if plant_code and plant_code not in allowed_plants:
        return False
    if _get_session_role() == ROLE_ADMIN:
        return True

    if bool(planning_session.get("is_sandbox")) != bool(_is_session_sandbox()):
        return False

    profile_name = (_get_session_profile_name() or "").strip()
    created_by = (planning_session.get("created_by") or "").strip()
    if not profile_name:
        return False
    return created_by.casefold() == profile_name.casefold()


def _load_access_failure_reason(load):
    if not load:
        return "missing"
    allowed_plants = set(_get_allowed_plants())
    if load.get("origin_plant") not in allowed_plants:
        return "plant"
    if _get_session_role() == ROLE_ADMIN:
        return None

    session_id = load.get("planning_session_id")
    if _is_session_sandbox():
        if not session_id:
            return "sandbox"
        planning_session = db.get_planning_session(session_id)
        if not planning_session or not bool(planning_session.get("is_sandbox")):
            return "sandbox"
        return None

    if session_id:
        planning_session = db.get_planning_session(session_id)
        if planning_session and bool(planning_session.get("is_sandbox")):
            return "sandbox"
    return None


def _load_access_error_message(reason):
    if reason == "sandbox":
        return "Sandbox accounts can only access sandbox planning-session loads."
    if reason == "plant":
        return "Not authorized for this plant."
    return "Not authorized for this load."


def _get_scoped_planning_session_or_404(session_id):
    planning_session = db.get_planning_session(session_id)
    if not planning_session or not _can_access_planning_session(planning_session):
        abort(404)
    return planning_session


def _year_suffix(value=None):
    value = value or date.today()
    return f"{value.year % 100:02d}"


def _format_load_number(plant_code, year_suffix, sequence, draft=False):
    base = f"{plant_code}{year_suffix}-{sequence:04d}"
    return f"{base}-D" if draft else base


def _normalize_load_number(load_number):
    if not load_number:
        return None, None
    trimmed = load_number.strip()
    parts = trimmed.split("-")
    if len(parts) < 2:
        return trimmed, None
    return trimmed, parts[-1]


def _planning_session_year_suffix(planning_session):
    if not planning_session:
        return _year_suffix()
    created_dt = _parse_datetime(planning_session.get("created_at"))
    created_date = created_dt.date() if created_dt else _parse_date(planning_session.get("created_at"))
    return _year_suffix(created_date or date.today())


def _reserve_session_load_number(planning_session, fallback_plant_code, starting_sequence=None):
    if not planning_session:
        return {"error": "session_not_found"}
    session_id = planning_session.get("id")
    plant_code = (
        (planning_session.get("plant_code") or "").strip().upper()
        or (fallback_plant_code or "").strip().upper()
    )
    year_suffix = _planning_session_year_suffix(planning_session)
    return db.reserve_planning_session_load_number(
        session_id,
        plant_code,
        year_suffix,
        starting_sequence=starting_sequence,
    )


def _normalize_tutorial_relpath(value):
    if value is None:
        return ""
    normalized = str(value).strip().replace("\\", "/").lstrip("/")
    if not normalized:
        return ""
    clean = os.path.normpath(normalized).replace("\\", "/")
    if clean in {"..", "."} or clean.startswith("../"):
        return ""
    return clean


def _tutorial_static_file_exists(static_relpath):
    normalized = _normalize_tutorial_relpath(static_relpath)
    if not normalized:
        return False
    static_root = os.path.abspath(app.static_folder or "static")
    candidate = os.path.abspath(os.path.join(static_root, normalized))
    try:
        if os.path.commonpath([static_root, candidate]) != static_root:
            return False
    except ValueError:
        return False
    return os.path.isfile(candidate)


def _coerce_tutorial_audience(value):
    if not isinstance(value, list):
        return ["all"]
    audience = []
    for entry in value:
        token = str(entry).strip().lower()
        if token in TUTORIAL_ALLOWED_AUDIENCE and token not in audience:
            audience.append(token)
    return audience or ["all"]


def _parse_tutorial_step_note(raw_note):
    if isinstance(raw_note, str):
        note_text = raw_note.strip()
        return ("tip", note_text) if note_text else ("", "")
    if isinstance(raw_note, dict):
        note_text = str(raw_note.get("text") or "").strip()
        note_tone = str(raw_note.get("tone") or "tip").strip().lower()
        if note_tone not in TUTORIAL_NOTE_LABELS:
            note_tone = "tip"
        return (note_tone, note_text) if note_text else ("", "")
    return "", ""


def _load_tutorial_manifest(manifest_path=None):
    resolved_path = manifest_path or TUTORIAL_MANIFEST_PATH
    manifest = {
        "version": 1,
        "modules": [],
        "error": "",
        "manifest_path": resolved_path,
    }
    if not os.path.isfile(resolved_path):
        logger.warning("Tutorial manifest not found at %s", resolved_path)
        manifest["error"] = "Tutorial content is unavailable right now."
        return manifest

    try:
        with open(resolved_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning("Failed to load tutorial manifest at %s: %s", resolved_path, exc)
        manifest["error"] = "Tutorial content is unavailable right now."
        return manifest

    if not isinstance(payload, dict):
        logger.warning("Tutorial manifest must be a JSON object: %s", resolved_path)
        manifest["error"] = "Tutorial content is unavailable right now."
        return manifest

    version = payload.get("version")
    if isinstance(version, int) and version > 0:
        manifest["version"] = version

    raw_modules = payload.get("modules")
    if not isinstance(raw_modules, list):
        logger.warning("Tutorial manifest is missing a valid modules list: %s", resolved_path)
        manifest["error"] = "Tutorial content is unavailable right now."
        return manifest

    for module_idx, raw_module in enumerate(raw_modules, start=1):
        if not isinstance(raw_module, dict):
            logger.warning("Skipping tutorial module #%s: expected object.", module_idx)
            continue
        slug = str(raw_module.get("slug") or "").strip().lower()
        title = str(raw_module.get("title") or "").strip()
        route_hint = str(raw_module.get("route_hint") or "").strip()
        summary = str(raw_module.get("summary") or "").strip()
        raw_steps = raw_module.get("steps")
        if not slug or not title or not summary:
            logger.warning("Skipping tutorial module #%s: missing slug/title/summary.", module_idx)
            continue
        if not isinstance(raw_steps, list) or not raw_steps:
            logger.warning("Skipping tutorial module '%s': steps must be a non-empty list.", slug)
            continue

        module = {
            "slug": slug,
            "title": title,
            "route_hint": route_hint,
            "summary": summary,
            "audience": _coerce_tutorial_audience(raw_module.get("audience")),
            "steps": [],
        }

        for step_idx, raw_step in enumerate(raw_steps, start=1):
            if not isinstance(raw_step, dict):
                logger.warning("Skipping tutorial step #%s in module '%s': expected object.", step_idx, slug)
                continue
            step_id = str(raw_step.get("id") or "").strip().lower()
            step_title = str(raw_step.get("title") or "").strip()
            instruction = str(raw_step.get("instruction") or "").strip()
            media = raw_step.get("media") if isinstance(raw_step.get("media"), dict) else {}
            media_type = str(media.get("type") or "").strip().lower()
            media_src = _normalize_tutorial_relpath(media.get("src"))
            if (
                not step_id
                or not step_title
                or not instruction
                or media_type not in TUTORIAL_ALLOWED_MEDIA_TYPES
                or not media_src
            ):
                logger.warning(
                    "Skipping tutorial step #%s in module '%s': invalid id/content/media.",
                    step_idx,
                    slug,
                )
                continue

            poster_src = _normalize_tutorial_relpath(media.get("poster"))
            media_alt = str(media.get("alt") or "").strip()
            if media_type == "image" and not media_alt:
                media_alt = f"{title} - {step_title}"
            note_tone, note_text = _parse_tutorial_step_note(raw_step.get("note"))
            module["steps"].append(
                {
                    "id": step_id,
                    "title": step_title,
                    "instruction": instruction,
                    "note_tone": note_tone,
                    "note_label": TUTORIAL_NOTE_LABELS.get(note_tone, ""),
                    "note_text": note_text,
                    "media": {
                        "type": media_type,
                        "src": media_src,
                        "alt": media_alt,
                        "caption": str(media.get("caption") or "").strip(),
                        "poster": poster_src,
                        "exists": _tutorial_static_file_exists(media_src),
                        "poster_exists": bool(poster_src and _tutorial_static_file_exists(poster_src)),
                    },
                }
            )

        if not module["steps"]:
            logger.warning("Skipping tutorial module '%s': no valid steps were found.", slug)
            continue

        manifest["modules"].append(module)

    if not manifest["modules"] and not manifest["error"]:
        manifest["error"] = "Tutorial content is unavailable right now."
    return manifest


def _coerce_filter_list(values):
    if values is None:
        return []
    if isinstance(values, (list, tuple, set)):
        source = values
    else:
        source = [values]
    cleaned = []
    for value in source:
        text = str(value or "").strip()
        if text:
            cleaned.append(text)
    return cleaned


def _reoptimize_form_data(plant_code, session_id=None):
    form_data = _default_optimize_form(plant_code=plant_code)
    form_data["origin_plant"] = plant_code

    def _finalize(data):
        trailer_key = stack_calculator.normalize_trailer_type(
            data.get("trailer_type"),
            default="STEP_DECK",
        )
        data["trailer_type"] = trailer_key
        data["capacity_feet"] = str(
            _capacity_for_trailer_setting(
                trailer_key,
                requested_capacity=data.get("capacity_feet"),
                default_capacity=_get_optimizer_default_settings().get("capacity_feet", 53.0),
            )
        )
        return data

    settings = db.get_optimizer_settings(plant_code) or {}
    if settings.get("capacity_feet") is not None:
        form_data["capacity_feet"] = str(settings.get("capacity_feet"))
    if settings.get("trailer_type"):
        form_data["trailer_type"] = stack_calculator.normalize_trailer_type(
            settings.get("trailer_type"),
            default=form_data.get("trailer_type"),
        )
    if settings.get("max_detour_pct") is not None:
        form_data["max_detour_pct"] = str(settings.get("max_detour_pct"))
    if settings.get("time_window_days") is not None:
        form_data["time_window_days"] = str(settings.get("time_window_days"))
    if settings.get("geo_radius") is not None:
        form_data["geo_radius"] = str(settings.get("geo_radius"))
    if settings.get("stack_overflow_max_height") is not None:
        form_data["stack_overflow_max_height"] = str(settings.get("stack_overflow_max_height"))
    if settings.get("max_back_overhang_ft") is not None:
        form_data["max_back_overhang_ft"] = str(settings.get("max_back_overhang_ft"))
    if settings.get("upper_two_across_max_length_ft") is not None:
        form_data["upper_two_across_max_length_ft"] = str(settings.get("upper_two_across_max_length_ft"))
    if settings.get("upper_deck_exception_max_length_ft") is not None:
        form_data["upper_deck_exception_max_length_ft"] = str(settings.get("upper_deck_exception_max_length_ft"))
    if settings.get("upper_deck_exception_overhang_allowance_ft") is not None:
        form_data["upper_deck_exception_overhang_allowance_ft"] = str(
            settings.get("upper_deck_exception_overhang_allowance_ft")
        )
    if settings.get("upper_deck_exception_categories") is not None:
        form_data["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
            settings.get("upper_deck_exception_categories"),
            default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
        )

    if not session_id:
        return _finalize(form_data)

    planning_session = db.get_planning_session(session_id)
    if not planning_session or not planning_session.get("config_json"):
        return _finalize(form_data)

    try:
        session_config = json.loads(planning_session.get("config_json") or "{}")
    except json.JSONDecodeError:
        return _finalize(form_data)

    if session_config.get("capacity_feet") is not None:
        form_data["capacity_feet"] = str(session_config.get("capacity_feet"))
    if session_config.get("trailer_type"):
        form_data["trailer_type"] = stack_calculator.normalize_trailer_type(
            session_config.get("trailer_type"),
            default=form_data.get("trailer_type"),
        )
    if session_config.get("max_detour_pct") is not None:
        form_data["max_detour_pct"] = str(session_config.get("max_detour_pct"))
    if session_config.get("time_window_days") is not None:
        form_data["time_window_days"] = str(session_config.get("time_window_days"))
    if session_config.get("geo_radius") is not None:
        form_data["geo_radius"] = str(session_config.get("geo_radius"))
    if session_config.get("stack_overflow_max_height") is not None:
        form_data["stack_overflow_max_height"] = str(session_config.get("stack_overflow_max_height"))
    if session_config.get("max_back_overhang_ft") is not None:
        form_data["max_back_overhang_ft"] = str(session_config.get("max_back_overhang_ft"))
    if session_config.get("upper_two_across_max_length_ft") is not None:
        form_data["upper_two_across_max_length_ft"] = str(session_config.get("upper_two_across_max_length_ft"))
    if session_config.get("upper_deck_exception_max_length_ft") is not None:
        form_data["upper_deck_exception_max_length_ft"] = str(
            session_config.get("upper_deck_exception_max_length_ft")
        )
    if session_config.get("upper_deck_exception_overhang_allowance_ft") is not None:
        form_data["upper_deck_exception_overhang_allowance_ft"] = str(
            session_config.get("upper_deck_exception_overhang_allowance_ft")
        )
    if session_config.get("upper_deck_exception_categories") is not None:
        form_data["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
            session_config.get("upper_deck_exception_categories"),
            default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
        )

    form_data["opt_toggles"] = "1"
    if session_config.get("enforce_time_window", True):
        form_data["enforce_time_window"] = "1"
    else:
        form_data["enforce_time_window"] = "0"

    batch_end_date = (session_config.get("batch_end_date") or "").strip()
    if session_config.get("batch_horizon_enabled") and batch_end_date:
        form_data["batch_horizon_enabled"] = "1"
        form_data["batch_end_date"] = batch_end_date
    else:
        form_data["batch_horizon_enabled"] = "0"
        form_data["batch_end_date"] = ""

    form_data["opt_states"] = _coerce_filter_list(session_config.get("state_filters"))
    form_data["opt_customers"] = _coerce_filter_list(session_config.get("customer_filters"))
    optimize_mode = (session_config.get("optimize_mode") or "auto").strip().lower()
    if optimize_mode not in {"auto", "manual"}:
        optimize_mode = "auto"
    form_data["optimize_mode"] = optimize_mode
    form_data["order_category_scope"] = order_categories.normalize_order_category_scope(
        session_config.get("order_category_scope"),
        default=form_data.get("order_category_scope", order_categories.ORDER_CATEGORY_SCOPE_ALL),
    )
    manual_order_input = (session_config.get("manual_order_input") or "").strip()
    if not manual_order_input and optimize_mode == "manual":
        selected_so_nums = session_config.get("selected_so_nums")
        if isinstance(selected_so_nums, (list, tuple, set)):
            manual_order_input = "\n".join(
                [str(value).strip() for value in selected_so_nums if str(value or "").strip()]
            )
    form_data["manual_order_input"] = manual_order_input
    ignore_due_date = bool(session_config.get("ignore_due_date", False))
    form_data["ignore_due_date"] = ignore_due_date
    orders_start_date = (session_config.get("orders_start_date") or "").strip()
    if not orders_start_date and session_config.get("ignore_past_due", True):
        orders_start_date = date.today().strftime("%Y-%m-%d")
    form_data["orders_start_date"] = orders_start_date or date.today().strftime("%Y-%m-%d")
    return _finalize(form_data)


def _reoptimize_for_plant(plant_code, session_id=None, created_by=None, speed_profile="standard"):
    if not plant_code:
        return {"errors": {"origin_plant": "Missing plant code."}}
    if not created_by:
        try:
            created_by = _get_session_profile_name() or _get_session_role()
        except RuntimeError:
            created_by = None
    if session_id:
        db.clear_unapproved_loads(session_id=session_id)
    else:
        db.clear_unapproved_loads(plant_code)
    form_data = _reoptimize_form_data(plant_code, session_id=session_id)
    speed_mode = (speed_profile or "").strip().lower()
    if speed_mode == "fast":
        form_data["__reopt_speed"] = "fast"
    return load_builder.build_loads(
        form_data,
        reset_proposed=False,
        store_settings=False,
        session_id=session_id,
        created_by=created_by,
        include_baseline=False,
    )


def _reopt_scope_key(plant_code, session_id=None):
    normalized_plant = (plant_code or "").strip().upper()
    if session_id:
        return f"{normalized_plant}:session:{int(session_id)}"
    return f"{normalized_plant}:plant"


def _trim_reopt_jobs(now_ts=None):
    current_ts = now_ts or datetime.utcnow().timestamp()
    active_ids = []
    finished_rows = []
    for job_id, job in _REOPT_JOBS.items():
        status = (job.get("status") or "").lower()
        if status == "running":
            active_ids.append(job_id)
            continue
        finished_at = job.get("finished_at")
        if finished_at:
            try:
                finished_ts = datetime.fromisoformat(finished_at).timestamp()
            except ValueError:
                finished_ts = current_ts
        else:
            finished_ts = current_ts
        age = current_ts - finished_ts
        if age > REOPT_JOB_RETENTION_SEC:
            finished_rows.append((job_id, finished_ts, True))
        else:
            finished_rows.append((job_id, finished_ts, False))

    for job_id, _, should_prune in finished_rows:
        if should_prune:
            _REOPT_JOBS.pop(job_id, None)

    active_count = len(active_ids)
    retained_finished = [
        (job_id, finished_ts)
        for job_id, finished_ts, should_prune in finished_rows
        if not should_prune
    ]
    if active_count + len(retained_finished) <= REOPT_JOB_MAX_ENTRIES:
        return

    retained_finished.sort(key=lambda item: item[1])
    while active_count + len(retained_finished) > REOPT_JOB_MAX_ENTRIES and retained_finished:
        stale_id, _ = retained_finished.pop(0)
        _REOPT_JOBS.pop(stale_id, None)


def _start_reopt_job(plant_code, session_id=None, created_by=None, speed_profile="standard"):
    normalized_plant = (plant_code or "").strip().upper()
    if not normalized_plant:
        raise ValueError("Missing plant code.")
    normalized_session = int(session_id) if session_id else None
    scope_key = _reopt_scope_key(normalized_plant, normalized_session)
    now_iso = datetime.utcnow().isoformat(timespec="seconds")

    with _REOPT_JOB_LOCK:
        _trim_reopt_jobs()
        for existing_id, existing in _REOPT_JOBS.items():
            if existing.get("scope_key") != scope_key:
                continue
            if (existing.get("status") or "").lower() == "running":
                return existing_id

        job_id = uuid.uuid4().hex
        _REOPT_JOBS[job_id] = {
            "id": job_id,
            "scope_key": scope_key,
            "plant_code": normalized_plant,
            "session_id": normalized_session,
            "speed_profile": (speed_profile or "standard").strip().lower() or "standard",
            "status": "running",
            "created_at": now_iso,
            "started_at": now_iso,
            "finished_at": "",
            "error": "",
            "success_message": "",
        }

    worker = threading.Thread(
        target=_run_reopt_job,
        args=(
            job_id,
            normalized_plant,
            normalized_session,
            created_by,
            (speed_profile or "standard").strip().lower() or "standard",
        ),
        daemon=True,
        name=f"reopt-{normalized_plant}-{job_id[:8]}",
    )
    worker.start()
    return job_id


def _run_reopt_job(job_id, plant_code, session_id=None, created_by=None, speed_profile="standard"):
    try:
        result = _reoptimize_for_plant(
            plant_code,
            session_id=session_id,
            created_by=created_by,
            speed_profile=speed_profile,
        ) or {}
        errors = result.get("errors") or {}
        success_message = (result.get("success_message") or "").strip()
        error_message = ""
        status = "done"
        if errors:
            status = "failed"
            try:
                first_error = next(iter(errors.values()))
            except StopIteration:
                first_error = "Re-optimization failed."
            error_message = str(first_error or "Re-optimization failed.").strip()

        finished_iso = datetime.utcnow().isoformat(timespec="seconds")
        with _REOPT_JOB_LOCK:
            job = _REOPT_JOBS.get(job_id)
            if job is not None:
                job.update(
                    {
                        "status": status,
                        "finished_at": finished_iso,
                        "error": error_message,
                        "success_message": success_message,
                    }
                )
                _trim_reopt_jobs()
    except Exception:  # pragma: no cover - defensive path
        logger.exception(
            "Background re-optimization failed for plant=%s session_id=%s",
            plant_code,
            session_id,
        )
        finished_iso = datetime.utcnow().isoformat(timespec="seconds")
        with _REOPT_JOB_LOCK:
            job = _REOPT_JOBS.get(job_id)
            if job is not None:
                job.update(
                    {
                        "status": "failed",
                        "finished_at": finished_iso,
                        "error": "Unexpected re-optimization error.",
                    }
                )
                _trim_reopt_jobs()


def _get_reopt_job(job_id):
    normalized = (job_id or "").strip()
    if not normalized:
        return None
    with _REOPT_JOB_LOCK:
        _trim_reopt_jobs()
        row = _REOPT_JOBS.get(normalized)
        if not row:
            return None
        return dict(row)


@app.context_processor
def inject_session_context():
    profile = _ensure_active_profile()
    role = _get_session_role()
    allowed_plants = _get_allowed_plants()
    selected_plants = _resolve_plant_filters(request.args.get("plants") or request.args.get("plant"))
    return {
        "session_role": role,
        "session_profile_name": _get_session_profile_name(),
        "session_is_sandbox": _is_session_sandbox(),
        "session_profile_id": session.get(SESSION_PROFILE_ID_KEY),
        "session_profile_default_plants": session.get(SESSION_PROFILE_DEFAULT_PLANTS_KEY) or [],
        "access_profiles": db.list_access_profiles(),
        "session_allowed_plants": allowed_plants,
        "session_plant_filter": session.get("plant_filter"),
        "session_plant_filters": selected_plants,
        "session_plant_filter_label": _format_plant_filter_label(selected_plants, allowed_plants),
        "plant_filter_cards": _build_plant_filter_cards(allowed_plants, selected_plants),
        "is_admin": role == ROLE_ADMIN,
        "show_tutorial_nav": TUTORIAL_NAV_ENABLED,
        "trailer_profile_options": TRAILER_PROFILE_OPTIONS,
        "app_release_label": APP_RELEASE_LABEL,
    }


def _build_command_center_dashboard_context():
    allowed_plants = _get_allowed_plants()
    plant_filters = _resolve_plant_filters(request.args.get("plants") or request.args.get("plant"))
    plant_scope = plant_filters or allowed_plants

    period_options = [
        ("last_30_days", "Last 30 Days"),
        ("last_7_days", "Last 7 Days"),
        ("this_month", "This Month"),
        ("today", "Today"),
    ]
    period_label_map = dict(period_options)
    period = (request.args.get("period") or "last_30_days").strip().lower()
    if period not in period_label_map:
        period = "last_30_days"

    today = date.today()
    start_date, end_date = _period_range(period, today)
    period_len_days = (end_date - start_date).days + 1
    prev_end = start_date - timedelta(days=1)
    prev_start = prev_end - timedelta(days=period_len_days - 1)

    def _iso(value):
        return value.strftime("%Y-%m-%d")

    start_iso = _iso(start_date)
    end_iso = _iso(end_date)
    prev_start_iso = _iso(prev_start)
    prev_end_iso = _iso(prev_end)

    plant_filter_param = ",".join(plant_filters) if plant_filters else ""
    plant_scope_label = (
        ", ".join([PLANT_NAMES.get(code, code) for code in plant_filters])
        if plant_filters
        else "All Plants"
    )
    period_label = period_label_map.get(period, "Last 30 Days")
    date_range_label = f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d')}"

    def _format_int(value):
        try:
            return f"{int(value or 0):,}"
        except (TypeError, ValueError):
            return "0"

    def _format_currency(amount, decimals=2):
        amount = float(amount or 0)
        return f"${amount:,.{decimals}f}"

    def _format_compact_currency(amount):
        amount = float(amount or 0)
        abs_amount = abs(amount)
        if abs_amount >= 1_000_000:
            return f"${amount / 1_000_000:.2f}M"
        if abs_amount >= 100_000:
            return f"${amount / 1_000:.0f}K"
        if abs_amount >= 10_000:
            return f"${amount / 1_000:.1f}K"
        return f"${amount:,.0f}"

    def _pct_change(current, previous):
        if current is None or previous is None:
            return None
        current = float(current or 0)
        previous = float(previous or 0)
        if previous == 0:
            return None
        return (current - previous) / previous * 100.0

    def _trend_meta(delta_pct):
        if delta_pct is None:
            return {"display": None, "icon": "trending_flat", "class": ""}
        icon = "trending_up" if delta_pct >= 0 else "trending_down"
        cls = "positive" if delta_pct >= 0 else "negative"
        return {"display": f"{delta_pct:+.1f}%", "icon": icon, "class": cls}

    def _point_delta_meta(delta_points):
        if delta_points is None:
            return {"display": None, "icon": "trending_flat", "class": ""}
        icon = "trending_up" if delta_points >= 0 else "trending_down"
        cls = "positive" if delta_points >= 0 else "negative"
        return {"display": f"{delta_points:+.1f} pts", "icon": icon, "class": cls}

    def _bucket_series(items, bucket_count=7):
        if not items:
            return []
        if len(items) <= bucket_count:
            return [
                {
                    "start": entry["date"],
                    "end": entry["date"],
                    "spend": float(entry.get("spend") or 0),
                    "qty": float(entry.get("qty") or 0),
                }
                for entry in items
            ]
        size = len(items)
        base = size // bucket_count
        remainder = size % bucket_count
        buckets = []
        idx = 0
        for bucket_index in range(bucket_count):
            take = base + (1 if bucket_index < remainder else 0)
            segment = items[idx:idx + take]
            idx += take
            if not segment:
                continue
            buckets.append(
                {
                    "start": segment[0]["date"],
                    "end": segment[-1]["date"],
                    "spend": sum(float(item.get("spend") or 0) for item in segment),
                    "qty": sum(float(item.get("qty") or 0) for item in segment),
                }
            )
        return buckets

    def _axis_label(value_date):
        if not value_date:
            return ""
        if value_date == today:
            return "Today"
        return value_date.strftime("%b %d")

    def _fetch_period_totals(connection, plants, start_value, end_value):
        if not plants:
            return {"loads": 0, "avg_util": 0.0, "spend": 0.0, "qty": 0.0}
        placeholders = ", ".join("?" for _ in plants)
        params = list(plants) + [start_value, end_value]
        loads_row = connection.execute(
            f"""
            SELECT
                COUNT(*) AS load_count,
                AVG(utilization_pct) AS avg_util,
                SUM(COALESCE(estimated_cost, 0)) AS total_spend
            FROM loads
            WHERE status = 'APPROVED'
              AND origin_plant IN ({placeholders})
              AND DATE(created_at) BETWEEN DATE(?) AND DATE(?)
            """,
            params,
        ).fetchone()
        qty_row = connection.execute(
            f"""
            SELECT
                SUM(COALESCE(ol.qty, 0)) AS total_qty
            FROM loads l
            JOIN load_lines ll ON ll.load_id = l.id
            JOIN order_lines ol ON ol.id = ll.order_line_id
            WHERE l.status = 'APPROVED'
              AND l.origin_plant IN ({placeholders})
              AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
              AND COALESCE(ol.is_excluded, 0) = 0
            """,
            params,
        ).fetchone()
        return {
            "loads": int(loads_row["load_count"] or 0) if loads_row else 0,
            "avg_util": float(loads_row["avg_util"] or 0) if loads_row else 0.0,
            "spend": float(loads_row["total_spend"] or 0) if loads_row else 0.0,
            "qty": float(qty_row["total_qty"] or 0) if qty_row else 0.0,
        }

    with db.get_connection() as connection:
        current_totals = _fetch_period_totals(connection, plant_scope, start_iso, end_iso)
        prev_totals = _fetch_period_totals(connection, plant_scope, prev_start_iso, prev_end_iso)

        current_cpu = (
            (current_totals["spend"] / current_totals["qty"]) if current_totals["qty"] else None
        )
        prev_cpu = (prev_totals["spend"] / prev_totals["qty"]) if prev_totals["qty"] else None

        loads_trend = _trend_meta(_pct_change(current_totals["loads"], prev_totals["loads"]))
        spend_trend = _trend_meta(_pct_change(current_totals["spend"], prev_totals["spend"]))
        cpu_trend = _trend_meta(_pct_change(current_cpu, prev_cpu))
        util_trend = _point_delta_meta(
            (current_totals["avg_util"] - prev_totals["avg_util"]) if prev_totals["loads"] else None
        )

        avg_util = float(current_totals["avg_util"] or 0)
        health_label = "NO DATA"
        health_class = "neutral"
        if current_totals["loads"]:
            if avg_util >= 80:
                health_label = "OPTIMAL"
                health_class = "success"
            elif avg_util >= 70:
                health_label = "MONITOR"
                health_class = "warning"
            else:
                health_label = "AT RISK"
                health_class = "danger"

        kpis = [
            {
                "label": "Total Loads Batched",
                "icon": "inventory_2",
                "value": _format_int(current_totals["loads"]),
                "unit": "",
                "delta_display": loads_trend["display"],
                "delta_icon": loads_trend["icon"],
                "delta_class": loads_trend["class"],
                "footer": "vs previous period",
            },
            {
                "label": "Average Utilization",
                "icon": "speed",
                "value": f"{avg_util:.1f}%",
                "unit": "",
                "delta_display": util_trend["display"],
                "delta_icon": util_trend["icon"],
                "delta_class": util_trend["class"],
                "footer": "vs previous period",
            },
            {
                "label": "Total Logistics Spend",
                "icon": "payments",
                "value": _format_compact_currency(current_totals["spend"]),
                "unit": "",
                "delta_display": spend_trend["display"],
                "delta_icon": spend_trend["icon"],
                "delta_class": spend_trend["class"],
                "footer": "vs previous period",
            },
            {
                "label": "Cost Per Unit",
                "icon": "straighten",
                "value": _format_currency(current_cpu, decimals=2) if current_cpu is not None else "—",
                "unit": "/ unit",
                "delta_display": cpu_trend["display"],
                "delta_icon": cpu_trend["icon"],
                "delta_class": cpu_trend["class"],
                "footer": "vs previous period",
            },
        ]

        # Plant cards.
        open_orders_by_plant = {plant: 0 for plant in allowed_plants}
        planned_loads_by_plant = {plant: 0 for plant in allowed_plants}
        plant_spend = {plant: 0.0 for plant in allowed_plants}
        plant_avg_util = {plant: None for plant in allowed_plants}
        plant_qty = {plant: 0.0 for plant in allowed_plants}

        allowed_placeholders = ", ".join("?" for _ in allowed_plants) if allowed_plants else "''"

        assigned_clause = """
            EXISTS (
                SELECT 1
                FROM order_lines ol
                JOIN load_lines ll ON ll.order_line_id = ol.id
                JOIN loads l ON l.id = ll.load_id
                LEFT JOIN planning_sessions ps ON ps.id = l.planning_session_id
                WHERE ol.so_num = orders.so_num
                  AND (
                    COALESCE(UPPER(l.status), '') = 'APPROVED'
                    OR COALESCE(UPPER(ps.status), 'DRAFT') IN ('DRAFT', 'ACTIVE')
                  )
                LIMIT 1
            )
        """

        for row in connection.execute(
            f"""
            SELECT plant, COUNT(*) AS open_orders
            FROM orders
            WHERE is_excluded = 0
              AND plant IN ({allowed_placeholders})
              AND COALESCE(UPPER(status), 'OPEN') != 'CLOSED'
              AND NOT {assigned_clause}
            GROUP BY plant
            """,
            list(allowed_plants),
        ).fetchall():
            open_orders_by_plant[row["plant"]] = int(row["open_orders"] or 0)

        for row in connection.execute(
            f"""
            SELECT l.origin_plant, COUNT(*) AS planned_loads
            FROM loads l
            LEFT JOIN planning_sessions ps ON ps.id = l.planning_session_id
            WHERE UPPER(l.status) IN ('PROPOSED', 'DRAFT')
              AND l.origin_plant IN ({allowed_placeholders})
              AND COALESCE(UPPER(ps.status), 'DRAFT') IN ('DRAFT', 'ACTIVE')
            GROUP BY l.origin_plant
            """,
            list(allowed_plants),
        ).fetchall():
            planned_loads_by_plant[row["origin_plant"]] = int(row["planned_loads"] or 0)

        for row in connection.execute(
            f"""
            SELECT
                origin_plant,
                AVG(utilization_pct) AS avg_util,
                SUM(COALESCE(estimated_cost, 0)) AS spend
            FROM loads
            WHERE status = 'APPROVED'
              AND origin_plant IN ({allowed_placeholders})
              AND DATE(created_at) BETWEEN DATE(?) AND DATE(?)
            GROUP BY origin_plant
            """,
            list(allowed_plants) + [start_iso, end_iso],
        ).fetchall():
            plant = row["origin_plant"]
            plant_spend[plant] = float(row["spend"] or 0)
            plant_avg_util[plant] = float(row["avg_util"]) if row["avg_util"] is not None else None

        for row in connection.execute(
            f"""
            SELECT
                l.origin_plant AS origin_plant,
                SUM(COALESCE(ol.qty, 0)) AS qty
            FROM loads l
            JOIN load_lines ll ON ll.load_id = l.id
            JOIN order_lines ol ON ol.id = ll.order_line_id
            WHERE l.status = 'APPROVED'
              AND l.origin_plant IN ({allowed_placeholders})
              AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
              AND COALESCE(ol.is_excluded, 0) = 0
            GROUP BY l.origin_plant
            """,
            list(allowed_plants) + [start_iso, end_iso],
        ).fetchall():
            plant_qty[row["origin_plant"]] = float(row["qty"] or 0)

        selected_set = set(plant_filters or [])
        plant_cards = []
        for plant in allowed_plants:
            avg_util_value = plant_avg_util.get(plant)
            status = "neutral"
            if avg_util_value is not None:
                if avg_util_value >= 80:
                    status = "success"
                elif avg_util_value >= 70:
                    status = "warning"
                else:
                    status = "error"

            qty_value = float(plant_qty.get(plant) or 0)
            spend_value = float(plant_spend.get(plant) or 0)
            cost_unit_value = (spend_value / qty_value) if qty_value else None
            plant_cards.append(
                {
                    "code": plant,
                    "name": PLANT_NAMES.get(plant, plant),
                    "open_orders": open_orders_by_plant.get(plant, 0),
                    "planned_loads": planned_loads_by_plant.get(plant, 0),
                    "cost_per_unit_display": _format_currency(cost_unit_value, decimals=2)
                    if cost_unit_value is not None
                    else "—",
                    "status": status,
                    "selected": plant in selected_set,
                }
            )

        # Spend vs Volume trend.
        spend_by_day = {}
        qty_by_day = {}
        if plant_scope:
            scope_placeholders = ", ".join("?" for _ in plant_scope)
            for row in connection.execute(
                f"""
                SELECT DATE(created_at) AS day, SUM(COALESCE(estimated_cost, 0)) AS spend
                FROM loads
                WHERE status = 'APPROVED'
                  AND origin_plant IN ({scope_placeholders})
                  AND DATE(created_at) BETWEEN DATE(?) AND DATE(?)
                GROUP BY day
                ORDER BY day
                """,
                list(plant_scope) + [start_iso, end_iso],
            ).fetchall():
                if row["day"]:
                    spend_by_day[row["day"]] = float(row["spend"] or 0)

            for row in connection.execute(
                f"""
                SELECT DATE(l.created_at) AS day, SUM(COALESCE(ol.qty, 0)) AS qty
                FROM loads l
                JOIN load_lines ll ON ll.load_id = l.id
                JOIN order_lines ol ON ol.id = ll.order_line_id
                WHERE l.status = 'APPROVED'
                  AND l.origin_plant IN ({scope_placeholders})
                  AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
                  AND COALESCE(ol.is_excluded, 0) = 0
                GROUP BY day
                ORDER BY day
                """,
                list(plant_scope) + [start_iso, end_iso],
            ).fetchall():
                if row["day"]:
                    qty_by_day[row["day"]] = float(row["qty"] or 0)

        daily_items = []
        for offset in range(period_len_days):
            day = start_date + timedelta(days=offset)
            day_key = _iso(day)
            daily_items.append(
                {
                    "date": day,
                    "spend": float(spend_by_day.get(day_key) or 0),
                    "qty": float(qty_by_day.get(day_key) or 0),
                }
            )

        raw_buckets = _bucket_series(daily_items, bucket_count=7)
        max_spend = max((bucket["spend"] for bucket in raw_buckets), default=0.0)
        max_qty = max((bucket["qty"] for bucket in raw_buckets), default=0.0)
        trend_buckets = []
        for bucket in raw_buckets:
            spend_height = 0
            volume_height = 0
            if max_spend and bucket["spend"]:
                spend_height = max(4, round((bucket["spend"] / max_spend) * 100))
            if max_qty and bucket["qty"]:
                volume_height = max(4, round((bucket["qty"] / max_qty) * 100))
            trend_buckets.append(
                {
                    "start": bucket["start"],
                    "end": bucket["end"],
                    "spend": bucket["spend"],
                    "qty": bucket["qty"],
                    "spend_height": spend_height,
                    "volume_height": volume_height,
                }
            )

        trend_axis = {"start": "", "mid1": "", "mid2": "", "end": ""}
        if trend_buckets:
            if len(trend_buckets) == 1:
                trend_axis["start"] = _axis_label(trend_buckets[0]["start"])
                trend_axis["end"] = _axis_label(trend_buckets[0]["end"])
            elif len(trend_buckets) == 2:
                trend_axis["start"] = _axis_label(trend_buckets[0]["start"])
                trend_axis["end"] = _axis_label(trend_buckets[1]["end"])
            elif len(trend_buckets) == 3:
                trend_axis["start"] = _axis_label(trend_buckets[0]["start"])
                trend_axis["mid1"] = _axis_label(trend_buckets[1]["start"])
                trend_axis["end"] = _axis_label(trend_buckets[2]["end"])
            else:
                idx1 = len(trend_buckets) // 3
                idx2 = (len(trend_buckets) * 2) // 3
                idx1 = min(max(idx1, 1), len(trend_buckets) - 2)
                idx2 = min(max(idx2, idx1 + 1), len(trend_buckets) - 2)
                trend_axis["start"] = _axis_label(trend_buckets[0]["start"])
                trend_axis["mid1"] = _axis_label(trend_buckets[idx1]["start"])
                trend_axis["mid2"] = _axis_label(trend_buckets[idx2]["start"])
                trend_axis["end"] = _axis_label(trend_buckets[-1]["end"])

        # Efficiency distribution bins.
        efficiency = {"a": 0, "b": 0, "c": 0, "df": 0, "total": 0}
        if plant_scope:
            scope_placeholders = ", ".join("?" for _ in plant_scope)
            row = connection.execute(
                f"""
                SELECT
                    SUM(CASE WHEN utilization_pct >= 95 THEN 1 ELSE 0 END) AS grade_a,
                    SUM(CASE WHEN utilization_pct >= 85 AND utilization_pct < 95 THEN 1 ELSE 0 END) AS grade_b,
                    SUM(CASE WHEN utilization_pct >= 70 AND utilization_pct < 85 THEN 1 ELSE 0 END) AS grade_c,
                    SUM(CASE WHEN utilization_pct < 70 THEN 1 ELSE 0 END) AS grade_df,
                    COUNT(*) AS total
                FROM loads
                WHERE status = 'APPROVED'
                  AND origin_plant IN ({scope_placeholders})
                  AND DATE(created_at) BETWEEN DATE(?) AND DATE(?)
                """,
                list(plant_scope) + [start_iso, end_iso],
            ).fetchone()
            if row:
                efficiency = {
                    "a": int(row["grade_a"] or 0),
                    "b": int(row["grade_b"] or 0),
                    "c": int(row["grade_c"] or 0),
                    "df": int(row["grade_df"] or 0),
                    "total": int(row["total"] or 0),
                }

        eff_total = efficiency["total"] or 0
        efficiency_rows = [
            {
                "label": "Grade A (95%+)",
                "count": efficiency["a"],
                "width": round((efficiency["a"] / eff_total) * 100) if eff_total else 0,
                "class": "success",
            },
            {
                "label": "Grade B (85-94%)",
                "count": efficiency["b"],
                "width": round((efficiency["b"] / eff_total) * 100) if eff_total else 0,
                "class": "info",
            },
            {
                "label": "Grade C (70-84%)",
                "count": efficiency["c"],
                "width": round((efficiency["c"] / eff_total) * 100) if eff_total else 0,
                "class": "neutral",
            },
            {
                "label": "Grade D/F (<70%)",
                "count": efficiency["df"],
                "width": round((efficiency["df"] / eff_total) * 100) if eff_total else 0,
                "class": "danger",
            },
        ]

        # Load review widgets.
        sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
        stop_color_palette = _get_stop_color_palette()

        def _fetch_loads(limit, ascending, max_util=None):
            if not plant_scope:
                return []
            scope_placeholders = ", ".join("?" for _ in plant_scope)
            util_clause = ""
            params = list(plant_scope)
            if max_util is not None:
                util_clause = " AND utilization_pct < ?"
                params.append(max_util)
            params.extend([start_iso, end_iso, limit])
            order_dir = "ASC" if ascending else "DESC"
            rows = connection.execute(
                f"""
                SELECT id, load_number, origin_plant, utilization_pct, created_at, trailer_type
                FROM loads
                WHERE status = 'APPROVED'
                  AND origin_plant IN ({scope_placeholders})
                  AND DATE(created_at) BETWEEN DATE(?) AND DATE(?)
                  {util_clause}
                ORDER BY utilization_pct {order_dir}, created_at DESC
                LIMIT ?
                """,
                params,
            ).fetchall()
            return [dict(row) for row in rows]

        def _format_load_rows(loads):
            formatted = []
            for load in loads:
                created_at = load.get("created_at") or ""
                ship_date = ""
                try:
                    ship_date = datetime.fromisoformat(created_at).strftime("%m/%d")
                except (TypeError, ValueError):
                    parsed = _parse_date(created_at)
                    ship_date = parsed.strftime("%m/%d") if parsed else ""
                formatted.append(
                    {
                        "id": load["id"],
                        "load_number": load.get("load_number") or f"Load #{load['id']}",
                        "origin_plant": load.get("origin_plant"),
                        "utilization_pct": round(load.get("utilization_pct") or 0, 1),
                        "ship_date": ship_date,
                        "thumbnail": _build_load_thumbnail(load, sku_specs, stop_color_palette),
                    }
                )
            return formatted

        top_loads = _format_load_rows(_fetch_loads(limit=5, ascending=False))
        problem_loads = _format_load_rows(_fetch_loads(limit=5, ascending=True, max_util=70))

    return {
        "period": period,
        "period_label": period_label,
        "period_options": period_options,
        "date_range_label": date_range_label,
        "start_date": start_date,
        "end_date": end_date,
        "plant_scope_label": plant_scope_label,
        "plant_filter_param": plant_filter_param,
        "network_health": {"label": health_label, "class": health_class},
        "kpis": kpis,
        "plant_cards": plant_cards,
        "trend_buckets": trend_buckets,
        "trend_axis": trend_axis,
        "efficiency_rows": efficiency_rows,
        "top_loads": top_loads,
        "problem_loads": problem_loads,
    }


def _build_performance_dashboard_context():
    allowed_plants = _get_allowed_plants()
    plant_filters = _resolve_plant_filters(request.args.get("plants") or request.args.get("plant"))
    plant_scope = plant_filters or allowed_plants

    period_options = [
        ("this_month", "This Month"),
        ("last_30_days", "Last 30 Days"),
        ("last_quarter", "Last Quarter"),
        ("ytd", "YTD"),
    ]
    period_label_map = dict(period_options)
    period = (request.args.get("period") or "last_30_days").strip().lower()
    if period not in period_label_map:
        period = "last_30_days"

    today = date.today()
    start_date, end_date = _period_range(period, today)

    def _shift_year(value, years=-1):
        target_year = value.year + years
        day = value.day
        while day >= 1:
            try:
                return date(target_year, value.month, day)
            except ValueError:
                day -= 1
        return date(target_year, value.month, 1)

    prior_start = _shift_year(start_date, years=-1)
    prior_end = _shift_year(end_date, years=-1)
    prior_year = prior_start.year

    plant_filter_param = ",".join(plant_filters) if plant_filters else ""
    plant_scope_label = (
        ", ".join([PLANT_NAMES.get(code, code) for code in plant_filters])
        if plant_filters
        else "All Plants"
    )
    period_label = period_label_map.get(period, "Last 30 Days")
    date_range_label = f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
    prior_date_range_label = f"{prior_start.strftime('%b %d, %Y')} - {prior_end.strftime('%b %d, %Y')}"

    def _format_int(value):
        try:
            return f"{int(value or 0):,}"
        except (TypeError, ValueError):
            return "0"

    def _format_number(value, decimals=1):
        if value is None:
            return "--"
        try:
            return f"{float(value):,.{decimals}f}"
        except (TypeError, ValueError):
            return "--"

    def _format_percent(value, decimals=1):
        if value is None:
            return "--"
        return f"{_format_number(value, decimals)}%"

    def _format_compact_currency(amount):
        if amount is None:
            return "--"
        amount = float(amount or 0.0)
        abs_amount = abs(amount)
        if abs_amount >= 1_000_000:
            return f"${amount / 1_000_000:.2f}M"
        if abs_amount >= 100_000:
            return f"${amount / 1_000:.0f}K"
        if abs_amount >= 10_000:
            return f"${amount / 1_000:.1f}K"
        return f"${amount:,.0f}"

    def _format_currency(amount, decimals=2):
        if amount is None:
            return "--"
        return f"${float(amount):,.{decimals}f}"

    def _delta_percent(current, previous):
        if current is None or previous is None:
            return None
        current = float(current or 0.0)
        previous = float(previous or 0.0)
        if abs(previous) <= 1e-9:
            return None
        return ((current - previous) / abs(previous)) * 100.0

    def _build_delta_payload(current, previous, lower_is_better):
        delta_pct = _delta_percent(current, previous)
        if delta_pct is None:
            return {"available": False, "display": "", "class": "neutral"}
        improved = delta_pct < 0 if lower_is_better else delta_pct > 0
        if abs(delta_pct) < 1e-6:
            return {"available": True, "display": "0.0% YoY", "class": "neutral"}
        return {
            "available": True,
            "display": f"{delta_pct:+.1f}% YoY",
            "class": "positive" if improved else "negative",
        }

    def _month_start(value):
        return date(value.year, value.month, 1)

    def _add_months(month_start, offset):
        index = (month_start.year * 12) + (month_start.month - 1) + offset
        year = index // 12
        month = (index % 12) + 1
        return date(year, month, 1)

    def _month_end(month_start):
        return _add_months(month_start, 1) - timedelta(days=1)

    def _build_svg_path(values, min_value, max_value, width, height):
        if not values:
            return ""
        span = max_value - min_value
        if span <= 1e-9:
            span = 1.0
        points = []
        count = len(values)
        for idx, raw in enumerate(values):
            if raw is None:
                points.append(None)
                continue
            x = (idx / max(count - 1, 1)) * float(width)
            y = float(height) - ((float(raw) - min_value) / span) * float(height)
            points.append((round(x, 2), round(y, 2)))
        commands = []
        drawing = False
        for point in points:
            if point is None:
                drawing = False
                continue
            commands.append(("L" if drawing else "M") + f"{point[0]},{point[1]}")
            drawing = True
        return " ".join(commands)

    approved_statuses = ("APPROVED", "SHIPPED")

    with db.get_connection() as connection:
        load_columns = {
            (row["name"] or "").strip().lower()
            for row in connection.execute("PRAGMA table_info(loads)").fetchall()
        }
        trailer_count_expr = (
            "SUM(COALESCE(l.trailer_count, 1))" if "trailer_count" in load_columns else "COUNT(*)"
        )
        status_placeholders = ", ".join("?" for _ in approved_statuses)
        metrics_cache = {}

        def _aggregate_metrics(plants, start_value, end_value):
            normalized_plants = [code for code in (plants or []) if code]
            key = (
                tuple(sorted(normalized_plants)),
                start_value.isoformat(),
                end_value.isoformat(),
            )
            if key in metrics_cache:
                return metrics_cache[key]
            if not normalized_plants:
                empty = {
                    "load_count": 0,
                    "avg_utilization": None,
                    "total_spend": 0.0,
                    "total_units": 0.0,
                    "total_miles": 0.0,
                    "total_trailers": 0.0,
                    "cost_per_unit": None,
                    "rate_per_mile": None,
                    "miles_per_load": None,
                    "trailers_per_load": None,
                }
                metrics_cache[key] = empty
                return empty

            plant_placeholders = ", ".join("?" for _ in normalized_plants)
            params = list(approved_statuses) + list(normalized_plants) + [
                start_value.isoformat(),
                end_value.isoformat(),
            ]
            load_row = connection.execute(
                f"""
                SELECT
                    COUNT(*) AS load_count,
                    AVG(COALESCE(l.utilization_pct, 0)) AS avg_utilization,
                    SUM(COALESCE(l.estimated_cost, 0)) AS total_spend,
                    SUM(
                        CASE
                            WHEN COALESCE(l.route_total_miles, 0) > 0 THEN l.route_total_miles
                            ELSE COALESCE(l.estimated_miles, 0)
                        END
                    ) AS total_miles,
                    {trailer_count_expr} AS total_trailers
                FROM loads l
                WHERE COALESCE(UPPER(l.status), '') IN ({status_placeholders})
                  AND l.origin_plant IN ({plant_placeholders})
                  AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
                """,
                params,
            ).fetchone()
            units_row = connection.execute(
                f"""
                SELECT SUM(COALESCE(ol.qty, 0)) AS total_units
                FROM loads l
                JOIN load_lines ll ON ll.load_id = l.id
                JOIN order_lines ol ON ol.id = ll.order_line_id
                WHERE COALESCE(UPPER(l.status), '') IN ({status_placeholders})
                  AND l.origin_plant IN ({plant_placeholders})
                  AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
                  AND COALESCE(ol.is_excluded, 0) = 0
                """,
                params,
            ).fetchone()

            load_count = int(load_row["load_count"] or 0) if load_row else 0
            avg_utilization = (
                float(load_row["avg_utilization"]) if load_row and load_row["avg_utilization"] is not None else None
            )
            total_spend = float(load_row["total_spend"] or 0.0) if load_row else 0.0
            total_miles = float(load_row["total_miles"] or 0.0) if load_row else 0.0
            total_trailers = float(load_row["total_trailers"] or 0.0) if load_row else 0.0
            total_units = float(units_row["total_units"] or 0.0) if units_row else 0.0

            metrics = {
                "load_count": load_count,
                "avg_utilization": avg_utilization,
                "total_spend": total_spend,
                "total_units": total_units,
                "total_miles": total_miles,
                "total_trailers": total_trailers,
                "cost_per_unit": (total_spend / total_units) if total_units > 0 else None,
                "rate_per_mile": (total_spend / total_miles) if total_miles > 0 else None,
                "miles_per_load": (total_miles / load_count) if load_count > 0 else None,
                "trailers_per_load": (total_trailers / load_count) if load_count > 0 else None,
            }
            metrics_cache[key] = metrics
            return metrics

        current_metrics = _aggregate_metrics(plant_scope, start_date, end_date)
        prior_metrics = _aggregate_metrics(plant_scope, prior_start, prior_end)

        cost_unit_delta = _build_delta_payload(
            current_metrics["cost_per_unit"],
            prior_metrics["cost_per_unit"],
            lower_is_better=True,
        )
        rate_mile_delta = _build_delta_payload(
            current_metrics["rate_per_mile"],
            prior_metrics["rate_per_mile"],
            lower_is_better=True,
        )
        miles_load_delta = _build_delta_payload(
            current_metrics["miles_per_load"],
            prior_metrics["miles_per_load"],
            lower_is_better=True,
        )
        trailers_load_delta = _build_delta_payload(
            current_metrics["trailers_per_load"],
            prior_metrics["trailers_per_load"],
            lower_is_better=False,
        )

        kpis = [
            {
                "label": "Loads Batched",
                "value": _format_int(current_metrics["load_count"]),
                "subtext": "Approved/shipped loads",
                "delta": None,
            },
            {
                "label": "Avg Utilization",
                "value": _format_percent(current_metrics["avg_utilization"], decimals=1),
                "subtext": "Mean space-based utilization",
                "delta": None,
            },
            {
                "label": "Total Freight Spend",
                "value": _format_compact_currency(current_metrics["total_spend"]),
                "subtext": "Estimated all-in freight cost",
                "delta": None,
            },
            {
                "label": "Cost / Unit",
                "value": _format_currency(current_metrics["cost_per_unit"], decimals=2),
                "subtext": f"vs same period {prior_year}",
                "delta": cost_unit_delta,
            },
            {
                "label": "Effective Rate / Mile",
                "value": _format_currency(current_metrics["rate_per_mile"], decimals=2),
                "subtext": "All-in cost per mile",
                "delta": rate_mile_delta,
            },
            {
                "label": "Miles / Load",
                "value": _format_number(current_metrics["miles_per_load"], decimals=1),
                "subtext": f"vs same period {prior_year}",
                "delta": miles_load_delta,
            },
        ]

        def _plant_color(code):
            if not code:
                return "var(--primary)"
            if code in PLANT_CODES and DEFAULT_STOP_COLOR_PALETTE:
                idx = PLANT_CODES.index(code) % len(DEFAULT_STOP_COLOR_PALETTE)
                return DEFAULT_STOP_COLOR_PALETTE[idx]
            return "var(--primary)"

        plant_placeholders = ", ".join("?" for _ in allowed_plants) if allowed_plants else "''"
        plant_params = list(approved_statuses) + list(allowed_plants) + [
            start_date.isoformat(),
            end_date.isoformat(),
        ]
        plant_rows = connection.execute(
            f"""
            SELECT
                l.origin_plant,
                COUNT(*) AS load_count,
                AVG(COALESCE(l.utilization_pct, 0)) AS avg_utilization
            FROM loads l
            WHERE COALESCE(UPPER(l.status), '') IN ({status_placeholders})
              AND l.origin_plant IN ({plant_placeholders})
              AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
            GROUP BY l.origin_plant
            """,
            plant_params,
        ).fetchall() if allowed_plants else []
        plant_stats = {}
        for row in plant_rows:
            plant_stats[row["origin_plant"]] = {
                "load_count": int(row["load_count"] or 0),
                "avg_utilization": (
                    float(row["avg_utilization"]) if row["avg_utilization"] is not None else None
                ),
            }
        selected_set = set(plant_filters or [])
        plant_cards = [
            {
                "code": "ALL",
                "name": "All Plants",
                "load_count": current_metrics["load_count"],
                "avg_utilization": current_metrics["avg_utilization"],
                "avg_util_display": _format_percent(current_metrics["avg_utilization"], decimals=1),
                "selected": not selected_set,
                "inactive": False,
                "interactive": True,
                "is_all": True,
                "color": "var(--primary)",
            }
        ]
        for plant in allowed_plants:
            stats = plant_stats.get(plant, {})
            load_count = int(stats.get("load_count") or 0)
            avg_util = stats.get("avg_utilization")
            inactive = load_count <= 0
            plant_cards.append(
                {
                    "code": plant,
                    "name": PLANT_NAMES.get(plant, plant),
                    "load_count": load_count,
                    "avg_utilization": avg_util,
                    "avg_util_display": _format_percent(avg_util, decimals=1),
                    "selected": plant in selected_set,
                    "inactive": inactive,
                    "interactive": not inactive,
                    "is_all": False,
                    "color": _plant_color(plant),
                }
            )

        scope_placeholders = ", ".join("?" for _ in plant_scope) if plant_scope else "''"
        scope_params = list(approved_statuses) + list(plant_scope) + [
            start_date.isoformat(),
            end_date.isoformat(),
        ]

        band_row = connection.execute(
            f"""
            SELECT
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) < 70 THEN 1 ELSE 0 END) AS under_70,
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) >= 70 AND COALESCE(l.utilization_pct, 0) < 80 THEN 1 ELSE 0 END) AS band_70_79,
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) >= 80 AND COALESCE(l.utilization_pct, 0) < 90 THEN 1 ELSE 0 END) AS band_80_89,
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) >= 90 AND COALESCE(l.utilization_pct, 0) < 100 THEN 1 ELSE 0 END) AS band_90_99,
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) >= 100 THEN 1 ELSE 0 END) AS over_100,
                COUNT(*) AS total
            FROM loads l
            WHERE COALESCE(UPPER(l.status), '') IN ({status_placeholders})
              AND l.origin_plant IN ({scope_placeholders})
              AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
            """,
            scope_params,
        ).fetchone() if plant_scope else None
        band_total = int(band_row["total"] or 0) if band_row else 0
        utilization_bands = [
            {
                "label": "<70%",
                "count": int(band_row["under_70"] or 0) if band_row else 0,
                "width": round(((int(band_row["under_70"] or 0) if band_row else 0) / band_total) * 100, 1)
                if band_total
                else 0,
                "class": "band-red",
            },
            {
                "label": "70-79%",
                "count": int(band_row["band_70_79"] or 0) if band_row else 0,
                "width": round(((int(band_row["band_70_79"] or 0) if band_row else 0) / band_total) * 100, 1)
                if band_total
                else 0,
                "class": "band-amber",
            },
            {
                "label": "80-89%",
                "count": int(band_row["band_80_89"] or 0) if band_row else 0,
                "width": round(((int(band_row["band_80_89"] or 0) if band_row else 0) / band_total) * 100, 1)
                if band_total
                else 0,
                "class": "band-blue",
            },
            {
                "label": "90-99%",
                "count": int(band_row["band_90_99"] or 0) if band_row else 0,
                "width": round(((int(band_row["band_90_99"] or 0) if band_row else 0) / band_total) * 100, 1)
                if band_total
                else 0,
                "class": "band-green",
            },
            {
                "label": "100%+",
                "count": int(band_row["over_100"] or 0) if band_row else 0,
                "width": round(((int(band_row["over_100"] or 0) if band_row else 0) / band_total) * 100, 1)
                if band_total
                else 0,
                "class": "band-purple",
            },
        ]

        grade_row = connection.execute(
            f"""
            SELECT
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) >= 95 THEN 1 ELSE 0 END) AS grade_a,
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) >= 85 AND COALESCE(l.utilization_pct, 0) < 95 THEN 1 ELSE 0 END) AS grade_b,
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) >= 70 AND COALESCE(l.utilization_pct, 0) < 85 THEN 1 ELSE 0 END) AS grade_c,
                SUM(CASE WHEN COALESCE(l.utilization_pct, 0) < 70 THEN 1 ELSE 0 END) AS grade_df,
                COUNT(*) AS total
            FROM loads l
            WHERE COALESCE(UPPER(l.status), '') IN ({status_placeholders})
              AND l.origin_plant IN ({scope_placeholders})
              AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
            """,
            scope_params,
        ).fetchone() if plant_scope else None
        grade_total = int(grade_row["total"] or 0) if grade_row else 0
        efficiency_grades = [
            {
                "label": "A (95%+)",
                "count": int(grade_row["grade_a"] or 0) if grade_row else 0,
                "width": round(((int(grade_row["grade_a"] or 0) if grade_row else 0) / grade_total) * 100, 1)
                if grade_total
                else 0,
                "class": "grade-a",
            },
            {
                "label": "B (85-94%)",
                "count": int(grade_row["grade_b"] or 0) if grade_row else 0,
                "width": round(((int(grade_row["grade_b"] or 0) if grade_row else 0) / grade_total) * 100, 1)
                if grade_total
                else 0,
                "class": "grade-b",
            },
            {
                "label": "C (70-84%)",
                "count": int(grade_row["grade_c"] or 0) if grade_row else 0,
                "width": round(((int(grade_row["grade_c"] or 0) if grade_row else 0) / grade_total) * 100, 1)
                if grade_total
                else 0,
                "class": "grade-c",
            },
            {
                "label": "D/F (<70%)",
                "count": int(grade_row["grade_df"] or 0) if grade_row else 0,
                "width": round(((int(grade_row["grade_df"] or 0) if grade_row else 0) / grade_total) * 100, 1)
                if grade_total
                else 0,
                "class": "grade-df",
            },
        ]

        trend_month_anchor = _month_start(end_date)
        trend_months = [_add_months(trend_month_anchor, offset) for offset in range(-11, 1)]
        current_series = []
        prior_series = []
        for month_start in trend_months:
            month_end = _month_end(month_start)
            prior_month_start = _shift_year(month_start, years=-1)
            prior_month_end = _shift_year(month_end, years=-1)
            month_current = _aggregate_metrics(plant_scope, month_start, month_end)
            month_prior = _aggregate_metrics(plant_scope, prior_month_start, prior_month_end)
            current_series.append(month_current["cost_per_unit"])
            prior_series.append(month_prior["cost_per_unit"])

        complete_month_anchor = _month_start(end_date)
        if end_date < _month_end(complete_month_anchor):
            complete_month_anchor = _add_months(complete_month_anchor, -1)
        complete_month_end = _month_end(complete_month_anchor)
        complete_prior_start = _shift_year(complete_month_anchor, years=-1)
        complete_prior_end = _shift_year(complete_month_end, years=-1)
        complete_current_metrics = _aggregate_metrics(
            plant_scope,
            complete_month_anchor,
            complete_month_end,
        )
        complete_prior_metrics = _aggregate_metrics(
            plant_scope,
            complete_prior_start,
            complete_prior_end,
        )
        trend_delta_chips = [
            {
                "label": "Cost/Unit",
                "value": _format_currency(complete_current_metrics["cost_per_unit"], decimals=2),
                "delta": _build_delta_payload(
                    complete_current_metrics["cost_per_unit"],
                    complete_prior_metrics["cost_per_unit"],
                    lower_is_better=True,
                ),
            },
            {
                "label": "Rate/Mi",
                "value": _format_currency(complete_current_metrics["rate_per_mile"], decimals=2),
                "delta": _build_delta_payload(
                    complete_current_metrics["rate_per_mile"],
                    complete_prior_metrics["rate_per_mile"],
                    lower_is_better=True,
                ),
            },
            {
                "label": "Miles/Load",
                "value": _format_number(complete_current_metrics["miles_per_load"], decimals=1),
                "delta": _build_delta_payload(
                    complete_current_metrics["miles_per_load"],
                    complete_prior_metrics["miles_per_load"],
                    lower_is_better=True,
                ),
            },
            {
                "label": "Trailers/Load",
                "value": _format_number(complete_current_metrics["trailers_per_load"], decimals=2),
                "delta": _build_delta_payload(
                    complete_current_metrics["trailers_per_load"],
                    complete_prior_metrics["trailers_per_load"],
                    lower_is_better=False,
                ),
            },
        ]
        line_values = [float(value) for value in (current_series + prior_series) if value is not None]
        line_min = min(line_values) if line_values else 0.0
        line_max = max(line_values) if line_values else 1.0
        if abs(line_max - line_min) <= 1e-9:
            line_max = line_min + 1.0
        chart_width = 860
        chart_height = 260
        trend_chart = {
            "metric_label": "Cost / Unit",
            "current_path": _build_svg_path(current_series, line_min, line_max, chart_width, chart_height),
            "prior_path": _build_svg_path(prior_series, line_min, line_max, chart_width, chart_height),
            "labels": [entry.strftime("%b") for entry in trend_months],
            "y_max_label": _format_currency(line_max, decimals=2),
            "y_mid_label": _format_currency((line_max + line_min) / 2, decimals=2),
            "y_min_label": _format_currency(line_min, decimals=2),
            "delta_chips": trend_delta_chips,
            "color": _plant_color(plant_filters[0]) if len(plant_filters) == 1 else "var(--primary)",
            "compare_label": f"vs same month {complete_prior_start.year}",
            "current_month_label": complete_month_anchor.strftime("%b %Y"),
            "prior_month_label": complete_prior_start.strftime("%b %Y"),
            "width": chart_width,
            "height": chart_height,
        }

        sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
        stop_color_palette = _get_stop_color_palette()

        def _fetch_review_loads(limit, ascending=False, max_utilization=None):
            if not plant_scope:
                return []
            util_clause = " AND COALESCE(l.utilization_pct, 0) < ?" if max_utilization is not None else ""
            rows = connection.execute(
                f"""
                SELECT l.id, l.load_number, l.origin_plant, l.utilization_pct, l.created_at, l.trailer_type
                FROM loads l
                WHERE COALESCE(UPPER(l.status), '') IN ({status_placeholders})
                  AND l.origin_plant IN ({scope_placeholders})
                  AND DATE(l.created_at) BETWEEN DATE(?) AND DATE(?)
                  {util_clause}
                ORDER BY COALESCE(l.utilization_pct, 0) {"ASC" if ascending else "DESC"}, DATE(l.created_at) DESC
                LIMIT ?
                """,
                list(approved_statuses)
                + list(plant_scope)
                + [start_date.isoformat(), end_date.isoformat()]
                + ([float(max_utilization)] if max_utilization is not None else [])
                + [int(limit)],
            ).fetchall()
            formatted = []
            for row in rows:
                load = dict(row)
                utilization_pct = round(float(load.get("utilization_pct") or 0.0), 1)
                parsed_date = _parse_date(load.get("created_at"))
                ship_date = parsed_date.strftime("%m/%d/%Y") if parsed_date else "--"
                thumbnail = _build_load_thumbnail(load, sku_specs, stop_color_palette=stop_color_palette)
                schematic_segments = []
                total_width = 0.0
                for position in thumbnail:
                    width = float(position.get("width_pct") or 0.0)
                    if width <= 0:
                        continue
                    colors = position.get("colors") or []
                    schematic_segments.append(
                        {
                            "width_pct": width,
                            "color": colors[0] if colors else "rgba(148, 163, 184, 0.55)",
                        }
                    )
                    total_width += width
                if total_width > 0:
                    for segment in schematic_segments:
                        segment["width_pct"] = round((segment["width_pct"] / total_width) * 100.0, 2)
                util_class = "positive"
                if utilization_pct < 70:
                    util_class = "negative"
                elif utilization_pct < 85:
                    util_class = "neutral"
                formatted.append(
                    {
                        "id": load["id"],
                        "load_number": load.get("load_number") or f"Load {load['id']}",
                        "origin_plant": load.get("origin_plant") or "--",
                        "utilization_pct": utilization_pct,
                        "utilization_class": util_class,
                        "ship_date": ship_date,
                        "schematic_segments": schematic_segments,
                        "has_schematic": bool(schematic_segments),
                        "fallback_width": max(0.0, min(100.0, utilization_pct)),
                    }
                )
            return formatted

        top_loads = _fetch_review_loads(limit=5, ascending=False)
        problem_loads = _fetch_review_loads(limit=5, ascending=True, max_utilization=70.0)

        latest_row = connection.execute(
            f"""
            SELECT MAX(l.created_at) AS latest_created_at
            FROM loads l
            WHERE COALESCE(UPPER(l.status), '') IN ({status_placeholders})
              AND l.origin_plant IN ({scope_placeholders})
            """,
            list(approved_statuses) + list(plant_scope),
        ).fetchone() if plant_scope else None
        latest_value = latest_row["latest_created_at"] if latest_row else None
        latest_label = _format_est_datetime_label(latest_value) if latest_value else "No approved load activity"

        has_prior_data = (
            prior_metrics["load_count"] > 0
            and prior_metrics["cost_per_unit"] is not None
            and prior_metrics["rate_per_mile"] is not None
        )
        savings = None
        if (
            has_prior_data
            and current_metrics["load_count"] > 0
            and current_metrics["cost_per_unit"] is not None
            and current_metrics["rate_per_mile"] is not None
            and current_metrics["miles_per_load"] is not None
            and current_metrics["total_units"] > 0
        ):
            utilization_gains = (
                (prior_metrics["cost_per_unit"] - current_metrics["cost_per_unit"])
                * current_metrics["total_units"]
            )
            rate_routing_efficiency = (
                (prior_metrics["rate_per_mile"] - current_metrics["rate_per_mile"])
                * current_metrics["miles_per_load"]
                * current_metrics["load_count"]
            )
            savings = {
                "utilization_gains": utilization_gains,
                "rate_routing_efficiency": rate_routing_efficiency,
                "total": utilization_gains + rate_routing_efficiency,
                "utilization_gains_display": _format_currency(utilization_gains, decimals=0),
                "rate_routing_efficiency_display": _format_currency(rate_routing_efficiency, decimals=0),
                "total_display": _format_currency(utilization_gains + rate_routing_efficiency, decimals=0),
                "utilization_formula": "(prior period cost/unit - current cost/unit) x units shipped",
                "rate_formula": "(prior period rate/mile - current rate/mile) x avg miles/load x load count",
                "tone": "positive" if (utilization_gains + rate_routing_efficiency) >= 0 else "negative",
            }

    return {
        "period": period,
        "period_options": period_options,
        "period_label": period_label,
        "date_range_label": date_range_label,
        "prior_date_range_label": prior_date_range_label,
        "plant_filter_param": plant_filter_param,
        "plant_scope_label": plant_scope_label,
        "prior_year": prior_year,
        "topbar_status": {"label": "Live", "detail": latest_label},
        "kpis": kpis,
        "plant_cards": plant_cards,
        "trend_chart": trend_chart,
        "utilization_bands": utilization_bands,
        "efficiency_grades": efficiency_grades,
        "top_loads": top_loads,
        "problem_loads": problem_loads,
        "savings": savings,
    }


@app.route("/")
def home():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    return redirect(url_for("orders"))


@app.route("/dashboard")
def dashboard():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    context = _build_performance_dashboard_context()
    if isinstance(context, Response):
        return context
    if not isinstance(context, dict):
        context = {}

    context.setdefault("period", "last_30_days")
    context.setdefault(
        "period_options",
        [
            ("this_month", "This Month"),
            ("last_30_days", "Last 30 Days"),
            ("last_quarter", "Last Quarter"),
            ("ytd", "YTD"),
        ],
    )
    context.setdefault("period_label", "Last 30 Days")
    context.setdefault("date_range_label", "--")
    context.setdefault("plant_filter_param", "")
    context.setdefault("plant_scope_label", "All Plants")

    for list_key in (
        "plant_cards",
        "kpis",
        "utilization_bands",
        "efficiency_grades",
        "top_loads",
        "problem_loads",
    ):
        if not isinstance(context.get(list_key), list):
            context[list_key] = []

    topbar_status = context.get("topbar_status")
    if not isinstance(topbar_status, dict):
        topbar_status = {}
    context["topbar_status"] = {
        "label": topbar_status.get("label") or "Live",
        "detail": topbar_status.get("detail") or "Awaiting refresh",
    }

    trend_defaults = {
        "metric_label": "Cost / Unit",
        "deltas": [],
        "months": [],
        "current_points": [],
        "prior_points": [],
        "current_path": "",
        "prior_path": "",
        "y_min": 0.0,
        "y_max": 1.0,
    }
    trend_chart = context.get("trend_chart")
    if not isinstance(trend_chart, dict):
        trend_chart = {}
    context["trend_chart"] = {**trend_defaults, **trend_chart}

    context.setdefault("savings", None)
    return render_template("dashboard.html", **context)


@app.route("/tutorial")
def tutorial():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    manifest = _load_tutorial_manifest()
    role = (_get_session_role() or "").strip().lower()
    modules = [
        module
        for module in manifest["modules"]
        if "all" in module.get("audience", []) or role in module.get("audience", [])
    ]
    selected_slug = (request.args.get("module") or "").strip().lower()
    selected_module = next((module for module in modules if module["slug"] == selected_slug), None)
    if not selected_module and modules:
        selected_module = modules[0]

    tutorial_error = manifest.get("error", "")
    if not tutorial_error and not modules:
        tutorial_error = "No tutorial modules are available for your current access profile."

    return render_template(
        "tutorial.html",
        tutorial_modules=modules,
        tutorial_selected_module=selected_module,
        tutorial_error=tutorial_error,
        tutorial_manifest_version=manifest.get("version", 1),
    )


@app.route("/upload", methods=["GET", "POST"])
def upload():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    if request.method == "GET":
        return redirect(url_for("orders"))
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            return render_template(
                "upload.html", error="Please choose a CSV file to upload.", summary=None
            )
        try:
            summary = _handle_order_upload(file)
        except Exception as exc:
            return render_template(
                "upload.html",
                error=f"Upload failed: {exc}",
                summary=None,
            )

        return render_template(
            "upload.html",
            error="",
            summary=summary,
        )
    return redirect(url_for("orders"))


@app.route("/orders/upload", methods=["POST"])
def orders_upload():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    file = request.files.get("file")
    if not file:
        return redirect(url_for("orders"))
    try:
        _handle_order_upload(file)
    except Exception:
        return redirect(url_for("orders"))

    return redirect(url_for("orders"))


@app.route("/orders/load-report/upload", methods=["POST"])
def orders_load_report_upload():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    file = request.files.get("file")
    if not file or not getattr(file, "filename", ""):
        return redirect(url_for("orders"))
    try:
        _handle_load_report_upload(file)
    except Exception as exc:
        logger.exception("Load report upload failed.")
        message = str(exc).strip() or "Load report upload failed."
        return redirect(
            url_for(
                "orders",
                load_report_status="error",
                load_report_message=message[:200],
            )
        )
    return redirect(url_for("orders", load_report_status="ok"))


@app.route("/api/orders/upload", methods=["POST"])
def api_orders_upload():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    file = request.files.get("file")
    if not file or not getattr(file, "filename", ""):
        return jsonify({"error": "Please choose a CSV file to upload."}), 400
    try:
        summary = _handle_order_upload(file)
    except UploadValidationError as exc:
        blocked_summary = exc.summary or {}
        unmapped_items = blocked_summary.get("unmapped_items") or []
        unmapped_suggestions = _build_unmapped_suggestions(unmapped_items)
        response = {
            "error": str(exc),
            "blocked": True,
            "total_rows": blocked_summary.get("total_rows") or 0,
            "total_orders": 0,
            "mapping_rate": round(blocked_summary.get("mapping_rate") or 0, 2),
            "unmapped_count": len(unmapped_items),
            "new_orders": 0,
            "changed_orders": 0,
            "unchanged_orders": 0,
            "reopened_orders": 0,
            "dropped_orders": 0,
            "unmapped_items": unmapped_suggestions,
        }
        return jsonify(response), 400
    except Exception as exc:
        return jsonify({"error": f"Upload failed: {exc}"}), 400
    unmapped_items = summary.get("unmapped_items") or []
    unmapped_suggestions = _build_unmapped_suggestions(unmapped_items)
    response = {
        "filename": getattr(file, "filename", ""),
        "total_rows": summary.get("total_rows"),
        "total_orders": len(summary.get("orders") or []),
        "mapping_rate": round(summary.get("mapping_rate") or 0, 2),
        "unmapped_count": len(summary.get("unmapped_items") or []),
        "new_orders": summary.get("new_orders"),
        "changed_orders": summary.get("changed_orders"),
        "unchanged_orders": summary.get("unchanged_orders"),
        "reopened_orders": summary.get("reopened_orders"),
        "dropped_orders": summary.get("dropped_orders"),
        "unmapped_items": unmapped_suggestions,
    }
    return jsonify(response)


@app.route("/api/skus/bulk-add", methods=["POST"])
def api_bulk_add_skus():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    payload = request.get_json(silent=True) or {}
    items = payload.get("items") or []
    if not isinstance(items, list) or not items:
        return jsonify({"error": "No SKU entries provided."}), 400

    cargo_pattern = re.compile(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)")

    def _cargo_length_from_text(value):
        if not value:
            return None
        match = cargo_pattern.search(str(value))
        if not match:
            return None
        try:
            length = float(match.group(2))
        except (TypeError, ValueError):
            return None
        return length + 4

    created = 0
    for entry in items:
        sku = (entry.get("sku") or "").strip().upper()
        if not sku:
            continue
        try:
            length = float(entry.get("length_with_tongue_ft") or 0)
        except (TypeError, ValueError):
            length = 0.0
        try:
            max_flat = int(entry.get("max_stack_flat_bed") or 1)
        except (TypeError, ValueError):
            max_flat = 1
        try:
            max_step = int(entry.get("max_stack_step_deck") or max_flat or 1)
        except (TypeError, ValueError):
            max_step = max_flat or 1

        category_value = (entry.get("bin") or entry.get("category") or "").strip()
        category_upper = category_value.upper() if category_value else ""
        if "CARGO" in category_upper:
            max_flat = 1
            max_step = 1
            if length <= 0:
                inferred = _cargo_length_from_text(sku) or _cargo_length_from_text(entry.get("description"))
                if inferred:
                    length = inferred

        if length <= 0:
            continue

        spec = {
            "sku": sku,
            "description": (entry.get("description") or "").strip(),
            "category": category_value or "UNKNOWN",
            "length_with_tongue_ft": length,
            "max_stack_step_deck": max_step,
            "max_stack_flat_bed": max_flat,
            "notes": (entry.get("notes") or "").strip(),
            "added_at": datetime.utcnow().isoformat(timespec="seconds"),
        }
        db.upsert_sku_spec(spec)
        created += 1

    return jsonify({"status": "ok", "created": created})


@app.route("/orders")
def orders():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    role = _get_session_role()
    profile_default_plants = session.get(SESSION_PROFILE_DEFAULT_PLANTS_KEY) or []
    allowed_plants = _get_allowed_plants()
    active_session_id = _get_active_planning_session_id()
    active_session = db.get_planning_session(active_session_id) if active_session_id else None
    if active_session and active_session.get("plant_code") not in allowed_plants:
        _set_active_planning_session_id(None)
        active_session = None
        active_session_id = None
    active_session_status = _normalize_session_status(active_session.get("status")) if active_session else ""
    needs_session = bool(request.args.get("needs_session"))
    needs_replace = bool(request.args.get("needs_replace"))
    plant_filters = _resolve_plant_filters(request.args.get("plants") or request.args.get("plant"))
    plant_scope = plant_filters or allowed_plants
    today_override = _resolve_today_override(request.args.get("today"))
    today = today_override or date.today()
    hide_past_due = _coerce_bool_value(request.args.get("hide_past_due"))

    due_filter = (request.args.get("due") or "").upper()
    due_start = request.args.get("due_start", "")
    due_end = request.args.get("due_end", "")
    if due_filter == "PAST_DUE":
        due_start = ""
        due_end = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    elif due_filter == "THIS_WEEK":
        start_week, end_week, _, _ = _week_bounds(today)
        due_start = start_week.strftime("%Y-%m-%d")
        due_end = end_week.strftime("%Y-%m-%d")
    elif due_filter == "NEXT_WEEK":
        _, _, next_start, next_end = _week_bounds(today)
        due_start = next_start.strftime("%Y-%m-%d")
        due_end = next_end.strftime("%Y-%m-%d")
    elif not due_filter and (due_start or due_end):
        due_filter = "CUSTOM"

    assignment_filter = (request.args.get("assigned") or "").upper()
    filters = {
        "plants": plant_scope,
        "state": request.args.get("state", ""),
        "cust_name": request.args.get("customer", ""),
        "due_start": due_start or None,
        "due_end": due_end or None,
        "assignment_status": assignment_filter,
    }
    sort_key = request.args.get("sort", "due_date")
    data = order_service.list_orders(filters=filters, sort_key=sort_key)
    orders_list = data["orders"]
    _annotate_orders_due_status(orders_list, today=today)
    if hide_past_due:
        orders_list = _filter_out_past_due_orders(orders_list)
        data["summary"] = order_service.summarize_orders(orders_list)
    snapshot_so_nums = [
        _normalize_so_num_for_load_report_match(order.get("so_num"))
        for order in orders_list
        if _normalize_so_num_for_load_report_match(order.get("so_num"))
    ]
    load_report_assignments = db.list_latest_load_report_assignments_by_so_nums(snapshot_so_nums)
    orders_snapshot = _build_orders_snapshot(
        orders_list,
        today=today,
        load_assignment_map=load_report_assignments,
    )

    plants = allowed_plants
    states = _distinct([order.get("state") for order in orders_list])
    customers = _distinct([order.get("cust_name") for order in orders_list])

    card_filters = {
        "plants": allowed_plants,
        "state": filters.get("state", ""),
        "cust_name": filters.get("cust_name", ""),
    }
    if hide_past_due:
        orders_by_plant = _count_active_orders_by_plant_from_rows(
            orders_list,
            plants=allowed_plants,
        )
    else:
        orders_by_plant = db.count_orders_by_plant(card_filters)
    plant_cards = []
    for plant in [code for code in plants if code != "CL"]:
        plant_cards.append(
            {
                "code": plant,
                "name": PLANT_NAMES.get(plant, plant),
                "orders": orders_by_plant.get(plant, 0),
            }
        )

    optimize_origin = ""
    if plant_filters:
        optimize_origin = plant_filters[0]
    elif plants:
        optimize_origin = plants[0]
    optimize_defaults = _default_optimize_form(plant_code=optimize_origin or None)
    if optimize_origin:
        optimize_defaults["origin_plant"] = optimize_origin
    optimize_defaults["state_filters"] = []
    optimize_defaults["customer_filters"] = []
    optimize_defaults["enforce_time_window"] = True
    optimize_defaults["batch_horizon_enabled"] = True
    optimize_defaults["batch_end_date"] = _default_batch_end_date().strftime("%Y-%m-%d")
    optimize_defaults["orders_start_date"] = date.today().strftime("%Y-%m-%d")
    optimize_defaults["ignore_due_date"] = False
    optimize_defaults["algorithm_version"] = "v2"
    optimize_defaults["compare_algorithms"] = False

    template_session_id = request.args.get("session_template_id")
    try:
        template_session_id = int(template_session_id) if template_session_id else None
    except (TypeError, ValueError):
        template_session_id = None
    if template_session_id:
        template_session = db.get_planning_session(template_session_id)
        if template_session and template_session.get("config_json"):
            try:
                config = json.loads(template_session.get("config_json") or "{}")
            except json.JSONDecodeError:
                config = {}
            optimize_defaults["origin_plant"] = config.get("origin_plant") or optimize_defaults["origin_plant"]
            optimize_defaults["capacity_feet"] = str(config.get("capacity_feet") or optimize_defaults["capacity_feet"])
            optimize_defaults["trailer_type"] = stack_calculator.normalize_trailer_type(
                config.get("trailer_type") or optimize_defaults["trailer_type"],
                default=optimize_defaults["trailer_type"],
            )
            optimize_defaults["max_detour_pct"] = str(config.get("max_detour_pct") or optimize_defaults["max_detour_pct"])
            optimize_defaults["time_window_days"] = str(config.get("time_window_days") or optimize_defaults["time_window_days"])
            optimize_defaults["geo_radius"] = str(config.get("geo_radius") or optimize_defaults["geo_radius"])
            optimize_defaults["stack_overflow_max_height"] = str(
                config.get("stack_overflow_max_height") or optimize_defaults["stack_overflow_max_height"]
            )
            optimize_defaults["max_back_overhang_ft"] = str(
                config.get("max_back_overhang_ft") or optimize_defaults["max_back_overhang_ft"]
            )
            optimize_defaults["upper_two_across_max_length_ft"] = str(
                config.get("upper_two_across_max_length_ft")
                or optimize_defaults.get("upper_two_across_max_length_ft", DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT)
            )
            optimize_defaults["upper_deck_exception_max_length_ft"] = str(
                config.get("upper_deck_exception_max_length_ft")
                or optimize_defaults.get(
                    "upper_deck_exception_max_length_ft",
                    DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
                )
            )
            optimize_defaults["upper_deck_exception_overhang_allowance_ft"] = str(
                config.get("upper_deck_exception_overhang_allowance_ft")
                or optimize_defaults.get(
                    "upper_deck_exception_overhang_allowance_ft",
                    DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
                )
            )
            optimize_defaults["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
                config.get("upper_deck_exception_categories")
                or optimize_defaults.get("upper_deck_exception_categories")
                or DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
                default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
            )
            optimize_defaults["enforce_time_window"] = bool(config.get("enforce_time_window", True))
            optimize_defaults["batch_horizon_enabled"] = bool(config.get("batch_horizon_enabled", False))
            optimize_defaults["batch_end_date"] = config.get("batch_end_date") or optimize_defaults["batch_end_date"]
            optimize_defaults["state_filters"] = config.get("state_filters") or []
            optimize_defaults["customer_filters"] = config.get("customer_filters") or []
            optimize_defaults["ignore_due_date"] = bool(config.get("ignore_due_date", False))
            config_start_date = (config.get("orders_start_date") or "").strip()
            if not config_start_date and config.get("ignore_past_due", True):
                config_start_date = date.today().strftime("%Y-%m-%d")
            optimize_defaults["orders_start_date"] = config_start_date or optimize_defaults["orders_start_date"]
            optimize_defaults["algorithm_version"] = "v2"
            optimize_defaults["compare_algorithms"] = False

    optimize_defaults["trailer_type"] = stack_calculator.normalize_trailer_type(
        optimize_defaults.get("trailer_type"),
        default="STEP_DECK",
    )
    optimize_defaults["capacity_feet"] = str(
        _capacity_for_trailer_setting(
            optimize_defaults.get("trailer_type"),
            requested_capacity=optimize_defaults.get("capacity_feet"),
            default_capacity=_get_optimizer_default_settings().get("capacity_feet", 53.0),
        )
    )
    last_upload = db.get_last_upload()
    upload_history = db.list_upload_history(limit=12)
    last_load_report_upload = db.get_last_load_report_upload()
    rejected_orders = sum(1 for order in orders_list if order.get("is_excluded"))
    due_dates = [
        _parse_date(order.get("due_date"))
        for order in orders_list
        if order.get("due_date")
    ]
    due_dates = [value for value in due_dates if value]
    ship_date_range = None
    if due_dates:
        ship_date_range = {
            "start": min(due_dates).strftime("%Y-%m-%d"),
            "end": max(due_dates).strftime("%Y-%m-%d"),
        }

    strategic_setting = _get_effective_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY)
    strategic_customers_raw = strategic_setting.get("value_text") or ""
    strategic_customers = _parse_strategic_customers(strategic_customers_raw)
    strategic_customer_groups = []
    eligible_strategic_customers = [
        entry for entry in strategic_customers if entry.get("include_in_optimizer_workbench", True)
    ]
    for entry in eligible_strategic_customers:
        matching_customers = [
            cust
            for cust in customers
            if customer_rules.matches_any_customer_pattern(cust, entry.get("patterns"))
        ]
        if matching_customers:
            strategic_customer_groups.append(
                {
                    "key": entry["key"],
                    "label": entry["label"],
                    "customers": matching_customers,
                }
            )
    strategic_orders = {entry["key"]: [] for entry in eligible_strategic_customers}
    other_orders = []
    for order in orders_list:
        cust_name = order.get("cust_name") or ""
        matched_key = None
        for entry in eligible_strategic_customers:
            if customer_rules.matches_any_customer_pattern(cust_name, entry.get("patterns")):
                matched_key = entry["key"]
                break
        if matched_key:
            strategic_orders[matched_key].append(order)
        else:
            other_orders.append(order)

    show_section = (request.args.get("show") or "").strip().lower()

    def _build_section(key, label, section_orders, default_limit):
        total_count = len(section_orders)
        render_all = show_section == key
        limit = None if render_all else default_limit
        # Keep all rows in the DOM so client-side search can match any order ID;
        # overflow rows stay collapsed via `section.limit` + `group-hidden`.
        render_orders = section_orders
        hidden_count = 0 if (render_all or limit is None) else max(total_count - limit, 0)

        show_url = None
        collapse_url = None
        if hidden_count and not render_all:
            params = request.args.to_dict(flat=True)
            params["show"] = key
            show_url = url_for("orders", **params)
        if render_all:
            params = request.args.to_dict(flat=True)
            params.pop("show", None)
            collapse_url = url_for("orders", **params)

        return {
            "key": key,
            "label": label,
            "orders": render_orders,
            "limit": limit,
            "hidden_count": hidden_count,
            "total_count": total_count,
            "show_all": render_all,
            "show_url": show_url,
            "collapse_url": collapse_url,
        }

    order_sections = []
    for entry in eligible_strategic_customers:
        section_orders = strategic_orders.get(entry["key"]) or []
        if not section_orders:
            continue
        order_sections.append(
            _build_section(entry["key"], entry["label"], section_orders, default_limit=5)
        )
    order_sections.append(
        _build_section("other", "Other Customers", other_orders, default_limit=5)
    )

    show_more_plants = False
    if profile_default_plants and role != ROLE_ADMIN:
        if not plant_filters:
            show_more_plants = True
        else:
            show_more_plants = any(code not in profile_default_plants for code in plant_filters)

    today_override_value = today_override.strftime("%Y-%m-%d") if today_override else ""
    today_override_label = today_override.strftime("%b %d, %Y") if today_override else ""
    today_display_label = (today_override or today).strftime("%b %d, %Y")
    load_report_status = (request.args.get("load_report_status") or "").strip().lower()
    if load_report_status not in {"ok", "error"}:
        load_report_status = ""
    load_report_message = (request.args.get("load_report_message") or "").strip()
    plant_filter_param = ",".join(plant_filters) if plant_filters else ""
    reset_args = {}
    if plant_filter_param:
        reset_args["plants"] = plant_filter_param
    if today_override_value:
        reset_args["today"] = today_override_value
    if hide_past_due:
        reset_args["hide_past_due"] = "1"
    orders_reset_url = url_for("orders", **reset_args)
    toggle_args = request.args.to_dict(flat=True)
    if hide_past_due:
        toggle_args.pop("hide_past_due", None)
    else:
        toggle_args["hide_past_due"] = "1"
    past_due_toggle_url = url_for("orders", **_clean_query_params(toggle_args))
    optimize_trailer_context = _build_optimizer_workbench_trailer_defaults(
        plants,
        optimizer_defaults=_get_optimizer_default_settings(),
    )

    return render_template(
        "orders.html",
        orders=orders_list,
        order_sections=order_sections,
        strategic_customers=strategic_customers,
        strategic_customer_groups=strategic_customer_groups,
        summary=data["summary"],
        filters=filters,
        plants=plants,
        states=states,
        customers=customers,
        trailer_profile_options=optimize_trailer_context["options"],
        optimize_plant_trailer_defaults=optimize_trailer_context["defaults_by_plant"],
        optimize_defaults=optimize_defaults,
        optimize_errors={},
        optimize_summary=None,
        last_upload=last_upload,
        upload_history=upload_history,
        last_load_report_upload=last_load_report_upload,
        rejected_orders=rejected_orders,
        ship_date_range=ship_date_range,
        plant_filters=plant_filters,
        plant_filter_param=plant_filter_param,
        due_filter=due_filter,
        due_start=due_start,
        due_end=due_end,
        assignment_filter=assignment_filter,
        plant_cards=plant_cards,
        orders_snapshot=orders_snapshot,
        hide_past_due=hide_past_due,
        past_due_toggle_url=past_due_toggle_url,
        orders_reset_url=orders_reset_url,
        today_override=today_override,
        today_override_value=today_override_value,
        today_override_label=today_override_label,
        today_display_label=today_display_label,
        load_report_status=load_report_status,
        load_report_message=load_report_message,
        profile_default_plants=profile_default_plants,
        show_more_plants=show_more_plants,
        active_session=active_session,
        active_session_status=active_session_status,
        active_session_id=active_session_id,
        needs_session=needs_session,
        needs_replace=needs_replace,
        is_admin=role == ROLE_ADMIN,
        optimizer_v2_enabled=OPTIMIZER_V2_ENABLED,
    )


@app.route("/orders/clear", methods=["POST"])
def clear_orders():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()
    db.clear_loads()
    db.clear_orders()
    db.mark_upload_history_deleted()
    db.clear_load_report_data()
    return redirect(url_for("orders"))


def _coerce_order_ids_for_scope(raw_ids):
    cleaned_ids = []
    for raw in raw_ids or []:
        try:
            parsed = int(raw)
        except (TypeError, ValueError):
            continue
        if parsed > 0:
            cleaned_ids.append(parsed)
    if not cleaned_ids:
        return []

    records = db.list_orders_by_ids(cleaned_ids)
    allowed_plants = set(_get_allowed_plants())
    unauthorized = [
        row.get("id")
        for row in records
        if _normalize_plant_code(row.get("plant")) not in allowed_plants
    ]
    if unauthorized:
        abort(403)
    return [row.get("id") for row in records if row.get("id")]


@app.route("/orders/exclude", methods=["POST"])
def exclude_orders():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    order_ids = _coerce_order_ids_for_scope(request.form.getlist("order_ids"))
    order_service.exclude_orders(order_ids)
    return redirect(url_for("orders"))


@app.route("/orders/include", methods=["POST"])
def include_orders():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    order_ids = _coerce_order_ids_for_scope(request.form.getlist("order_ids"))
    order_service.include_orders(order_ids)
    return redirect(url_for("orders"))


@app.route("/orders/optimize", methods=["POST"])
def orders_optimize():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    role = _get_session_role()
    profile_default_plants = session.get(SESSION_PROFILE_DEFAULT_PLANTS_KEY) or []
    allowed_plants = _get_allowed_plants()
    active_session_id = _get_active_planning_session_id()
    active_session = db.get_planning_session(active_session_id) if active_session_id else None
    if active_session and active_session.get("plant_code") not in allowed_plants:
        _set_active_planning_session_id(None)
        active_session = None
        active_session_id = None
    active_session_status = _normalize_session_status(active_session.get("status")) if active_session else ""
    optimize_form = request.form.copy()
    origin_plant = _normalize_plant_code(optimize_form.get("origin_plant"))
    optimize_mode = (optimize_form.get("optimize_mode") or "auto").strip().lower()
    if optimize_mode not in {"auto", "manual"}:
        optimize_mode = "auto"
    manual_mode_error = ""

    if optimize_mode == "manual":
        # Paste mode ignores state/customer filters, but still honors origin/trailer preference.
        if hasattr(optimize_form, "setlist"):
            optimize_form.setlist("opt_states", [])
            optimize_form.setlist("opt_customers", [])
        optimize_form["batch_horizon_enabled"] = ""
        optimize_form["batch_end_date"] = ""

        raw_manual = str(optimize_form.get("manual_order_input") or "").strip()
        manual_so_nums = _parse_manual_so_nums(raw_manual)

        if manual_so_nums:
            manual_lines = db.list_order_lines_by_so_nums(manual_so_nums)
            matched_plants = sorted(
                {
                    _normalize_plant_code(line.get("plant"))
                    for line in manual_lines
                    if not _coerce_bool_value(line.get("is_excluded"))
                    and _normalize_plant_code(line.get("plant"))
                }
            )
            if not matched_plants:
                manual_mode_error = "No eligible open orders were found for the pasted order numbers."
            elif len(matched_plants) > 1:
                manual_mode_error = (
                    "Pasted orders span multiple plants. Paste orders from one plant only."
                )
            else:
                inferred_plant = matched_plants[0]
                if origin_plant and inferred_plant != origin_plant:
                    manual_mode_error = (
                        f"Pasted orders resolve to {inferred_plant}, but selected origin is {origin_plant}."
                    )
                elif inferred_plant not in allowed_plants:
                    manual_mode_error = "Pasted orders do not match your allowed plant scope."
                else:
                    origin_plant = inferred_plant
        optimize_form["origin_plant"] = origin_plant or ""

    if origin_plant:
        plant_defaults = _get_optimizer_settings_for_plant(origin_plant)
        requested_trailer = stack_calculator.normalize_trailer_type(
            optimize_form.get("trailer_type"),
            default=plant_defaults["trailer_type"],
        )
        optimize_form["trailer_type"] = requested_trailer
        optimize_form["capacity_feet"] = str(
            _capacity_for_trailer_setting(
                requested_trailer,
                requested_capacity=optimize_form.get("capacity_feet"),
                default_capacity=plant_defaults["capacity_feet"],
            )
        )

    replace_session = optimize_form.get("replace_session") == "1"

    def _collect_form_data():
        form_data = _default_optimize_form()
        form_data["origin_plant"] = origin_plant
        form_data["trailer_type"] = stack_calculator.normalize_trailer_type(
            optimize_form.get("trailer_type"),
            default=form_data.get("trailer_type", "STEP_DECK"),
        )
        form_data["time_window_days"] = optimize_form.get(
            "time_window_days", form_data.get("time_window_days", "7")
        )
        form_data["geo_radius"] = optimize_form.get("geo_radius", form_data.get("geo_radius", "100"))
        form_data["max_detour_pct"] = optimize_form.get("max_detour_pct", form_data.get("max_detour_pct", "15"))
        form_data["capacity_feet"] = optimize_form.get("capacity_feet", form_data.get("capacity_feet", "53"))
        form_data["capacity_feet"] = str(
            _capacity_for_trailer_setting(
                form_data.get("trailer_type"),
                requested_capacity=form_data.get("capacity_feet"),
                default_capacity=_get_optimizer_default_settings().get("capacity_feet", 53.0),
            )
        )
        form_data["stack_overflow_max_height"] = optimize_form.get(
            "stack_overflow_max_height",
            form_data.get("stack_overflow_max_height", str(DEFAULT_STACK_OVERFLOW_MAX_HEIGHT)),
        )
        form_data["max_back_overhang_ft"] = optimize_form.get(
            "max_back_overhang_ft",
            form_data.get("max_back_overhang_ft", str(DEFAULT_MAX_BACK_OVERHANG_FT)),
        )
        form_data["upper_two_across_max_length_ft"] = optimize_form.get(
            "upper_two_across_max_length_ft",
            form_data.get(
                "upper_two_across_max_length_ft",
                str(DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT),
            ),
        )
        form_data["upper_deck_exception_max_length_ft"] = optimize_form.get(
            "upper_deck_exception_max_length_ft",
            form_data.get(
                "upper_deck_exception_max_length_ft",
                str(DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT),
            ),
        )
        form_data["upper_deck_exception_overhang_allowance_ft"] = optimize_form.get(
            "upper_deck_exception_overhang_allowance_ft",
            form_data.get(
                "upper_deck_exception_overhang_allowance_ft",
                str(DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT),
            ),
        )
        form_data["upper_deck_exception_categories"] = stack_calculator.normalize_upper_deck_exception_categories(
            optimize_form.getlist("upper_deck_exception_categories")
            or form_data.get("upper_deck_exception_categories")
            or DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
            default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
        )
        form_data["state_filters"] = [
            value.strip().upper()
            for value in optimize_form.getlist("opt_states")
            if value and value.strip()
        ]
        form_data["customer_filters"] = [
            value.strip()
            for value in optimize_form.getlist("opt_customers")
            if value and value.strip()
        ]
        ui_toggles = "opt_toggles" in optimize_form
        form_data["ignore_due_date"] = bool(optimize_form.get("ignore_due_date")) if ui_toggles else False
        form_data["enforce_time_window"] = bool(optimize_form.get("enforce_time_window")) if ui_toggles else True
        if form_data["ignore_due_date"]:
            form_data["enforce_time_window"] = False
        form_data["batch_horizon_enabled"] = bool(optimize_form.get("batch_horizon_enabled")) if ui_toggles else False
        form_data["batch_end_date"] = optimize_form.get("batch_end_date") or _default_batch_end_date().strftime(
            "%Y-%m-%d"
        )
        form_data["orders_start_date"] = optimize_form.get("orders_start_date") or date.today().strftime("%Y-%m-%d")
        form_data["algorithm_version"] = "v2"
        form_data["compare_algorithms"] = False
        mode = (optimize_form.get("optimize_mode") or "auto").strip().lower()
        form_data["optimize_mode"] = mode if mode in {"auto", "manual"} else "auto"
        form_data["manual_order_input"] = optimize_form.get("manual_order_input", "")
        form_data["order_category_scope"] = order_categories.normalize_order_category_scope(
            optimize_form.get("order_category_scope"),
            default=form_data.get("order_category_scope", order_categories.ORDER_CATEGORY_SCOPE_ALL),
        )
        return form_data

    if active_session and active_session_status == "DRAFT" and not replace_session:
        form_data = _collect_form_data()
        result = {
            "errors": {"session": "Replace existing load planning session to continue."},
            "form_data": form_data,
            "success_message": "",
            "summary": None,
        }
    elif manual_mode_error:
        form_data = _collect_form_data()
        result = {
            "errors": {"manual_order_input": manual_mode_error},
            "form_data": form_data,
            "success_message": "",
            "summary": None,
        }
    elif optimize_mode == "auto" and not _collect_form_data().get("customer_filters"):
        form_data = _collect_form_data()
        result = {
            "errors": {"customer_filters": "Select at least one customer to run optimization."},
            "form_data": form_data,
            "success_message": "",
            "summary": None,
        }
    elif origin_plant and origin_plant not in allowed_plants:
        form_data = _collect_form_data()
        result = {
            "errors": {"origin_plant": "Select a plant within your scope."},
            "form_data": form_data,
            "success_message": "",
            "summary": None,
        }
    else:
        if replace_session and active_session:
            _archive_session_and_release_loads(active_session_id)
        created_by = _get_session_profile_name() or _get_session_role()

        def _session_factory(form_data, params):
            session_plant = (
                _normalize_plant_code(params.get("origin_plant"))
                or _normalize_plant_code(form_data.get("origin_plant"))
                or origin_plant
            )
            return _create_planning_session(created_by, session_plant, form_data, params)
        result = load_builder.build_loads(
            optimize_form,
            session_factory=_session_factory,
            created_by=created_by,
        )

    resolved_origin_plant = _normalize_plant_code(
        (result.get("form_data") or {}).get("origin_plant") or origin_plant
    )

    if not result["errors"]:
        session_id = result.get("session_id")
        _set_active_planning_session_id(session_id)
        redirect_args = {"plants": resolved_origin_plant, "session_id": session_id}
        comparison = result.get("algorithm_comparison") or {}
        v1 = comparison.get("v1") or {}
        v2 = comparison.get("v2") or {}
        if v1 and v2:
            redirect_args.update(
                {
                    "alg_cmp": 1,
                    "alg_sel": comparison.get("selected") or "v2",
                    "v1_loads": int(v1.get("total_loads") or 0),
                    "v1_util": round(v1.get("avg_utilization") or 0.0, 1),
                    "v1_cost": round(v1.get("est_cost") or 0.0, 2),
                    "v2_loads": int(v2.get("total_loads") or 0),
                    "v2_util": round(v2.get("avg_utilization") or 0.0, 1),
                    "v2_cost": round(v2.get("est_cost") or 0.0, 2),
                }
            )
        return redirect(url_for("loads", **redirect_args))

    if result["errors"].get("order_lines") and resolved_origin_plant:
        resumable = _find_resumable_planning_session(resolved_origin_plant)
        if resumable and resumable.get("id"):
            resume_id = resumable["id"]
            _set_active_planning_session_id(resume_id)
            return redirect(url_for("loads", plants=resolved_origin_plant, session_id=resume_id))

    plant_filters = _resolve_plant_filters(request.args.get("plants") or request.args.get("plant"))
    plant_scope = plant_filters or allowed_plants

    due_filter = (request.args.get("due") or "").upper()
    due_start = request.args.get("due_start", "")
    due_end = request.args.get("due_end", "")
    today_override = _resolve_today_override(request.values.get("today"))
    today = today_override or date.today()
    hide_past_due = _coerce_bool_value(request.values.get("hide_past_due"))
    if due_filter == "PAST_DUE":
        due_start = ""
        due_end = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    elif due_filter == "THIS_WEEK":
        start_week, end_week, _, _ = _week_bounds(today)
        due_start = start_week.strftime("%Y-%m-%d")
        due_end = end_week.strftime("%Y-%m-%d")
    elif due_filter == "NEXT_WEEK":
        _, _, next_start, next_end = _week_bounds(today)
        due_start = next_start.strftime("%Y-%m-%d")
        due_end = next_end.strftime("%Y-%m-%d")
    elif not due_filter and (due_start or due_end):
        due_filter = "CUSTOM"

    assignment_filter = (request.args.get("assigned") or "").upper()
    filters = {
        "plants": plant_scope,
        "state": request.args.get("state", ""),
        "cust_name": request.args.get("customer", ""),
        "due_start": due_start or None,
        "due_end": due_end or None,
        "assignment_status": assignment_filter,
    }
    data = order_service.list_orders(filters=filters, sort_key="due_date")
    orders_list = data["orders"]
    _annotate_orders_due_status(orders_list, today=today)
    if hide_past_due:
        orders_list = _filter_out_past_due_orders(orders_list)
        data["summary"] = order_service.summarize_orders(orders_list)
    snapshot_so_nums = [
        _normalize_so_num_for_load_report_match(order.get("so_num"))
        for order in orders_list
        if _normalize_so_num_for_load_report_match(order.get("so_num"))
    ]
    load_report_assignments = db.list_latest_load_report_assignments_by_so_nums(snapshot_so_nums)
    orders_snapshot = _build_orders_snapshot(
        orders_list,
        today=today,
        load_assignment_map=load_report_assignments,
    )
    plants = allowed_plants
    states = _distinct([order.get("state") for order in orders_list])
    customers = _distinct([order.get("cust_name") for order in orders_list])
    last_upload = db.get_last_upload()
    upload_history = db.list_upload_history(limit=12)
    last_load_report_upload = db.get_last_load_report_upload()

    card_filters = {
        "plants": allowed_plants,
        "state": filters.get("state", ""),
        "cust_name": filters.get("cust_name", ""),
    }
    if hide_past_due:
        orders_by_plant = _count_active_orders_by_plant_from_rows(
            orders_list,
            plants=allowed_plants,
        )
    else:
        orders_by_plant = db.count_orders_by_plant(card_filters)
    plant_cards = [
        {
            "code": plant,
            "name": PLANT_NAMES.get(plant, plant),
            "orders": orders_by_plant.get(plant, 0),
        }
        for plant in plants
        if plant != "CL"
    ]

    rejected_orders = sum(1 for order in orders_list if order.get("is_excluded"))
    due_dates = [
        _parse_date(order.get("due_date"))
        for order in orders_list
        if order.get("due_date")
    ]
    due_dates = [value for value in due_dates if value]
    ship_date_range = None
    if due_dates:
        ship_date_range = {
            "start": min(due_dates).strftime("%Y-%m-%d"),
            "end": max(due_dates).strftime("%Y-%m-%d"),
        }

    strategic_setting = _get_effective_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY)
    strategic_customers_raw = strategic_setting.get("value_text") or ""
    strategic_customers = _parse_strategic_customers(strategic_customers_raw)
    strategic_customer_groups = []
    eligible_strategic_customers = [
        entry for entry in strategic_customers if entry.get("include_in_optimizer_workbench", True)
    ]
    for entry in eligible_strategic_customers:
        matching_customers = [
            cust
            for cust in customers
            if customer_rules.matches_any_customer_pattern(cust, entry.get("patterns"))
        ]
        if matching_customers:
            strategic_customer_groups.append(
                {
                    "key": entry["key"],
                    "label": entry["label"],
                    "customers": matching_customers,
                }
            )
    strategic_orders = {entry["key"]: [] for entry in eligible_strategic_customers}
    other_orders = []
    for order in orders_list:
        cust_name = order.get("cust_name") or ""
        matched_key = None
        for entry in eligible_strategic_customers:
            if customer_rules.matches_any_customer_pattern(cust_name, entry.get("patterns")):
                matched_key = entry["key"]
                break
        if matched_key:
            strategic_orders[matched_key].append(order)
        else:
            other_orders.append(order)

    show_section = (request.args.get("show") or "").strip().lower()

    def _build_section(key, label, section_orders, default_limit):
        total_count = len(section_orders)
        render_all = show_section == key
        limit = None if render_all else default_limit
        # Keep all rows in the DOM so client-side search can match any order ID;
        # overflow rows stay collapsed via `section.limit` + `group-hidden`.
        render_orders = section_orders
        hidden_count = 0 if (render_all or limit is None) else max(total_count - limit, 0)

        show_url = None
        collapse_url = None
        if hidden_count and not render_all:
            params = request.args.to_dict(flat=True)
            params["show"] = key
            show_url = url_for("orders", **params)
        if render_all:
            params = request.args.to_dict(flat=True)
            params.pop("show", None)
            collapse_url = url_for("orders", **params)

        return {
            "key": key,
            "label": label,
            "orders": render_orders,
            "limit": limit,
            "hidden_count": hidden_count,
            "total_count": total_count,
            "show_all": render_all,
            "show_url": show_url,
            "collapse_url": collapse_url,
        }

    order_sections = []
    for entry in eligible_strategic_customers:
        section_orders = strategic_orders.get(entry["key"]) or []
        if not section_orders:
            continue
        order_sections.append(
            _build_section(entry["key"], entry["label"], section_orders, default_limit=5)
        )
    order_sections.append(
        _build_section("other", "Other Customers", other_orders, default_limit=5)
    )

    show_more_plants = False
    if profile_default_plants and role != ROLE_ADMIN:
        if not plant_filters:
            show_more_plants = True
        else:
            show_more_plants = any(code not in profile_default_plants for code in plant_filters)

    today_override_value = today_override.strftime("%Y-%m-%d") if today_override else ""
    today_override_label = today_override.strftime("%b %d, %Y") if today_override else ""
    today_display_label = (today_override or today).strftime("%b %d, %Y")
    load_report_status = (request.args.get("load_report_status") or "").strip().lower()
    if load_report_status not in {"ok", "error"}:
        load_report_status = ""
    load_report_message = (request.args.get("load_report_message") or "").strip()
    plant_filter_param = ",".join(plant_filters) if plant_filters else ""
    reset_args = {}
    if plant_filter_param:
        reset_args["plants"] = plant_filter_param
    if today_override_value:
        reset_args["today"] = today_override_value
    if hide_past_due:
        reset_args["hide_past_due"] = "1"
    orders_reset_url = url_for("orders", **reset_args)
    toggle_args = request.args.to_dict(flat=True)
    if hide_past_due:
        toggle_args.pop("hide_past_due", None)
    else:
        toggle_args["hide_past_due"] = "1"
    past_due_toggle_url = url_for("orders", **_clean_query_params(toggle_args))
    optimize_trailer_context = _build_optimizer_workbench_trailer_defaults(
        plants,
        optimizer_defaults=_get_optimizer_default_settings(),
    )

    return render_template(
        "orders.html",
        orders=orders_list,
        order_sections=order_sections,
        strategic_customers=strategic_customers,
        strategic_customer_groups=strategic_customer_groups,
        summary=data["summary"],
        filters=filters,
        plants=plants,
        states=states,
        customers=customers,
        trailer_profile_options=optimize_trailer_context["options"],
        optimize_plant_trailer_defaults=optimize_trailer_context["defaults_by_plant"],
        optimize_defaults=result["form_data"],
        optimize_errors=result["errors"],
        optimize_summary=result["summary"],
        last_upload=last_upload,
        upload_history=upload_history,
        last_load_report_upload=last_load_report_upload,
        rejected_orders=rejected_orders,
        ship_date_range=ship_date_range,
        plant_filters=plant_filters,
        plant_filter_param=plant_filter_param,
        due_filter=due_filter,
        due_start=due_start,
        due_end=due_end,
        assignment_filter=assignment_filter,
        plant_cards=plant_cards,
        orders_snapshot=orders_snapshot,
        hide_past_due=hide_past_due,
        past_due_toggle_url=past_due_toggle_url,
        orders_reset_url=orders_reset_url,
        today_override=today_override,
        today_override_value=today_override_value,
        today_override_label=today_override_label,
        today_display_label=today_display_label,
        load_report_status=load_report_status,
        load_report_message=load_report_message,
        profile_default_plants=profile_default_plants,
        show_more_plants=show_more_plants,
        active_session=active_session,
        active_session_status=active_session_status,
        active_session_id=active_session_id,
        needs_session=False,
        needs_replace=False,
        is_admin=role == ROLE_ADMIN,
        optimizer_v2_enabled=OPTIMIZER_V2_ENABLED,
    )


@app.route("/api/orders/manual-validate", methods=["POST"])
def api_manual_order_validate():
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    payload = request.get_json(silent=True) or {}
    manual_order_input = payload.get("manual_order_input") or ""
    snapshot = _manual_order_scope_snapshot(
        manual_order_input,
        allowed_plants=_get_allowed_plants(),
    )
    return jsonify(snapshot)


@app.route("/api/orders/scope-count", methods=["POST"])
def api_orders_scope_count():
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    payload = request.get_json(silent=True) or {}
    optimize_mode = (payload.get("optimize_mode") or "auto").strip().lower()
    if optimize_mode not in {"auto", "manual"}:
        optimize_mode = "auto"
    selected_order_category_scope = order_categories.normalize_order_category_scope(
        payload.get("order_category_scope"),
        default=order_categories.ORDER_CATEGORY_SCOPE_ALL,
    )

    allowed_plants = _get_allowed_plants()
    origin_plant = _normalize_plant_code(payload.get("origin_plant"))
    if origin_plant and origin_plant not in allowed_plants:
        return jsonify(
            {
                "mode": optimize_mode,
                "order_category_scope": selected_order_category_scope,
                "category_counts": order_categories.empty_category_counts(),
                "orders_in_scope": 0,
                "lines_in_scope": 0,
            }
        )

    if optimize_mode == "manual":
        snapshot = _manual_order_scope_snapshot(
            payload.get("manual_order_input") or "",
            origin_plant=origin_plant,
            allowed_plants=allowed_plants,
            order_category_scope=selected_order_category_scope,
        )
        return jsonify(snapshot)

    snapshot = _auto_order_scope_snapshot(
        origin_plant=origin_plant,
        state_filters=_coerce_filter_list(payload.get("opt_states")),
        customer_filters=_coerce_filter_list(payload.get("opt_customers")),
        orders_start_date=payload.get("orders_start_date"),
        batch_end_date=payload.get("batch_end_date"),
        batch_horizon_enabled=_coerce_bool_value(payload.get("batch_horizon_enabled")),
        order_category_scope=selected_order_category_scope,
    )
    return jsonify(snapshot)


@app.route("/orders/export")
def export_orders():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    plant_filters = _resolve_plant_filters(request.args.get("plants") or request.args.get("plant"))
    plant_scope = plant_filters or _get_allowed_plants()
    filters = {"plants": plant_scope} if plant_scope else {}
    data = order_service.list_orders(filters=filters)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(data["orders"][0].keys() if data["orders"] else [])
    for order in data["orders"]:
        writer.writerow(order.values())
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=orders_export.csv"},
    )


@app.route("/optimize", methods=["GET"])
def optimize():
    return redirect(url_for("orders"))


@app.route("/optimize/build", methods=["POST"])
def build_optimize():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    allowed_plants = _get_allowed_plants()
    origin_plant = _normalize_plant_code(request.form.get("origin_plant"))
    if origin_plant and origin_plant not in allowed_plants:
        return redirect(url_for("orders"))

    active_session_id = _get_active_planning_session_id()
    active_session = db.get_planning_session(active_session_id) if active_session_id else None
    active_session_status = _normalize_session_status(active_session.get("status")) if active_session else ""
    replace_session = request.form.get("replace_session") == "1"
    if active_session and active_session_status == "DRAFT" and not replace_session:
        return redirect(url_for("orders", needs_replace=1))
    if replace_session and active_session:
        _archive_session_and_release_loads(active_session_id)

    created_by = _get_session_profile_name() or _get_session_role()

    def _session_factory(form_data, params):
        return _create_planning_session(created_by, origin_plant, form_data, params)

    result = load_builder.build_loads(
        request.form,
        session_factory=_session_factory,
        created_by=created_by,
    )
    if result["errors"]:
        return redirect(url_for("orders"))
    session_id = result.get("session_id")
    _set_active_planning_session_id(session_id)
    return redirect(url_for("loads", session_id=session_id))


@app.route("/loads")
def loads():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    allowed_plants = _get_allowed_plants()
    plant_filters = _resolve_plant_filters(request.args.get("plants") or request.args.get("plant"))
    plant_scope = plant_filters or allowed_plants
    replay_mode = False
    replay_context = {}
    replay_run_id = request.args.get("replay_run_id")
    replay_date_created = (request.args.get("replay_date_created") or "").strip()
    replay_plant_code = (request.args.get("replay_plant_code") or "").strip().upper()
    replay_scenario = (request.args.get("replay_scenario") or "OPTIMIZED").strip().upper()
    if replay_scenario not in {"ACTUAL", "OPTIMIZED"}:
        replay_scenario = "OPTIMIZED"
    try:
        replay_run_id = int(replay_run_id) if replay_run_id else None
    except (TypeError, ValueError):
        replay_run_id = None

    session_id = None
    active_session = None
    all_loads = []
    if replay_run_id:
        _require_admin()
        replay_mode = True
        run = db.get_replay_eval_run(replay_run_id)
        if not run:
            abort(404)
        run_status = (run.get("status") or "").upper()
        if run_status != "COMPLETED":
            return redirect(
                url_for(
                    "planning_sessions_replay",
                    run_id=replay_run_id,
                    replay_error="Replay run is not completed yet.",
                )
            )

        metrics = db.list_replay_eval_load_metrics(replay_run_id, scenario=replay_scenario)
        metrics = [
            row
            for row in metrics
            if (row.get("plant_code") or "").strip().upper() in allowed_plants
        ]
        day_values = sorted({row.get("date_created") for row in metrics if row.get("date_created")})
        plant_values = sorted({(row.get("plant_code") or "").strip().upper() for row in metrics if row.get("plant_code")})

        if replay_date_created:
            metrics = [
                row for row in metrics if (row.get("date_created") or "") == replay_date_created
            ]
        if replay_plant_code:
            metrics = [
                row
                for row in metrics
                if (row.get("plant_code") or "").strip().upper() == replay_plant_code
            ]

        all_loads = _build_replay_simulation_loads(metrics)
        replay_plants_in_scope = sorted(
            {
                (load.get("origin_plant") or "").strip().upper()
                for load in all_loads
                if (load.get("origin_plant") or "").strip().upper()
            }
        )
        if replay_plants_in_scope:
            allowed_plants = [plant for plant in allowed_plants if plant in replay_plants_in_scope]
        if replay_plant_code and replay_plant_code in allowed_plants:
            plant_scope = [replay_plant_code]
            plant_filters = [replay_plant_code]
        else:
            scoped_filters = [plant for plant in plant_filters if plant in allowed_plants]
            plant_scope = scoped_filters or allowed_plants
            plant_filters = scoped_filters

        replay_context = {
            "run_id": replay_run_id,
            "run_filename": run.get("filename") or "",
            "scenario": replay_scenario,
            "date_created": replay_date_created,
            "plant_code": replay_plant_code,
            "day_values": day_values,
            "plant_values": plant_values,
        }
    else:
        session_id = request.args.get("session_id")
        try:
            session_id = int(session_id) if session_id else None
        except (TypeError, ValueError):
            session_id = None
        if session_id:
            _set_active_planning_session_id(session_id)
        else:
            session_id = _get_active_planning_session_id()
        if not session_id:
            return redirect(url_for("orders", needs_session=1))

        active_session = db.get_planning_session(session_id)
        if not active_session:
            _set_active_planning_session_id(None)
            return redirect(url_for("orders", needs_session=1))
        if not _can_access_planning_session(active_session):
            _set_active_planning_session_id(None)
            return redirect(url_for("orders", needs_session=1))
        if active_session.get("plant_code") and active_session.get("plant_code") not in allowed_plants:
            _set_active_planning_session_id(None)
            return redirect(url_for("orders", needs_session=1))
        if active_session.get("plant_code"):
            plant_scope = [active_session.get("plant_code")]
            plant_filters = list(plant_scope)
    tab = (request.args.get("tab") or "").strip().lower()
    status_filter = (request.args.get("status") or "").strip().upper()
    sort_mode = (request.args.get("sort") or "flow").strip().lower()
    if sort_mode not in {"flow", "util"}:
        sort_mode = "flow"
    today_override = _resolve_today_override(request.args.get("today"))
    today = today_override or date.today()
    reopt_status = request.args.get("reopt", "")
    reopt_job_id = (request.args.get("reopt_job") or "").strip()
    if reopt_job_id:
        reopt_job = _get_reopt_job(reopt_job_id)
        if reopt_job:
            if reopt_job.get("plant_code") not in allowed_plants:
                reopt_job_id = ""
            else:
                current_status = (reopt_job.get("status") or "").strip().lower()
                if current_status in {"running", "done", "failed"}:
                    reopt_status = current_status
        else:
            reopt_job_id = ""
    feedback_error = request.args.get("feedback_error") or ""
    feedback_target = request.args.get("feedback_target") or ""
    manual_error = request.args.get("manual_error") or ""
    algorithm_comparison = None
    if request.args.get("alg_cmp"):
        def _to_float(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        def _to_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return None

        v1_cost = _to_float(request.args.get("v1_cost"))
        v2_cost = _to_float(request.args.get("v2_cost"))
        v1_util = _to_float(request.args.get("v1_util"))
        v2_util = _to_float(request.args.get("v2_util"))
        v1_loads = _to_int(request.args.get("v1_loads"))
        v2_loads = _to_int(request.args.get("v2_loads"))
        if None not in {v1_cost, v2_cost, v1_util, v2_util, v1_loads, v2_loads}:
            cost_delta = v2_cost - v1_cost
            util_delta = v2_util - v1_util
            load_delta = v2_loads - v1_loads
            algorithm_comparison = {
                "selected": (request.args.get("alg_sel") or "v2").lower(),
                "v1": {"loads": v1_loads, "avg_util": v1_util, "cost": v1_cost},
                "v2": {"loads": v2_loads, "avg_util": v2_util, "cost": v2_cost},
                "delta": {
                    "cost": cost_delta,
                    "util": util_delta,
                    "loads": load_delta,
                    "cost_direction": "down" if cost_delta < 0 else ("up" if cost_delta > 0 else "flat"),
                },
            }
    if not replay_mode:
        all_loads = load_builder.list_loads(
            None,
            session_id=session_id,
            include_stack_metrics=False,
        )
    all_loads = [load for load in all_loads if load.get("origin_plant") in allowed_plants]
    loads_data = [load for load in all_loads if load.get("origin_plant") in plant_scope]
    if not replay_mode:
        session_status = _sync_planning_session_status(session_id, loads=all_loads)
        if active_session and session_status:
            active_session["status"] = session_status
    zip_coords = geo_utils.load_zip_coordinates()
    plant_names = {row["plant_code"]: row["name"] for row in db.list_plants()}
    sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
    stop_color_palette = _get_stop_color_palette()
    sku_color_palette = [
        "#137fec",
        "#10b981",
        "#f59e0b",
        "#ef4444",
        "#8b5cf6",
        "#22d3ee",
        "#f472b6",
        "#f97316",
    ]
    stop_fee_amount = _get_stop_fee_amount()
    fuel_surcharge_per_mile = _get_fuel_surcharge_per_mile()
    load_minimum_amount = _get_load_minimum_amount()
    trailer_assignment_rules = _get_effective_trailer_assignment_rules()
    hotshot_enabled_by_plant = {}
    strategic_setting = _get_effective_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY)
    strategic_customers = _parse_strategic_customers(strategic_setting.get("value_text") or "")
    carrier_pricing_context = _build_load_carrier_pricing_context()

    for load in loads_data:
        lines = load.get("lines", [])
        trailer_type = stack_calculator.normalize_trailer_type(load.get("trailer_type"), default="STEP_DECK")
        load["trailer_type"] = trailer_type
        load["total_units"] = sum((line.get("qty") or 0) for line in lines)
        load["total_sales"] = sum((line.get("sales") or 0) for line in lines)
        stops = []
        stop_map = {}
        due_dates = []
        for line in lines:
            due_date = _parse_date(line.get("due_date"))
            if due_date:
                due_dates.append(due_date)
            zip_code = geo_utils.normalize_zip(line.get("zip"))
            state = line.get("state") or ""
            city = line.get("city") or ""
            key = f"{zip_code}|{state}"
            if key not in stop_map:
                coords = zip_coords.get(zip_code) if zip_code else None
                stop_map[key] = {
                    "zip": zip_code,
                    "state": state,
                    "city": city,
                    "lat": coords[0] if coords else None,
                    "lng": coords[1] if coords else None,
                    "customers": set(),
                }
            elif city and not stop_map[key].get("city"):
                stop_map[key]["city"] = city
            if line.get("cust_name"):
                stop_map[key]["customers"].add(line.get("cust_name"))

        anchor_date = min(due_dates) if due_dates else None
        load["ship_date"] = anchor_date.strftime("%Y-%m-%d") if anchor_date else ""
        load["ship_date_status"] = _due_status(anchor_date, today=today)
        for line in lines:
            status_label = ""
            due_date = _parse_date(line.get("due_date"))
            if due_date and due_date < today:
                status_label = "Past Due"
            elif due_date and anchor_date:
                delta_days = (due_date - anchor_date).days
                if delta_days > 0:
                    status_label = f"Early {delta_days}d"
            line["status_label"] = status_label

        manifest_groups = []
        group_map = {}
        for line in lines:
            order_id = _normalize_order_identifier(line.get("so_num"), fallback="UNKNOWN")
            group = group_map.get(order_id)
            if not group:
                group = {
                    "order_id": order_id,
                    "due_date": line.get("due_date") or "",
                    "cust_name": line.get("cust_name") or "",
                    "city": line.get("city") or "",
                    "state": line.get("state") or "",
                    "zip": line.get("zip") or "",
                    "total_qty": 0,
                    "status_label": "",
                    "sku_set": set(),
                    "lines": [],
                }
                group_map[order_id] = group
                manifest_groups.append(group)

            group["total_qty"] += line.get("qty") or 0
            group["lines"].append(line)
            sku_value = line.get("sku")
            if sku_value:
                group["sku_set"].add(sku_value)

            group_due = _parse_date(group.get("due_date"))
            line_due = _parse_date(line.get("due_date"))
            if line_due and (not group_due or line_due < group_due):
                group["due_date"] = line.get("due_date")

            if not group["city"] and line.get("city"):
                group["city"] = line.get("city")
            if not group["state"] and line.get("state"):
                group["state"] = line.get("state")
            if not group["zip"] and line.get("zip"):
                group["zip"] = line.get("zip")

            line_status = line.get("status_label") or ""
            if "Past Due" in line_status:
                group["status_label"] = "Past Due"
            elif line_status and not group["status_label"]:
                group["status_label"] = line_status

        for group in manifest_groups:
            group["sku_list"] = sorted(group.get("sku_set") or [])
            group_due = _parse_date(group.get("due_date"))
            group["due_status"] = _due_status(group_due, today=today)
            early_days = (group_due - anchor_date).days if group_due and anchor_date else 0
            group["early_days"] = early_days if early_days > 0 else 0

        for stop in stop_map.values():
            coords = None
            if stop.get("lat") is not None and stop.get("lng") is not None:
                coords = (stop.get("lat"), stop.get("lng"))
            stops.append(
                {
                    "zip": stop["zip"],
                    "state": stop["state"],
                    "city": stop.get("city") or "",
                    "city_abbr": _city_abbr(stop.get("city")),
                    "lat": stop.get("lat"),
                    "lng": stop.get("lng"),
                    "coords": coords,
                    "customers": sorted(stop.get("customers") or []),
                }
            )

        origin_code = load.get("origin_plant")
        origin_coords = geo_utils.plant_coords_for_code(origin_code)
        requires_return_to_origin = _requires_return_to_origin(lines)
        if _alternate_requires_return_hint(
            lines,
            trailer_type,
            origin_code,
            carrier_pricing_context,
        ):
            requires_return_to_origin = True
        if _load_has_lowes_order(lines):
            requires_return_to_origin = True
        reverse_route = _is_load_route_reversed(load)
        ordered_stops = (
            tsp_solver.solve_route(
                origin_coords,
                stops,
                return_to_origin=requires_return_to_origin,
            )
            if origin_coords
            else list(stops)
        )
        ordered_stops = _apply_load_route_direction(ordered_stops, reverse_route=reverse_route)
        stop_sequence_map = _stop_sequence_map_from_ordered_stops(ordered_stops)
        order_colors = _build_order_colors_for_lines(
            lines,
            stop_sequence_map=stop_sequence_map,
            stop_palette=stop_color_palette,
        )
        load["order_colors"] = order_colors
        load["order_count"] = len(order_colors)

        for group in manifest_groups:
            group["stop_sequence"] = stop_sequence_map.get(
                _line_stop_key(group.get("state"), group.get("zip"))
            )
            group["color"] = order_colors.get(
                group["order_id"],
                _color_for_stop_sequence(group.get("stop_sequence"), stop_color_palette),
            )
        manifest_groups.sort(
            key=lambda group: (
                int(group.get("stop_sequence") or 999),
                group.get("due_date") or "9999-12-31",
                group.get("order_id") or "",
            )
        )
        origin_name = plant_names.get(origin_code, PLANT_NAMES.get(origin_code, origin_code))
        if requires_return_to_origin and ordered_stops:
            return_stop_sequence = len(ordered_stops) + 1
            manifest_groups.append(
                {
                    "order_id": "RETURN",
                    "due_date": "",
                    "cust_name": origin_name or origin_code or "Plant",
                    "city": "",
                    "state": origin_code or "",
                    "zip": "",
                    "total_qty": 0,
                    "status_label": "",
                    "sku_set": set(),
                    "lines": [],
                    "sku_list": [],
                    "due_status": "",
                    "early_days": 0,
                    "stop_sequence": return_stop_sequence,
                    "color": _color_for_stop_sequence(return_stop_sequence, stop_color_palette),
                    "is_terminal_stop": True,
                    "terminal_label": "Return to Plant",
                }
            )
        route_nodes = [
            {
                "type": "origin",
                "label": origin_code or "",
                "subtitle": origin_name or "",
                "icon": "home",
                "coords": origin_coords,
                "sequence": 0,
            }
        ]
        for idx, stop in enumerate(ordered_stops, start=1):
            coords = None
            if stop.get("lat") is not None and stop.get("lng") is not None:
                coords = (stop.get("lat"), stop.get("lng"))
            city = stop.get("city") or ""
            state = stop.get("state") or ""
            subtitle = ", ".join([part for part in [city, state] if part]).strip()
            route_nodes.append(
                {
                    "type": "stop",
                    "label": stop.get("city_abbr") or state or stop.get("zip") or "",
                    "subtitle": subtitle,
                    "icon": "place",
                    "coords": coords,
                    "sequence": idx,
                }
            )
        if requires_return_to_origin and origin_coords and len(route_nodes) > 1:
            route_nodes.append(
                {
                    "type": "final",
                    "label": origin_code or "",
                    "subtitle": origin_name or "",
                    "icon": "flag",
                    "coords": origin_coords,
                    "sequence": len(route_nodes),
                }
            )
        elif len(route_nodes) > 1:
            route_nodes[-1]["type"] = "final"
            route_nodes[-1]["icon"] = "flag"

        for node in route_nodes:
            node_type = (node.get("type") or "").strip().lower()
            is_return_origin = (
                node_type == "final"
                and requires_return_to_origin
                and origin_coords
                and node.get("coords") == origin_coords
            )
            if node_type == "origin" or is_return_origin:
                color = "#38bdf8"
            else:
                color = _color_for_stop_sequence(node.get("sequence"), stop_color_palette)
            node["color"] = color
            node["bg"] = f"{color}22"

        route_metrics = _load_route_display_metrics(
            load,
            route_nodes,
            use_cached_route=not reverse_route,
        )
        route_legs = route_metrics["route_legs"]
        total_route_distance = route_metrics["route_distance"]
        route_geometry = route_metrics["route_geometry"]
        load["estimated_miles"] = total_route_distance

        # Surface per-stop connector mileage for the manifest timeline.
        # route_legs index 0 is origin->stop1, index 1 is stop1->stop2, etc.
        group_count = len(manifest_groups)
        for index, group in enumerate(manifest_groups):
            group["leg_to_next_miles"] = None
            current_seq = _coerce_int_value(group.get("stop_sequence"), None)
            if current_seq is None:
                continue
            if index + 1 < group_count:
                next_same_seq = _coerce_int_value(
                    manifest_groups[index + 1].get("stop_sequence"),
                    None,
                )
                if next_same_seq == current_seq:
                    continue
            has_next = index + 1 < group_count
            if not has_next:
                continue
            leg_index = int(current_seq)
            if leg_index < 0 or leg_index >= len(route_legs):
                continue
            try:
                leg_miles = int(round(float(route_legs[leg_index] or 0)))
            except (TypeError, ValueError):
                leg_miles = 0
            if leg_miles > 0:
                group["leg_to_next_miles"] = leg_miles

        load["manifest_groups"] = manifest_groups

        schematic, line_items, order_numbers = _calculate_load_schematic(
            lines,
            sku_specs,
            trailer_type,
            stop_sequence_map=stop_sequence_map,
        )
        if origin_code not in hotshot_enabled_by_plant:
            hotshot_enabled_by_plant[origin_code] = _resolve_auto_hotshot_enabled_for_plant(
                origin_code,
                trailer_rules=trailer_assignment_rules,
            )
        load_trailer_rules = dict(trailer_assignment_rules or {})
        load_trailer_rules["auto_assign_hotshot_enabled"] = bool(
            hotshot_enabled_by_plant.get(origin_code, True)
        )
        auto_label, auto_reason = _auto_trailer_rule_annotation(
            load=load,
            lines=lines,
            trailer_type=trailer_type,
            schematic=schematic,
            sku_specs=sku_specs,
            stop_sequence_map=stop_sequence_map,
            trailer_assignment_rules=load_trailer_rules,
            strategic_customers=strategic_customers,
        )
        load["auto_trailer_label"] = auto_label
        load["auto_trailer_reason"] = auto_reason
        sku_colors = {}
        for idx, item in enumerate(line_items):
            sku = item.get("sku") or f"item-{idx}"
            if sku not in sku_colors:
                sku_colors[sku] = sku_color_palette[len(sku_colors) % len(sku_color_palette)]

        utilization_pct = schematic.get("utilization_pct", load.get("utilization_pct", 0)) or 0
        load["utilization_pct"] = utilization_pct
        exceeds_capacity = schematic.get("exceeds_capacity", False)
        load["over_capacity"] = exceeds_capacity and len(order_numbers) <= 1
        load["display_utilization_pct"] = utilization_pct
        load["utilization_display_note"] = ""
        load["raw_total_length_ft"] = round(
            sum(
                float(line.get("total_length_ft") or line.get("line_total_feet") or 0.0)
                for line in lines
            ),
            1,
        )
        load["utilization_credit_ft"] = float(schematic.get("utilization_credit_ft") or 0.0)
        load["total_linear_feet"] = float(schematic.get("total_linear_feet") or 0.0)
        load["schematic"] = schematic
        load["stops"] = ordered_stops
        load["stop_count"] = len(ordered_stops)
        load["sku_colors"] = sku_colors
        load["route_nodes"] = route_nodes
        load["route_legs"] = route_legs
        load["route_distance"] = total_route_distance
        load["route_geometry"] = route_geometry
        load["route_reversed"] = bool(reverse_route)
        load["return_to_origin"] = bool(requires_return_to_origin)
        carrier_pricing = _resolve_load_carrier_pricing(
            lines=lines,
            trailer_type=trailer_type,
            origin_plant=origin_code,
            ordered_stops=ordered_stops,
            route_legs=route_legs,
            total_miles=total_route_distance,
            stop_count=load.get("stop_count"),
            requires_return_to_origin=requires_return_to_origin,
            carrier_context=carrier_pricing_context,
            forced_carrier_key=load.get("carrier_override_key"),
        )
        load["carrier_pricing"] = carrier_pricing
        load["carrier_key"] = carrier_pricing.get("carrier_key")
        load["carrier_label"] = carrier_pricing.get("carrier_label")
        load["carrier_source_label"] = carrier_pricing.get("rate_source_label")
        load["carrier_selection_reason"] = carrier_pricing.get("selection_reason")
        load["rate_per_mile"] = carrier_pricing.get("rate_per_mile", load.get("rate_per_mile"))
        load["estimated_cost"] = carrier_pricing.get("total_cost", load.get("estimated_cost"))
        load["freight_breakdown"] = _build_freight_breakdown(
            load,
            stop_fee_amount=stop_fee_amount,
            fuel_surcharge_per_mile=fuel_surcharge_per_mile,
            load_minimum_amount=load_minimum_amount,
        )
        map_stops = []
        if origin_coords:
            origin_color = route_nodes[0]["color"] if route_nodes else "#38bdf8"
            map_stops.append(
                {
                    "type": "origin",
                    "lat": origin_coords[0],
                    "lng": origin_coords[1],
                    "label": origin_name or origin_code,
                    "sequence": 0,
                    "color": origin_color,
                }
            )
        for idx, stop in enumerate(ordered_stops, start=1):
            stop_color = route_nodes[idx]["color"] if idx < len(route_nodes) else "#f59e0b"
            map_stops.append(
                {
                    "type": "stop",
                    "lat": stop.get("lat"),
                    "lng": stop.get("lng"),
                    "label": stop.get("city") or stop.get("state") or stop.get("zip") or "",
                    "sequence": idx,
                    "color": stop_color,
                }
            )
        if requires_return_to_origin and origin_coords and map_stops:
            final_color = route_nodes[-1]["color"] if route_nodes else "#f59e0b"
            map_stops.append(
                {
                    "type": "final",
                    "lat": origin_coords[0],
                    "lng": origin_coords[1],
                    "label": origin_name or origin_code,
                    "sequence": len(map_stops),
                    "color": final_color,
                }
            )
        elif map_stops:
            map_stops[-1]["type"] = "final"
        load["map_stops"] = map_stops

    optimized_loads = []
    for load in loads_data:
        build_source = (load.get("build_source") or "OPTIMIZED").upper()
        if build_source == "MANUAL":
            continue
        optimized_loads.append(load)
    optimized_order_ids = {
        line.get("so_num")
        for load in optimized_loads
        for line in load.get("lines", [])
        if line.get("so_num")
    }
    optimized_total_spend = sum((load.get("estimated_cost") or 0) for load in optimized_loads)
    optimized_total_units = sum((load.get("total_units") or 0) for load in optimized_loads)
    optimized_util_values = [load.get("utilization_pct") or 0 for load in optimized_loads]
    optimized_avg_util = (
        round(sum(optimized_util_values) / len(optimized_util_values), 1)
        if optimized_util_values
        else 0.0
    )
    grade_counts = {"A": 0, "B": 0, "C": 0, "D": 0, "F": 0}
    for load in optimized_loads:
        grade = (load.get("schematic") or {}).get("utilization_grade")
        if not grade:
            grade = _utilization_grade(load.get("utilization_pct") or 0)
        grade = (grade or "F").upper()
        if grade not in grade_counts:
            grade_counts[grade] = 0
        grade_counts[grade] += 1

    baseline_cost = None
    baseline_set_at = None
    baseline_delta = None
    baseline_direction = ""
    if len(plant_scope) == 1:
        baseline_info = db.get_optimizer_baseline(plant_scope[0])
        baseline_cost = baseline_info.get("baseline_cost")
        baseline_set_at = baseline_info.get("baseline_set_at")
        if baseline_cost is None and optimized_total_spend:
            db.set_optimizer_baseline(plant_scope[0], optimized_total_spend)
            baseline_cost = optimized_total_spend
            baseline_set_at = datetime.utcnow().isoformat(timespec="seconds")
        if baseline_cost is not None:
            baseline_delta = optimized_total_spend - baseline_cost
            if baseline_delta < 0:
                baseline_direction = "below"
            elif baseline_delta > 0:
                baseline_direction = "above"

    spend_per_unit = (
        (optimized_total_spend / optimized_total_units) if optimized_total_units else 0.0
    )
    optimization_summary = {
        "total_orders": len(optimized_order_ids),
        "total_loads": len(optimized_loads),
        "total_spend": optimized_total_spend,
        "total_units": optimized_total_units,
        "spend_per_unit": spend_per_unit,
        "avg_utilization": optimized_avg_util,
        "grade_counts": grade_counts,
        "baseline_cost": baseline_cost,
        "baseline_set_at": baseline_set_at,
        "baseline_delta": baseline_delta,
        "baseline_direction": baseline_direction,
    }

    metrics_loads = list(loads_data)
    util_values = [load.get("utilization_pct") for load in metrics_loads if load.get("utilization_pct") is not None]
    avg_utilization = round(sum(util_values) / len(util_values), 1) if util_values else 0.0
    total_planned_cost = sum((load.get("estimated_cost") or 0) for load in metrics_loads)
    pending_exceptions = sum(
        1
        for load in metrics_loads
        if load.get("over_capacity") or (load.get("schematic") or {}).get("exceeds_capacity")
    )

    all_statuses = sorted({(load.get("status") or STATUS_PROPOSED).upper() for load in loads_data})
    if tab not in {"draft", "final"}:
        if status_filter == STATUS_APPROVED:
            tab = "final"
        elif status_filter in {STATUS_DRAFT, STATUS_PROPOSED}:
            tab = "draft"
        else:
            tab = "draft"

    if tab == "draft":
        loads_data = [
            load
            for load in loads_data
            if (load.get("status") or STATUS_PROPOSED).upper() in {STATUS_PROPOSED, STATUS_DRAFT}
        ]
    elif tab == "final":
        loads_data = [
            load
            for load in loads_data
            if (load.get("status") or STATUS_PROPOSED).upper() == STATUS_APPROVED
        ]
    elif status_filter:
        loads_data = [
            load
            for load in loads_data
            if (load.get("status") or STATUS_PROPOSED).upper() == status_filter
        ]

    if tab == "draft":
        for load in loads_data:
            standalone_cost = load.get("standalone_cost")
            consolidation_savings = load.get("consolidation_savings")
            if not standalone_cost:
                if consolidation_savings is not None:
                    standalone_cost = (load.get("estimated_cost") or 0) + consolidation_savings
                else:
                    standalone_cost = load.get("estimated_cost") or 0
            if consolidation_savings is None:
                consolidation_savings = (standalone_cost or 0) - (load.get("estimated_cost") or 0)

            fragility_score = (consolidation_savings / standalone_cost) if standalone_cost else 0.0
            load["standalone_cost"] = standalone_cost
            load["consolidation_savings"] = consolidation_savings
            load["fragility_score"] = fragility_score
            load["stop_count"] = load.get("stop_count") or len(load.get("stops") or [])

        def _approval_sort_key(load):
            utilization = load.get("utilization_pct") or 0
            fragility = load.get("fragility_score") or 0
            stop_count = load.get("stop_count") or 0
            if stop_count <= 1:
                tier = 1
                primary = -utilization
            elif fragility < 0.10:
                tier = 2
                primary = fragility
            else:
                tier = 3
                primary = -fragility
            return (
                tier,
                primary,
                -utilization,
                -(load.get("estimated_cost") or 0),
                -(load.get("id") or 0),
            )

        if sort_mode == "util":
            loads_data.sort(
                key=lambda load: (
                    -(load.get("utilization_pct") or 0),
                    -(load.get("estimated_cost") or 0),
                    -(load.get("id") or 0),
                )
            )
        else:
            loads_data.sort(key=_approval_sort_key)
    elif sort_mode == "util":
        loads_data.sort(
            key=lambda load: (
                -(load.get("utilization_pct") or 0),
                -(load.get("estimated_cost") or 0),
                -(load.get("id") or 0),
            )
        )

    for idx, load in enumerate(loads_data, start=1):
        load["display_sequence"] = idx

    full_truckloads = []
    past_due_loads = []
    manual_loads = []
    other_loads = list(loads_data)
    if tab == "draft":
        manual_loads = [
            load
            for load in loads_data
            if (load.get("build_source") or "OPTIMIZED").upper() == "MANUAL"
        ]
        other_loads = [
            load
            for load in loads_data
            if (load.get("build_source") or "OPTIMIZED").upper() != "MANUAL"
        ]
        full_ids = set()
        for load in other_loads:
            if _is_full_truckload(load):
                full_truckloads.append(load)
                if load.get("id") is not None:
                    full_ids.add(load.get("id"))
        other_loads = [load for load in other_loads if load.get("id") not in full_ids]
        past_due_ids = set()
        for load in other_loads:
            if (load.get("ship_date_status") or "").upper() == "PAST_DUE":
                past_due_loads.append(load)
                if load.get("id") is not None:
                    past_due_ids.add(load.get("id"))
        other_loads = [load for load in other_loads if load.get("id") not in past_due_ids]

    load_sections = []
    if tab == "draft":
        if manual_loads:
            load_sections.append(
                {"title": "Manually Built Loads", "loads": manual_loads, "kind": "manual"}
            )
        if full_truckloads:
            load_sections.append(
                {"title": "Full Truckload", "loads": full_truckloads, "kind": "full"}
            )
        if past_due_loads:
            load_sections.append(
                {"title": "Past Due Orders", "loads": past_due_loads, "kind": "past_due"}
            )
        load_sections.append(
            {"title": "Remaining Draft Loads", "loads": other_loads, "kind": "draft"}
        )
    else:
        load_sections.append({"title": "Loads", "loads": loads_data, "kind": "all"})

    plant_cards = []
    for plant in allowed_plants:
        plant_loads = [load for load in all_loads if load.get("origin_plant") == plant]
        load_count = len(plant_loads)
        order_count = len(
            {
                line.get("so_num")
                for load in plant_loads
                for line in load.get("lines", [])
                if line.get("so_num")
            }
        )
        plant_cards.append(
            {
                "code": plant,
                "name": PLANT_NAMES.get(plant, plant),
                "orders": order_count,
                "loads": load_count,
            }
        )

    plant_scope = plant_scope or allowed_plants
    progress_snapshot = _compute_load_progress_snapshot(
        plant_scope=plant_scope,
        all_loads=all_loads,
        allowed_plants=allowed_plants,
    )
    order_status_counts = progress_snapshot["order_status_counts"]
    load_status_counts = progress_snapshot["load_status_counts"]
    total_orders = progress_snapshot["total_orders"]
    approved_orders = progress_snapshot["approved_orders"]
    remaining_orders = max(total_orders - approved_orders, 0)
    progress_pct = progress_snapshot["progress_pct"]
    draft_tab_count = progress_snapshot["draft_tab_count"]
    final_tab_count = progress_snapshot["final_tab_count"]

    today_override_value = today_override.strftime("%Y-%m-%d") if today_override else ""
    today_override_label = today_override.strftime("%b %d, %Y") if today_override else ""
    session_config = {}
    if active_session and active_session.get("config_json"):
        try:
            session_config = json.loads(active_session.get("config_json") or "{}")
        except json.JSONDecodeError:
            session_config = {}
    replay_link_params = {}
    if replay_mode:
        replay_link_params = {
            "replay_run_id": replay_context.get("run_id"),
            "replay_scenario": replay_context.get("scenario"),
            "replay_date_created": replay_context.get("date_created") or None,
            "replay_plant_code": replay_context.get("plant_code") or None,
        }

    return render_template(
        "loads.html",
        loads=loads_data,
        plants=allowed_plants,
        plant_filters=plant_filters,
        plant_filter_param=",".join(plant_filters) if plant_filters else "",
        plant_cards=plant_cards,
        statuses=all_statuses,
        status_filter=status_filter,
        tab=tab,
        sort_mode=sort_mode,
        draft_tab_count=draft_tab_count,
        final_tab_count=final_tab_count,
        reopt_status=reopt_status,
        reopt_job_id=reopt_job_id,
        load_sections=load_sections,
        order_status_counts=order_status_counts,
        load_status_counts=load_status_counts,
        progress_pct=progress_pct,
        total_orders=total_orders,
        approved_orders=approved_orders,
        remaining_orders=remaining_orders,
        avg_utilization=avg_utilization,
        total_planned_cost=total_planned_cost,
        pending_exceptions=pending_exceptions,
        optimization_summary=optimization_summary,
        feedback_error=feedback_error,
        feedback_target=feedback_target,
        manual_error=manual_error,
        today_override=today_override,
        today_override_value=today_override_value,
        today_override_label=today_override_label,
        active_session=active_session,
        active_session_config=session_config,
        active_session_id=session_id,
        order_removal_reasons=ORDER_REMOVAL_REASONS,
        load_rejection_reasons=LOAD_REJECTION_REASONS,
        is_admin=_get_session_role() == ROLE_ADMIN,
        algorithm_comparison=algorithm_comparison,
        replay_mode=replay_mode,
        replay_context=replay_context,
        replay_link_params=replay_link_params,
        carrier_selection_options=LOAD_CARRIER_SELECTION_OPTIONS,
    )


@app.route("/planning-sessions")
def planning_sessions():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    plant_code = (request.args.get("plant") or "").strip().upper()
    planner = (request.args.get("planner") or "").strip()
    start_date = (request.args.get("start") or "").strip()
    end_date = (request.args.get("end") or "").strip()
    role = _get_session_role()
    can_manage_sessions = role == ROLE_ADMIN
    allowed_plants = set(_get_allowed_plants())
    profile_name = (_get_session_profile_name() or "").strip()
    scoped_planner = planner if can_manage_sessions else profile_name

    if plant_code and plant_code not in allowed_plants:
        abort(403)

    archived_session_param = (request.args.get("archived_session_id") or "").strip()
    archived_all_param = (request.args.get("archived_all_count") or "").strip()
    try:
        archived_session_id = int(archived_session_param) if archived_session_param else None
    except ValueError:
        archived_session_id = None
    try:
        archived_all_count = int(archived_all_param) if archived_all_param else None
    except ValueError:
        archived_all_count = None

    sessions = db.list_planning_sessions(
        {
            "plant_code": plant_code or None,
            "created_by": scoped_planner or None,
            "start_date": start_date or None,
            "end_date": end_date or None,
        }
    )

    visible_sessions = []
    for session in sessions:
        session_plant = _normalize_plant_code(session.get("plant_code"))
        if session_plant and session_plant not in allowed_plants:
            continue
        if not can_manage_sessions and not _can_access_planning_session(session):
            continue
        config = {}
        if session.get("config_json"):
            try:
                config = json.loads(session.get("config_json") or "{}")
            except json.JSONDecodeError:
                config = {}
        session["config"] = config
        session["status"] = _normalize_session_status(session.get("status"))
        session["created_at_label"] = _format_est_datetime_label(session.get("created_at"))
        visible_sessions.append(session)

    sessions = visible_sessions

    total_sessions = len(sessions)
    avg_efficiency = 0.0
    loads_optimized = 0
    active_session_id = _get_active_planning_session_id()
    active_session_label = None
    if sessions:
        util_values = [
            (session.get("avg_utilization") or 0) for session in sessions if session.get("avg_utilization") is not None
        ]
        avg_efficiency = round(sum(util_values) / len(util_values), 1) if util_values else 0.0
        loads_optimized = sum((session.get("load_count") or 0) for session in sessions)
    if active_session_id:
        active = next((session for session in sessions if session.get("id") == active_session_id), None)
        if not active:
            fetched = db.get_planning_session(active_session_id)
            active = fetched if _can_access_planning_session(fetched) else None
        if active and _can_access_planning_session(active):
            active_session_label = f"{active.get('plant_code') or ''} - {active.get('session_code') or ''}".strip(" -")

    planner_options = (
        sorted({session.get("created_by") for session in sessions if session.get("created_by")})
        if can_manage_sessions
        else ([profile_name] if profile_name else [])
    )
    plant_options = sorted({session.get("plant_code") for session in sessions if session.get("plant_code")})

    return render_template(
        "planning_sessions.html",
        sessions=sessions,
        total_sessions=total_sessions,
        avg_efficiency=avg_efficiency,
        loads_optimized=loads_optimized,
        active_session_label=active_session_label,
        active_session_id=active_session_id,
        planner_options=planner_options,
        plant_options=plant_options,
        archived_session_id=archived_session_id,
        archived_all_count=archived_all_count,
        can_manage_sessions=can_manage_sessions,
        filters={
            "plant": plant_code,
            "planner": planner if can_manage_sessions else profile_name,
            "start": start_date,
            "end": end_date,
        },
    )


def _build_replay_totals(network_rows):
    actual_loads = sum(int(row.get("actual_loads") or 0) for row in network_rows)
    optimized_loads = sum(int(row.get("optimized_loads") or 0) for row in network_rows)
    actual_util_num = sum(
        float(row.get("actual_avg_utilization") or 0.0) * int(row.get("actual_loads") or 0)
        for row in network_rows
    )
    optimized_util_num = sum(
        float(row.get("optimized_avg_utilization") or 0.0) * int(row.get("optimized_loads") or 0)
        for row in network_rows
    )
    actual_avg_util = (actual_util_num / actual_loads) if actual_loads else 0.0
    optimized_avg_util = (optimized_util_num / optimized_loads) if optimized_loads else 0.0
    actual_total_cost = sum(float(row.get("actual_total_cost") or 0.0) for row in network_rows)
    optimized_total_cost = sum(float(row.get("optimized_total_cost") or 0.0) for row in network_rows)
    delta_total_cost = optimized_total_cost - actual_total_cost
    return {
        "days": len(network_rows),
        "matched_orders": sum(int(row.get("matched_orders") or 0) for row in network_rows),
        "missing_orders": sum(int(row.get("missing_orders") or 0) for row in network_rows),
        "actual_loads": actual_loads,
        "optimized_loads": optimized_loads,
        "actual_avg_utilization": actual_avg_util,
        "optimized_avg_utilization": optimized_avg_util,
        "actual_total_miles": sum(float(row.get("actual_total_miles") or 0.0) for row in network_rows),
        "optimized_total_miles": sum(float(row.get("optimized_total_miles") or 0.0) for row in network_rows),
        "actual_total_cost": actual_total_cost,
        "optimized_total_cost": optimized_total_cost,
        "delta_total_cost": delta_total_cost,
        "delta_cost_pct": ((delta_total_cost / actual_total_cost) * 100.0) if actual_total_cost else None,
    }


def _parse_replay_load_json(raw_json):
    if not raw_json:
        return {}
    try:
        parsed = json.loads(raw_json)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _parse_replay_order_numbers(raw_json):
    if not raw_json:
        return []
    try:
        parsed = json.loads(raw_json)
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
    values = []
    for value in parsed if isinstance(parsed, list) else []:
        text = str(value or "").strip()
        if text:
            values.append(text)
    return values


def _build_replay_simulation_loads(load_metrics):
    stop_fee_amount = _get_stop_fee_amount()
    fuel_surcharge_per_mile = _get_fuel_surcharge_per_mile()
    load_minimum_amount = _get_load_minimum_amount()

    loads = []
    for idx, metric in enumerate(load_metrics or [], start=1):
        load_data = _parse_replay_load_json(metric.get("load_json"))
        lines = load_data.get("lines") if isinstance(load_data.get("lines"), list) else []
        order_numbers = _parse_replay_order_numbers(metric.get("order_numbers_json"))
        if not lines and order_numbers:
            lines = [{"so_num": so_num} for so_num in order_numbers]

        load_number = (metric.get("load_key") or "").strip() or f"SIM-{idx:03d}"
        stop_count = int(load_data.get("stop_count") or 0)
        if not stop_count:
            stop_keys = {
                f"{(str(line.get('state') or '').strip().upper())}|{str(line.get('zip') or '').strip()}"
                for line in lines
                if str(line.get("state") or "").strip() or str(line.get("zip") or "").strip()
            }
            stop_count = len(stop_keys)

        load = {
            "id": idx,
            "load_number": load_number,
            "status": STATUS_PROPOSED,
            "simulation_status": "SIMULATED",
            "build_source": "OPTIMIZED",
            "trailer_type": stack_calculator.normalize_trailer_type(load_data.get("trailer_type"), default="STEP_DECK"),
            "utilization_pct": float(metric.get("utilization_pct") or 0.0),
            "estimated_miles": float(metric.get("estimated_miles") or 0.0),
            "estimated_cost": float(metric.get("estimated_cost") or 0.0),
            "rate_per_mile": float(load_data.get("rate_per_mile") or 0.0),
            "stop_count": stop_count,
            "return_to_origin": bool(load_data.get("return_to_origin")),
            "return_miles": float(load_data.get("return_miles") or 0.0),
            "return_cost": float(load_data.get("return_cost") or 0.0),
            "origin_plant": (load_data.get("origin_plant") or metric.get("plant_code") or "").strip().upper(),
            "destination_state": (load_data.get("destination_state") or "").strip().upper(),
            "route": load_data.get("route") if isinstance(load_data.get("route"), list) else [],
            "route_legs": load_data.get("route_legs") if isinstance(load_data.get("route_legs"), list) else [],
            "lines": lines,
        }
        load["freight_breakdown"] = _build_freight_breakdown(
            load,
            stop_fee_amount=stop_fee_amount,
            fuel_surcharge_per_mile=fuel_surcharge_per_mile,
            load_minimum_amount=load_minimum_amount,
        )
        loads.append(load)

    return loads


def _build_replay_workbook(run, network_rows, day_rows, issues, load_metrics):
    workbook = Workbook()
    network_sheet = workbook.active
    network_sheet.title = "Network Daily"
    network_headers = [
        "Replay Date / Period",
        "Plants",
        "Matched Orders",
        "Missing Orders",
        "Actual Loads",
        "Actual Avg Util %",
        "Actual Miles",
        "Actual Cost",
        "Optimized Loads",
        "Optimized Avg Util %",
        "Optimized Miles",
        "Optimized Cost",
        "Delta Loads",
        "Delta Util Pts",
        "Delta Miles",
        "Delta Cost",
        "Delta Cost %",
        "Report Ref Cost",
        "Report Ref Miles",
    ]
    network_sheet.append(network_headers)
    for row in network_rows:
        network_sheet.append(
            [
                row.get("date_created") or "",
                int(row.get("plants") or 0),
                int(row.get("matched_orders") or 0),
                int(row.get("missing_orders") or 0),
                int(row.get("actual_loads") or 0),
                round(float(row.get("actual_avg_utilization") or 0.0), 2),
                round(float(row.get("actual_total_miles") or 0.0), 2),
                round(float(row.get("actual_total_cost") or 0.0), 2),
                int(row.get("optimized_loads") or 0),
                round(float(row.get("optimized_avg_utilization") or 0.0), 2),
                round(float(row.get("optimized_total_miles") or 0.0), 2),
                round(float(row.get("optimized_total_cost") or 0.0), 2),
                int(row.get("delta_loads") or 0),
                round(float(row.get("delta_avg_utilization") or 0.0), 2),
                round(float(row.get("delta_total_miles") or 0.0), 2),
                round(float(row.get("delta_total_cost") or 0.0), 2),
                round(float(row.get("delta_cost_pct") or 0.0), 4) if row.get("delta_cost_pct") is not None else "",
                round(float(row.get("report_ref_cost") or 0.0), 2),
                round(float(row.get("report_ref_miles") or 0.0), 2),
            ]
        )

    plant_sheet = workbook.create_sheet("Plant Daily")
    plant_headers = [
        "Replay Date / Period",
        "Plant",
        "Report Rows",
        "Report Loads",
        "Report Orders",
        "Matched Orders",
        "Missing Orders",
        "Actual Loads",
        "Actual Avg Util %",
        "Actual Miles",
        "Actual Cost",
        "Optimized Loads",
        "Optimized Strategy",
        "Optimized Avg Util %",
        "Optimized Miles",
        "Optimized Cost",
        "Delta Loads",
        "Delta Util Pts",
        "Delta Miles",
        "Delta Cost",
        "Delta Cost %",
        "Report Ref Cost",
        "Report Ref Miles",
        "Report Ref Avg Truck Use",
    ]
    plant_sheet.append(plant_headers)
    for row in day_rows:
        plant_sheet.append(
            [
                row.get("date_created") or "",
                row.get("plant_code") or "",
                int(row.get("report_rows") or 0),
                int(row.get("report_loads") or 0),
                int(row.get("report_orders") or 0),
                int(row.get("matched_orders") or 0),
                int(row.get("missing_orders") or 0),
                int(row.get("actual_loads") or 0),
                round(float(row.get("actual_avg_utilization") or 0.0), 2),
                round(float(row.get("actual_total_miles") or 0.0), 2),
                round(float(row.get("actual_total_cost") or 0.0), 2),
                int(row.get("optimized_loads") or 0),
                row.get("optimized_strategy") or "",
                round(float(row.get("optimized_avg_utilization") or 0.0), 2),
                round(float(row.get("optimized_total_miles") or 0.0), 2),
                round(float(row.get("optimized_total_cost") or 0.0), 2),
                int(row.get("delta_loads") or 0),
                round(float(row.get("delta_avg_utilization") or 0.0), 2),
                round(float(row.get("delta_total_miles") or 0.0), 2),
                round(float(row.get("delta_total_cost") or 0.0), 2),
                round(float(row.get("delta_cost_pct") or 0.0), 4) if row.get("delta_cost_pct") is not None else "",
                round(float(row.get("report_ref_cost") or 0.0), 2) if row.get("report_ref_cost") is not None else "",
                round(float(row.get("report_ref_miles") or 0.0), 2) if row.get("report_ref_miles") is not None else "",
                round(float(row.get("report_ref_avg_truck_use") or 0.0), 2)
                if row.get("report_ref_avg_truck_use") is not None
                else "",
            ]
        )

    issues_sheet = workbook.create_sheet("Issues")
    issues_sheet.append(
        [
            "Replay Date / Period",
            "Plant",
            "Load Number",
            "Order Number",
            "Issue Type",
            "Severity",
            "Message",
            "Meta JSON",
        ]
    )
    for issue in issues:
        issues_sheet.append(
            [
                issue.get("date_created") or "",
                issue.get("plant_code") or "",
                issue.get("load_number") or "",
                issue.get("order_number") or "",
                issue.get("issue_type") or "",
                issue.get("severity") or "",
                issue.get("message") or "",
                issue.get("meta_json") or "",
            ]
        )

    metrics_sheet = workbook.create_sheet("Load Metrics")
    metrics_sheet.append(
        [
            "Replay Date / Period",
            "Plant",
            "Scenario",
            "Load Key",
            "Order Count",
            "Utilization %",
            "Estimated Miles",
            "Estimated Cost",
            "Order Numbers JSON",
        ]
    )
    for row in load_metrics:
        metrics_sheet.append(
            [
                row.get("date_created") or "",
                row.get("plant_code") or "",
                row.get("scenario") or "",
                row.get("load_key") or "",
                int(row.get("order_count") or 0),
                round(float(row.get("utilization_pct") or 0.0), 2),
                round(float(row.get("estimated_miles") or 0.0), 2),
                round(float(row.get("estimated_cost") or 0.0), 2),
                row.get("order_numbers_json") or "",
            ]
        )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for idx, _ in enumerate(sheet[1], start=1):
            sheet.column_dimensions[chr(64 + min(idx, 26))].width = 18
    return workbook


@app.route("/planning-sessions/replay", methods=["GET", "POST"])
def planning_sessions_replay():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()

    preset = _get_replay_eval_preset()
    error = (request.args.get("replay_error") or "").strip()
    replay_success = (request.args.get("replay_success") or "").strip()
    source_run_id = (request.args.get("source_run_id") or "").strip()
    run_id_param = (request.args.get("run_id") or "").strip()
    selected_day = (request.args.get("day") or "").strip()
    scope_param = (request.args.get("evaluation_scope") or "").strip()
    requested_scope = replay_evaluator.normalize_evaluation_scope(scope_param)
    upload_scope = requested_scope
    upload_ops_parity_enabled = _coerce_bool_value(preset.get("ops_parity_enabled"))
    try:
        selected_run_id = int(run_id_param) if run_id_param else None
    except (TypeError, ValueError):
        selected_run_id = None
    run_id = None
    if request.method == "POST":
        upload_scope = replay_evaluator.normalize_evaluation_scope(request.form.get("evaluation_scope"))
        upload_ops_parity_enabled = _coerce_bool_value(request.form.get("ops_parity_enabled"))
        report_file = request.files.get("report_file")
        if not report_file or not getattr(report_file, "filename", ""):
            error = "Choose a .csv or .xlsx report file."
        else:
            try:
                run_preset = dict(preset)
                run_preset["ops_parity_enabled"] = upload_ops_parity_enabled
                run_id = replay_evaluator.run_replay_evaluation(
                    report_file,
                    preset=run_preset,
                    created_by=_get_session_profile_name() or _get_session_role(),
                    evaluation_scope=upload_scope,
                )
            except Exception as exc:
                error = str(exc)
        if run_id:
            return redirect(url_for("planning_sessions_replay", run_id=run_id, evaluation_scope=upload_scope))

    runs = db.list_replay_eval_runs(limit=20)
    for entry in runs:
        entry["summary"] = _parse_replay_summary(entry.get("summary_json"))
        entry["status"] = (entry.get("status") or "").upper()

    active_run = None
    if selected_run_id:
        active_run = next((entry for entry in runs if int(entry.get("id") or 0) == selected_run_id), None)
    if not active_run:
        active_run = runs[0] if runs else None

    active_summary = {}
    active_scope = requested_scope
    date_basis = "shipped_date"
    active_ops_parity_enabled = upload_ops_parity_enabled
    if active_run:
        active_summary = active_run.get("summary") or {}
        active_scope = replay_evaluator.normalize_evaluation_scope(
            scope_param or active_summary.get("evaluation_scope")
        )
        upload_scope = active_scope
        date_basis = (active_summary.get("date_basis") or "shipped_date").strip().lower()
        raw_params = (active_run.get("params_json") or "").strip()
        if raw_params:
            try:
                parsed_params = json.loads(raw_params)
            except (TypeError, ValueError, json.JSONDecodeError):
                parsed_params = None
            if isinstance(parsed_params, dict):
                active_ops_parity_enabled = _coerce_bool_value(
                    parsed_params.get("ops_parity_enabled")
                )
        upload_ops_parity_enabled = active_ops_parity_enabled

    day_values = []
    selected_day_row = {}
    selected_day_plants = []
    selected_day_issues = []
    issues_by_type = {}
    kpis = {}
    if active_run and active_run.get("status") == "COMPLETED":
        active_run_id = int(active_run.get("id") or 0)
        day_rows = db.list_replay_eval_day_plant(active_run_id)
        network_rows = replay_evaluator.build_network_daily_rollup(day_rows)
        day_values = sorted({row.get("date_created") for row in network_rows if row.get("date_created")})
        if selected_day not in day_values:
            selected_day = day_values[-1] if day_values else ""
        selected_day_row = next(
            (row for row in network_rows if (row.get("date_created") or "") == selected_day),
            {},
        )
        selected_day_plants = sorted(
            [row for row in day_rows if (row.get("date_created") or "") == selected_day],
            key=lambda row: row.get("plant_code") or "",
        )
        issues = db.list_replay_eval_issues(active_run_id)
        selected_day_issues = [
            issue for issue in issues if (issue.get("date_created") or "") == selected_day
        ] if selected_day else list(issues)
        for issue in selected_day_issues:
            issue_type = issue.get("issue_type") or "unknown"
            issues_by_type[issue_type] = issues_by_type.get(issue_type, 0) + 1

        actual_loads = int(selected_day_row.get("actual_loads") or 0)
        optimized_loads = int(selected_day_row.get("optimized_loads") or 0)
        actual_util = float(selected_day_row.get("actual_avg_utilization") or 0.0)
        optimized_util = float(selected_day_row.get("optimized_avg_utilization") or 0.0)
        actual_cost = float(selected_day_row.get("actual_total_cost") or 0.0)
        optimized_cost = float(selected_day_row.get("optimized_total_cost") or 0.0)
        actual_miles = float(selected_day_row.get("actual_total_miles") or 0.0)
        optimized_miles = float(selected_day_row.get("optimized_total_miles") or 0.0)
        load_delta = optimized_loads - actual_loads
        load_reduction_pct = (
            ((actual_loads - optimized_loads) / actual_loads) * 100.0
            if actual_loads
            else None
        )
        util_gain_pts = optimized_util - actual_util
        util_gain_pct = ((util_gain_pts / actual_util) * 100.0) if actual_util else None
        cost_delta = optimized_cost - actual_cost
        cost_savings = -cost_delta
        cost_savings_pct = ((cost_savings / actual_cost) * 100.0) if actual_cost else None
        miles_delta = optimized_miles - actual_miles
        miles_savings = -miles_delta
        miles_savings_pct = ((miles_savings / actual_miles) * 100.0) if actual_miles else None
        kpis = {
            "actual_loads": actual_loads,
            "optimized_loads": optimized_loads,
            "load_delta": load_delta,
            "load_reduction_pct": load_reduction_pct,
            "util_gain_pts": util_gain_pts,
            "util_gain_pct": util_gain_pct,
            "cost_delta": cost_delta,
            "cost_savings": cost_savings,
            "cost_savings_pct": cost_savings_pct,
            "miles_delta": miles_delta,
            "miles_savings": miles_savings,
            "miles_savings_pct": miles_savings_pct,
            "actual_util": actual_util,
            "optimized_util": optimized_util,
            "actual_cost": actual_cost,
            "optimized_cost": optimized_cost,
            "actual_miles": actual_miles,
            "optimized_miles": optimized_miles,
            "matched_orders": int(selected_day_row.get("matched_orders") or 0),
            "missing_orders": int(selected_day_row.get("missing_orders") or 0),
        }

    return render_template(
        "replay_eval.html",
        preset=preset,
        active_run=active_run,
        evaluation_scope=active_scope,
        date_basis=date_basis,
        day_label=("Replay Period" if active_scope == replay_evaluator.EVAL_SCOPE_WEEKLY_POOLED else "Shipped Day"),
        upload_scope=upload_scope,
        selected_day=selected_day,
        day_values=day_values,
        selected_day_row=selected_day_row,
        selected_day_plants=selected_day_plants,
        selected_day_issues=selected_day_issues,
        issues_by_type=issues_by_type,
        kpis=kpis,
        error=error,
        replay_success=replay_success,
        source_run_id=source_run_id,
        upload_ops_parity_enabled=upload_ops_parity_enabled,
        active_ops_parity_enabled=active_ops_parity_enabled,
        active_summary=active_summary,
    )


@app.route("/planning-sessions/replay/<int:run_id>")
def planning_sessions_replay_detail(run_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()

    replay_error = (request.args.get("replay_error") or "").strip()
    replay_success = (request.args.get("replay_success") or "").strip()
    source_run_id = (request.args.get("source_run_id") or "").strip()
    day = (request.args.get("day") or "").strip()
    evaluation_scope = (request.args.get("evaluation_scope") or "").strip()
    return redirect(
        url_for(
            "planning_sessions_replay",
            run_id=run_id,
            day=day or None,
            evaluation_scope=evaluation_scope or None,
            replay_error=replay_error or None,
            replay_success=replay_success or None,
            source_run_id=source_run_id or None,
        )
    )


@app.route("/planning-sessions/replay/<int:run_id>/loads")
def planning_sessions_replay_loads(run_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()

    scenario = (request.args.get("scenario") or "OPTIMIZED").strip().upper()
    if scenario not in {"OPTIMIZED", "ACTUAL"}:
        scenario = "OPTIMIZED"
    date_created = (request.args.get("date_created") or "").strip()
    plant_code = (request.args.get("plant_code") or "").strip().upper()
    return redirect(
        url_for(
            "loads",
            replay_run_id=run_id,
            replay_scenario=scenario,
            replay_date_created=date_created or None,
            replay_plant_code=plant_code or None,
        )
    )


@app.route("/planning-sessions/replay/<int:run_id>/reproduce", methods=["POST"])
def planning_sessions_replay_reproduce(run_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()

    date_created = (request.form.get("date_created") or "").strip()
    plant_code = (request.form.get("plant_code") or "").strip().upper()
    evaluation_scope = replay_evaluator.normalize_evaluation_scope(request.form.get("evaluation_scope"))
    if not date_created or not plant_code:
        return redirect(
            url_for(
                "planning_sessions_replay",
                run_id=run_id,
                evaluation_scope=evaluation_scope,
                replay_error="Select a valid day and plant bucket to reproduce.",
            )
        )

    try:
        reproduced_run_id = replay_evaluator.reproduce_replay_bucket(
            source_run_id=run_id,
            date_created=date_created,
            plant_code=plant_code,
            created_by=_get_session_profile_name() or _get_session_role(),
        )
    except Exception as exc:
        return redirect(
            url_for(
                "planning_sessions_replay",
                run_id=run_id,
                evaluation_scope=evaluation_scope,
                replay_error=str(exc),
            )
        )

    return redirect(
        url_for(
            "planning_sessions_replay",
            run_id=reproduced_run_id,
            day=date_created,
            evaluation_scope=evaluation_scope,
            replay_success=f"Reproduced bucket {date_created} / {plant_code} from run #{run_id}.",
            source_run_id=run_id,
        )
    )


@app.route("/planning-sessions/replay/<int:run_id>/export.xlsx")
def planning_sessions_replay_export(run_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()

    run = db.get_replay_eval_run(run_id)
    if not run:
        abort(404)
    day_rows = db.list_replay_eval_day_plant(run_id)
    network_rows = replay_evaluator.build_network_daily_rollup(day_rows)
    issues = db.list_replay_eval_issues(run_id)
    load_metrics = db.list_replay_eval_load_metrics(run_id)
    workbook = _build_replay_workbook(run, network_rows, day_rows, issues, load_metrics)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"replay_eval_run_{run_id}_{date.today().isoformat()}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/planning-sessions/replay/<int:run_id>/issues.csv")
def planning_sessions_replay_issues_csv(run_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_admin()

    run = db.get_replay_eval_run(run_id)
    if not run:
        abort(404)
    issues = db.list_replay_eval_issues(run_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "date_created",
            "plant_code",
            "load_number",
            "order_number",
            "issue_type",
            "severity",
            "message",
            "meta_json",
        ]
    )
    for issue in issues:
        writer.writerow(
            [
                issue.get("date_created") or "",
                issue.get("plant_code") or "",
                issue.get("load_number") or "",
                issue.get("order_number") or "",
                issue.get("issue_type") or "",
                issue.get("severity") or "",
                issue.get("message") or "",
                issue.get("meta_json") or "",
            ]
        )
    output.seek(0)
    filename = f"replay_eval_issues_run_{run_id}_{date.today().isoformat()}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _planning_session_filter_values(source):
    return {
        "plant": ((source.get("plant") if source else "") or "").strip().upper(),
        "planner": ((source.get("planner") if source else "") or "").strip(),
        "start": ((source.get("start") if source else "") or "").strip(),
        "end": ((source.get("end") if source else "") or "").strip(),
    }


def _planning_sessions_redirect_args(filters):
    return {
        "plant": filters.get("plant") or "",
        "planner": filters.get("planner") or "",
        "start": filters.get("start") or "",
        "end": filters.get("end") or "",
    }


def _session_plant_scope(session_id):
    plants = set()
    if not session_id:
        return []

    planning_session = db.get_planning_session(session_id)
    if planning_session:
        plant_code = _normalize_plant_code(planning_session.get("plant_code"))
        if plant_code in PLANT_CODES:
            plants.add(plant_code)

    for load in db.list_loads(None, session_id=session_id):
        plant_code = _normalize_plant_code(load.get("origin_plant"))
        if plant_code:
            plants.add(plant_code)

    return sorted(plants)


def _reintroduce_orders_to_pool(plants):
    cleaned = []
    seen = set()
    for plant in (plants or []):
        code = _normalize_plant_code(plant)
        if not code or code in seen:
            continue
        seen.add(code)
        cleaned.append(code)
    if not cleaned:
        return
    db.include_orders_for_plants(cleaned)


def _archive_session_and_release_loads(session_id):
    if not session_id:
        return False
    planning_session = db.get_planning_session(session_id)
    if not planning_session:
        return False
    if not bool(planning_session.get("is_sandbox")):
        _reintroduce_orders_to_pool(_session_plant_scope(session_id))
    db.archive_planning_session(session_id)
    if _get_active_planning_session_id() == session_id:
        _set_active_planning_session_id(None)
    return True


def _build_planning_session_rollup(loads):
    load_summaries = []
    order_map = {}

    for load in loads or []:
        lines = load.get("lines") or []
        stop_keys = set()
        load_order_map = {}
        total_feet = 0.0

        for line in lines:
            so_num = (line.get("so_num") or "").strip()
            if not so_num:
                continue

            total_feet += float(line.get("total_length_ft") or line.get("line_total_feet") or 0)
            state = (line.get("state") or "").strip().upper()
            zip_code = (line.get("zip") or "").strip()
            stop_keys.add(f"{state}|{zip_code}")

            line_due = _parse_date(line.get("due_date"))
            due_label = line_due.isoformat() if line_due else (line.get("due_date") or "")
            line_qty = float(line.get("qty") or 0)
            line_length = float(line.get("total_length_ft") or line.get("line_total_feet") or 0)

            load_order = load_order_map.get(so_num)
            if not load_order:
                load_order = {
                    "so_num": so_num,
                    "cust_name": (line.get("cust_name") or "").strip(),
                    "state": state,
                    "city": (line.get("city") or "").strip(),
                    "zip": zip_code,
                    "due_date": due_label,
                    "due_date_obj": line_due,
                    "line_count": 0,
                    "total_qty": 0.0,
                    "total_length_ft": 0.0,
                }
                load_order_map[so_num] = load_order
            load_order["line_count"] += 1
            load_order["total_qty"] += line_qty
            load_order["total_length_ft"] += line_length
            if line_due and (not load_order.get("due_date_obj") or line_due < load_order["due_date_obj"]):
                load_order["due_date_obj"] = line_due
                load_order["due_date"] = line_due.isoformat()

            session_order = order_map.get(so_num)
            if not session_order:
                session_order = {
                    "so_num": so_num,
                    "cust_name": (line.get("cust_name") or "").strip(),
                    "state": state,
                    "city": (line.get("city") or "").strip(),
                    "zip": zip_code,
                    "due_date": due_label,
                    "due_date_obj": line_due,
                    "line_count": 0,
                    "total_qty": 0.0,
                    "total_length_ft": 0.0,
                    "loads": set(),
                }
                order_map[so_num] = session_order
            session_order["line_count"] += 1
            session_order["total_qty"] += line_qty
            session_order["total_length_ft"] += line_length
            if line_due and (not session_order.get("due_date_obj") or line_due < session_order["due_date_obj"]):
                session_order["due_date_obj"] = line_due
                session_order["due_date"] = line_due.isoformat()
            load_number = load.get("load_number") or f"Load #{load.get('id')}"
            session_order["loads"].add(load_number)

        load_orders = []
        for entry in load_order_map.values():
            entry.pop("due_date_obj", None)
            entry["total_qty"] = round(entry["total_qty"], 1)
            entry["total_length_ft"] = round(entry["total_length_ft"], 1)
            load_orders.append(entry)
        load_orders.sort(key=lambda entry: (entry.get("due_date") or "9999-12-31", entry.get("so_num") or ""))

        load_summaries.append(
            {
                "id": load.get("id"),
                "load_number": load.get("load_number") or f"Load #{load.get('id')}",
                "status": (load.get("status") or "PROPOSED").upper(),
                "trailer_type": stack_calculator.normalize_trailer_type(load.get("trailer_type"), default="STEP_DECK"),
                "utilization_pct": round(float(load.get("utilization_pct") or 0), 1),
                "estimated_cost": round(float(load.get("estimated_cost") or 0), 2),
                "estimated_miles": round(float(load.get("estimated_miles") or 0), 1),
                "stop_count": len(stop_keys),
                "line_count": len(lines),
                "order_count": len(load_orders),
                "total_length_ft": round(total_feet, 1),
                "created_at": load.get("created_at"),
                "orders": load_orders,
                "return_to_origin": bool(load.get("return_to_origin")),
                "return_miles": round(float(load.get("return_miles") or 0.0), 1),
                "return_cost": round(float(load.get("return_cost") or 0.0), 2),
                "freight_breakdown": load.get("freight_breakdown") or {},
            }
        )

    load_summaries.sort(key=lambda entry: (entry.get("load_number") or "", entry.get("id") or 0))

    session_orders = []
    for entry in order_map.values():
        entry.pop("due_date_obj", None)
        entry["loads"] = sorted(entry.get("loads") or [])
        entry["load_count"] = len(entry["loads"])
        entry["total_qty"] = round(entry["total_qty"], 1)
        entry["total_length_ft"] = round(entry["total_length_ft"], 1)
        session_orders.append(entry)
    session_orders.sort(key=lambda entry: (entry.get("due_date") or "9999-12-31", entry.get("so_num") or ""))

    return {
        "loads": load_summaries,
        "orders": session_orders,
    }


def _build_load_report_rows(loads):
    sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
    zip_coords = geo_utils.load_zip_coordinates()
    stop_color_palette = _get_stop_color_palette()
    rows = []

    for load in loads or []:
        lines = load.get("lines") or []
        trailer_type = stack_calculator.normalize_trailer_type(load.get("trailer_type"), default="STEP_DECK")
        ordered_stops = _ordered_stops_for_lines(lines, load.get("origin_plant"), zip_coords)
        ordered_stops = _apply_load_route_direction(ordered_stops, load=load)
        stop_sequence_map = _stop_sequence_map_from_ordered_stops(ordered_stops)
        schematic, _, _ = _calculate_load_schematic(
            lines,
            sku_specs,
            trailer_type,
            stop_sequence_map=stop_sequence_map,
        )
        trailer_config = _trailer_config_for_type(trailer_type)
        schematic = dict(schematic or {})
        schematic.setdefault("positions", [])
        schematic.setdefault("warnings", [])
        schematic.setdefault("trailer_type", trailer_config["type"])
        schematic.setdefault("capacity_feet", trailer_config["capacity"])
        schematic.setdefault("lower_deck_length", trailer_config["lower"])
        schematic.setdefault("upper_deck_length", trailer_config["upper"])
        schematic.setdefault(
            "utilization_pct",
            round(float(load.get("utilization_pct") or 0), 1),
        )
        schematic.setdefault(
            "utilization_grade",
            _utilization_grade(float(schematic.get("utilization_pct") or 0)),
        )
        schematic_warnings = list(schematic.get("warnings") or [])
        display_utilization_pct = round(
            float(schematic.get("utilization_pct") or load.get("utilization_pct") or 0),
            1,
        )

        order_map = {}
        customers = set()
        for idx, line in enumerate(lines):
            so_num = _normalize_order_identifier(line.get("so_num"), fallback=f"UNASSIGNED-{idx + 1}")
            state = (line.get("state") or "").strip().upper()
            city = (line.get("city") or "").strip()
            zip_code = (line.get("zip") or "").strip()
            cust_name = (line.get("cust_name") or "").strip()
            stop_order = stop_sequence_map.get(_line_stop_key(state, zip_code))
            due_obj = _parse_date(line.get("due_date"))
            due_label = due_obj.isoformat() if due_obj else (line.get("due_date") or "")
            qty = float(line.get("qty") or 0)
            line_length = float(line.get("total_length_ft") or line.get("line_total_feet") or 0)
            if cust_name:
                customers.add(cust_name)

            entry = order_map.get(so_num)
            if not entry:
                entry = {
                    "so_num": so_num,
                    "cust_name": cust_name,
                    "state": state,
                    "city": city,
                    "zip": zip_code,
                    "due_date": due_label,
                    "due_date_obj": due_obj,
                    "stop_order": stop_order,
                    "line_count": 0,
                    "total_qty": 0.0,
                    "total_length_ft": 0.0,
                }
                order_map[so_num] = entry
            entry["line_count"] += 1
            entry["total_qty"] += qty
            entry["total_length_ft"] += line_length
            if stop_order and (not entry.get("stop_order") or stop_order < entry["stop_order"]):
                entry["stop_order"] = stop_order
            if due_obj and (not entry.get("due_date_obj") or due_obj < entry["due_date_obj"]):
                entry["due_date_obj"] = due_obj
                entry["due_date"] = due_obj.isoformat()
            if not entry.get("cust_name") and cust_name:
                entry["cust_name"] = cust_name
            if not entry.get("city") and city:
                entry["city"] = city
            if not entry.get("state") and state:
                entry["state"] = state
            if not entry.get("zip") and zip_code:
                entry["zip"] = zip_code

        ship_date_obj = None
        order_rows = []
        for order in order_map.values():
            due_obj = order.get("due_date_obj")
            if due_obj and (not ship_date_obj or due_obj < ship_date_obj):
                ship_date_obj = due_obj
            order_rows.append(order)
        order_rows.sort(
            key=lambda entry: (
                int(entry.get("stop_order") or 999),
                entry.get("due_date") or "9999-12-31",
                entry.get("so_num") or "",
            )
        )
        order_colors = _build_order_colors_for_lines(
            lines,
            stop_sequence_map=stop_sequence_map,
            stop_palette=stop_color_palette,
        )

        early_orders = []
        for order in order_rows:
            due_obj = order.pop("due_date_obj", None)
            early_days = (due_obj - ship_date_obj).days if due_obj and ship_date_obj else 0
            order["early_days"] = early_days if early_days > 0 else 0
            order["is_early_delivery"] = order["early_days"] > 0
            order["early_flag"] = "YES" if order["is_early_delivery"] else "NO"
            order["total_qty"] = round(order["total_qty"], 1)
            order["total_length_ft"] = round(order["total_length_ft"], 1)
            order["stop_order_display"] = (
                f"{int(order.get('stop_order')):02d}" if order.get("stop_order") else "--"
            )
            destination_bits = []
            if order.get("city"):
                destination_bits.append(order.get("city"))
            if order.get("state"):
                destination_bits.append(order.get("state"))
            order["destination_label"] = ", ".join(destination_bits) if destination_bits else "--"
            if order["is_early_delivery"]:
                early_orders.append(order)

        deck_blocks = {"lower": [], "upper": []}
        for deck_key in ("lower", "upper"):
            positions = [
                pos for pos in (schematic.get("positions") or [])
                if (pos.get("deck") or "lower") == deck_key
            ]
            total_length = sum(float(pos.get("length_ft") or 0) for pos in positions) or 1.0
            for pos in positions:
                length_ft = float(pos.get("length_ft") or 0)
                order_ids = []
                for item in pos.get("items", []) or []:
                    order_id = (item.get("order_id") or "").strip()
                    if order_id and order_id not in order_ids:
                        order_ids.append(order_id)
                if order_ids:
                    if len(order_ids) > 2:
                        label = f"{order_ids[0]}, {order_ids[1]} +{len(order_ids) - 2}"
                    else:
                        label = ", ".join(order_ids)
                else:
                    label = "Open"
                width_pct = max((length_ft / total_length) * 100.0, 8.0) if length_ft else 8.0
                deck_blocks[deck_key].append(
                    {
                        "length_ft": round(length_ft, 1),
                        "order_ids": order_ids,
                        "label": label,
                        "width_pct": round(width_pct, 1),
                    }
                )

        schematic_segments = []
        if deck_blocks["lower"]:
            lower_summary = " > ".join(
                [
                    f"{block['label']} ({block['length_ft']:.1f} ft)"
                    for block in deck_blocks["lower"]
                ]
            )
            schematic_segments.append(f"Lower: {lower_summary}")
        if deck_blocks["upper"]:
            upper_summary = " > ".join(
                [
                    f"{block['label']} ({block['length_ft']:.1f} ft)"
                    for block in deck_blocks["upper"]
                ]
            )
            schematic_segments.append(f"Upper: {upper_summary}")

        early_callout = ""
        if early_orders:
            details = ", ".join(
                [f"SO {entry['so_num']} ({entry['early_days']}d early)" for entry in early_orders[:4]]
            )
            if len(early_orders) > 4:
                details = f"{details}, +{len(early_orders) - 4} more"
            early_callout = f"Customer notification required: {details}."

        route_cities = [order.get("city") for order in order_rows if order.get("city")]
        route_states = [order.get("state") for order in order_rows if order.get("state")]
        route_city = route_cities[0] if route_cities else ""
        route_state = route_states[0] if route_states else ((load.get("destination_state") or "").strip().upper())
        route_label = (
            f"{route_city}, {route_state}".strip(", ")
            if route_city or route_state
            else (load.get("destination_state") or "--")
        )
        unique_cities = sorted({city for city in route_cities if city})
        if len(unique_cities) > 1:
            route_label = f"{route_label} (+{len(unique_cities) - 1} more)"

        total_units = sum((order.get("total_qty") or 0) for order in order_rows)

        rows.append(
            {
                "id": load.get("id"),
                "load_number": load.get("load_number") or f"Load #{load.get('id')}",
                "status": _normalize_session_status(load.get("status") or "PROPOSED"),
                "created_at": load.get("created_at"),
                "origin_plant": (load.get("origin_plant") or "").strip().upper(),
                "destination_state": (load.get("destination_state") or "").strip().upper(),
                "route_label": route_label,
                "trailer_type": trailer_type,
                "utilization_pct": round(float(load.get("utilization_pct") or 0), 1),
                "estimated_cost": round(float(load.get("estimated_cost") or 0), 2),
                "estimated_miles": round(float(load.get("route_distance") or load.get("estimated_miles") or 0), 1),
                "stop_count": len(ordered_stops),
                "line_count": len(lines),
                "order_count": len(order_rows),
                "total_units": round(total_units, 1),
                "ship_date": ship_date_obj.isoformat() if ship_date_obj else "",
                "orders": order_rows,
                "lines": lines,
                "order_numbers": [entry.get("so_num") for entry in order_rows if entry.get("so_num")],
                "customers": sorted(customers),
                "early_orders": early_orders,
                "has_early_delivery": bool(early_orders),
                "early_delivery_callout": early_callout,
                "schematic_utilization_pct": round(float(schematic.get("utilization_pct") or 0), 1),
                "schematic_grade": schematic.get("utilization_grade") or _utilization_grade(float(load.get("utilization_pct") or 0)),
                "schematic_blocks": deck_blocks,
                "schematic_summary_text": " | ".join(schematic_segments),
                "schematic": schematic,
                "display_utilization_pct": display_utilization_pct,
                "over_capacity": bool(schematic.get("exceeds_capacity")),
                "schematic_warnings": schematic_warnings,
                "schematic_warning_count": len(schematic_warnings),
                "has_custom_schematic": False,
                "order_colors": order_colors,
                "auto_trailer_label": "",
                "auto_trailer_reason": "",
            }
        )

    rows.sort(key=lambda entry: (entry.get("load_number") or "", entry.get("id") or 0))
    for idx, row in enumerate(rows, start=1):
        display_id = f"L#{idx:02d}"
        row["display_load_id"] = display_id
        for order in row.get("orders") or []:
            order["display_load_id"] = display_id
    return rows


def _build_load_report_preview_rows(report_rows, limit=8):
    preview = []
    for load in report_rows or []:
        for idx, order in enumerate(load.get("orders") or []):
            preview.append(
                {
                    "load_id": load.get("load_number") or load.get("display_load_id") or "",
                    "stop_order": order.get("stop_order_display") or "--",
                    "so_number": order.get("so_num") or "",
                    "customer_name": order.get("cust_name") or "",
                    "destination_city": order.get("city") or "",
                    "state": order.get("state") or "",
                    "ship_date": load.get("ship_date") or "",
                    "due_date": order.get("due_date") or "",
                    "total_units": order.get("total_qty") or 0,
                    "early_flag": order.get("early_flag") or "NO",
                    "is_group_start": idx == 0,
                }
            )
            if len(preview) >= limit:
                return preview
    return preview


def _hex_to_rgb_tuple(hex_value, fallback):
    raw = str(hex_value or "").strip().lstrip("#")
    if len(raw) == 3:
        raw = "".join([ch * 2 for ch in raw])
    if len(raw) != 6:
        return fallback
    try:
        return tuple(int(raw[idx:idx + 2], 16) for idx in (0, 2, 4))
    except ValueError:
        return fallback


def _blend_rgb(base_rgb, target_rgb, ratio):
    ratio = max(0.0, min(float(ratio), 1.0))
    return tuple(
        int(base_rgb[idx] + (target_rgb[idx] - base_rgb[idx]) * ratio)
        for idx in range(3)
    )


def _measure_text(draw_ctx, text):
    try:
        left, top, right, bottom = draw_ctx.textbbox((0, 0), text)
        return max(right - left, 1), max(bottom - top, 1)
    except AttributeError:
        return draw_ctx.textsize(text)


def _build_excel_schematic_image(load):
    if not (PILImage and ImageDraw and OpenPyxlImage):
        return None

    schematic = load.get("schematic") or {}
    positions = schematic.get("positions") or []
    if not positions:
        return None

    image_width = 360
    image_height = 118
    border_rgb = (177, 194, 216)
    text_rgb = (44, 62, 92)
    muted_text_rgb = (92, 112, 142)
    track_fill_rgb = (233, 241, 252)
    deck_title_rgb = (62, 89, 129)

    image = PILImage.new("RGB", (image_width, image_height), (248, 251, 255))
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle(
        (4, 4, image_width - 5, image_height - 5),
        radius=10,
        fill=(243, 248, 255),
        outline=border_rgb,
        width=1,
    )

    trailer_label = (schematic.get("trailer_type") or load.get("trailer_type") or "TRAILER")
    trailer_label = trailer_label.replace("_", " ")
    draw.text((12, 10), f"{trailer_label} schematic", fill=text_rgb)

    deck_meta = [("lower", "Lower Deck", 34, 67)]
    has_upper = any((pos.get("deck") or "lower") == "upper" for pos in positions)
    if has_upper:
        deck_meta.append(("upper", "Upper Deck", 74, 106))

    order_color_map = load.get("order_colors") or {}
    for deck_key, deck_label, top_y, bottom_y in deck_meta:
        deck_positions = [
            pos for pos in positions
            if (pos.get("deck") or "lower") == deck_key
        ]
        if not deck_positions:
            continue

        draw.text((12, top_y - 13), deck_label, fill=deck_title_rgb)
        left_x = 12
        right_x = image_width - 12
        draw.rounded_rectangle(
            (left_x, top_y, right_x, bottom_y),
            radius=6,
            fill=track_fill_rgb,
            outline=border_rgb,
            width=1,
        )

        usable_width = right_x - left_x - 4
        lengths = [max(float(pos.get("length_ft") or 0), 0.0) for pos in deck_positions]
        total_length = sum(lengths) or float(len(deck_positions))
        cursor_x = left_x + 2

        for idx, pos in enumerate(deck_positions):
            segment_ratio = (lengths[idx] / total_length) if total_length else (1.0 / len(deck_positions))
            segment_width = max(int(round(usable_width * segment_ratio)), 18)
            segment_right = (
                right_x - 2
                if idx == len(deck_positions) - 1
                else min(cursor_x + segment_width, right_x - 2)
            )

            item_order_ids = []
            for item in pos.get("items") or []:
                order_id = str(item.get("order_id") or "").strip()
                if order_id and order_id not in item_order_ids:
                    item_order_ids.append(order_id)

            primary_order = item_order_ids[0] if item_order_ids else ""
            color_hex = order_color_map.get(primary_order, "#94A3B8")
            base_rgb = _hex_to_rgb_tuple(color_hex, (148, 163, 184))
            fill_rgb = _blend_rgb(base_rgb, (255, 255, 255), 0.72)
            outline_rgb = _blend_rgb(base_rgb, (32, 43, 63), 0.25)

            draw.rounded_rectangle(
                (cursor_x, top_y + 2, segment_right, bottom_y - 2),
                radius=4,
                fill=fill_rgb,
                outline=outline_rgb,
                width=1,
            )

            if primary_order:
                label = primary_order
                if len(item_order_ids) > 1:
                    label = f"{primary_order}+{len(item_order_ids) - 1}"
            else:
                label = "OPEN"
            if len(label) > 12:
                label = f"{label[:11]}~"

            text_w, text_h = _measure_text(draw, label)
            label_x = cursor_x + max(((segment_right - cursor_x) - text_w) // 2, 2)
            label_y = top_y + max(((bottom_y - top_y) - text_h) // 2, 1)
            draw.text((label_x, label_y), label, fill=text_rgb)
            cursor_x = segment_right + 2

    draw.text((12, image_height - 17), "Generated from current load stacking plan", fill=muted_text_rgb)

    image_stream = io.BytesIO()
    image.save(image_stream, format="PNG")
    image_stream.seek(0)
    excel_image = OpenPyxlImage(image_stream)
    excel_image.width = 270
    excel_image.height = 88 if has_upper else 72
    return excel_image


def _hex_to_excel_argb(hex_color, fallback="#94A3B8"):
    normalized = _normalize_hex_color(hex_color, fallback)
    return f"FF{normalized.lstrip('#')}"


def _lighten_hex_color(hex_color, ratio=0.84):
    base_rgb = _hex_to_rgb_tuple(_normalize_hex_color(hex_color, "#94A3B8"), (148, 163, 184))
    lighter = _blend_rgb(base_rgb, (255, 255, 255), ratio)
    return "#{:02X}{:02X}{:02X}".format(*lighter)


def _format_street_address(line):
    address1 = str(line.get("address1") or "").strip()
    address2 = str(line.get("address2") or "").strip()
    if address1 and address2:
        return f"{address1} {address2}".strip()
    return address1 or address2


def _build_load_sheet_stops(load):
    lines = load.get("lines") or []
    if not lines:
        return []

    stops_by_key = {}
    for line in lines:
        key = _line_stop_key(line.get("state"), line.get("zip"))
        if key not in stops_by_key:
            stops_by_key[key] = {
                "stop_key": key,
                "stop_order": None,
                "state": (line.get("state") or "").strip().upper(),
                "zip": (line.get("zip") or "").strip(),
                "city": (line.get("city") or "").strip(),
                "address": _format_street_address(line),
                "customers": [],
                "sku_entries": [],
            }

        stop = stops_by_key[key]
        customer = (line.get("cust_name") or "").strip()
        if customer and customer not in stop["customers"]:
            stop["customers"].append(customer)
        if not stop.get("city"):
            stop["city"] = (line.get("city") or "").strip()
        if not stop.get("address"):
            stop["address"] = _format_street_address(line)

        so_num = (line.get("so_num") or "").strip()
        sku = (line.get("sku") or "").strip()
        item = (line.get("item") or "").strip()
        qty_value = float(line.get("qty") or 0)
        qty_text = "{:,.0f}".format(qty_value) if qty_value.is_integer() else "{:,.1f}".format(qty_value)
        descriptor = sku or item or "SKU"
        entry = f"{so_num} / {descriptor} x{qty_text}" if so_num else f"{descriptor} x{qty_text}"
        if entry not in stop["sku_entries"]:
            stop["sku_entries"].append(entry)

    order_stops = {}
    for order in load.get("orders") or []:
        order_key = _line_stop_key(order.get("state"), order.get("zip"))
        raw_sequence = order.get("stop_order")
        if raw_sequence:
            order_stops[order_key] = int(raw_sequence)

    stop_rows = []
    for stop in stops_by_key.values():
        stop_order = order_stops.get(stop["stop_key"])
        stop["stop_order"] = stop_order or 999
        stop_rows.append(stop)

    stop_rows.sort(
        key=lambda entry: (
            int(entry.get("stop_order") or 999),
            entry.get("city") or "",
            entry.get("zip") or "",
        )
    )
    return stop_rows


def _color_luminance(hex_color):
    raw = _normalize_hex_color(hex_color, "#94A3B8").lstrip("#")
    r = int(raw[0:2], 16)
    g = int(raw[2:4], 16)
    b = int(raw[4:6], 16)
    return (0.299 * r) + (0.587 * g) + (0.114 * b)


def _expand_schematic_units_for_position(position):
    expanded = []
    for item in position.get("items") or []:
        units = max(_coerce_int_value(item.get("units"), 0), 0)
        if units <= 0:
            continue
        label = (item.get("sku") or item.get("item") or "SKU").strip() or "SKU"
        stop_sequence = _coerce_int_value(item.get("stop_sequence"), 0)
        for _ in range(units):
            expanded.append(
                {
                    "label": label,
                    "stop_sequence": stop_sequence,
                }
            )
    return expanded


def _write_load_sheet_schematic_grid(
    ws,
    start_row,
    load,
    stop_palette,
    medium_side,
    thin_side,
):
    schematic = load.get("schematic") or {}
    positions = list(schematic.get("positions") or [])
    if not positions:
        return start_row

    deck_groups = [
        ("upper", "Upper Deck"),
        ("lower", "Lower Deck"),
    ]
    columns = list(range(1, 9))
    chunk_size = len(columns)
    row_cursor = start_row

    for deck_key, deck_label in deck_groups:
        deck_positions = [pos for pos in positions if (pos.get("deck") or "lower") == deck_key]
        if not deck_positions:
            continue

        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
        title_cell = ws.cell(row=row_cursor, column=1, value=f"{deck_label} Stacking Schematic")
        title_cell.font = Font(bold=True, color="FF334155")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        row_cursor += 1

        position_chunks = [deck_positions[idx : idx + chunk_size] for idx in range(0, len(deck_positions), chunk_size)]
        for chunk in position_chunks:
            expanded_columns = [_expand_schematic_units_for_position(pos) for pos in chunk]
            max_stack = max((len(column_units) for column_units in expanded_columns), default=1)
            max_stack = max(max_stack, 1)

            # Stack/position row
            ws.row_dimensions[row_cursor].height = 20
            for col_idx in columns:
                cell = ws.cell(row=row_cursor, column=col_idx, value="")
                if col_idx <= len(chunk):
                    cell.value = f"Stack {col_idx}"
                cell.font = Font(bold=True, color="FF475569")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(fill_type="solid", fgColor="FFF8FAFC")
                cell.border = Border(
                    left=medium_side if col_idx == 1 else thin_side,
                    right=medium_side if col_idx == columns[-1] else thin_side,
                    top=thin_side,
                    bottom=thin_side,
                )
            row_cursor += 1

            # Unit cells (one SKU per cell), bottom-aligned within each stack.
            for stack_row in range(max_stack):
                excel_row = row_cursor + stack_row
                ws.row_dimensions[excel_row].height = 22
                for col_idx in columns:
                    cell = ws.cell(row=excel_row, column=col_idx, value="")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=medium_side if col_idx == 1 else thin_side,
                        right=medium_side if col_idx == columns[-1] else thin_side,
                        top=thin_side,
                        bottom=thin_side,
                    )
                    if col_idx > len(expanded_columns):
                        continue
                    col_units = expanded_columns[col_idx - 1]
                    blank_lead = max_stack - len(col_units)
                    unit_idx = stack_row - blank_lead
                    if unit_idx < 0 or unit_idx >= len(col_units):
                        continue
                    unit = col_units[unit_idx]
                    stop_sequence = _coerce_int_value(unit.get("stop_sequence"), 0)
                    stop_color = _color_for_stop_sequence(stop_sequence, stop_palette)
                    cell.value = unit.get("label") or "SKU"
                    cell.fill = PatternFill(fill_type="solid", fgColor=_hex_to_excel_argb(stop_color, fallback="#94A3B8"))
                    text_color = "FFFFFFFF" if _color_luminance(stop_color) < 138 else "FF0F172A"
                    cell.font = Font(bold=True, color=text_color)
            row_cursor += max_stack

            # Stop-sequence legend row aligned to stacks.
            ws.row_dimensions[row_cursor].height = 19
            for col_idx in columns:
                legend = ws.cell(row=row_cursor, column=col_idx, value="")
                legend.alignment = Alignment(horizontal="center", vertical="center")
                legend.border = Border(
                    left=medium_side if col_idx == 1 else thin_side,
                    right=medium_side if col_idx == columns[-1] else thin_side,
                    top=thin_side,
                    bottom=medium_side,
                )
                legend.fill = PatternFill(fill_type="solid", fgColor="FFF8FAFC")
                if col_idx > len(expanded_columns):
                    continue
                col_units = expanded_columns[col_idx - 1]
                top_sequence = _coerce_int_value((col_units[-1] if col_units else {}).get("stop_sequence"), 0)
                if top_sequence > 0:
                    legend.value = f"Stop {top_sequence}"
                    legend.font = Font(bold=True, color="FF334155")
            row_cursor += 1
            row_cursor += 1
    return row_cursor


def _write_load_sheet_block(
    ws,
    start_row,
    load,
    stop_palette,
    medium_side,
    thin_side,
    header_font,
    body_font,
):
    columns = list(range(1, 9))
    has_medium_right_col = columns[-1]
    route_stops = _build_load_sheet_stops(load)

    start_zip = route_stops[0].get("zip") if route_stops else ""
    stop_zip = route_stops[-1].get("zip") if route_stops else ""
    ship_budget = round(float(load.get("estimated_cost") or 0), 2)
    ship_date = load.get("ship_date") or ""

    trailer_label = str(load.get("trailer_type") or "").replace("_", " ").title() or "Trailer"
    load_title_values = [
        "COT Loadsheet",
        load.get("origin_plant") or "",
        load.get("load_number") or load.get("display_load_id") or "",
        trailer_label,
        "",
        "",
        "",
        "",
    ]
    meta_headers = [
        "Start ZIP",
        "Stop ZIP",
        "Carrier",
        "Trailer Type",
        "Trailer Count",
        "Total Miles",
        "Ship Budget",
        "Must Ship By",
    ]
    meta_values = [
        start_zip or "",
        stop_zip or "",
        "TBD",
        trailer_label,
        1,
        round(float(load.get("estimated_miles") or 0), 1),
        ship_budget,
        ship_date,
    ]

    for col_idx in columns:
        cell = ws.cell(row=start_row, column=col_idx, value=load_title_values[col_idx - 1])
        cell.font = Font(bold=True, color="FF1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
        cell.border = Border(
            left=medium_side if col_idx == 1 else thin_side,
            right=medium_side if col_idx == has_medium_right_col else thin_side,
            top=medium_side,
            bottom=medium_side,
        )

    for col_idx in columns:
        header = ws.cell(row=start_row + 1, column=col_idx, value=meta_headers[col_idx - 1])
        header.font = header_font
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.fill = PatternFill(fill_type="solid", fgColor="FFF8FAFC")
        header.border = Border(
            left=medium_side if col_idx == 1 else thin_side,
            right=medium_side if col_idx == has_medium_right_col else thin_side,
            top=thin_side,
            bottom=thin_side,
        )
        value_cell = ws.cell(row=start_row + 2, column=col_idx, value=meta_values[col_idx - 1])
        value_cell.font = body_font
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = Border(
            left=medium_side if col_idx == 1 else thin_side,
            right=medium_side if col_idx == has_medium_right_col else thin_side,
            top=thin_side,
            bottom=thin_side,
        )
        if col_idx == 7:
            value_cell.number_format = "$#,##0.00"
        if col_idx == 8:
            header.fill = PatternFill(fill_type="solid", fgColor="FFFFEB99")
            value_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFF3BF")

    ws.merge_cells(start_row=start_row + 3, start_column=1, end_row=start_row + 3, end_column=8)
    instructions_cell = ws.cell(
        row=start_row + 3,
        column=1,
        value="Check Special Instructions",
    )
    instructions_cell.font = Font(bold=True, color="FFB91C1C")
    instructions_cell.alignment = Alignment(horizontal="center", vertical="center")
    instructions_cell.fill = PatternFill(fill_type="solid", fgColor="FFFEF2F2")
    instructions_cell.border = Border(left=medium_side, right=medium_side, top=thin_side, bottom=thin_side)

    chunk_size = 8
    route_rows_per_chunk = 10
    route_start_row = start_row + 4
    stop_chunks = [route_stops[i : i + chunk_size] for i in range(0, len(route_stops), chunk_size)] or [[]]

    for chunk_idx, stop_chunk in enumerate(stop_chunks):
        chunk_row = route_start_row + (chunk_idx * route_rows_per_chunk)
        row_labels = [
            "Route",
            "Stop #",
            "Customer",
            "Address",
            "City",
            "State",
            "ZIP",
            "Phone",
            "Group",
            "SKU / SO",
        ]
        for offset, label in enumerate(row_labels):
            row_number = chunk_row + offset
            ws.row_dimensions[row_number].height = 26
            if offset in {8, 9}:
                ws.row_dimensions[row_number].height = 42
            for col_idx in columns:
                cell = ws.cell(row=row_number, column=col_idx)
                stop_data = stop_chunk[col_idx - 1] if col_idx <= len(stop_chunk) else None
                stop_sequence = (stop_data or {}).get("stop_order") or 0
                color_hex = _color_for_stop_sequence(stop_sequence, stop_palette)
                light_hex = _lighten_hex_color(color_hex, ratio=0.7 if offset == 1 else 0.84)
                cell.fill = PatternFill(fill_type="solid", fgColor=_hex_to_excel_argb(light_hex, fallback="#EEF2FF"))
                cell.border = Border(
                    left=medium_side if col_idx == 1 else thin_side,
                    right=medium_side if col_idx == has_medium_right_col else thin_side,
                    top=thin_side,
                    bottom=medium_side if offset == len(row_labels) - 1 else thin_side,
                )
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if not stop_data:
                    continue
                if offset == 0:
                    cell.value = f"Stop {int(stop_sequence)}"
                    cell.font = Font(bold=True, color="FF1E293B")
                elif offset == 1:
                    cell.value = int(stop_sequence)
                    cell.font = header_font
                elif offset == 2:
                    customers = stop_data.get("customers") or []
                    cell.value = ", ".join(customers[:2]) + (f" (+{len(customers)-2})" if len(customers) > 2 else "")
                elif offset == 3:
                    cell.value = stop_data.get("address") or "(address unavailable)"
                elif offset == 4:
                    cell.value = stop_data.get("city") or ""
                elif offset == 5:
                    cell.value = stop_data.get("state") or ""
                elif offset == 6:
                    cell.value = stop_data.get("zip") or ""
                elif offset == 7:
                    cell.value = "(blank)"
                elif offset == 8:
                    cell.value = "COT Stickers"
                    cell.font = header_font
                elif offset in {8, 9}:
                    sku_entries = stop_data.get("sku_entries") or []
                    if offset == 9:
                        cell.value = "\n".join(sku_entries[:8])

    after_routes_row = route_start_row + (len(stop_chunks) * route_rows_per_chunk) + 1
    row_after_schematic = _write_load_sheet_schematic_grid(
        ws,
        start_row=after_routes_row,
        load=load,
        stop_palette=stop_palette,
        medium_side=medium_side,
        thin_side=thin_side,
    )
    return max(row_after_schematic + 1, after_routes_row + 3)


def _build_single_load_sheet_workbook(load):
    workbook = Workbook()
    sheet_name = (load.get("load_number") or "Load Sheet")[:31]
    ws = workbook.active
    ws.title = sheet_name

    for col_letter in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col_letter].width = 26
    ws.sheet_view.showGridLines = True

    _write_load_sheet_block(
        ws,
        start_row=1,
        load=load,
        stop_palette=_get_stop_color_palette(),
        medium_side=Side(style="medium", color="FF475569"),
        thin_side=Side(style="thin", color="FFCBD5E1"),
        header_font=Font(bold=True, color="FF1F2937"),
        body_font=Font(color="FF111827"),
    )
    return workbook


def _build_load_report_workbook(planning_session, report_rows):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Load Summary"

    header_fill = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
    header_font = Font(bold=True, color="FF1F2937")
    body_font = Font(color="FF111827")
    muted_font = Font(color="FF94A3B8")
    total_fill = PatternFill(fill_type="solid", fgColor="FFD1D5DB")
    yes_fill = PatternFill(fill_type="solid", fgColor="FFFDE68A")
    yes_font = Font(bold=True, color="FF92400E")
    no_font = Font(color="FF64748B")
    link_font = Font(color="FF2563EB", underline="single")
    thin_side = Side(style="thin", color="FFCBD5E1")
    separator_side = Side(style="medium", color="FF9CA3AF")
    all_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    alternating_fill_a = PatternFill(fill_type="solid", fgColor="FFF3F8FF")
    alternating_fill_b = PatternFill(fill_type="solid", fgColor="FFFFFFFF")

    summary_headers = [
        "Load ID",
        "Stop Order",
        "SO Number",
        "Customer Name",
        "Destination City",
        "State",
        "Ship Date",
        "Due Date",
        "Total Units",
        "Early Delivery Flag",
    ]
    summary.append(summary_headers)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = all_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    all_preview_rows = _build_load_report_preview_rows(report_rows, limit=200000)

    current_row = 2
    group_index = -1
    previous_load_id = None
    for row_idx, row in enumerate(all_preview_rows):
        load_id = str(row.get("load_id") or "").strip()
        if load_id != previous_load_id:
            group_index += 1
        next_load_id = (
            str(all_preview_rows[row_idx + 1].get("load_id") or "").strip()
            if row_idx + 1 < len(all_preview_rows)
            else None
        )
        is_group_end = load_id != next_load_id
        row_fill = alternating_fill_a if (group_index % 2 == 0) else alternating_fill_b

        summary.append(
            [
                load_id,
                row.get("stop_order") or "--",
                row.get("so_number") or "",
                row.get("customer_name") or "",
                row.get("destination_city") or "",
                row.get("state") or "",
                row.get("ship_date") or "",
                row.get("due_date") or "",
                row.get("total_units") or 0,
                row.get("early_flag") or "NO",
            ]
        )
        for col_idx in range(1, 11):
            cell = summary.cell(row=current_row, column=col_idx)
            cell.fill = row_fill
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=separator_side if is_group_end else thin_side,
            )
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal="right" if col_idx in {9} else "left",
                vertical="center",
            )
        if row.get("is_group_start"):
            summary.cell(row=current_row, column=1).font = Font(bold=True, color="FF111827")
        else:
            summary.cell(row=current_row, column=1).font = muted_font
        summary.cell(row=current_row, column=3).font = link_font
        if row.get("early_flag") == "YES":
            summary.cell(row=current_row, column=10).fill = yes_fill
            summary.cell(row=current_row, column=10).font = yes_font
            summary.cell(row=current_row, column=10).alignment = Alignment(horizontal="center")
        else:
            summary.cell(row=current_row, column=10).font = no_font
            summary.cell(row=current_row, column=10).alignment = Alignment(horizontal="center")
        if isinstance(summary.cell(row=current_row, column=9).value, (int, float)):
            summary.cell(row=current_row, column=9).number_format = "#,##0"
        current_row += 1
        previous_load_id = load_id

    totals_row = current_row
    total_units = sum((row.get("total_units") or 0) for row in all_preview_rows)
    summary.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=8)
    summary.cell(row=totals_row, column=1, value="SESSION TOTALS")
    summary.cell(row=totals_row, column=9, value=round(total_units, 0))
    summary.cell(row=totals_row, column=10, value="")
    for col_idx in range(1, 11):
        cell = summary.cell(row=totals_row, column=col_idx)
        cell.fill = total_fill
        cell.font = Font(bold=True, color="FF111827")
        cell.border = all_border
        if col_idx in {9}:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in {9} and isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0"

    summary.column_dimensions["A"].width = 13
    summary.column_dimensions["B"].width = 12
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 24
    summary.column_dimensions["E"].width = 18
    summary.column_dimensions["F"].width = 8
    summary.column_dimensions["G"].width = 13
    summary.column_dimensions["H"].width = 13
    summary.column_dimensions["I"].width = 12
    summary.column_dimensions["J"].width = 16
    summary.freeze_panes = "A2"

    stop_details = workbook.create_sheet(title="Stop Details")
    stop_headers = [
        "Load ID",
        "Stop Order",
        "SO Number",
        "Customer Name",
        "Destination City",
        "State",
        "ZIP",
        "Ship Date",
        "Due Date",
        "Early Delivery Flag",
        "Total Units",
        "Total Length (ft)",
    ]
    stop_details.append(stop_headers)
    for cell in stop_details[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = all_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for load in report_rows or []:
        for order in load.get("orders") or []:
            stop_details.append(
                [
                    load.get("load_number") or load.get("display_load_id") or "",
                    order.get("stop_order_display") or "--",
                    order.get("so_num") or "",
                    order.get("cust_name") or "",
                    order.get("city") or "",
                    order.get("state") or "",
                    order.get("zip") or "",
                    load.get("ship_date") or "",
                    order.get("due_date") or "",
                    order.get("early_flag") or "NO",
                    order.get("total_qty") or 0,
                    order.get("total_length_ft") or 0,
                ]
            )
            for col_idx in range(1, 13):
                cell = stop_details.cell(row=row_idx, column=col_idx)
                cell.border = all_border
                cell.font = body_font
                cell.alignment = Alignment(
                    horizontal="right" if col_idx in {11, 12} else "left",
                    vertical="center",
                )
            stop_details.cell(row=row_idx, column=3).font = link_font
            stop_details.cell(row=row_idx, column=10).alignment = Alignment(horizontal="center")
            if (order.get("early_flag") or "NO") == "YES":
                stop_details.cell(row=row_idx, column=10).fill = yes_fill
                stop_details.cell(row=row_idx, column=10).font = yes_font
            else:
                stop_details.cell(row=row_idx, column=10).font = no_font
            row_idx += 1

    stop_details.column_dimensions["A"].width = 13
    stop_details.column_dimensions["B"].width = 12
    stop_details.column_dimensions["C"].width = 14
    stop_details.column_dimensions["D"].width = 26
    stop_details.column_dimensions["E"].width = 18
    stop_details.column_dimensions["F"].width = 8
    stop_details.column_dimensions["G"].width = 10
    stop_details.column_dimensions["H"].width = 13
    stop_details.column_dimensions["I"].width = 13
    stop_details.column_dimensions["J"].width = 16
    stop_details.column_dimensions["K"].width = 11
    stop_details.column_dimensions["L"].width = 14
    stop_details.freeze_panes = "A2"

    sku_breakdown = workbook.create_sheet(title="SKU Breakdown")
    sku_headers = [
        "Load ID",
        "SO Number",
        "SKU",
        "Item",
        "Item Description",
        "QTY",
        "Unit Length (ft)",
        "Total Length (ft)",
    ]
    sku_breakdown.append(sku_headers)
    for cell in sku_breakdown[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = all_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sku_row = 2
    for load in report_rows or []:
        for line in load.get("lines") or []:
            sku_breakdown.append(
                [
                    load.get("load_number") or load.get("display_load_id") or "",
                    line.get("so_num") or "",
                    line.get("sku") or "",
                    line.get("item") or "",
                    line.get("item_desc") or "",
                    line.get("qty") or 0,
                    line.get("unit_length_ft") or 0,
                    line.get("total_length_ft") or line.get("line_total_feet") or 0,
                ]
            )
            for col_idx in range(1, 9):
                cell = sku_breakdown.cell(row=sku_row, column=col_idx)
                cell.border = all_border
                cell.font = body_font
                cell.alignment = Alignment(
                    horizontal="right" if col_idx in {6, 7, 8} else "left",
                    vertical="center",
                )
            sku_breakdown.cell(row=sku_row, column=2).font = link_font
            sku_row += 1

    sku_breakdown.column_dimensions["A"].width = 13
    sku_breakdown.column_dimensions["B"].width = 14
    sku_breakdown.column_dimensions["C"].width = 12
    sku_breakdown.column_dimensions["D"].width = 12
    sku_breakdown.column_dimensions["E"].width = 30
    sku_breakdown.column_dimensions["F"].width = 9
    sku_breakdown.column_dimensions["G"].width = 15
    sku_breakdown.column_dimensions["H"].width = 15
    sku_breakdown.freeze_panes = "A2"

    load_sheets = workbook.create_sheet(title="Load Sheets")
    for col_letter in ("A", "B", "C", "D", "E", "F", "G", "H"):
        load_sheets.column_dimensions[col_letter].width = 26
    load_sheets.sheet_view.showGridLines = True

    row_pointer = 1
    stop_palette = _get_stop_color_palette()
    medium_side = Side(style="medium", color="FF475569")
    thin_side = Side(style="thin", color="FFCBD5E1")
    for load in report_rows or []:
        row_pointer = _write_load_sheet_block(
            load_sheets,
            row_pointer,
            load,
            stop_palette=stop_palette,
            medium_side=medium_side,
            thin_side=thin_side,
            header_font=header_font,
            body_font=body_font,
        )
        row_pointer += 1

    return workbook


def _get_session_report_data(session_id):
    planning_session = _get_scoped_planning_session_or_404(session_id)

    loads = load_builder.list_loads(None, session_id=session_id)
    session_status = _sync_planning_session_status(session_id, loads=loads)
    session_status = _normalize_session_status(session_status or planning_session.get("status"))
    planning_session["status"] = session_status
    approved_loads = [
        load for load in loads
        if (load.get("status") or "").strip().upper() == STATUS_APPROVED
    ]
    report_rows = _build_load_report_rows(approved_loads)
    return planning_session, report_rows


@app.route("/load-report/<int:session_id>")
def load_report(session_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    planning_session, report_rows = _get_session_report_data(session_id)
    total_miles = round(sum((row.get("estimated_miles") or 0) for row in report_rows), 1)
    total_cost = round(sum((row.get("estimated_cost") or 0) for row in report_rows), 2)
    early_count = sum(1 for row in report_rows if row.get("has_early_delivery"))
    avg_utilization = round(
        (sum((row.get("utilization_pct") or 0) for row in report_rows) / len(report_rows)),
        1,
    ) if report_rows else 0.0
    completion_candidates = [
        _parse_datetime(row.get("created_at"))
        for row in report_rows
        if row.get("created_at")
    ]
    completion_candidates = [value for value in completion_candidates if value]
    completion_dt = max(completion_candidates) if completion_candidates else _parse_datetime(planning_session.get("created_at"))
    completion_label = _format_datetime_label(completion_dt)
    export_preview_rows = _build_load_report_preview_rows(report_rows, limit=8)

    return render_template(
        "load_report.html",
        planning_session=planning_session,
        report_rows=report_rows,
        export_preview_rows=export_preview_rows,
        load_count=len(report_rows),
        total_orders=sum(len(row.get("orders") or []) for row in report_rows),
        total_miles=total_miles,
        total_cost=total_cost,
        avg_utilization=avg_utilization,
        completion_label=completion_label,
        early_load_count=early_count,
    )


@app.route("/load-report/<int:session_id>/export.xlsx")
def load_report_export(session_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    planning_session, report_rows = _get_session_report_data(session_id)
    workbook = _build_load_report_workbook(planning_session, report_rows)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    safe_code = (planning_session.get("session_code") or f"session_{session_id}").replace(" ", "_")
    filename = f"load_report_{safe_code}_{date.today().isoformat()}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/load-report/<int:session_id>/load/<int:load_id>/sheet.xlsx")
def load_report_load_sheet_export(session_id, load_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    planning_session, report_rows = _get_session_report_data(session_id)
    matched = next((row for row in (report_rows or []) if int(row.get("id") or 0) == int(load_id)), None)
    if not matched:
        abort(404)

    workbook = _build_single_load_sheet_workbook(matched)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    raw_load_number = str(matched.get("load_number") or f"load_{load_id}").strip()
    safe_load_number = re.sub(r"[^A-Za-z0-9_-]+", "_", raw_load_number).strip("_") or f"load_{load_id}"
    safe_code = (planning_session.get("session_code") or f"session_{session_id}").replace(" ", "_")
    filename = f"load_sheet_{safe_code}_{safe_load_number}_{date.today().isoformat()}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/planning-sessions/<int:session_id>")
def planning_session_detail(session_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    planning_session = _get_scoped_planning_session_or_404(session_id)
    planning_session["status"] = _normalize_session_status(planning_session.get("status"))

    session_config = {}
    if planning_session.get("config_json"):
        try:
            session_config = json.loads(planning_session.get("config_json") or "{}")
        except json.JSONDecodeError:
            session_config = {}

    loads = load_builder.list_loads(None, session_id=session_id)
    session_status = _sync_planning_session_status(session_id, loads=loads)
    planning_session["status"] = _normalize_session_status(session_status or planning_session.get("status"))
    rollup = _build_planning_session_rollup(loads)
    avg_util = round(
        sum((load.get("utilization_pct") or 0) for load in loads) / len(loads), 1
    ) if loads else 0.0

    return render_template(
        "planning_session_detail.html",
        planning_session=planning_session,
        session_config=session_config,
        loads=rollup["loads"],
        session_orders=rollup["orders"],
        load_count=len(rollup["loads"]),
        order_count=len(rollup["orders"]),
        avg_utilization=avg_util,
        can_manage_sessions=_get_session_role() == ROLE_ADMIN,
    )


@app.route("/planning-sessions/<int:session_id>/summary")
def planning_session_summary(session_id):
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401
    planning_session = db.get_planning_session(session_id)
    if not planning_session or not _can_access_planning_session(planning_session):
        return jsonify({"error": "Session not found"}), 404

    loads = load_builder.list_loads(None, session_id=session_id)
    rollup = _build_planning_session_rollup(loads)
    avg_util = round(
        sum((load.get("utilization_pct") or 0) for load in loads) / len(loads), 1
    ) if loads else 0.0
    total_cost = round(sum((load.get("estimated_cost") or 0) for load in loads), 2)

    return jsonify(
        {
            "session_id": session_id,
            "session_code": planning_session.get("session_code"),
            "status": _normalize_session_status(planning_session.get("status")),
            "load_count": len(rollup["loads"]),
            "order_count": len(rollup["orders"]),
            "avg_utilization": avg_util,
            "total_cost": total_cost,
            "loads": rollup["loads"],
            "orders": rollup["orders"],
        }
    )


@app.route("/planning-sessions/<int:session_id>/archive", methods=["POST"])
def planning_session_archive(session_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _get_scoped_planning_session_or_404(session_id)
    filters = _planning_session_filter_values(request.form)
    if not _archive_session_and_release_loads(session_id):
        abort(404)
    redirect_args = _planning_sessions_redirect_args(filters)
    redirect_args["archived_session_id"] = session_id
    return redirect(url_for("planning_sessions", **redirect_args))


@app.route("/planning-sessions/archive-all", methods=["POST"])
def planning_sessions_archive_all():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    filters = _planning_session_filter_values(request.form)
    role = _get_session_role()
    profile_name = (_get_session_profile_name() or "").strip()
    created_by_filter = filters.get("planner") or None
    if role != ROLE_ADMIN:
        created_by_filter = profile_name
    sessions = db.list_planning_sessions(
        {
            "plant_code": filters.get("plant") or None,
            "created_by": created_by_filter,
            "start_date": filters.get("start") or None,
            "end_date": filters.get("end") or None,
        }
    )
    archived_count = 0
    for entry in sessions:
        if not _can_access_planning_session(entry):
            continue
        status = _normalize_session_status(entry.get("status"))
        if status == "ARCHIVED":
            continue
        session_id = entry.get("id")
        if _archive_session_and_release_loads(session_id):
            archived_count += 1

    redirect_args = _planning_sessions_redirect_args(filters)
    redirect_args["archived_all_count"] = archived_count
    return redirect(url_for("planning_sessions", **redirect_args))


@app.route("/planning-sessions/<int:session_id>/delete", methods=["POST"])
def planning_session_delete(session_id):
    abort(404)


@app.route("/planning-sessions/<int:session_id>/revise")
def planning_session_revise(session_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _get_scoped_planning_session_or_404(session_id)
    return redirect(url_for("orders", session_template_id=session_id))


@app.route("/planning-sessions/<int:session_id>/resume")
def planning_session_resume(session_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _get_scoped_planning_session_or_404(session_id)
    _set_active_planning_session_id(session_id)
    return redirect(url_for("loads", session_id=session_id))


@app.route("/loads/manual/search")
def manual_load_search():
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    allowed_plants = _get_allowed_plants()
    plant_code = (request.args.get("plant") or "").strip().upper()
    if not plant_code or plant_code not in allowed_plants:
        return jsonify({"error": "Invalid plant"}), 400

    q = (request.args.get("q") or "").strip()
    orders = db.list_eligible_manual_orders(plant_code, search=q, limit=25)
    return jsonify({"orders": orders})


@app.route("/loads/manual/suggest")
def manual_load_suggest():
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    allowed_plants = _get_allowed_plants()
    plant_code = (request.args.get("plant") or "").strip().upper()
    if not plant_code or plant_code not in allowed_plants:
        return jsonify({"error": "Invalid plant"}), 400

    seed_so_num = (request.args.get("seed") or "").strip()
    if not seed_so_num:
        return jsonify({"error": "Missing seed order"}), 400

    settings = db.get_optimizer_settings(plant_code) or {}
    try:
        geo_radius = float(settings.get("geo_radius") or load_builder.DEFAULT_BUILD_PARAMS.get("geo_radius") or 0)
    except (TypeError, ValueError):
        geo_radius = float(load_builder.DEFAULT_BUILD_PARAMS.get("geo_radius") or 0)
    try:
        time_window_days = int(settings.get("time_window_days") or load_builder.DEFAULT_BUILD_PARAMS.get("time_window_days") or 0)
    except (TypeError, ValueError):
        time_window_days = int(load_builder.DEFAULT_BUILD_PARAMS.get("time_window_days") or 0)

    candidates = db.list_eligible_manual_orders(plant_code, search=None, limit=None)
    candidate_map = {str(order.get("so_num") or "").strip(): order for order in candidates if order.get("so_num")}
    seed = candidate_map.get(seed_so_num)
    if not seed:
        return jsonify({"error": "Seed order not available in draft scope"}), 404

    zip_coords = geo_utils.load_zip_coordinates()
    seed_zip = geo_utils.normalize_zip(seed.get("zip"))
    seed_coords = zip_coords.get(seed_zip) if seed_zip else None
    seed_due = _parse_date(seed.get("due_date"))

    suggestions = []
    for order in candidates:
        so_num = str(order.get("so_num") or "").strip()
        if not so_num or so_num == seed_so_num:
            continue

        order_due = _parse_date(order.get("due_date"))
        if seed_due and order_due and time_window_days and time_window_days > 0:
            if abs((order_due - seed_due).days) > time_window_days:
                continue

        dist = None
        order_zip = geo_utils.normalize_zip(order.get("zip"))
        order_coords = zip_coords.get(order_zip) if order_zip else None
        if seed_coords and order_coords:
            dist = geo_utils.haversine_distance_coords(seed_coords, order_coords)
            if geo_radius and geo_radius > 0 and dist > geo_radius:
                continue

        suggestions.append(
            {
                "so_num": so_num,
                "cust_name": order.get("cust_name") or "",
                "due_date": order.get("due_date") or "",
                "city": order.get("city") or "",
                "state": order.get("state") or "",
                "zip": order.get("zip") or "",
                "total_length_ft": order.get("total_length_ft") or 0,
                "total_qty": order.get("total_qty") or 0,
                "utilization_pct": order.get("utilization_pct") or 0,
                "distance_miles": round(dist, 1) if dist is not None else None,
            }
        )

    suggestions.sort(
        key=lambda item: (
            item["distance_miles"] is None,
            item["distance_miles"] if item["distance_miles"] is not None else 0,
            item.get("due_date") or "",
        )
    )

    return jsonify(
        {
            "seed": {
                "so_num": seed_so_num,
                "cust_name": seed.get("cust_name") or "",
                "due_date": seed.get("due_date") or "",
                "city": seed.get("city") or "",
                "state": seed.get("state") or "",
                "zip": seed.get("zip") or "",
                "total_length_ft": seed.get("total_length_ft") or 0,
                "total_qty": seed.get("total_qty") or 0,
                "utilization_pct": seed.get("utilization_pct") or 0,
            },
            "suggestions": suggestions[:25],
            "params": {
                "geo_radius": geo_radius,
                "time_window_days": time_window_days,
            },
        }
    )


@app.route("/loads/manual/create", methods=["POST"])
def manual_load_create():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    allowed_plants = _get_allowed_plants()
    session_id = request.form.get("session_id")
    try:
        session_id = int(session_id) if session_id else None
    except (TypeError, ValueError):
        session_id = None
    redirect_session_id = session_id or _get_active_planning_session_id()
    planning_session = db.get_planning_session(session_id) if session_id else None
    if session_id and not planning_session:
        return redirect(
            url_for("loads", manual_error="Select a valid planning session.", session_id=redirect_session_id)
        )
    if planning_session and not _can_access_planning_session(planning_session):
        return redirect(
            url_for("loads", manual_error="Not authorized for this planning session.", session_id=redirect_session_id)
        )

    plant_code = (request.form.get("plant") or "").strip().upper()
    if not plant_code or plant_code not in allowed_plants:
        return redirect(
            url_for("loads", manual_error="Select a valid plant.", session_id=redirect_session_id)
        )
    if planning_session and (planning_session.get("plant_code") or "").strip().upper() != plant_code:
        return redirect(
            url_for(
                "loads",
                manual_error="Selected plant does not match the active planning session.",
                session_id=redirect_session_id,
            )
        )
    if _is_session_sandbox() and not session_id:
        return redirect(
            url_for(
                "loads",
                plants=plant_code,
                manual_error="Sandbox accounts can only create loads inside a sandbox planning session.",
                session_id=redirect_session_id,
            )
        )

    so_nums = list(
        dict.fromkeys(
            [value.strip() for value in request.form.getlist("so_nums") if (value or "").strip()]
        )
    )
    if not so_nums:
        return redirect(
            url_for(
                "loads",
                plants=plant_code,
                manual_error="Select at least one order.",
                session_id=redirect_session_id,
            )
        )

    eligible = db.filter_eligible_manual_so_nums(plant_code, so_nums)
    if eligible != set(so_nums) or len(eligible) != len(so_nums):
        return redirect(
            url_for(
                "loads",
                plants=plant_code,
                manual_error="Some selected orders are no longer available in Draft Loads.",
                session_id=redirect_session_id,
            )
        )

    trailer_type_raw = (request.form.get("trailer_type") or "").strip().upper()
    trailer_type = trailer_type_raw if stack_calculator.is_valid_trailer_type(trailer_type_raw) else None

    # Clear any non-manual draft loads so selected orders can be reassigned.
    if session_id:
        db.clear_unapproved_loads(session_id=session_id)
    else:
        db.clear_unapproved_loads(plant_code)

    result = load_builder.create_manual_load(
        plant_code,
        so_nums,
        trailer_type=trailer_type,
        created_by=_get_session_profile_name() or _get_session_role(),
        session_id=session_id,
    )
    if result.get("errors"):
        _reoptimize_for_plant(plant_code, session_id=session_id)
        message = next(iter(result["errors"].values()))
        return redirect(url_for("loads", plants=plant_code, manual_error=message, session_id=session_id))

    _reoptimize_for_plant(plant_code, session_id=session_id)
    return redirect(url_for("loads", plants=plant_code, tab="draft", reopt="done", session_id=session_id))


def _capacity_for_trailer(trailer_type):
    trailer_key = stack_calculator.normalize_trailer_type(trailer_type, default="STEP_DECK")
    config = stack_calculator.TRAILER_CONFIGS.get(trailer_key, stack_calculator.TRAILER_CONFIGS["STEP_DECK"])
    try:
        return float(config.get("capacity") or 53)
    except (TypeError, ValueError):
        return 53.0


@app.route("/loads/<int:load_id>/manual_add/suggestions")
def manual_add_suggestions(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    load = db.get_load(load_id)
    if not load:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    status = (load.get("status") or STATUS_PROPOSED).upper()
    if status == STATUS_APPROVED:
        return jsonify({"error": "Approved loads cannot be modified."}), 400

    plant_code = load.get("origin_plant")
    trailer_type = stack_calculator.normalize_trailer_type(load.get("trailer_type"), default="STEP_DECK")

    lines = db.list_load_lines(load_id)
    existing_so_nums = {
        _normalize_order_identifier(line.get("so_num"))
        for line in lines
        if _normalize_order_identifier(line.get("so_num"))
    }
    line_totals = {}
    for line in lines:
        so_num = _normalize_order_identifier(line.get("so_num"))
        if not so_num:
            continue
        line_totals[so_num] = line_totals.get(so_num, 0) + float(line.get("total_length_ft") or 0)

    order_rows = db.list_orders_by_so_nums(plant_code, list(existing_so_nums)) if existing_so_nums else []
    order_map = {
        _normalize_order_identifier(row.get("so_num")): row
        for row in order_rows
        if _normalize_order_identifier(row.get("so_num"))
    }

    used_ft = 0.0
    for so_num in existing_so_nums:
        order = order_map.get(so_num)
        if order and order.get("total_length_ft") is not None:
            used_ft += float(order.get("total_length_ft") or 0)
        else:
            used_ft += float(line_totals.get(so_num) or 0)

    zip_coords = geo_utils.load_zip_coordinates()
    stop_coords = []
    for line in lines:
        zip_code = geo_utils.normalize_zip(line.get("zip"))
        coords = zip_coords.get(zip_code) if zip_code else None
        if coords and coords not in stop_coords:
            stop_coords.append(coords)

    load_payload = _build_load_schematic_payload(load_id) or {}
    payload_schematic = load_payload.get("schematic") or {}
    if load_payload.get("trailer_type"):
        trailer_type = stack_calculator.normalize_trailer_type(
            load_payload.get("trailer_type"),
            default=trailer_type,
        )

    lower_deck_length = float(payload_schematic.get("lower_deck_length") or 0.0)
    upper_deck_length = float(payload_schematic.get("upper_deck_length") or 0.0)
    if lower_deck_length > 0 or upper_deck_length > 0:
        capacity_ft = lower_deck_length + upper_deck_length
    else:
        capacity_ft = float(payload_schematic.get("capacity_feet") or _capacity_for_trailer(trailer_type))
    lower_used_ft = float(payload_schematic.get("lower_deck_used_length_ft") or 0.0)
    upper_used_ft = float(payload_schematic.get("upper_deck_effective_length_ft") or 0.0)
    schematic_used_ft = lower_used_ft + upper_used_ft
    if schematic_used_ft <= 0.0:
        schematic_used_ft = float(payload_schematic.get("total_linear_feet") or 0.0)
    if schematic_used_ft > 0.0:
        used_ft = schematic_used_ft

    remaining_ft = max(capacity_ft - used_ft, 0)
    util_pct = round((used_ft / capacity_ft) * 100, 1) if capacity_ft else 0.0

    existing_schematic = {
        "trailer_type": trailer_type,
        "lower_deck_length": 0.0,
        "upper_deck_length": 0.0,
        "lower_remaining_ft": 0.0,
        "upper_remaining_ft": 0.0,
        "decks": {"lower": [], "upper": []},
    }
    if payload_schematic:
        deck_rows = {"lower": [], "upper": []}
        for pos in payload_schematic.get("positions") or []:
            deck = (pos.get("deck") or "lower").strip().lower()
            if deck not in {"lower", "upper"}:
                deck = "lower"
            length_ft = float(pos.get("length_ft") or 0.0)
            effective_length_ft = float(pos.get("effective_length_ft") or length_ft)
            display_length_ft = effective_length_ft if deck == "upper" else length_ft
            capacity_used = max(float(pos.get("capacity_used") or 0.0), 0.0)
            vertical_fill_ratio = min(capacity_used, 1.0)
            vertical_open_ratio = max(1.0 - vertical_fill_ratio, 0.0)
            deck_rows.setdefault(deck, []).append(
                {
                    "position_id": pos.get("position_id") or "",
                    "length_ft": round(length_ft, 1),
                    "effective_length_ft": round(effective_length_ft, 1),
                    "display_length_ft": round(display_length_ft, 1),
                    "vertical_fill_ratio": round(vertical_fill_ratio, 4),
                    "vertical_open_ratio": round(vertical_open_ratio, 4),
                    "capacity_used": round(capacity_used, 4),
                    "overflow_applied": bool(pos.get("overflow_applied")),
                }
            )
        existing_schematic = {
            "trailer_type": trailer_type,
            "lower_deck_length": round(lower_deck_length, 1),
            "upper_deck_length": round(upper_deck_length, 1),
            "lower_remaining_ft": round(max(lower_deck_length - lower_used_ft, 0.0), 1),
            "upper_remaining_ft": round(max(upper_deck_length - upper_used_ft, 0.0), 1),
            "decks": deck_rows,
        }

    search_query = (request.args.get("q") or "").strip()
    candidates = db.list_eligible_manual_orders(
        plant_code,
        search=search_query or None,
        limit=None,
    )
    strategic_setting = _get_effective_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY)
    strategic_customers = _parse_strategic_customers(strategic_setting.get("value_text") or "")
    strategic_groups = [
        {"key": entry.get("key"), "label": entry.get("label")}
        for entry in (strategic_customers or [])
        if entry.get("key")
        and entry.get("label")
        and entry.get("include_in_optimizer_workbench", True)
    ]
    strategic_lookup = {
        str(entry["key"]): entry
        for entry in (strategic_customers or [])
        if entry.get("key") and entry.get("include_in_optimizer_workbench", True)
    }
    candidate_so_nums = []
    seen_candidate_so_nums = set()
    for order in candidates:
        so_num = str(order.get("so_num") or "").strip()
        if not so_num or so_num in existing_so_nums or so_num in seen_candidate_so_nums:
            continue
        seen_candidate_so_nums.add(so_num)
        candidate_so_nums.append(so_num)

    sku_lines = db.list_order_lines_for_so_nums(
        plant_code,
        list(existing_so_nums) + candidate_so_nums,
    )
    sku_rollups = {}
    for line in sku_lines:
        so_num = str(line.get("so_num") or "").strip()
        if not so_num:
            continue
        sku_key = str(line.get("sku") or line.get("item") or "").strip() or "UNKNOWN"
        so_rollup = sku_rollups.setdefault(so_num, {})
        sku_entry = so_rollup.setdefault(
            sku_key,
            {
                "sku": sku_key,
                "ft": 0.0,
                "qty": 0,
            },
        )
        sku_entry["ft"] += float(line.get("total_length_ft") or 0)
        try:
            sku_entry["qty"] += int(float(line.get("qty") or 0))
        except (TypeError, ValueError):
            pass

    def _serialize_sku_rollup(so_num):
        rollup = sku_rollups.get(so_num, {})
        entries = []
        for entry in rollup.values():
            entries.append(
                {
                    "sku": entry.get("sku") or "UNKNOWN",
                    "ft": round(float(entry.get("ft") or 0), 1),
                    "qty": int(entry.get("qty") or 0),
                }
            )
        entries.sort(key=lambda item: (-(item.get("ft") or 0), item.get("sku") or ""))
        return entries

    existing_segments = []
    existing_order_sort = sorted(
        existing_so_nums,
        key=lambda so_num: (
            (order_map.get(so_num) or {}).get("due_date") or "9999-12-31",
            so_num,
        ),
    )
    for so_num in existing_order_sort:
        order = order_map.get(so_num) or {}
        length_ft = (
            float(order.get("total_length_ft") or 0)
            if order.get("total_length_ft") is not None
            else float(line_totals.get(so_num) or 0)
        )
        if length_ft <= 0:
            continue
        existing_segments.append(
            {
                "so_num": so_num,
                "total_length_ft": round(length_ft, 1),
                "sku_breakdown": _serialize_sku_rollup(so_num),
            }
        )

    suggestions = []
    for order in candidates:
        so_num = str(order.get("so_num") or "").strip()
        if not so_num or so_num in existing_so_nums:
            continue
        customer_name = str(order.get("cust_name") or "").strip()
        matched_strategic = None
        for group_entry in strategic_lookup.values():
            if customer_rules.matches_any_customer_pattern(customer_name, group_entry.get("patterns")):
                matched_strategic = group_entry
                break

        dist = None
        order_zip = geo_utils.normalize_zip(order.get("zip"))
        order_coords = zip_coords.get(order_zip) if order_zip else None
        if order_coords and stop_coords:
            dist = min(
                geo_utils.haversine_distance_coords(order_coords, stop)
                for stop in stop_coords
            )

        total_length_ft = float(order.get("total_length_ft") or 0)
        sku_breakdown = _serialize_sku_rollup(so_num)

        suggestions.append(
            {
                "so_num": so_num,
                "cust_name": customer_name,
                "due_date": order.get("due_date") or "",
                "city": order.get("city") or "",
                "state": order.get("state") or "",
                "zip": order.get("zip") or "",
                "total_length_ft": total_length_ft,
                "stack_added_ft": round(total_length_ft, 1),
                "utilization_pct": order.get("utilization_pct") or 0,
                "distance_miles": round(dist, 1) if dist is not None else None,
                "sku_breakdown": sku_breakdown,
                "strategic_key": str((matched_strategic or {}).get("key") or ""),
                "strategic_label": str((matched_strategic or {}).get("label") or ""),
            }
        )

    suggestions.sort(
        key=lambda item: (
            item.get("due_date") or "9999-12-31",
            item["distance_miles"] is None,
            item["distance_miles"] if item["distance_miles"] is not None else 0,
            item.get("so_num") or "",
        )
    )

    return jsonify(
        {
            "load": {
                "id": load_id,
                "origin_plant": plant_code,
                "trailer_type": trailer_type,
            },
            "space": {
                "capacity_ft": round(capacity_ft, 1),
                "used_ft": round(used_ft, 1),
                "remaining_ft": round(remaining_ft, 1),
                "util_pct": util_pct,
            },
            "existing_schematic": existing_schematic,
            "existing_segments": existing_segments,
            "params": {},
            "strategic_groups": strategic_groups,
            "suggestions": suggestions,
        }
    )


@app.route("/loads/<int:load_id>/manual_add", methods=["POST"])
def manual_add_orders(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    load = db.get_load(load_id)
    if not load:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    status = (load.get("status") or STATUS_PROPOSED).upper()
    if status == STATUS_APPROVED:
        return jsonify({"error": "Approved loads cannot be modified."}), 400

    selected = list(
        dict.fromkeys(
            [value.strip() for value in request.form.getlist("so_nums") if (value or "").strip()]
        )
    )
    if not selected:
        return jsonify({"error": "Select at least one order."}), 400

    existing_lines = db.list_load_lines(load_id)
    existing_so_nums = {line.get("so_num") for line in existing_lines if line.get("so_num")}
    if any(so_num in existing_so_nums for so_num in selected):
        return jsonify({"error": "Some selected orders are already in this load."}), 400

    plant_code = load.get("origin_plant")
    eligible = db.filter_eligible_manual_so_nums(plant_code, selected)
    if eligible != set(selected) or len(eligible) != len(selected):
        return jsonify({"error": "Some selected orders are no longer available in Draft Loads."}), 400

    order_lines = db.list_order_lines_for_so_nums(plant_code, selected)
    if not order_lines:
        return jsonify({"error": "No eligible order lines found."}), 400

    db.update_load_build_source(load_id, "MANUAL")
    for line in order_lines:
        db.create_load_line(load_id, line["id"], line.get("total_length_ft") or 0)
    db.delete_load_schematic_override(load_id)

    session_id = load.get("planning_session_id")
    reopt_job_id = _start_reopt_job(
        plant_code,
        session_id=session_id,
        created_by=_get_session_profile_name() or _get_session_role(),
        speed_profile="fast",
    )
    return jsonify(
        {
            "redirect_url": url_for(
                "loads",
                plants=plant_code,
                tab="draft",
                reopt="done",
                session_id=session_id,
            ),
            "status_url": url_for("loads_reopt_job_status", job_id=reopt_job_id),
            "reopt_job_id": reopt_job_id,
        }
    )


@app.route("/loads/reopt_jobs/<job_id>")
def loads_reopt_job_status(job_id):
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    job = _get_reopt_job(job_id)
    if not job:
        return jsonify({"error": "Re-optimization job not found."}), 404

    if job.get("plant_code") not in _get_allowed_plants():
        return jsonify({"error": "Not authorized for this re-optimization job."}), 403

    return jsonify(
        {
            "job_id": job.get("id"),
            "status": job.get("status") or "unknown",
            "plant_code": job.get("plant_code") or "",
            "session_id": job.get("session_id"),
            "created_at": job.get("created_at") or "",
            "started_at": job.get("started_at") or "",
            "finished_at": job.get("finished_at") or "",
            "error": job.get("error") or "",
            "success_message": job.get("success_message") or "",
        }
    )


@app.route("/feedback")
def feedback_log():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    filters = {
        "start_date": request.args.get("start_date") or "",
        "end_date": request.args.get("end_date") or "",
        "planner_id": request.args.get("planner_id") or "",
        "action_type": request.args.get("action_type") or "",
        "reason_category": request.args.get("reason_category") or "",
        "search": request.args.get("search") or "",
        "sort": request.args.get("sort") or "timestamp_desc",
    }
    entries = db.list_load_feedback(filters)
    options = db.list_feedback_filter_options()
    return render_template(
        "feedback_log.html",
        entries=entries,
        filters=filters,
        options=options,
    )


@app.route("/feedback/app")
def app_feedback_log():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    filters = {
        "start_date": request.args.get("start_date") or "",
        "end_date": request.args.get("end_date") or "",
        "planner_id": request.args.get("planner_id") or "",
        "category": request.args.get("category") or "",
        "status": request.args.get("status") or "",
        "search": request.args.get("search") or "",
        "sort": request.args.get("sort") or "timestamp_desc",
    }
    entries = db.list_app_feedback(filters)
    options = db.list_app_feedback_filter_options()
    return render_template(
        "app_feedback_log.html",
        entries=entries,
        filters=filters,
        options=options,
    )


@app.route("/feedback/app", methods=["POST"])
def submit_app_feedback():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    category = (request.form.get("category") or "").strip()
    title = (request.form.get("title") or "").strip()
    message = (request.form.get("message") or "").strip()
    page = (request.form.get("page") or "").strip()
    planner_id = _get_session_profile_name() or _get_session_role()

    next_url = request.form.get("next") or request.referrer or url_for("app_feedback_log")
    if not category or not title or not message:
        return redirect(next_url)

    db.add_app_feedback(
        category=category,
        title=title,
        message=message,
        page=page,
        planner_id=planner_id,
    )
    return redirect(next_url)


@app.route("/feedback/app/<int:feedback_id>/resolve", methods=["POST"])
def resolve_app_feedback(feedback_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    resolved_by = _get_session_profile_name() or _get_session_role()
    db.resolve_app_feedback(feedback_id, resolved_by=resolved_by)
    return redirect(request.form.get("next") or request.referrer or url_for("app_feedback_log"))


@app.route("/loads/<int:load_id>")
def load_detail(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    route_error = (request.args.get("route_error") or "").strip().lower()

    with db.get_connection() as connection:
        load = connection.execute(
            "SELECT * FROM loads WHERE id = ?",
            (load_id,),
        ).fetchone()
    if not load:
        return redirect(url_for("loads", session_id=_get_active_planning_session_id()))

    access_reason = _load_access_failure_reason(dict(load))
    if access_reason:
        return redirect(
            url_for("loads", session_id=load.get("planning_session_id") or _get_active_planning_session_id())
        )

    load_data = dict(load)
    trailer_type = stack_calculator.normalize_trailer_type(load_data.get("trailer_type"), default="STEP_DECK")
    load_data["trailer_type"] = trailer_type
    stop_color_palette = _get_stop_color_palette()
    lines = db.list_load_lines(load_id)
    plant_names = {row["plant_code"]: row["name"] for row in db.list_plants()}
    sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
    stops = []
    stop_map = {}
    zip_coords = geo_utils.load_zip_coordinates()
    for line in lines:
        zip_code = geo_utils.normalize_zip(line.get("zip"))
        state = line.get("state") or ""
        key = f"{zip_code}|{state}"
        if key not in stop_map:
            coords = zip_coords.get(zip_code) if zip_code else None
            stop_map[key] = {
                "zip": zip_code,
                "state": state,
                "customers": set(),
                "lat": coords[0] if coords else None,
                "lng": coords[1] if coords else None,
            }
        if line.get("cust_name"):
            stop_map[key]["customers"].add(line.get("cust_name"))

    for stop in stop_map.values():
        coords = None
        if stop.get("lat") is not None and stop.get("lng") is not None:
            coords = (stop.get("lat"), stop.get("lng"))
        stops.append(
            {
                "zip": stop["zip"],
                "state": stop["state"],
                "customers": sorted(stop["customers"]),
                "lat": stop.get("lat"),
                "lng": stop.get("lng"),
                "coords": coords,
            }
        )

    origin_code = load_data.get("origin_plant")
    origin_coords = geo_utils.plant_coords_for_code(origin_code)
    requires_return_to_origin = _requires_return_to_origin(lines)
    reverse_route = _is_load_route_reversed(load_data)
    ordered_stops = (
        tsp_solver.solve_route(
            origin_coords,
            stops,
            return_to_origin=requires_return_to_origin,
        )
        if origin_coords
        else list(stops)
    )
    ordered_stops = _apply_load_route_direction(ordered_stops, reverse_route=reverse_route)
    stop_sequence_map = _stop_sequence_map_from_ordered_stops(ordered_stops)

    origin_name = plant_names.get(origin_code, PLANT_NAMES.get(origin_code, origin_code))
    route_nodes = [
        {
            "type": "origin",
            "label": origin_code or "",
            "subtitle": origin_name or "",
            "icon": "home",
            "coords": origin_coords,
            "sequence": 0,
        }
    ]
    for idx, stop in enumerate(ordered_stops, start=1):
        coords = None
        if stop.get("lat") is not None and stop.get("lng") is not None:
            coords = (stop.get("lat"), stop.get("lng"))
        route_nodes.append(
            {
                "type": "customer",
                "label": f"{stop.get('state') or ''} {stop.get('zip') or ''}".strip(),
                "subtitle": ", ".join(stop.get("customers") or []),
                "icon": "person_pin_circle",
                "coords": coords,
                "sequence": idx,
            }
        )
    if requires_return_to_origin and origin_coords and len(route_nodes) > 1:
        route_nodes.append(
            {
                "type": "final",
                "label": origin_code or "",
                "subtitle": origin_name or "",
                "icon": "home",
                "coords": origin_coords,
                "sequence": len(route_nodes),
            }
        )

    for node in route_nodes:
        node_type = (node.get("type") or "").strip().lower()
        is_return_origin = (
            node_type == "final"
            and requires_return_to_origin
            and origin_coords
            and node.get("coords") == origin_coords
        )
        if node_type == "origin" or is_return_origin:
            color = "#38bdf8"
        else:
            color = _color_for_stop_sequence(node.get("sequence"), stop_color_palette)
        node["color"] = color
        node["bg"] = f"{color}22"

    route_metrics = _load_route_display_metrics(
        load_data,
        route_nodes,
        use_cached_route=not reverse_route,
    )
    route_legs = route_metrics["route_legs"]
    route_geometry = route_metrics["route_geometry"]

    map_stops = []
    if origin_coords:
        origin_color = route_nodes[0]["color"] if route_nodes else "#38bdf8"
        map_stops.append(
            {
                "type": "origin",
                "lat": origin_coords[0],
                "lng": origin_coords[1],
                "label": origin_name or origin_code,
                "color": origin_color,
            }
        )
    for idx, stop in enumerate(ordered_stops, start=1):
        stop_color = route_nodes[idx]["color"] if idx < len(route_nodes) else "#f59e0b"
        map_stops.append(
            {
                "type": "customer",
                "lat": stop.get("lat"),
                "lng": stop.get("lng"),
                "label": f"{stop.get('state') or ''} {stop.get('zip') or ''}".strip(),
                "color": stop_color,
            }
        )
    if requires_return_to_origin and origin_coords and map_stops:
        final_color = route_nodes[-1]["color"] if route_nodes else "#f59e0b"
        map_stops.append(
            {
                "type": "final",
                "lat": origin_coords[0],
                "lng": origin_coords[1],
                "label": origin_name or origin_code,
                "color": final_color,
            }
        )

    schematic, line_items, order_numbers = _calculate_load_schematic(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
    )
    utilization_pct = schematic.get("utilization_pct", load_data.get("utilization_pct", 0)) or 0
    exceeds_capacity = schematic.get("exceeds_capacity", False)
    over_capacity = exceeds_capacity and len(order_numbers) <= 1
    sku_color_palette = [
        "#137fec",
        "#10b981",
        "#f59e0b",
        "#ef4444",
        "#8b5cf6",
        "#22d3ee",
        "#f472b6",
        "#f97316",
    ]
    sku_colors = {}
    for idx, item in enumerate(line_items):
        sku = item.get("sku") or f"item-{idx}"
        if sku not in sku_colors:
            sku_colors[sku] = sku_color_palette[len(sku_colors) % len(sku_color_palette)]

    load_data["schematic"] = schematic
    load_data["sku_colors"] = sku_colors
    load_data["stops"] = ordered_stops
    load_data["route_nodes"] = route_nodes
    load_data["route_legs"] = route_legs
    load_data["route_distance"] = route_metrics["route_distance"]
    load_data["route_geometry"] = route_geometry
    load_data["map_stops"] = map_stops
    load_data["route_reversed"] = bool(reverse_route)

    due_dates = [ _parse_date(line.get("due_date")) for line in lines ]
    due_dates = [value for value in due_dates if value]
    anchor_date = min(due_dates) if due_dates else None
    today = date.today()

    manifest_rows = []
    for line in lines:
        due_date = _parse_date(line.get("due_date"))
        max_stack = line.get("max_stack_height") or 1
        qty = line.get("qty") or 0
        positions_required = math.ceil(qty / max_stack) if max_stack else qty
        linear_feet = (line.get("unit_length_ft") or 0) * positions_required
        status = ""
        early_days = None
        if due_date and due_date < today:
            status = "Past Due"
        elif due_date and anchor_date:
            delta_days = (due_date - anchor_date).days
            if delta_days > 0:
                early_days = delta_days
                status = f"Early {delta_days}d"

        manifest_rows.append(
            {
                "so_num": line.get("so_num"),
                "cust_name": line.get("cust_name"),
                "destination": f"{line.get('city') or ''}, {line.get('state') or ''} {line.get('zip') or ''}".strip(", "),
                "due_date": line.get("due_date") or "",
                "sku": line.get("sku"),
                "qty": qty,
                "linear_feet": round(linear_feet, 1),
                "status": status,
                "early_days": early_days,
                "stop_sequence": stop_sequence_map.get(
                    _line_stop_key(line.get("state"), line.get("zip"))
                ),
            }
        )

    manifest_rows.sort(
        key=lambda row: (
            int(row.get("stop_sequence") or 999),
            row.get("due_date") or "9999-12-31",
            row.get("so_num") or "",
            row.get("sku") or "",
        )
    )

    utilization_grade = schematic.get("utilization_grade") or "F"

    return render_template(
        "load_detail.html",
        load=load_data,
        lines=lines,
        manifest_rows=manifest_rows,
        schematic=schematic,
        sku_colors=sku_colors,
        stops=ordered_stops,
        over_capacity=over_capacity,
        utilization_pct=utilization_pct,
        utilization_grade=utilization_grade,
        utilization_credit_ft=schematic.get("utilization_credit_ft", 0),
        route_error=route_error,
    )


@app.route("/api/loads/<int:load_id>/route-geometry")
def load_route_geometry(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    load = db.get_load(load_id)
    if not load:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    if _is_load_route_reversed(load):
        return jsonify(
            {
                "load_id": load_id,
                "route_provider": "manual",
                "route_profile": "",
                "route_fallback": True,
                "route_total_miles": load.get("route_total_miles") or load.get("estimated_miles") or 0.0,
                "route_legs": [],
                "route_geometry": [],
            }
        )

    existing_geometry = load.get("route_geometry") or []
    existing_provider = (load.get("route_provider") or "").strip().lower()
    has_cached_road_geometry = bool(existing_geometry) and not bool(load.get("route_fallback")) and existing_provider not in {
        "",
        "none",
        "haversine",
    }
    force_refresh = _coerce_bool_value(request.args.get("force"))
    if has_cached_road_geometry and not force_refresh:
        return jsonify(
            {
                "load_id": load_id,
                "route_provider": load.get("route_provider"),
                "route_profile": load.get("route_profile"),
                "route_fallback": bool(load.get("route_fallback")),
                "route_total_miles": load.get("route_total_miles"),
                "route_legs": load.get("route_legs") or [],
                "route_geometry": existing_geometry,
            }
        )

    lines = db.list_load_lines(load_id)
    zip_coords = geo_utils.load_zip_coordinates()
    stops = _build_route_stops_for_lines(lines, zip_coords)
    origin_code = load.get("origin_plant")
    origin_coords = geo_utils.plant_coords_for_code(origin_code)
    if not origin_coords or not stops:
        return jsonify(
            {
                "load_id": load_id,
                "route_provider": "none",
                "route_profile": "",
                "route_fallback": True,
                "route_total_miles": 0.0,
                "route_legs": [],
                "route_geometry": [],
            }
        )

    requires_return_to_origin = _requires_return_to_origin(lines)
    route_result = routing_service.get_routing_service().build_route(
        origin_coords,
        stops,
        return_to_origin=requires_return_to_origin,
        objective="distance",
        include_geometry=True,
    )

    existing_total_miles = load.get("route_total_miles")
    if existing_total_miles is None:
        existing_total_miles = load.get("estimated_miles")
    try:
        existing_total_miles = float(existing_total_miles or 0.0)
    except (TypeError, ValueError):
        existing_total_miles = 0.0
    route_payload = {
        "route_provider": route_result.get("provider"),
        "route_profile": route_result.get("profile"),
        # Preserve optimization/cost miles from load build (haversine mode).
        "route_total_miles": existing_total_miles,
        "route_legs": load.get("route_legs") or [],
        "route_geometry": route_result.get("geometry_latlng") or [],
        "route_fallback": bool(route_result.get("used_fallback")),
    }
    db.update_load_route_data(load_id, route_payload)

    return jsonify(
        {
            "load_id": load_id,
            "route_provider": route_payload["route_provider"],
            "route_profile": route_payload["route_profile"],
            "route_fallback": route_payload["route_fallback"],
            "route_total_miles": route_payload["route_total_miles"],
            "route_legs": route_payload["route_legs"],
            "route_geometry": route_payload["route_geometry"],
        }
    )


@app.route("/loads/<int:load_id>/carrier", methods=["POST"])
def update_load_carrier(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return _json_session_expired_response()

    body_data = request.get_json(silent=True) or {}
    carrier_key = (
        request.form.get("carrier_key")
        or body_data.get("carrier_key")
        or ""
    ).strip().lower()
    if carrier_key in {"auto", "default"}:
        carrier_key = ""

    valid_keys = {option["key"] for option in LOAD_CARRIER_SELECTION_OPTIONS}
    if carrier_key and carrier_key not in valid_keys:
        return jsonify({"error": "Invalid carrier selection."}), 400

    load = db.get_load(load_id)
    if not load:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    status = (load.get("status") or STATUS_PROPOSED).upper()
    if status == STATUS_APPROVED:
        return jsonify({"error": "Approved loads cannot be modified."}), 400

    updater = _get_session_profile_name() or _get_session_role()
    requested_key = carrier_key or None
    db.update_load_carrier_override(load_id, requested_key, updated_by=updater)

    load_data = _build_load_schematic_payload(load_id)
    if not load_data:
        return jsonify({"error": "Load not found"}), 404

    resolved_key = str(load_data.get("carrier_key") or "").strip().lower()
    if requested_key and resolved_key != requested_key:
        db.update_load_carrier_override(load_id, None, updated_by=updater)
        requested_label = next(
            (
                option.get("label")
                for option in LOAD_CARRIER_SELECTION_OPTIONS
                if option.get("key") == requested_key
            ),
            requested_key.upper(),
        )
        return jsonify(
            {
                "error": (
                    f"{requested_label} is not available for this load's lanes/trailer. "
                    "Carrier override was cleared."
                )
            }
        ), 400

    tab = (request.args.get("tab") or request.form.get("tab") or "").strip().lower()
    payload = _build_schematic_fragment_payload(
        load_data,
        status=(load_data.get("status") or STATUS_PROPOSED).upper(),
        tab=tab,
    )
    payload["ok"] = True
    payload["carrier_override_key"] = requested_key or ""
    return jsonify(payload)


@app.route("/loads/<int:load_id>/trailer", methods=["POST"])
def update_load_trailer(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    body_data = request.get_json(silent=True) or {}
    trailer_type = (
        request.form.get("trailer_type")
        or body_data.get("trailer_type")
        or ""
    ).strip().upper()
    if not stack_calculator.is_valid_trailer_type(trailer_type):
        return jsonify({"error": "Invalid trailer type"}), 400

    load = db.get_load(load_id)
    if not load:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    status = (load.get("status") or STATUS_PROPOSED).upper()
    if status == STATUS_APPROVED:
        return jsonify({"error": "Approved loads cannot be modified."}), 400

    confirm_violation = _coerce_bool_value(
        request.form.get("confirm_violation") or body_data.get("confirm_violation")
    )
    assumptions = _get_stack_capacity_assumptions()

    lines = db.list_load_lines(load_id)
    sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
    zip_coords = geo_utils.load_zip_coordinates()
    ordered_stops = _ordered_stops_for_lines(lines, load["origin_plant"], zip_coords)
    ordered_stops = _apply_load_route_direction(ordered_stops, load=load)
    stop_sequence_map = _stop_sequence_map_from_ordered_stops(ordered_stops)
    order_colors = _build_order_colors_for_lines(
        lines,
        stop_sequence_map=stop_sequence_map,
        stop_palette=_get_stop_color_palette(),
    )

    override = db.get_load_schematic_override(load_id)
    if override:
        units = _build_schematic_units(
            lines,
            sku_specs,
            trailer_type,
            stop_sequence_map=stop_sequence_map,
            order_colors=order_colors,
        )
        units_by_id = {unit["unit_id"]: unit for unit in units}
        base_schematic, _, _ = _calculate_load_schematic(
            lines,
            sku_specs,
            trailer_type,
            stop_sequence_map=stop_sequence_map,
            assumptions=assumptions,
        )
        normalized_layout = _layout_from_schematic(base_schematic, units)

        remapped_schematic, warnings = _build_schematic_from_layout(
            normalized_layout,
            units_by_id,
            trailer_type,
            assumptions=assumptions,
        )
        if warnings and not confirm_violation:
            return jsonify(
                {
                    "requires_confirmation": True,
                    "warnings": warnings,
                    "warning_count": len(warnings),
                    "remap_preview": {
                        "layout": normalized_layout,
                        "metrics": {
                            "utilization_pct": remapped_schematic.get("utilization_pct") or 0,
                            "utilization_grade": remapped_schematic.get("utilization_grade") or "F",
                            "total_linear_feet": remapped_schematic.get("total_linear_feet") or 0,
                            "exceeds_capacity": bool(remapped_schematic.get("exceeds_capacity")),
                        },
                    },
                }
            ), 409

        db.update_load_trailer_type(load_id, trailer_type)
        db.upsert_load_schematic_override(
            load_id,
            trailer_type,
            json.dumps(normalized_layout),
            warnings_json=json.dumps(warnings),
            is_invalid=bool(warnings),
            updated_by=_get_session_profile_name() or _get_session_role(),
        )
        load_data = _build_load_schematic_payload(load_id)
        if not load_data:
            return jsonify({"error": "Load not found"}), 404
        tab = (request.args.get("tab") or request.form.get("tab") or "").strip().lower()
        payload = _build_schematic_fragment_payload(
            load_data,
            status=(load_data.get("status") or STATUS_PROPOSED).upper(),
            tab=tab,
        )
        payload["ok"] = True
        return jsonify(payload)

    order_numbers = {
        _normalize_order_identifier(line.get("so_num"))
        for line in lines
        if _normalize_order_identifier(line.get("so_num"))
    }
    if len(order_numbers) > 1:
        schematic, _, _ = _calculate_load_schematic(
            lines,
            sku_specs,
            trailer_type,
            stop_sequence_map=stop_sequence_map,
            assumptions=assumptions,
        )
        if schematic.get("exceeds_capacity") and not confirm_violation:
            return jsonify(
                {
                    "requires_confirmation": True,
                    "warnings": [
                        {
                            "code": "TRAILER_EXCEEDS_CAPACITY",
                            "message": (
                                "This trailer selection exceeds deck capacity/overhang limits for the current mix. "
                                "Continue anyway to keep the selected trailer."
                            ),
                            "level": "warning",
                        }
                    ],
                    "warning_count": 1,
                }
            ), 409

    best_schematic, _, _ = _calculate_load_schematic(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
        assumptions=assumptions,
    )
    best_units = _build_schematic_units(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
        order_colors=order_colors,
    )
    best_layout = _layout_from_schematic(best_schematic, best_units)
    units_by_id = {unit["unit_id"]: unit for unit in best_units}
    remapped_schematic, warnings = _build_schematic_from_layout(
        best_layout,
        units_by_id,
        trailer_type,
        assumptions=assumptions,
    )
    if warnings and not confirm_violation:
        return jsonify(
            {
                "requires_confirmation": True,
                "warnings": warnings,
                "warning_count": len(warnings),
                "remap_preview": {
                    "layout": best_layout,
                    "metrics": {
                        "utilization_pct": remapped_schematic.get("utilization_pct") or 0,
                        "utilization_grade": remapped_schematic.get("utilization_grade") or "F",
                        "total_linear_feet": remapped_schematic.get("total_linear_feet") or 0,
                        "exceeds_capacity": bool(remapped_schematic.get("exceeds_capacity")),
                    },
                },
            }
        ), 409

    db.update_load_trailer_type(load_id, trailer_type)
    db.upsert_load_schematic_override(
        load_id,
        trailer_type,
        json.dumps(best_layout),
        warnings_json=json.dumps(warnings),
        is_invalid=bool(warnings),
        updated_by=_get_session_profile_name() or _get_session_role(),
    )
    return ("", 204)


@app.route("/loads/<int:load_id>/reverse-order", methods=["POST"])
def reverse_load_order(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    next_url = _safe_next_url(request.form.get("next") or request.args.get("next")) or ""
    load = db.get_load(load_id)
    if not load:
        wants_json = request.is_json or request.accept_mimetypes.best == "application/json"
        if wants_json:
            return jsonify({"error": "Load not found"}), 404
        fallback = url_for("loads", session_id=_get_active_planning_session_id())
        return redirect(next_url or fallback)

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        wants_json = request.is_json or request.accept_mimetypes.best == "application/json"
        if wants_json:
            return jsonify({"error": _load_access_error_message(access_reason)}), 403
        fallback = url_for(
            "loads",
            session_id=load.get("planning_session_id") or _get_active_planning_session_id(),
        )
        return redirect(next_url or fallback)

    status = (load.get("status") or STATUS_PROPOSED).upper()
    if status == STATUS_APPROVED:
        wants_json = request.is_json or request.accept_mimetypes.best == "application/json"
        if wants_json:
            return jsonify({"error": "Approved loads cannot be modified."}), 400
        fallback = url_for(
            "loads",
            session_id=load.get("planning_session_id") or _get_active_planning_session_id(),
        )
        return redirect(next_url or fallback)

    next_value = not _is_load_route_reversed(load)
    db.update_load_route_reversed(load_id, next_value)

    wants_json = request.is_json or request.accept_mimetypes.best == "application/json"
    if wants_json:
        return jsonify({"ok": True, "route_reversed": bool(next_value)})
    fallback = url_for(
        "loads",
        session_id=load.get("planning_session_id") or _get_active_planning_session_id(),
    )
    return redirect(next_url or fallback)


def _build_load_schematic_payload(load_id):
    load = db.get_load(load_id)
    if not load:
        return None

    trailer_type = stack_calculator.normalize_trailer_type(load.get("trailer_type"), default="STEP_DECK")
    assumptions = _get_stack_capacity_assumptions()
    load["trailer_type"] = trailer_type
    lines = db.list_load_lines(load_id)
    sku_specs = {spec["sku"]: spec for spec in db.list_sku_specs()}
    carrier_pricing_context = _build_load_carrier_pricing_context()

    zip_coords = geo_utils.load_zip_coordinates()
    ordered_stops = _ordered_stops_for_lines(lines, load.get("origin_plant"), zip_coords)
    ordered_stops = _apply_load_route_direction(ordered_stops, load=load)
    requires_return_to_origin = _requires_return_to_origin(lines)
    if _alternate_requires_return_hint(
        lines,
        trailer_type,
        load.get("origin_plant"),
        carrier_pricing_context,
    ):
        requires_return_to_origin = True
    if _load_has_lowes_order(lines):
        requires_return_to_origin = True

    carrier_pricing = _resolve_load_carrier_pricing(
        lines=lines,
        trailer_type=trailer_type,
        origin_plant=load.get("origin_plant"),
        ordered_stops=ordered_stops,
        route_legs=load.get("route_legs") or [],
        total_miles=load.get("route_total_miles") or load.get("estimated_miles") or 0.0,
        stop_count=len(ordered_stops),
        requires_return_to_origin=requires_return_to_origin,
        carrier_context=carrier_pricing_context,
        forced_carrier_key=load.get("carrier_override_key"),
    )
    load["carrier_pricing"] = carrier_pricing
    load["carrier_key"] = carrier_pricing.get("carrier_key")
    load["carrier_label"] = carrier_pricing.get("carrier_label")
    load["carrier_source_label"] = carrier_pricing.get("rate_source_label")
    load["carrier_selection_reason"] = carrier_pricing.get("selection_reason")
    load["rate_per_mile"] = carrier_pricing.get("rate_per_mile", load.get("rate_per_mile"))
    load["estimated_cost"] = carrier_pricing.get("total_cost", load.get("estimated_cost"))
    load["freight_breakdown"] = _build_freight_breakdown(
        load,
        stop_fee_amount=_get_stop_fee_amount(),
        fuel_surcharge_per_mile=_get_fuel_surcharge_per_mile(),
        load_minimum_amount=_get_load_minimum_amount(),
    )
    stop_sequence_map = _stop_sequence_map_from_ordered_stops(ordered_stops)
    order_colors = _build_order_colors_for_lines(
        lines,
        stop_sequence_map=stop_sequence_map,
        stop_palette=_get_stop_color_palette(),
    )
    schematic, _, order_numbers = _calculate_load_schematic(
        lines,
        sku_specs,
        trailer_type,
        stop_sequence_map=stop_sequence_map,
        assumptions=assumptions,
    )
    schematic_warnings = list(schematic.get("warnings") or [])
    has_custom_schematic = False
    override = db.get_load_schematic_override(load_id)
    if override and (override.get("trailer_type") or "").strip().upper() == trailer_type:
        units = _build_schematic_units(
            lines,
            sku_specs,
            trailer_type,
            stop_sequence_map=stop_sequence_map,
            order_colors=order_colors,
        )
        units_by_id = {unit["unit_id"]: unit for unit in units}
        try:
            override_layout = json.loads(override.get("layout_json") or "{}")
            normalized_layout = _normalize_edit_layout(override_layout, units_by_id, trailer_type)
            schematic, schematic_warnings = _build_schematic_from_layout(
                normalized_layout,
                units_by_id,
                trailer_type,
                assumptions=assumptions,
            )
            has_custom_schematic = True
        except (json.JSONDecodeError, ValueError):
            has_custom_schematic = False
            schematic_warnings = []

    if (
        not schematic_warnings
        and override
        and (override.get("trailer_type") or "").strip().upper() == trailer_type
        and override.get("warnings_json")
    ):
        try:
            parsed_warnings = json.loads(override.get("warnings_json") or "[]")
        except json.JSONDecodeError:
            parsed_warnings = []
        if isinstance(parsed_warnings, list):
            schematic_warnings = parsed_warnings

    trailer_assignment_rules = _get_effective_trailer_assignment_rules()
    origin_code = _normalize_plant_code(load.get("origin_plant") or load.get("plant"))
    trailer_assignment_rules = dict(trailer_assignment_rules or {})
    trailer_assignment_rules["auto_assign_hotshot_enabled"] = _resolve_auto_hotshot_enabled_for_plant(
        origin_code,
        trailer_rules=trailer_assignment_rules,
    )
    strategic_setting = _get_effective_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY)
    strategic_customers = _parse_strategic_customers(strategic_setting.get("value_text") or "")
    auto_label, auto_reason = _auto_trailer_rule_annotation(
        load=load,
        lines=lines,
        trailer_type=trailer_type,
        schematic=schematic,
        sku_specs=sku_specs,
        stop_sequence_map=stop_sequence_map,
        assumptions=assumptions,
        trailer_assignment_rules=trailer_assignment_rules,
        strategic_customers=strategic_customers,
    )
    load["auto_trailer_label"] = auto_label
    load["auto_trailer_reason"] = auto_reason
    utilization_pct = schematic.get("utilization_pct", load.get("utilization_pct", 0)) or 0
    exceeds_capacity = schematic.get("exceeds_capacity", False)
    over_capacity = exceeds_capacity and len(order_numbers) <= 1

    load["schematic"] = schematic
    load["order_colors"] = order_colors
    load["over_capacity"] = over_capacity
    load["utilization_pct"] = utilization_pct
    # Schematic partial expects this key during async refresh/save responses.
    load["display_utilization_pct"] = utilization_pct
    load["has_custom_schematic"] = has_custom_schematic
    load["schematic_warnings"] = schematic_warnings
    load["schematic_warning_count"] = len(schematic_warnings)
    load["schematic_is_invalid"] = bool((override or {}).get("is_invalid")) if override else False
    load["stack_assumptions"] = assumptions
    return load


@app.route("/loads/<int:load_id>/schematic")
def load_schematic_fragment(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return _json_session_expired_response()

    try:
        load_data = _build_load_schematic_payload(load_id)
    except Exception as exc:
        logger.exception("Failed to build schematic fragment for load_id=%s", load_id)
        return jsonify({"error": f"Unable to load schematic: {exc}"}), 500
    if not load_data:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(load_data)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    status = (load_data.get("status") or STATUS_PROPOSED).upper()
    tab = (request.args.get("tab") or "").strip().lower()
    return jsonify(_build_schematic_fragment_payload(load_data, status=status, tab=tab))


@app.route("/loads/<int:load_id>/schematic/edit")
def load_schematic_edit(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return _json_session_expired_response()

    try:
        payload = _build_load_schematic_edit_payload(load_id)
    except Exception as exc:
        logger.exception("Failed to build schematic edit payload for load_id=%s", load_id)
        return jsonify({"error": f"Unable to load schematic editor: {exc}"}), 500
    if not payload:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(payload.get("load") or {})
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    return jsonify(
        {
            "load_id": payload.get("load_id"),
            "status": payload.get("status"),
            "trailer_type": payload.get("trailer_type"),
            "can_edit": bool(payload.get("can_edit")),
            "units": payload.get("units") or [],
            "layout": payload.get("layout") or {"positions": []},
            "base_layout": payload.get("base_layout") or {"positions": []},
            "metrics": payload.get("metrics") or {},
            "warnings": payload.get("warnings") or [],
            "warning_count": int(payload.get("warning_count") or 0),
            "assumptions": payload.get("assumptions") or {},
        }
    )


@app.route("/loads/<int:load_id>/schematic/edit/save", methods=["POST"])
def save_schematic_edit(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return jsonify({"error": "Session expired"}), 401

    try:
        payload = _build_load_schematic_edit_payload(load_id)
    except Exception as exc:
        logger.exception("Failed to prepare schematic save payload for load_id=%s", load_id)
        return jsonify({"error": f"Unable to save schematic editor state: {exc}"}), 500
    if not payload:
        return jsonify({"error": "Load not found"}), 404

    access_reason = _load_access_failure_reason(payload.get("load") or {})
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403
    if not payload.get("can_edit"):
        return jsonify({"error": "Approved loads are read-only."}), 403

    try:
        data = request.get_json(silent=True) or {}
        requested_trailer_raw = (data.get("trailer_type") or payload.get("trailer_type") or "").strip().upper()
        if not stack_calculator.is_valid_trailer_type(requested_trailer_raw):
            return jsonify({"error": "Invalid trailer type"}), 400
        requested_trailer = stack_calculator.normalize_trailer_type(
            requested_trailer_raw,
            default=payload.get("trailer_type") or "STEP_DECK",
        )
        if requested_trailer != payload.get("trailer_type"):
            return jsonify({"error": "Trailer type changed. Update trailer first, then edit schematic."}), 400

        units_by_id = {unit["unit_id"]: unit for unit in (payload.get("units") or [])}
        try:
            normalized_layout = _normalize_edit_layout(
                data.get("layout") or {},
                units_by_id,
                requested_trailer,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        assumptions = payload.get("assumptions") or _get_stack_capacity_assumptions()
        normalized_layout = _auto_stack_upper_two_across_layout(
            normalized_layout,
            units_by_id,
            requested_trailer,
            assumptions=assumptions,
        )
        schematic, warnings = _build_schematic_from_layout(
            normalized_layout,
            units_by_id,
            requested_trailer,
            assumptions=assumptions,
        )
        confirm_violation = _coerce_bool_value(data.get("confirm_violation"))
        if warnings and not confirm_violation:
            return jsonify(
                {
                    "ok": False,
                    "requires_confirmation": True,
                    "warnings": warnings,
                    "warning_count": len(warnings),
                }
            ), 409

        db.upsert_load_schematic_override(
            load_id,
            requested_trailer,
            json.dumps(normalized_layout),
            warnings_json=json.dumps(warnings),
            is_invalid=bool(warnings),
            updated_by=_get_session_profile_name() or _get_session_role(),
        )

        load_data = _build_load_schematic_payload(load_id)
        if not load_data:
            return jsonify({"error": "Load not found"}), 404

        tab = (request.args.get("tab") or data.get("tab") or "").strip().lower()
        response_payload = _build_schematic_fragment_payload(
            load_data,
            status=(load_data.get("status") or STATUS_PROPOSED).upper(),
            tab=tab,
        )
        response_payload.update(
            {
                "ok": True,
                "warning_count": len(warnings),
                "warnings": warnings,
                "metrics": {
                    "utilization_pct": schematic.get("utilization_pct") or 0,
                    "utilization_grade": schematic.get("utilization_grade") or "F",
                    "total_linear_feet": schematic.get("total_linear_feet") or 0,
                    "exceeds_capacity": bool(schematic.get("exceeds_capacity")),
                },
            }
        )
        return jsonify(response_payload)
    except Exception as exc:
        logger.exception("Unhandled schematic save error for load_id=%s", load_id)
        return jsonify({"error": f"Unhandled schematic save error: {exc.__class__.__name__}: {exc}"}), 500


@app.route("/loads/<int:load_id>/status", methods=["POST"], strict_slashes=False)
def update_load_status(load_id):
    is_async = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    session_redirect = _require_session()
    if session_redirect:
        if is_async:
            return jsonify({"error": "Session expired"}), 401
        return session_redirect

    action = (request.form.get("action") or "").strip().lower()
    if action not in {"approve_draft", "approve_lock", "propose", "unapprove"}:
        return jsonify({"error": "Invalid action"}), 400

    load = db.get_load(load_id)
    if not load:
        return jsonify({"error": "Load not found"}), 404

    allowed_plants = _get_allowed_plants()
    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    plant_filters = _parse_plant_filters(request.form.get("plants"))
    if plant_filters is None:
        plant_filters = []
    plant_scope = [code for code in plant_filters if code in allowed_plants] if plant_filters else []
    if not plant_scope:
        plant_scope = allowed_plants

    current_status = (load.get("status") or STATUS_PROPOSED).upper()
    load_number = load.get("load_number")
    plant_code = load.get("origin_plant")
    year_suffix = _year_suffix()
    planning_session_id = load.get("planning_session_id")
    planning_session = db.get_planning_session(planning_session_id) if planning_session_id else None
    if _is_session_sandbox() and action in {"approve_draft", "approve_lock", "unapprove"}:
        message = "Sandbox accounts cannot approve or finalize loads."
        if is_async:
            return jsonify({"error": message}), 403
        return redirect(
            url_for(
                "loads",
                session_id=planning_session_id or _get_active_planning_session_id(),
                manual_error=message,
            )
        )

    redirect_target = request.referrer or url_for(
        "loads", session_id=_get_active_planning_session_id()
    )

    if action == "propose":
        if current_status != STATUS_PROPOSED:
            db.update_load_status(load_id, STATUS_PROPOSED, load_number)
            if planning_session_id:
                _sync_planning_session_status(planning_session_id)
        if is_async:
            snapshot_loads = (
                load_builder.list_loads(
                    None,
                    session_id=planning_session_id,
                    include_stack_metrics=False,
                )
                if planning_session_id
                else None
            )
            snapshot = _compute_load_progress_snapshot(
                plant_scope=plant_scope,
                all_loads=snapshot_loads,
                allowed_plants=plant_scope if planning_session_id else None,
            )
            return jsonify(
                {
                    "status": STATUS_PROPOSED,
                    "load_id": load_id,
                    "progress": {
                        "approved_orders": snapshot["approved_orders"],
                        "total_orders": snapshot["total_orders"],
                        "progress_pct": snapshot["progress_pct"],
                    },
                    "tab_counts": {
                        "draft": snapshot["draft_tab_count"],
                        "final": snapshot["final_tab_count"],
                    },
                }
            )
        return redirect(redirect_target)

    if action == "unapprove":
        if current_status == STATUS_APPROVED:
            if not load_number:
                seq = db.get_next_load_sequence(plant_code, year_suffix)
                load_number = _format_load_number(plant_code, year_suffix, seq, draft=True)
            else:
                normalized, suffix = _normalize_load_number(load_number)
                if suffix != "D":
                    load_number = f"{normalized}-D"
            db.update_load_status(load_id, STATUS_DRAFT, load_number)
            if planning_session_id:
                _sync_planning_session_status(planning_session_id)
        if is_async:
            snapshot_loads = (
                load_builder.list_loads(
                    None,
                    session_id=planning_session_id,
                    include_stack_metrics=False,
                )
                if planning_session_id
                else None
            )
            snapshot = _compute_load_progress_snapshot(
                plant_scope=plant_scope,
                all_loads=snapshot_loads,
                allowed_plants=plant_scope if planning_session_id else None,
            )
            return jsonify(
                {
                    "status": STATUS_DRAFT,
                    "load_id": load_id,
                    "load_number": load_number,
                    "progress": {
                        "approved_orders": snapshot["approved_orders"],
                        "total_orders": snapshot["total_orders"],
                        "progress_pct": snapshot["progress_pct"],
                    },
                    "tab_counts": {
                        "draft": snapshot["draft_tab_count"],
                        "final": snapshot["final_tab_count"],
                    },
                }
            )
        return redirect(redirect_target)

    if action == "approve_draft":
        if current_status == STATUS_APPROVED:
            return redirect(
                url_for(
                    "loads",
                    session_id=planning_session_id or _get_active_planning_session_id(),
                )
            )
        if not load_number:
            seq = db.get_next_load_sequence(plant_code, year_suffix)
            load_number = _format_load_number(plant_code, year_suffix, seq, draft=True)
        else:
            normalized, suffix = _normalize_load_number(load_number)
            if suffix != "D":
                load_number = f"{normalized}-D"
        db.update_load_status(load_id, STATUS_DRAFT, load_number)
        if planning_session_id:
            _sync_planning_session_status(planning_session_id)
        if is_async:
            snapshot_loads = (
                load_builder.list_loads(
                    None,
                    session_id=planning_session_id,
                    include_stack_metrics=False,
                )
                if planning_session_id
                else None
            )
            snapshot = _compute_load_progress_snapshot(
                plant_scope=plant_scope,
                all_loads=snapshot_loads,
                allowed_plants=plant_scope if planning_session_id else None,
            )
            return jsonify(
                {
                    "status": STATUS_DRAFT,
                    "load_id": load_id,
                    "load_number": load_number,
                    "progress": {
                        "approved_orders": snapshot["approved_orders"],
                        "total_orders": snapshot["total_orders"],
                        "progress_pct": snapshot["progress_pct"],
                    },
                    "tab_counts": {
                        "draft": snapshot["draft_tab_count"],
                        "final": snapshot["final_tab_count"],
                    },
                }
            )
        return redirect(redirect_target)

    if action == "approve_lock":
        should_redirect_to_report = False
        if not load_number:
            if planning_session:
                start_value = (request.form.get("load_number_start") or "").strip()
                starting_sequence = None
                if start_value:
                    if not LOAD_NUMBER_START_PATTERN.fullmatch(start_value):
                        return jsonify(
                            {
                                "error": "Starting load number must be exactly 4 digits.",
                                "invalid_load_number_seed": True,
                            }
                        ), 400
                    starting_sequence = int(start_value)
                reservation = _reserve_session_load_number(
                    planning_session,
                    plant_code,
                    starting_sequence=starting_sequence,
                )
                if reservation.get("needs_start"):
                    return jsonify(
                        {
                            "error": "Enter the first 4-digit load number for this planning session.",
                            "requires_load_number_seed": True,
                            "load_number_prefix": reservation.get("prefix") or "",
                        }
                    ), 428
                reserved_number = reservation.get("load_number")
                if reserved_number:
                    load_number = reserved_number
            if not load_number:
                seq = db.get_next_load_sequence(plant_code, year_suffix)
                load_number = _format_load_number(plant_code, year_suffix, seq, draft=False)
        else:
            normalized, suffix = _normalize_load_number(load_number)
            if suffix == "D":
                load_number = normalized[:-2] if normalized.endswith("-D") else normalized
        db.update_load_status(load_id, STATUS_APPROVED, load_number)
        if planning_session_id:
            next_session_status = _normalize_session_status(
                _sync_planning_session_status(planning_session_id)
            )
            should_redirect_to_report = next_session_status == "COMPLETED"
        if is_async:
            snapshot_loads = (
                load_builder.list_loads(
                    None,
                    session_id=planning_session_id,
                    include_stack_metrics=False,
                )
                if planning_session_id
                else None
            )
            snapshot = _compute_load_progress_snapshot(
                plant_scope=plant_scope,
                all_loads=snapshot_loads,
                allowed_plants=plant_scope if planning_session_id else None,
            )
            return jsonify(
                {
                    "status": STATUS_APPROVED,
                    "load_id": load_id,
                    "load_number": load_number,
                    "redirect_to_report": bool(should_redirect_to_report and planning_session_id),
                    "session_report_url": (
                        url_for("load_report", session_id=planning_session_id)
                        if should_redirect_to_report and planning_session_id
                        else ""
                    ),
                    "progress": {
                        "approved_orders": snapshot["approved_orders"],
                        "total_orders": snapshot["total_orders"],
                        "progress_pct": snapshot["progress_pct"],
                    },
                    "tab_counts": {
                        "draft": snapshot["draft_tab_count"],
                        "final": snapshot["final_tab_count"],
                    },
                }
            )
        if should_redirect_to_report and planning_session_id:
            return redirect(url_for("load_report", session_id=planning_session_id))
        return redirect(redirect_target)

    return redirect(redirect_target)


@app.route("/loads/<int:load_id>/remove_order", methods=["GET", "POST"])
def remove_order_from_load(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    load = db.get_load(load_id)
    if not load:
        return redirect(url_for("loads", session_id=_get_active_planning_session_id()))

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return redirect(url_for("loads", session_id=_get_active_planning_session_id()))

    order_id = (request.values.get("order_id") or "").strip()
    load_data = dict(load)
    plant_code = load_data.get("origin_plant")
    load_label = load_data.get("load_number") or f"Load #{load_id}"
    load_status = (load_data.get("status") or STATUS_PROPOSED).upper()
    session_id = load_data.get("planning_session_id")

    def build_order_summary(order_lines):
        if not order_lines:
            return {}
        due_dates = [_parse_date(line.get("due_date")) for line in order_lines if line.get("due_date")]
        due_date = min(due_dates).strftime("%Y-%m-%d") if due_dates else (order_lines[0].get("due_date") or "")
        city = order_lines[0].get("city") or ""
        state = order_lines[0].get("state") or ""
        zip_code = order_lines[0].get("zip") or ""
        location = ", ".join([part for part in [city, state] if part]).strip()
        if zip_code:
            location = f"{location} {zip_code}".strip()
        return {
            "customer": order_lines[0].get("cust_name") or "",
            "due_date": due_date,
            "location": location,
            "total_qty": sum((line.get("qty") or 0) for line in order_lines),
        }

    reasons_options = ORDER_REMOVAL_REASONS

    lines = db.list_load_lines(load_id)
    order_lines = [line for line in lines if line.get("so_num") == order_id]

    if request.method == "GET":
        if not order_id or not order_lines:
            return redirect(url_for("loads", plants=plant_code, session_id=session_id))
        order_summary = build_order_summary(order_lines)
        return render_template(
            "remove_feedback.html",
            load=load_data,
            load_label=load_label,
            load_status=load_status,
            order_id=order_id,
            order_summary=order_summary,
            reasons_options=reasons_options,
            selected_reasons=[],
            notes="",
            error=None,
            plant_filters=[plant_code],
            plant_filter_param=plant_code,
        )

    if not order_id or not order_lines:
        return redirect(url_for("loads", plants=plant_code, session_id=session_id))

    reason_category = (request.form.get("reason_category") or "").strip()
    details = (request.form.get("details") or "").strip()
    selected_reasons = request.form.getlist("reasons")
    notes = (request.form.get("notes") or "").strip()
    if not reason_category and selected_reasons:
        reason_category = ", ".join([reason for reason in selected_reasons if reason])
    if not details:
        details = notes

    if not reason_category or len(details or "") < 10:
        error_message = "Select a reason and add at least 10 characters before removing this order."
        return redirect(
            url_for(
                "loads",
                plants=plant_code,
                feedback_error=error_message,
                feedback_target=f"order-{load_id}-{order_id}",
                session_id=session_id,
            )
        )

    db.add_load_feedback(
        load_id,
        order_id=order_id,
        action_type="order_removed",
        reason_category=reason_category,
        details=details,
        planner_id=_get_session_profile_name() or _get_session_role(),
    )
    db.remove_order_from_load(load_id, order_id)
    if db.count_load_lines(load_id) == 0:
        db.delete_load(load_id)

    session_id = load_data.get("planning_session_id") if load_data else None
    _reoptimize_for_plant(plant_code, session_id=session_id)
    return redirect(url_for("loads", plants=plant_code, reopt="done", session_id=session_id))


@app.route("/loads/<int:load_id>/reject", methods=["POST"])
def reject_load(load_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    load = db.get_load(load_id)
    if not load:
        if request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"error": "Load not found"}), 404
        return redirect(
            url_for(
                "loads",
                feedback_error="Load no longer exists. It may have already been returned or refreshed.",
                session_id=_get_active_planning_session_id(),
            )
        )

    access_reason = _load_access_failure_reason(load)
    if access_reason:
        return jsonify({"error": _load_access_error_message(access_reason)}), 403

    session_id = load.get("planning_session_id")
    reason_category = (request.form.get("reason_category") or "").strip()
    details = (request.form.get("details") or "").strip()
    if not reason_category or len(details or "") < 10:
        error_message = "Select a reason and add at least 10 characters before returning these orders to the pool."
        return redirect(
            url_for(
                "loads",
                plants=load["origin_plant"],
                feedback_error=error_message,
                feedback_target=f"load-return-{load_id}",
                session_id=session_id,
            )
        )

    db.add_load_feedback(
        load_id,
        order_id=None,
        action_type="orders_returned_to_pool",
        reason_category=reason_category,
        details=details,
        planner_id=_get_session_profile_name() or _get_session_role(),
    )
    db.delete_load(load_id)
    if session_id:
        _sync_planning_session_status(session_id)
    return redirect(url_for("loads", plants=load["origin_plant"], tab="draft", session_id=session_id))


@app.route("/loads/clear", methods=["POST"])
def clear_loads():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    plant_filters = _resolve_plant_filters(request.form.get("plants") or request.form.get("plant"))
    plant_scope = plant_filters or _get_allowed_plants()
    tab = (request.form.get("tab") or "").strip().lower()
    sort_mode = (request.form.get("sort") or "").strip().lower()
    today_param = request.form.get("today")
    session_id = request.form.get("session_id")
    try:
        session_id = int(session_id) if session_id else None
    except (TypeError, ValueError):
        session_id = None
    redirect_session_id = session_id
    if session_id:
        planning_session = db.get_planning_session(session_id)
        if not planning_session or not _can_access_planning_session(planning_session):
            abort(404)
    elif _is_session_sandbox():
        return redirect(
            url_for(
                "loads",
                plants=",".join(plant_filters) if plant_filters else None,
                tab=tab or None,
                sort=sort_mode or None,
                today=today_param or None,
                session_id=_get_active_planning_session_id(),
                manual_error="Sandbox accounts can only clear loads inside an active sandbox session.",
            )
        )

    if session_id:
        _archive_session_and_release_loads(session_id)
        redirect_session_id = None
    elif plant_filters:
        _reintroduce_orders_to_pool(plant_scope)
        for plant in plant_scope:
            db.clear_loads_for_plant(plant)
    else:
        _reintroduce_orders_to_pool(plant_scope)
        for plant in plant_scope:
            db.clear_loads_for_plant(plant)

    return redirect(
        url_for(
            "loads",
            plants=",".join(plant_filters) if plant_filters else None,
            tab=tab or None,
            sort=sort_mode or None,
            today=today_param or None,
            session_id=redirect_session_id,
        )
    )


@app.route("/loads/approve_all", methods=["POST"])
def approve_all_loads():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    plant_filters = _resolve_plant_filters(request.form.get("plants") or request.form.get("plant"))
    plant_scope = plant_filters or _get_allowed_plants()
    status_filter = (request.form.get("status") or "").strip().upper()
    tab = (request.form.get("tab") or "").strip().lower()
    sort_mode = (request.form.get("sort") or "").strip().lower()
    today_param = request.form.get("today")
    session_id = request.form.get("session_id")
    try:
        session_id = int(session_id) if session_id else None
    except (TypeError, ValueError):
        session_id = None
    if _is_session_sandbox():
        return redirect(
            url_for(
                "loads",
                plants=",".join(plant_filters) if plant_filters else None,
                status=status_filter or None,
                tab=tab or None,
                sort=sort_mode or None,
                today=today_param or None,
                session_id=session_id or _get_active_planning_session_id(),
                manual_error="Sandbox accounts cannot approve or finalize loads.",
            )
        )
    if session_id:
        planning_session = db.get_planning_session(session_id)
        if not planning_session or not _can_access_planning_session(planning_session):
            abort(404)

    if not plant_scope:
        return redirect(url_for("loads", session_id=session_id or _get_active_planning_session_id()))

    placeholders = ", ".join("?" for _ in plant_scope)
    params = list(plant_scope)
    status_clause = ""
    if status_filter:
        status_clause = " AND UPPER(status) = ?"
        params.append(status_filter)
    session_clause = ""
    if session_id:
        session_clause = " AND planning_session_id = ?"
        params.append(session_id)

    with db.get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id, load_number, origin_plant, status
            FROM loads
            WHERE origin_plant IN ({placeholders})
            {status_clause}
            {session_clause}
            """,
            params,
        ).fetchall()

    planning_session = db.get_planning_session(session_id) if session_id else None
    year_suffix = _year_suffix()
    seed_year_suffix = _planning_session_year_suffix(planning_session) if planning_session else year_suffix
    for row in rows:
        current_status = (row["status"] or STATUS_PROPOSED).upper()
        if current_status == STATUS_APPROVED:
            continue
        plant_code = row["origin_plant"]
        load_number = row["load_number"]
        if not load_number:
            if planning_session:
                reservation = _reserve_session_load_number(planning_session, plant_code)
                if reservation.get("needs_start"):
                    seed_seq = db.get_next_load_sequence(plant_code, seed_year_suffix)
                    reservation = _reserve_session_load_number(
                        planning_session,
                        plant_code,
                        starting_sequence=seed_seq,
                    )
                reserved_number = reservation.get("load_number")
                if reserved_number:
                    load_number = reserved_number
            if not load_number:
                seq = db.get_next_load_sequence(plant_code, year_suffix)
                load_number = _format_load_number(plant_code, year_suffix, seq, draft=False)
        else:
            normalized, suffix = _normalize_load_number(load_number)
            if suffix == "D":
                load_number = normalized[:-2] if normalized.endswith("-D") else normalized
        db.update_load_status(row["id"], STATUS_APPROVED, load_number)

    if session_id:
        _sync_planning_session_status(session_id)
        return redirect(url_for("load_report", session_id=session_id))

    return redirect(
        url_for(
            "loads",
            plants=",".join(plant_filters) if plant_filters else None,
            status=status_filter or None,
            tab=tab or None,
            sort=sort_mode or None,
            today=today_param or None,
            session_id=session_id,
        )
    )


@app.route("/loads/approve_full", methods=["POST"])
def approve_full_truckloads():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    plant_filters = _resolve_plant_filters(request.form.get("plants") or request.form.get("plant"))
    plant_scope = plant_filters or _get_allowed_plants()
    tab = (request.form.get("tab") or "").strip().lower()
    sort_mode = (request.form.get("sort") or "").strip().lower()
    today_param = request.form.get("today")
    session_id = request.form.get("session_id")
    try:
        session_id = int(session_id) if session_id else None
    except (TypeError, ValueError):
        session_id = None
    if _is_session_sandbox():
        return redirect(
            url_for(
                "loads",
                plants=",".join(plant_filters) if plant_filters else None,
                tab=tab or None,
                sort=sort_mode or None,
                today=today_param or None,
                session_id=session_id or _get_active_planning_session_id(),
                manual_error="Sandbox accounts cannot approve or finalize loads.",
            )
        )
    if session_id:
        planning_session = db.get_planning_session(session_id)
        if not planning_session or not _can_access_planning_session(planning_session):
            abort(404)

    if not plant_scope:
        return redirect(url_for("loads", session_id=session_id or _get_active_planning_session_id()))

    all_loads = load_builder.list_loads(None, session_id=session_id)
    candidates = [
        load
        for load in all_loads
        if load.get("origin_plant") in plant_scope
        and (load.get("status") or STATUS_PROPOSED).upper() in {STATUS_PROPOSED, STATUS_DRAFT}
        and (load.get("build_source") or "OPTIMIZED").upper() != "MANUAL"
        and _is_full_truckload(load)
    ]

    planning_session = db.get_planning_session(session_id) if session_id else None
    year_suffix = _year_suffix()
    seed_year_suffix = _planning_session_year_suffix(planning_session) if planning_session else year_suffix
    for load in candidates:
        current_status = (load.get("status") or STATUS_PROPOSED).upper()
        if current_status == STATUS_APPROVED:
            continue
        plant_code = load.get("origin_plant")
        load_number = load.get("load_number")
        if not load_number:
            if planning_session:
                reservation = _reserve_session_load_number(planning_session, plant_code)
                if reservation.get("needs_start"):
                    seed_seq = db.get_next_load_sequence(plant_code, seed_year_suffix)
                    reservation = _reserve_session_load_number(
                        planning_session,
                        plant_code,
                        starting_sequence=seed_seq,
                    )
                reserved_number = reservation.get("load_number")
                if reserved_number:
                    load_number = reserved_number
            if not load_number:
                seq = db.get_next_load_sequence(plant_code, year_suffix)
                load_number = _format_load_number(plant_code, year_suffix, seq, draft=False)
        else:
            normalized, suffix = _normalize_load_number(load_number)
            if suffix == "D":
                load_number = normalized[:-2] if normalized.endswith("-D") else normalized
        db.update_load_status(load["id"], STATUS_APPROVED, load_number)

    if session_id:
        _sync_planning_session_status(session_id)

    return redirect(
        url_for(
            "loads",
            plants=",".join(plant_filters) if plant_filters else None,
            tab=tab or None,
            sort=sort_mode or None,
            today=today_param or None,
            session_id=session_id,
        )
    )


@app.route("/loads/reject_all", methods=["POST"])
def reject_all_loads():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    plant_filters = _resolve_plant_filters(request.form.get("plants") or request.form.get("plant"))
    plant_scope = plant_filters or _get_allowed_plants()
    status_filter = (request.form.get("status") or "").strip().upper()
    tab = (request.form.get("tab") or "").strip().lower()
    session_id = request.form.get("session_id")
    try:
        session_id = int(session_id) if session_id else None
    except (TypeError, ValueError):
        session_id = None
    if session_id:
        planning_session = db.get_planning_session(session_id)
        if not planning_session or not _can_access_planning_session(planning_session):
            abort(404)
    elif _is_session_sandbox():
        return redirect(
            url_for(
                "loads",
                plants=",".join(plant_filters) if plant_filters else None,
                status=status_filter or None,
                tab=tab or None,
                session_id=_get_active_planning_session_id(),
                manual_error="Sandbox accounts can only return loads inside an active sandbox session.",
            )
        )

    if not plant_scope:
        return redirect(url_for("loads", session_id=session_id or _get_active_planning_session_id()))

    if tab == "final":
        return redirect(
            url_for(
                "loads",
                plants=",".join(plant_filters) if plant_filters else None,
                status=status_filter or None,
                tab=tab,
            )
        )

    placeholders = ", ".join("?" for _ in plant_scope)
    params = list(plant_scope)
    status_clause = ""
    if status_filter:
        status_clause = " AND UPPER(status) = ?"
        params.append(status_filter)
    else:
        status_clause = " AND UPPER(status) != ?"
        params.append(STATUS_APPROVED)
    session_clause = ""
    if session_id:
        session_clause = " AND planning_session_id = ?"
        params.append(session_id)

    with db.get_connection() as connection:
        rows = connection.execute(
            f"""
              SELECT id, origin_plant
              FROM loads
              WHERE origin_plant IN ({placeholders})
              {status_clause}
              {session_clause}
              """,
              params,
          ).fetchall()

        for row in rows:
            connection.execute(
                "DELETE FROM load_lines WHERE load_id = ?",
                (row["id"],),
            )
            connection.execute(
                "DELETE FROM loads WHERE id = ?",
                (row["id"],),
            )
        connection.commit()

    plants_to_reopt = sorted({row["origin_plant"] for row in rows})
    for plant in plants_to_reopt:
        _reoptimize_for_plant(plant, session_id=session_id)

    return redirect(
        url_for(
            "loads",
              plants=",".join(plant_filters) if plant_filters else None,
              status=status_filter or None,
              tab=tab or None,
              reopt="done" if plants_to_reopt else None,
              session_id=session_id,
          )
      )


@app.route("/rates")
def rates():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    if _get_session_role() != ROLE_ADMIN:
        return redirect(url_for("settings", tab="rates"))
    rates_data = db.list_rate_matrix()
    plants, states, matrix = _build_rate_matrix(rates_data)
    return render_template(
        "rates.html",
        rates=rates_data,
        plants=plants,
        states=states,
        matrix=matrix,
    )


@app.route("/settings")
def settings():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    current_role = _get_session_role()
    is_admin = current_role == ROLE_ADMIN
    can_edit_settings = current_role in {ROLE_ADMIN, ROLE_PLANNER}
    current_profile_name = (_get_session_profile_name() or "").strip()
    tab = (request.args.get("tab") or "overview").strip().lower()
    valid_tabs = {"overview", "rates", "skus", "lookups"}
    if tab not in valid_tabs:
        tab = "overview"

    rates_data = []
    rate_plants = []
    rate_states = []
    rate_matrix = {}
    rate_matrix_serialized = {}
    specs = []
    recent_specs = []
    sku_categories = []
    sku_lengths = []
    sku_step_decks = []
    sku_flat_beds = []
    optimizer_exception_category_options = []
    planner_specs = []
    system_specs = []
    source_led_specs = []
    source_led_mapped_count = 0
    source_led_unmapped_count = 0
    source_led_missing_spec_count = 0
    source_led_unique_sku_count = 0
    lookups_data = []
    plants_data = []
    strategic_customers_raw = ""
    strategic_customers = []
    stop_color_rows = []
    trailer_assignment_rules_admin = _get_trailer_assignment_rules()
    planner_trailer_overrides = (
        _get_planner_trailer_rule_overrides(current_profile_name)
        if current_role == ROLE_PLANNER
        else {}
    )
    trailer_assignment_rules = _merge_trailer_assignment_rules(
        trailer_assignment_rules_admin,
        planner_trailer_overrides,
    )
    trailer_assignment_rules_override_active = bool(planner_trailer_overrides)
    rate_table_contexts = _get_rate_table_contexts()
    ryder_dedicated_rate_table = {}
    default_rate_change_metadata = {"changed_index": {}, "changed_count": 0}
    lst_rate_matrix = {}
    alternate_trailer_rates = {}
    rates_v2_payload = {}
    optimizer_defaults = _default_optimize_form()
    optimizer_defaults.update(_get_optimizer_default_settings())
    optimizer_plant_defaults = []
    util_grade_thresholds = []
    fuel_surcharge_per_mile = _get_fuel_surcharge_per_mile()
    global_rate_metrics = _get_rates_overview_metrics()

    if tab in {"overview", "rates"}:
        rates_data = db.list_rate_matrix()
        rate_plants, rate_states, rate_matrix = _build_rate_matrix_records(rates_data)
        rate_matrix_serialized = _build_serialized_rate_matrix(rate_plants, rate_states, rate_matrix)
        ryder_dedicated_rate_table = _get_ryder_dedicated_rate_table()
        default_rate_change_metadata = _get_default_rate_change_metadata()
        lst_rate_matrix = _get_lst_rate_matrix()
        alternate_trailer_rates = _get_alternate_trailer_rates()
        rates_v2_payload = _build_rates_v2_payload(
            rate_plants,
            rate_states,
            rate_matrix_serialized,
            global_rate_metrics,
            fuel_surcharge_per_mile,
            ryder_dedicated_rate_table,
            lst_rate_matrix,
            alternate_trailer_rates,
        )
    if tab in {"overview", "skus", "lookups"}:
        specs = db.list_sku_specs()
        category_source = {
            (spec.get("category") or "").strip().upper()
            for spec in specs
            if (spec.get("category") or "").strip()
        }
        category_source.update(DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES)
        category_source.update(
            stack_calculator.normalize_upper_deck_exception_categories(
                optimizer_defaults.get("upper_deck_exception_categories"),
                default=DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
            )
        )
        optimizer_exception_category_options = sorted(category_source)
    if tab == "overview":
        recent_specs = sorted(
            specs,
            key=lambda spec: spec.get("added_at") or spec.get("created_at") or "",
            reverse=True,
        )[:5]
    if tab == "skus":
        def _collect_numeric_filters(field_name, cast=int):
            values = set()
            for spec in specs:
                raw_value = spec.get(field_name)
                if raw_value in (None, ""):
                    continue
                try:
                    if cast is int:
                        normalized = str(int(float(raw_value)))
                    else:
                        normalized = f"{float(raw_value):g}"
                except (TypeError, ValueError):
                    continue
                values.add(normalized)
            return sorted(values, key=lambda value: float(value))

        planner_specs = [spec for spec in specs if _sku_is_planner_input(spec)]
        system_specs = [spec for spec in specs if not _sku_is_planner_input(spec)]
        raw_categories = {
            (spec.get("category") or "").strip().upper()
            for spec in specs
        }
        sku_categories = sorted([category for category in raw_categories if category])
        sku_lengths = _collect_numeric_filters("length_with_tongue_ft", cast=float)
        sku_step_decks = _collect_numeric_filters("max_stack_step_deck", cast=int)
        sku_flat_beds = _collect_numeric_filters("max_stack_flat_bed", cast=int)
        source_led_specs = _build_source_led_cheat_sheet_rows(specs)
        source_led_mapped_count = sum(
            1 for row in source_led_specs if row.get("mapping_status") in {"Mapped", "Mapped (Cargo Rule)"}
        )
        source_led_unmapped_count = sum(
            1 for row in source_led_specs if row.get("mapping_status") == "Unmapped"
        )
        source_led_missing_spec_count = sum(
            1 for row in source_led_specs if row.get("mapping_status") == "Mapped SKU Missing Spec"
        )
        source_led_unique_sku_count = len(
            {
                sku
                for row in source_led_specs
                for sku in (row.get("mapped_sku_list") or [])
                if str(sku or "").strip()
            }
        )
    elif tab == "lookups":
        lookups_data = db.list_item_lookups()
    if tab in {"overview", "plants"}:
        plants_data = db.list_plants()
    if tab == "overview":
        optimizer_plant_defaults = _build_optimizer_plant_defaults(
            optimizer_defaults=optimizer_defaults,
            trailer_rules=trailer_assignment_rules,
            plant_rows=plants_data,
        )
    if tab in {"overview", "planning_tools"}:
        setting = _get_effective_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY)
        strategic_customers_raw = setting.get("value_text") or ""
        strategic_customers = _parse_strategic_customers(strategic_customers_raw)
        util_grade_thresholds = _build_utilization_grade_rows(_get_utilization_grade_thresholds())
        stop_color_rows = [
            {"sequence": idx, "color": color}
            for idx, color in enumerate(_get_stop_color_palette(), start=1)
        ]

    return render_template(
        "settings.html",
        tab=tab,
        global_rate_metrics=global_rate_metrics,
        rates=rates_data,
        rate_plants=rate_plants,
        rate_states=rate_states,
        rate_matrix=rate_matrix,
        rate_matrix_serialized=rate_matrix_serialized,
        specs=specs,
        planner_specs=planner_specs,
        system_specs=system_specs,
        recent_specs=recent_specs,
        lookups=lookups_data,
        plants_data=plants_data,
        strategic_customers_raw=strategic_customers_raw,
        strategic_customers=strategic_customers,
        trailer_assignment_rules=trailer_assignment_rules,
        rate_table_contexts=rate_table_contexts,
        rate_table_key_options=RATE_TABLE_KEY_OPTIONS,
        ryder_dedicated_rate_table=ryder_dedicated_rate_table,
        default_rate_change_metadata=default_rate_change_metadata,
        default_rate_changed_index=default_rate_change_metadata.get("changed_index") or {},
        default_rate_changed_count=default_rate_change_metadata.get("changed_count") or 0,
        lst_rate_matrix=lst_rate_matrix,
        alternate_trailer_rates=alternate_trailer_rates,
        rates_v2_payload=rates_v2_payload,
        optimizer_defaults=optimizer_defaults,
        optimizer_plant_defaults=optimizer_plant_defaults,
        util_grade_thresholds=util_grade_thresholds,
        sku_categories=sku_categories,
        sku_lengths=sku_lengths,
        sku_step_decks=sku_step_decks,
        sku_flat_beds=sku_flat_beds,
        source_led_specs=source_led_specs,
        source_led_mapped_count=source_led_mapped_count,
        source_led_unmapped_count=source_led_unmapped_count,
        source_led_missing_spec_count=source_led_missing_spec_count,
        source_led_unique_sku_count=source_led_unique_sku_count,
        fuel_surcharge_per_mile=fuel_surcharge_per_mile,
        stop_color_rows=stop_color_rows,
        optimizer_exception_category_options=optimizer_exception_category_options,
        trailer_assignment_rules_admin=trailer_assignment_rules_admin,
        trailer_assignment_rules_override_active=trailer_assignment_rules_override_active,
        current_profile_name=current_profile_name,
        is_admin=is_admin,
        can_edit_settings=can_edit_settings,
    )


@app.route("/planning-tools/save", methods=["POST"])
def save_planning_tools():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    current_role = _get_session_role()
    is_admin = current_role == ROLE_ADMIN
    can_edit_settings = current_role in {ROLE_ADMIN, ROLE_PLANNER}
    if not can_edit_settings:
        abort(403)

    if "strategic_customers" in request.form:
        strategic_customers_raw = request.form.get("strategic_customers") or ""
        parsed = customer_rules.parse_strategic_customers(strategic_customers_raw)
        serialized = customer_rules.serialize_strategic_customers(parsed)
        _upsert_scoped_planning_setting(STRATEGIC_CUSTOMERS_SETTING_KEY, serialized)

    has_trailer_rules_payload = request.form.get("trailer_rules_form") == "1"
    if has_trailer_rules_payload:
        livestock_wedge_enabled = _coerce_bool_value(
            request.form.get("livestock_wedge_enabled")
        )
        if is_admin:
            trailer_rules = _get_trailer_assignment_rules()
            trailer_rules["livestock_wedge_enabled"] = livestock_wedge_enabled
            trailer_rules.pop("auto_assign_hotshot_enabled", None)
            trailer_rules.pop("auto_assign_hotshot_utilization_threshold_pct", None)
            db.upsert_planning_setting(
                TRAILER_ASSIGNMENT_RULES_SETTING_KEY,
                json.dumps(trailer_rules),
            )
        else:
            profile_name = _get_session_profile_name() or _get_session_role()
            if not profile_name:
                abort(403)
            base_rules = _get_trailer_assignment_rules()
            override_payload = {}
            if livestock_wedge_enabled != _coerce_bool_value(base_rules.get("livestock_wedge_enabled")):
                override_payload["livestock_wedge_enabled"] = livestock_wedge_enabled
            db.upsert_planning_setting(
                _planner_trailer_rules_override_setting_key(profile_name),
                json.dumps(override_payload, separators=(",", ":")),
            )
    target_tab = (request.form.get("tab") or "overview").strip().lower()
    if target_tab != "overview":
        target_tab = "overview"
    return redirect(url_for("settings", tab=target_tab))


@app.route("/settings/stop-colors/save", methods=["POST"])
def save_stop_colors():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    palette = []
    for idx, default_color in enumerate(DEFAULT_STOP_COLOR_PALETTE, start=1):
        raw = request.form.get(f"stop_color_{idx}")
        palette.append(_normalize_hex_color(raw, default_color))
    _upsert_scoped_planning_setting(STOP_COLOR_PALETTE_SETTING_KEY, json.dumps(palette))

    target_tab = (request.form.get("tab") or "overview").strip().lower()
    if target_tab != "overview":
        target_tab = "overview"
    return redirect(url_for("settings", tab=target_tab))


@app.route("/rates/fuel-surcharge/save", methods=["POST"])
def save_fuel_surcharge():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or request.form
    value = round(
        _coerce_non_negative_float(
            payload.get("fuel_surcharge_per_mile"),
            DEFAULT_FUEL_SURCHARGE_PER_MILE,
        ),
        2,
    )
    _upsert_scoped_planning_setting(FUEL_SURCHARGE_SETTING_KEY, f"{value:.2f}")

    if request.is_json:
        return jsonify({"fuel_surcharge_per_mile": value})
    target_tab = (payload.get("tab") or request.form.get("tab") or "rates").strip().lower()
    if target_tab not in {"overview", "rates"}:
        target_tab = "rates"
    return redirect(url_for("settings", tab=target_tab))


@app.route("/settings/rates/contexts/save", methods=["POST"])
def save_rate_table_contexts():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or request.form
    contexts = _get_rate_table_contexts()
    contexts["default_rate_table_key"] = _normalize_rate_table_key(
        payload.get("default_rate_table_key"),
        contexts.get("default_rate_table_key", "FLS"),
    )
    contexts["carrier_dedicated_ryder_rate_table_key"] = _normalize_rate_table_key(
        payload.get("carrier_dedicated_ryder_rate_table_key"),
        contexts.get("carrier_dedicated_ryder_rate_table_key", "LST"),
    )
    contexts["trailer_hotshot_rate_table_key"] = _normalize_rate_table_key(
        payload.get("trailer_hotshot_rate_table_key"),
        contexts.get("trailer_hotshot_rate_table_key", "ALTERNATE_TRAILERS"),
    )
    _upsert_scoped_planning_setting(RATE_TABLE_CONTEXTS_SETTING_KEY, json.dumps(contexts))

    if request.is_json:
        return jsonify({"rate_table_contexts": contexts})
    target_tab = (payload.get("tab") or request.form.get("tab") or "rates").strip().lower()
    if target_tab not in {"overview", "rates"}:
        target_tab = "rates"
    return redirect(url_for("settings", tab=target_tab))


@app.route("/settings/rates/lst/save", methods=["POST"])
def save_lst_rate_settings():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or request.form
    data = _get_lst_rate_matrix()
    plants = _context_plants()
    data["plants"] = plants
    data.setdefault("states", [])
    data.setdefault("state_lane_labels", {})
    data.setdefault("matrix", {})

    destination_state = str(payload.get("destination_state") or "").strip().upper()
    origin_plant = str(payload.get("origin_plant") or payload.get("plant") or "").strip().upper()
    if destination_state and origin_plant in plants and "rate_per_mile" in payload:
        matrix_row = data["matrix"].setdefault(destination_state, {})
        matrix_row[origin_plant] = _serialize_optional_money(payload.get("rate_per_mile"))
        if destination_state not in data["states"]:
            data["states"].append(destination_state)
            data["states"] = sorted({str(state).strip().upper() for state in data["states"] if str(state).strip()})

    for key in ("fuel_surcharge", "per_stop", "load_minimum"):
        if key in payload:
            data[key] = _serialize_optional_money(payload.get(key))

    _upsert_scoped_planning_setting(LST_RATE_TABLE_SETTING_KEY, json.dumps(data))
    return jsonify(
        {
            "lst_rate_matrix": data,
            "saved_rate": data.get("matrix", {}).get(destination_state, {}).get(origin_plant),
            "destination_state": destination_state,
            "origin_plant": origin_plant,
        }
    )


@app.route("/settings/rates/ryder/save", methods=["POST"])
def save_ryder_dedicated_rate_settings():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or request.form
    data = _get_ryder_dedicated_rate_table()
    plants = _context_plants()
    data["plants"] = plants
    data.setdefault("rates_by_plant", {plant: None for plant in plants})

    origin_plant = str(payload.get("origin_plant") or payload.get("plant") or "").strip().upper()
    if origin_plant in plants and "rate_per_mile" in payload:
        data["rates_by_plant"][origin_plant] = _serialize_optional_money(payload.get("rate_per_mile"))

    for key in ("fuel_surcharge", "per_stop", "load_minimum"):
        if key in payload:
            data[key] = _serialize_optional_money(payload.get(key))

    _upsert_scoped_planning_setting(RYDER_DEDICATED_RATE_TABLE_SETTING_KEY, json.dumps(data))
    return jsonify(
        {
            "ryder_dedicated_rate_table": data,
            "saved_rate": data.get("rates_by_plant", {}).get(origin_plant),
            "origin_plant": origin_plant,
        }
    )


@app.route("/settings/rates/alternate/save", methods=["POST"])
def save_alternate_trailer_rate_settings():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or request.form
    data = _get_alternate_trailer_rates()
    trailer_code = _normalize_alternate_trailer_code(
        payload.get("trailer_type_code") or payload.get("trailer_type") or payload.get("code")
    )
    origin_plant = str(payload.get("origin_plant") or payload.get("plant") or "").strip().upper()
    raw_field = str(payload.get("field") or "").strip().lower()
    field = "apply_fuel_surcharge" if raw_field == "fuel_surcharge_per_mile" else raw_field
    plants = _context_plants()
    saved_value = None

    if trailer_code and origin_plant in plants and field in {"rate_per_mile", "apply_fuel_surcharge", "per_stop", "load_minimum", "requires_return_miles"}:
        for section in data.get("sections") or []:
            if section.get("code") != trailer_code:
                continue
            section.setdefault("rates_by_plant", {})
            section.setdefault("placeholders_by_plant", {})
            section["rates_by_plant"].setdefault(origin_plant, None)
            section["placeholders_by_plant"].setdefault(
                origin_plant,
                _default_alternate_trailer_placeholder(),
            )
            if field == "requires_return_miles":
                saved_value = bool(_coerce_bool_value(payload.get("value")))
            elif field == "apply_fuel_surcharge":
                if raw_field == "fuel_surcharge_per_mile":
                    saved_value = _coerce_alternate_apply_fuel_surcharge(
                        {"fuel_surcharge_per_mile": payload.get("value")},
                        default=True,
                    )
                else:
                    saved_value = _coerce_alternate_apply_fuel_surcharge(
                        {"apply_fuel_surcharge": payload.get("value")},
                        default=True,
                    )
            else:
                saved_value = _serialize_optional_money(payload.get("value"))
            if field == "rate_per_mile":
                section["rates_by_plant"][origin_plant] = saved_value
            else:
                section["placeholders_by_plant"][origin_plant][field] = saved_value
            break

    _upsert_scoped_planning_setting(ALTERNATE_TRAILER_RATES_SETTING_KEY, json.dumps(data))
    return jsonify(
        {
            "alternate_trailer_rates": data,
            "trailer_type_code": trailer_code,
            "origin_plant": origin_plant,
            "field": field,
            "saved_value": saved_value,
        }
    )


@app.route("/settings/rates/batch-save", methods=["POST"])
def save_rates_batch():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or {}
    pending = payload.get("pending") if isinstance(payload.get("pending"), dict) else {}
    plants = _context_plants()
    saved_counts = {
        "fls_lanes": 0,
        "ryder_lanes": 0,
        "lst_lanes": 0,
        "alternate_cells": 0,
    }

    fls_accessorial = pending.get("fls_accessorial") if isinstance(pending.get("fls_accessorial"), dict) else {}
    if fls_accessorial:
        current = _get_rates_overview_metrics()
        stop_fee = round(
            _coerce_non_negative_float(
                fls_accessorial.get("per_stop"),
                current.get("stop_fee", DEFAULT_STOP_FEE),
            ),
            2,
        )
        load_minimum = round(
            _coerce_non_negative_float(
                fls_accessorial.get("load_minimum"),
                current.get("load_minimum", DEFAULT_MIN_LOAD_COST),
            ),
            2,
        )
        fuel_surcharge = round(
            _coerce_non_negative_float(
                fls_accessorial.get("fuel_surcharge"),
                current.get("fuel_surcharge", DEFAULT_FUEL_SURCHARGE_PER_MILE),
            ),
            2,
        )
        _upsert_scoped_planning_setting(STOP_FEE_SETTING_KEY, f"{stop_fee:.2f}")
        _upsert_scoped_planning_setting(MIN_LOAD_COST_SETTING_KEY, f"{load_minimum:.2f}")
        _upsert_scoped_planning_setting(FUEL_SURCHARGE_SETTING_KEY, f"{fuel_surcharge:.2f}")

    fls_lane_updates = pending.get("fls_lanes") if isinstance(pending.get("fls_lanes"), list) else []
    for update in fls_lane_updates:
        if not isinstance(update, dict):
            continue
        origin_plant = str(update.get("origin_plant") or "").strip().upper()
        destination_state = str(update.get("destination_state") or "").strip().upper()
        parsed_rate = _coerce_optional_non_negative_float(update.get("rate_per_mile"))
        if origin_plant not in plants or not destination_state or parsed_rate is None:
            continue
        effective_year = int(update.get("effective_year") or datetime.now().year)
        rate_payload = {
            "origin_plant": origin_plant,
            "destination_state": destination_state,
            "rate_per_mile": round(parsed_rate, 2),
            "effective_year": effective_year,
            "notes": str(update.get("notes") or "").strip(),
        }
        rate_id = update.get("rate_id")
        if rate_id:
            try:
                db.update_rate(int(rate_id), rate_payload)
            except (TypeError, ValueError):
                db.upsert_rate(rate_payload)
        else:
            db.upsert_rate(rate_payload)
        saved_counts["fls_lanes"] += 1

    ryder_table = _get_ryder_dedicated_rate_table()
    ryder_changed = False
    ryder_lane_updates = pending.get("ryder_lanes") if isinstance(pending.get("ryder_lanes"), list) else []
    for update in ryder_lane_updates:
        if not isinstance(update, dict):
            continue
        origin_plant = str(update.get("origin_plant") or "").strip().upper()
        if origin_plant not in plants:
            continue
        ryder_table.setdefault("rates_by_plant", {})
        ryder_table["rates_by_plant"][origin_plant] = _serialize_optional_money(update.get("rate_per_mile"))
        ryder_changed = True
        saved_counts["ryder_lanes"] += 1
    ryder_accessorial = pending.get("ryder_accessorial") if isinstance(pending.get("ryder_accessorial"), dict) else {}
    for key in ("per_stop", "load_minimum", "fuel_surcharge"):
        if key not in ryder_accessorial:
            continue
        parsed = _serialize_optional_money(ryder_accessorial.get(key))
        ryder_table[key] = parsed if parsed is not None else 0.0
        ryder_changed = True
    if ryder_changed:
        _upsert_scoped_planning_setting(RYDER_DEDICATED_RATE_TABLE_SETTING_KEY, json.dumps(ryder_table))

    lst_table = _get_lst_rate_matrix()
    lst_changed = False
    lst_lane_updates = pending.get("lst_lanes") if isinstance(pending.get("lst_lanes"), list) else []
    for update in lst_lane_updates:
        if not isinstance(update, dict):
            continue
        origin_plant = str(update.get("origin_plant") or "").strip().upper()
        destination_state = str(update.get("destination_state") or "").strip().upper()
        if origin_plant not in plants or not destination_state:
            continue
        lst_table.setdefault("matrix", {})
        lst_table.setdefault("states", [])
        row = lst_table["matrix"].setdefault(destination_state, {})
        row[origin_plant] = _serialize_optional_money(update.get("rate_per_mile"))
        if destination_state not in lst_table["states"]:
            lst_table["states"].append(destination_state)
        lst_changed = True
        saved_counts["lst_lanes"] += 1
    if lst_changed:
        lst_table["states"] = sorted(
            {str(state).strip().upper() for state in lst_table.get("states") or [] if str(state or "").strip()}
        )
    lst_accessorial = pending.get("lst_accessorial") if isinstance(pending.get("lst_accessorial"), dict) else {}
    for key in ("per_stop", "load_minimum", "fuel_surcharge"):
        if key not in lst_accessorial:
            continue
        lst_table[key] = _serialize_optional_money(lst_accessorial.get(key))
        lst_changed = True
    if lst_changed:
        _upsert_scoped_planning_setting(LST_RATE_TABLE_SETTING_KEY, json.dumps(lst_table))

    alternate_table = _get_alternate_trailer_rates()
    alternate_changed = False
    alternate_updates = pending.get("alternate_cells") if isinstance(pending.get("alternate_cells"), list) else []
    sections = alternate_table.get("sections") if isinstance(alternate_table.get("sections"), list) else []
    section_index = {
        _normalize_alternate_trailer_code(section.get("code")): section
        for section in sections
        if isinstance(section, dict)
    }
    for update in alternate_updates:
        if not isinstance(update, dict):
            continue
        trailer_code = _normalize_alternate_trailer_code(update.get("trailer_type_code") or update.get("code"))
        origin_plant = str(update.get("origin_plant") or "").strip().upper()
        raw_field = str(update.get("field") or "").strip().lower()
        field = "apply_fuel_surcharge" if raw_field == "fuel_surcharge_per_mile" else raw_field
        if trailer_code not in section_index or origin_plant not in plants:
            continue
        if field not in {"rate_per_mile", "per_stop", "apply_fuel_surcharge", "load_minimum", "requires_return_miles"}:
            continue
        section = section_index[trailer_code]
        section.setdefault("rates_by_plant", {})
        section.setdefault("placeholders_by_plant", {})
        section["placeholders_by_plant"].setdefault(
            origin_plant,
            _default_alternate_trailer_placeholder(),
        )
        if field == "requires_return_miles":
            saved_value = bool(_coerce_bool_value(update.get("value")))
        elif field == "apply_fuel_surcharge":
            if raw_field == "fuel_surcharge_per_mile":
                saved_value = _coerce_alternate_apply_fuel_surcharge(
                    {"fuel_surcharge_per_mile": update.get("value")},
                    default=True,
                )
            else:
                saved_value = _coerce_alternate_apply_fuel_surcharge(
                    {"apply_fuel_surcharge": update.get("value")},
                    default=True,
                )
        else:
            saved_value = _serialize_optional_money(update.get("value"))
        if field == "rate_per_mile":
            section["rates_by_plant"][origin_plant] = saved_value
        else:
            section["placeholders_by_plant"][origin_plant][field] = saved_value
        alternate_changed = True
        saved_counts["alternate_cells"] += 1
    if alternate_changed:
        _upsert_scoped_planning_setting(ALTERNATE_TRAILER_RATES_SETTING_KEY, json.dumps(alternate_table))

    return jsonify({"ok": True, "saved_counts": saved_counts})


@app.route("/settings/global-metrics/save", methods=["POST"])
def save_global_metrics():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or request.form
    stop_fee = round(
        _coerce_non_negative_float(
            payload.get("stop_fee"),
            DEFAULT_STOP_FEE,
        ),
        2,
    )
    load_minimum = round(
        _coerce_non_negative_float(
            payload.get("load_minimum"),
            DEFAULT_MIN_LOAD_COST,
        ),
        2,
    )
    fuel_surcharge = round(
        _coerce_non_negative_float(
            payload.get("fuel_surcharge"),
            DEFAULT_FUEL_SURCHARGE_PER_MILE,
        ),
        2,
    )

    _upsert_scoped_planning_setting(STOP_FEE_SETTING_KEY, f"{stop_fee:.2f}")
    _upsert_scoped_planning_setting(MIN_LOAD_COST_SETTING_KEY, f"{load_minimum:.2f}")
    _upsert_scoped_planning_setting(FUEL_SURCHARGE_SETTING_KEY, f"{fuel_surcharge:.2f}")

    if request.is_json:
        return jsonify(
            {
                "stop_fee": stop_fee,
                "load_minimum": load_minimum,
                "fuel_surcharge": fuel_surcharge,
            }
        )

    target_tab = (payload.get("tab") or request.form.get("tab") or "overview").strip().lower()
    if target_tab not in {"overview", "rates"}:
        target_tab = "overview"
    return redirect(url_for("settings", tab=target_tab))


@app.route("/settings/optimizer-defaults/save", methods=["POST"])
def save_optimizer_defaults():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    is_admin = _get_session_role() == ROLE_ADMIN

    payload = request.get_json(silent=True) or request.form
    current = _get_optimizer_default_settings()
    raw_exception_categories = payload.get("upper_deck_exception_categories")
    if not request.is_json and hasattr(request.form, "getlist"):
        form_categories = request.form.getlist("upper_deck_exception_categories")
        if form_categories:
            raw_exception_categories = form_categories
    optimized = {
        # Trailer type/capacity are controlled per plant in optimizer_settings.
        # Keep these global keys stable as legacy fallbacks only.
        "trailer_type": stack_calculator.normalize_trailer_type(
            current.get("trailer_type"),
            default="STEP_DECK",
        ),
        "capacity_feet": _capacity_for_trailer_setting(
            current.get("trailer_type"),
            requested_capacity=current.get("capacity_feet"),
            default_capacity=current.get("capacity_feet", 53.0),
        ),
        "max_detour_pct": round(
            _coerce_non_negative_float(payload.get("max_detour_pct"), current["max_detour_pct"]),
            2,
        ),
        "time_window_days": _coerce_non_negative_int(
            payload.get("time_window_days"),
            current["time_window_days"],
        ),
        "geo_radius": round(
            _coerce_non_negative_float(payload.get("geo_radius"), current["geo_radius"]),
            2,
        ),
        "stack_overflow_max_height": _coerce_non_negative_int(
            payload.get("stack_overflow_max_height"),
            current.get("stack_overflow_max_height", DEFAULT_STACK_OVERFLOW_MAX_HEIGHT),
        ),
        "max_back_overhang_ft": round(
            _coerce_non_negative_float(
                payload.get("max_back_overhang_ft"),
                current.get("max_back_overhang_ft", DEFAULT_MAX_BACK_OVERHANG_FT),
            ),
            2,
        ),
        "upper_two_across_max_length_ft": round(
            _coerce_non_negative_float(
                payload.get("upper_two_across_max_length_ft"),
                current.get(
                    "upper_two_across_max_length_ft",
                    DEFAULT_UPPER_TWO_ACROSS_MAX_LENGTH_FT,
                ),
            ),
            2,
        ),
        "upper_deck_exception_max_length_ft": round(
            _coerce_non_negative_float(
                payload.get("upper_deck_exception_max_length_ft"),
                current.get(
                    "upper_deck_exception_max_length_ft",
                    DEFAULT_UPPER_DECK_EXCEPTION_MAX_LENGTH_FT,
                ),
            ),
            2,
        ),
        "upper_deck_exception_overhang_allowance_ft": round(
            _coerce_non_negative_float(
                payload.get("upper_deck_exception_overhang_allowance_ft"),
                current.get(
                    "upper_deck_exception_overhang_allowance_ft",
                    DEFAULT_UPPER_DECK_EXCEPTION_OVERHANG_ALLOWANCE_FT,
                ),
            ),
            2,
        ),
        "upper_deck_exception_categories": stack_calculator.normalize_upper_deck_exception_categories(
            raw_exception_categories,
            default=current.get(
                "upper_deck_exception_categories",
                DEFAULT_UPPER_DECK_EXCEPTION_CATEGORIES,
            ),
        ),
        "equal_length_deck_length_order_enabled": (
            _coerce_bool_value(payload.get("equal_length_deck_length_order_enabled"))
            if payload.get("equal_length_deck_length_order_enabled") is not None
            else bool(
                current.get(
                    "equal_length_deck_length_order_enabled",
                    DEFAULT_EQUAL_LENGTH_DECK_LENGTH_ORDER_ENABLED,
                )
            )
        ),
    }

    _upsert_scoped_planning_setting(OPTIMIZER_DEFAULTS_SETTING_KEY, json.dumps(optimized))
    if is_admin:
        trailer_rules = _get_trailer_assignment_rules()

        def _save_plant_optimizer_default(plant_code, trailer_value, hotshot_value, hotshot_present):
            normalized_plant = _normalize_plant_code(plant_code)
            if not normalized_plant:
                return
            current_row = db.get_optimizer_settings(normalized_plant) or {}
            current_defaults = _get_optimizer_settings_for_plant(
                normalized_plant,
                optimizer_defaults=optimized,
                trailer_rules=trailer_rules,
                settings_row=current_row,
            )
            trailer_key = stack_calculator.normalize_trailer_type(
                trailer_value,
                default=current_defaults["trailer_type"],
            )
            capacity_feet = _capacity_for_trailer_setting(
                trailer_key,
                requested_capacity=current_row.get("capacity_feet", optimized.get("capacity_feet")),
                default_capacity=optimized.get("capacity_feet", 53.0),
            )
            auto_hotshot_enabled = (
                _coerce_bool_value(hotshot_value)
                if hotshot_present
                else bool(current_defaults["auto_hotshot_enabled"])
            )
            db.upsert_optimizer_settings(
                {
                    "origin_plant": normalized_plant,
                    "capacity_feet": capacity_feet,
                    "trailer_type": trailer_key,
                    "max_detour_pct": _coerce_non_negative_float(
                        current_row.get("max_detour_pct"),
                        optimized.get("max_detour_pct", 15.0),
                    ),
                    "time_window_days": _coerce_non_negative_int(
                        current_row.get("time_window_days"),
                        optimized.get("time_window_days", 7),
                    ),
                    "geo_radius": _coerce_non_negative_float(
                        current_row.get("geo_radius"),
                        optimized.get("geo_radius", 100.0),
                    ),
                    "auto_hotshot_enabled": 1 if auto_hotshot_enabled else 0,
                }
            )

        if request.is_json:
            plant_defaults_payload = payload.get("plant_defaults")
            if isinstance(plant_defaults_payload, dict):
                for plant_code, plant_payload in plant_defaults_payload.items():
                    if not isinstance(plant_payload, dict):
                        continue
                    _save_plant_optimizer_default(
                        plant_code,
                        plant_payload.get("trailer_type"),
                        plant_payload.get("auto_hotshot_enabled"),
                        "auto_hotshot_enabled" in plant_payload,
                    )
        else:
            for plant_code in PLANT_CODES:
                if request.form.get(f"plant_defaults_present_{plant_code}") != "1":
                    continue
                _save_plant_optimizer_default(
                    plant_code,
                    request.form.get(f"plant_default_trailer_{plant_code}"),
                    request.form.get(f"plant_auto_hotshot_enabled_{plant_code}"),
                    True,
                )
    stack_calculator.invalidate_stack_assumptions_cache()

    if request.is_json:
        return jsonify({"optimizer_defaults": optimized})

    target_tab = (payload.get("tab") or request.form.get("tab") or "overview").strip().lower()
    if target_tab != "overview":
        target_tab = "overview"
    return redirect(url_for("settings", tab=target_tab))


@app.route("/settings/utilization-grades/save", methods=["POST"])
def save_utilization_grades():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()

    payload = request.get_json(silent=True) or request.form
    current = _get_utilization_grade_thresholds()
    raw_d_min = payload.get("grade_d_min")
    raw_f_max = payload.get("grade_f_max")
    d_value = raw_d_min if raw_d_min not in (None, "") else current["D"]
    if raw_d_min in (None, "") and raw_f_max not in (None, ""):
        d_value = _coerce_non_negative_int(raw_f_max, max(current["D"] - 1, 0)) + 1
    thresholds = _coerce_utilization_grade_thresholds(
        {
            "A": payload.get("grade_a_min", current["A"]),
            "B": payload.get("grade_b_min", current["B"]),
            "C": payload.get("grade_c_min", current["C"]),
            "D": d_value,
        }
    )
    _upsert_scoped_planning_setting(UTILIZATION_GRADE_THRESHOLDS_SETTING_KEY, json.dumps(thresholds))
    stack_calculator.invalidate_utilization_grade_thresholds_cache()

    if request.is_json:
        return jsonify({"utilization_grade_thresholds": thresholds})

    target_tab = (payload.get("tab") or request.form.get("tab") or "overview").strip().lower()
    if target_tab != "overview":
        target_tab = "overview"
    return redirect(url_for("settings", tab=target_tab))


@app.route("/rates/save", methods=["POST"])
def save_rate():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    payload = request.get_json(silent=True) or request.form

    rate_id = payload.get("id") or payload.get("rate_id")
    origin_plant = (payload.get("origin_plant") or "").strip().upper()
    destination_state = (payload.get("destination_state") or "").strip().upper()
    rate_per_mile = float(payload.get("rate_per_mile") or 0)
    effective_year = int(payload.get("effective_year") or 2026)
    notes = (payload.get("notes") or "").strip()

    rate_payload = {
        "origin_plant": origin_plant,
        "destination_state": destination_state,
        "rate_per_mile": rate_per_mile,
        "effective_year": effective_year,
        "notes": notes,
    }

    if rate_id:
        db.update_rate(int(rate_id), rate_payload)
        saved_rate = db.get_rate_by_id(int(rate_id))
    else:
        db.upsert_rate(rate_payload)
        saved_rate = db.get_rate_by_lane(origin_plant, destination_state, effective_year)

    if request.is_json:
        return jsonify({"rate": saved_rate})

    return redirect(request.referrer or url_for("settings", tab="rates"))


@app.route("/rates/add", methods=["POST"])
def add_rate():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    rate = {
        "origin_plant": request.form.get("origin_plant", "").strip().upper(),
        "destination_state": request.form.get("destination_state", "").strip().upper(),
        "rate_per_mile": float(request.form.get("rate_per_mile", 0) or 0),
        "effective_year": int(request.form.get("effective_year", 2026) or 2026),
        "notes": request.form.get("notes", "").strip(),
    }
    db.upsert_rate(rate)
    return redirect(request.referrer or url_for("rates"))


@app.route("/rates/delete/<int:rate_id>", methods=["POST"])
def delete_rate(rate_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    db.delete_rate(rate_id)
    return redirect(url_for("rates"))


@app.route("/skus/save", methods=["POST"])
def save_sku():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    payload = request.get_json(silent=True) or {}
    spec_id = payload.get("id")
    if not spec_id:
        return jsonify({"error": "Missing SKU id"}), 400

    spec = {
        "sku": (payload.get("sku") or "").strip(),
        "description": (payload.get("description") or "").strip(),
        "category": (payload.get("category") or "").strip(),
        "length_with_tongue_ft": float(payload.get("length_with_tongue_ft") or 0),
        "max_stack_step_deck": int(payload.get("max_stack_step_deck") or 1),
        "max_stack_flat_bed": int(payload.get("max_stack_flat_bed") or 1),
        "notes": (payload.get("notes") or "").strip(),
    }
    db.update_sku_spec(int(spec_id), spec)
    return jsonify({"status": "ok"})


@app.route("/skus/source-led/save", methods=["POST"])
def save_source_led_sku():
    session_redirect = _require_session()
    if session_redirect:
        return _json_session_expired_response()
    _require_settings_editor()

    payload = request.get_json(silent=True) or {}
    sku = str(payload.get("sku") or "").strip().upper()
    if not sku:
        return jsonify({"error": "Mapped SKU is required."}), 400

    current = db.get_sku_spec_by_sku(sku)
    if not current:
        return jsonify({"error": "Mapped SKU does not exist in canonical specs."}), 404

    length_with_tongue_ft = _coerce_optional_non_negative_float(payload.get("length_with_tongue_ft"))
    max_stack_step_deck = _coerce_int_value(payload.get("max_stack_step_deck"), current.get("max_stack_step_deck") or 1)
    max_stack_flat_bed = _coerce_int_value(payload.get("max_stack_flat_bed"), current.get("max_stack_flat_bed") or 1)
    if length_with_tongue_ft is None:
        return jsonify({"error": "Length must be a valid non-negative number."}), 400
    if max_stack_step_deck <= 0 or max_stack_flat_bed <= 0:
        return jsonify({"error": "Stack limits must be whole numbers greater than zero."}), 400

    updated_by = _get_session_profile_name() or _get_session_role()
    db.update_sku_spec_by_sku(
        sku,
        {
            "description": current.get("description") or "",
            "category": current.get("category") or "",
            "length_with_tongue_ft": round(float(length_with_tongue_ft), 2),
            "max_stack_step_deck": int(max_stack_step_deck),
            "max_stack_flat_bed": int(max_stack_flat_bed),
            "notes": current.get("notes") or "",
        },
        updated_by=updated_by,
    )
    refreshed = db.get_sku_spec_by_sku(sku) or {}
    source_text = _source_led_source_text(refreshed)

    return jsonify(
        {
            "ok": True,
            "sku": sku,
            "length_with_tongue_ft": float(refreshed.get("length_with_tongue_ft") or 0.0),
            "max_stack_step_deck": int(refreshed.get("max_stack_step_deck") or 1),
            "max_stack_flat_bed": int(refreshed.get("max_stack_flat_bed") or 1),
            "mapped_source": source_text,
        }
    )


@app.route("/lookups/save", methods=["POST"])
def save_lookup():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    payload = request.get_json(silent=True) or {}
    entry_id = payload.get("id")
    if not entry_id:
        return jsonify({"error": "Missing lookup id"}), 400

    entry = {
        "plant": (payload.get("plant") or "").strip().upper(),
        "bin": (payload.get("bin") or "").strip().upper(),
        "item_pattern": (payload.get("item_pattern") or "").strip(),
        "sku": (payload.get("sku") or "").strip(),
    }
    db.update_item_lookup(int(entry_id), entry)
    return jsonify({"status": "ok"})


@app.route("/plants/save", methods=["POST"])
def save_plant():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    payload = request.get_json(silent=True) or {}
    plant_id = payload.get("id")
    if not plant_id:
        return jsonify({"error": "Missing plant id"}), 400

    plant = {
        "plant_code": (payload.get("plant_code") or "").strip().upper(),
        "name": (payload.get("name") or "").strip(),
        "lat": float(payload.get("lat") or 0),
        "lng": float(payload.get("lng") or 0),
        "address": (payload.get("address") or "").strip(),
    }
    db.update_plant(int(plant_id), plant)
    geo_utils.invalidate_coordinate_caches(plant_coords=True)
    return jsonify({"status": "ok"})


@app.route("/skus")
def skus():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    if _get_session_role() != ROLE_ADMIN:
        return redirect(url_for("settings", tab="skus"))
    specs = db.list_sku_specs()
    return render_template("skus.html", specs=specs)


@app.route("/skus/export-cheat-sheet.xlsx")
def export_sku_cheat_sheet():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    specs = db.list_sku_specs()
    ordered_specs = sorted(
        specs,
        key=lambda spec: (
            0 if _sku_is_planner_input(spec) else 1,
            str(spec.get("sku") or "").upper(),
        ),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SKU Cheat Sheet"

    headers = [
        "Source",
        "SKU",
        "Description",
        "Category",
        "Length w/ Tongue (ft)",
        "Max Stack Step Deck",
        "Max Stack Flat Bed",
        "Notes",
        "Added",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, _ in enumerate(headers, start=1):
        header_cell = sheet.cell(row=1, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment

    for spec in ordered_specs:
        sheet.append(
            [
                _sku_source_label(spec),
                spec.get("sku") or "",
                spec.get("description") or "",
                spec.get("category") or "",
                spec.get("length_with_tongue_ft"),
                spec.get("max_stack_step_deck"),
                spec.get("max_stack_flat_bed"),
                spec.get("notes") or "",
                spec.get("added_at") or spec.get("created_at") or "",
            ]
        )

    widths = {
        "A": 16,
        "B": 18,
        "C": 38,
        "D": 18,
        "E": 22,
        "F": 20,
        "G": 19,
        "H": 30,
        "I": 21,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(sheet.max_row, 1)}"

    data_alignment_left = Alignment(horizontal="left", vertical="center")
    data_alignment_center = Alignment(horizontal="center", vertical="center")
    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in [1, 2, 3, 4, 8, 9]:
            sheet.cell(row=row_idx, column=col_idx).alignment = data_alignment_left
        for col_idx in [5, 6, 7]:
            sheet.cell(row=row_idx, column=col_idx).alignment = data_alignment_center

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"sku_cheat_sheet_{date.today().isoformat()}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/skus/add", methods=["POST"])
def add_sku():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    spec = {
        "sku": request.form.get("sku", "").strip(),
        "description": request.form.get("description", "").strip(),
        "category": request.form.get("category", "").strip(),
        "length_with_tongue_ft": float(request.form.get("length_with_tongue_ft", 0) or 0),
        "max_stack_step_deck": int(request.form.get("max_stack_step_deck", 1) or 1),
        "max_stack_flat_bed": int(request.form.get("max_stack_flat_bed", 1) or 1),
        "notes": request.form.get("notes", "").strip(),
    }
    db.upsert_sku_spec(spec)
    return redirect(request.referrer or url_for("settings", tab="skus"))


@app.route("/skus/delete/<int:spec_id>", methods=["POST"])
def delete_sku(spec_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    db.delete_sku_spec(spec_id)
    return redirect(request.referrer or url_for("settings", tab="skus"))


@app.route("/lookups")
def lookups():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    if _get_session_role() != ROLE_ADMIN:
        return redirect(url_for("settings", tab="lookups"))
    lookups_data = db.list_item_lookups()
    specs = db.list_sku_specs()
    return render_template("lookups.html", lookups=lookups_data, specs=specs)


@app.route("/lookups/add", methods=["POST"])
def add_lookup():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    entry = {
        "plant": request.form.get("plant", "").strip().upper(),
        "bin": request.form.get("bin", "").strip().upper(),
        "item_pattern": request.form.get("item_pattern", "").strip().upper(),
        "sku": request.form.get("sku", "").strip(),
    }
    db.add_item_lookup(entry)
    return redirect(url_for("lookups"))


@app.route("/lookups/delete/<int:entry_id>", methods=["POST"])
def delete_lookup(entry_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect
    _require_settings_editor()
    db.delete_item_lookup(entry_id)
    return redirect(url_for("lookups"))


@app.route("/api/orders/<so_num>/stack-config")
def order_stack_config(so_num):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    with db.get_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                ol.plant,
                ol.item,
                ol.item_desc,
                ol.sku,
                ol.qty,
                ol.unit_length_ft,
                ol.max_stack_height,
                COALESCE(ss.category, ol.bin, '') AS category
            FROM order_lines ol
            LEFT JOIN sku_specifications ss ON ol.sku = ss.sku
            WHERE ol.so_num = ?
            ORDER BY ol.id ASC
            """,
            (so_num,),
        ).fetchall()
        order_row = connection.execute(
            """
            SELECT
                orders.so_num,
                orders.plant,
                orders.cust_name,
                orders.due_date,
                orders.state,
                orders.zip,
                orders.total_qty,
                orders.total_length_ft,
                orders.utilization_pct,
                orders.line_count,
                orders.is_excluded,
                (
                    SELECT city
                    FROM order_lines
                    WHERE order_lines.so_num = orders.so_num
                      AND city IS NOT NULL
                      AND city != ''
                    LIMIT 1
                ) AS city
            FROM orders
            WHERE orders.so_num = ?
            LIMIT 1
            """,
            (so_num,),
        ).fetchone()

    rows = [dict(row) for row in rows]
    order_row = dict(order_row) if order_row else None

    if not rows:
        return jsonify({"error": "Order not found"}), 404

    allowed_plants = _get_allowed_plants()
    if rows[0]["plant"] not in allowed_plants:
        return jsonify({"error": "Not authorized for this plant."}), 403

    line_items = []
    for row in rows:
        max_stack = row["max_stack_height"] or 1
        qty = row["qty"] or 0
        positions_required = math.ceil(qty / max_stack) if max_stack else qty
        linear_feet = (row["unit_length_ft"] or 0) * positions_required
        line_items.append(
            {
                "item": row["item"],
                "item_desc": row.get("item_desc"),
                "sku": row["sku"],
                "qty": qty,
                "unit_length_ft": row["unit_length_ft"] or 0,
                "max_stack_height": max_stack,
                "positions_required": positions_required,
                "linear_feet": linear_feet,
                "category": row["category"] or "",
                "stop_sequence": 1,
            }
        )

    # Orders page schematic should show strict stack capacity without singleton overflow allowance.
    # Keep overflow logic enabled for optimization and Loads page workflows.
    config = stack_calculator.calculate_stack_configuration(
        line_items,
        stack_overflow_max_height=0,
    )
    config["order_id"] = so_num
    config["line_items"] = line_items
    config["positions_count"] = len(config["positions"])

    plant_code = (order_row["plant"] if order_row else None) or rows[0]["plant"]
    dest_zip = order_row["zip"] if order_row else None
    dest_state = order_row["state"] if order_row else None
    dest_city = order_row["city"] if order_row else None

    zip_coords = geo_utils.load_zip_coordinates()
    origin_coords = geo_utils.plant_coords_for_code(plant_code) if plant_code else None
    dest_coords = (
        zip_coords.get(geo_utils.normalize_zip(dest_zip)) if dest_zip else None
    )

    destination_label_parts = []
    if dest_city:
        destination_label_parts.append(str(dest_city))
    if dest_state:
        destination_label_parts.append(str(dest_state))
    if dest_zip:
        destination_label_parts.append(str(dest_zip))
    destination_label = " ".join(part for part in destination_label_parts if part).strip()

    map_stops = []
    map_stops.append(
        {
            "label": f"{plant_code} Plant" if plant_code else "Origin Plant",
            "lat": origin_coords[0] if origin_coords else None,
            "lng": origin_coords[1] if origin_coords else None,
            "type": "origin",
            "sequence": 1,
        }
    )
    map_stops.append(
        {
            "label": destination_label or "Destination",
            "lat": dest_coords[0] if dest_coords else None,
            "lng": dest_coords[1] if dest_coords else None,
            "type": "final",
            "sequence": 2,
        }
    )
    config["map_stops"] = map_stops

    if order_row:
        config["order_meta"] = {
            "so_num": order_row["so_num"],
            "plant": order_row["plant"],
            "cust_name": order_row["cust_name"],
            "due_date": order_row["due_date"],
            "city": order_row["city"],
            "state": order_row["state"],
            "zip": order_row["zip"],
            "total_qty": order_row["total_qty"],
            "total_length_ft": order_row["total_length_ft"],
            "utilization_pct": order_row["utilization_pct"],
            "line_count": order_row["line_count"],
            "is_excluded": order_row["is_excluded"],
        }

    return jsonify(config)


@app.route("/api/optimize", methods=["POST"])
def run_optimization():
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    data = request.get_json(silent=True) or {}
    plant_code = data.get("plant_code") or data.get("origin_plant")
    plant_code = _normalize_plant_code(plant_code)
    if plant_code and plant_code not in _get_allowed_plants():
        return jsonify({"error": "Plant not in scope"}), 403
    flexibility_days = data.get("flexibility_days", 7)
    proximity_miles = data.get("proximity_miles", data.get("geo_radius", 200))
    capacity_feet = data.get("capacity_feet", 53)
    trailer_type = data.get("trailer_type", "STEP_DECK")

    engine = OptimizerEngine()
    result = engine.run_optimization(
        plant_code,
        flexibility_days=flexibility_days,
        proximity_miles=proximity_miles,
        capacity_feet=capacity_feet,
        trailer_type=trailer_type,
    )
    return jsonify(result)


@app.route("/api/optimize/<int:run_id>/loads")
def get_optimization_loads(run_id):
    session_redirect = _require_session()
    if session_redirect:
        return session_redirect

    with db.get_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                l.id,
                l.load_number,
                l.plant_code,
                l.total_util,
                l.total_miles,
                l.total_cost,
                l.num_orders,
                l.status,
                GROUP_CONCAT(a.order_so_num) as order_nums
            FROM optimized_loads l
            LEFT JOIN load_order_assignments a ON l.id = a.load_id
            WHERE l.run_id = ?
            GROUP BY l.id
            ORDER BY l.load_number
            """,
            (run_id,),
        ).fetchall()

    if rows:
        allowed_plants = _get_allowed_plants()
        if rows[0]["plant_code"] not in allowed_plants:
            return jsonify({"error": "Not authorized for this plant."}), 403

    loads = []
    for row in rows:
        order_numbers = row["order_nums"].split(",") if row["order_nums"] else []
        loads.append(
            {
                "id": row["id"],
                "load_number": row["load_number"],
                "total_util": (row["total_util"] or 0) * 100,
                "total_miles": row["total_miles"],
                "total_cost": row["total_cost"],
                "num_orders": row["num_orders"],
                "status": row["status"],
                "order_numbers": order_numbers,
            }
        )

    return jsonify({"loads": loads})


if __name__ == "__main__":
    app.run(debug=True)


