import argparse
import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import db

DEFAULT_SOURCE_DIR = Path(
    r"c:\Users\gabramowitz\OneDrive - Council Advisors\Bain Capital - ATW - ATW Operations Value Creation\03 - Phase 2\04 - Carry On MFO\Phase 2 - COT Freight Optimization"
)
DEFAULT_FLS_FILE = DEFAULT_SOURCE_DIR / "FLS ATL COT Rate Matrix Schedule B effective 03-01-2026.xlsx"
DEFAULT_LST_FILE = DEFAULT_SOURCE_DIR / "LST Rates All Plants November 2025.xlsx"
DEFAULT_SSA_FILE = DEFAULT_SOURCE_DIR / "SSA Dedicated Wedge Hot Shot Pricing.ods"

RATE_TABLE_CONTEXTS_SETTING_KEY = "rate_table_contexts"
LST_RATE_TABLE_SETTING_KEY = "lst_rate_matrix"
ALTERNATE_TRAILER_RATES_SETTING_KEY = "alternate_trailer_rates"
DEFAULT_RATE_CHANGE_METADATA_SETTING_KEY = "default_rate_change_metadata"

_STATE_DEST_PATTERN = re.compile(r"^([A-Z]{2})-")
_NUMBER_PATTERN = re.compile(r"-?\d+(?:\.\d+)?")
_SPECIAL_ROWS = {"DETENTION", "MIN", "STOP", "TONU"}
_SKIP_ROWS = {"FUEL", "HS FUEL", "ATL", "COT"}
_ODS_NAMESPACES = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}


def _clean_text(raw_value):
    if raw_value is None:
        return ""
    return str(raw_value).strip()


def _parse_float(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    text = _clean_text(raw_value)
    if not text:
        return None
    compact = text.replace(",", "")
    match = _NUMBER_PATTERN.search(compact)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def _normalize_destination(raw_label):
    label = _clean_text(raw_label).upper()
    if not label:
        return None
    if label in _SKIP_ROWS:
        return None
    if label in _SPECIAL_ROWS:
        return label
    if label.startswith("NY ZIP 100"):
        return "NY ZIP 100"
    match = _STATE_DEST_PATTERN.match(label)
    if match:
        return match.group(1)
    return None


def _read_matrix_workbook(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    plants = []
    for col_idx in range(2, worksheet.max_column + 1):
        raw = _clean_text(worksheet.cell(4, col_idx).value).upper()
        if not raw:
            continue
        plants.append(raw)

    states_in_order = OrderedDict()
    state_lane_labels = {}
    state_values = {}
    special_values = {}

    for row_idx in range(6, worksheet.max_row + 1):
        lane_label = _clean_text(worksheet.cell(row_idx, 1).value).upper()
        if not lane_label:
            continue
        destination = _normalize_destination(lane_label)
        if not destination:
            continue
        if len(destination) == 2:
            states_in_order.setdefault(destination, True)
            state_lane_labels[destination] = lane_label

        for offset, plant in enumerate(plants):
            raw_value = worksheet.cell(row_idx, 2 + offset).value
            parsed = _parse_float(raw_value)
            if parsed is None:
                continue
            lane_key = (destination, plant)
            if len(destination) == 2:
                # Lane subzones exist in source files (CA-N/CA-S/etc). Keep the last seen row.
                state_values[lane_key] = parsed
            else:
                special_values[lane_key] = parsed

    states = list(states_in_order.keys())
    matrix = {}
    for state in states:
        matrix[state] = {}
        for plant in plants:
            matrix[state][plant] = state_values.get((state, plant))

    specials = {}
    for destination, plant in sorted(special_values):
        specials.setdefault(destination, {})
        specials[destination][plant] = special_values[(destination, plant)]

    return {
        "plants": plants,
        "states": states,
        "matrix": matrix,
        "state_lane_labels": state_lane_labels,
        "specials": specials,
    }


def _ods_cell_text(cell):
    chunks = []
    for paragraph in cell.findall(".//text:p", _ODS_NAMESPACES):
        text = "".join(paragraph.itertext()).strip()
        if text:
            chunks.append(text)
    return "\n".join(chunks).strip()


def _expand_ods_row(row):
    expanded = []
    for cell in row.findall("table:table-cell", _ODS_NAMESPACES):
        repeated = int(
            cell.get(
                "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}number-columns-repeated",
                "1",
            )
        )
        text = _ods_cell_text(cell)
        for _ in range(repeated):
            expanded.append(text)
    return expanded


def _infer_trailer_type(carrier, comments):
    text = f"{carrier} {comments}".upper()
    if "HOT SHOT" in text or "HOTSHOT" in text:
        return "Hot Shot"
    if "WEDGE" in text:
        return "Wedge"
    if "DEDICATED" in text or "FLAT BED" in text or "FLATBED" in text:
        return "Dedicated"
    return "Other"


def _parse_alternate_trailer_rates(path):
    with zipfile.ZipFile(path, "r") as archive:
        content = archive.read("content.xml")
    root = ET.fromstring(content)
    table = root.find(".//table:table", _ODS_NAMESPACES)
    if table is None:
        return {"sections": []}

    sections = OrderedDict()
    row_index = 0
    for row in table.findall("table:table-row", _ODS_NAMESPACES):
        row_index += 1
        values = _expand_ods_row(row)
        if row_index == 1:
            continue
        if not values:
            continue
        carrier = _clean_text(values[0] if len(values) > 0 else "")
        if not carrier:
            continue
        plant = _clean_text(values[3] if len(values) > 3 else "").upper()
        minimum_text = _clean_text(values[4] if len(values) > 4 else "")
        rate_text = _clean_text(values[5] if len(values) > 5 else "")
        round_trip = _clean_text(values[6] if len(values) > 6 else "")
        fuel_flag = _clean_text(values[7] if len(values) > 7 else "")
        stop_text = _clean_text(values[8] if len(values) > 8 else "")
        comments = _clean_text(values[9] if len(values) > 9 else "")
        phone = _clean_text(values[10] if len(values) > 10 else "")
        email = _clean_text(values[11] if len(values) > 11 else "")
        trailer_type = _infer_trailer_type(carrier, comments)
        sections.setdefault(trailer_type, [])
        sections[trailer_type].append(
            {
                "carrier": carrier,
                "plant": plant,
                "minimum": _parse_float(minimum_text),
                "minimum_text": minimum_text,
                "rate_per_mile": _parse_float(rate_text),
                "rate_per_mile_text": rate_text,
                "dedicated_round_trip": round_trip,
                "cot_fuel_surcharge": fuel_flag,
                "stop_charge": _parse_float(stop_text),
                "stop_charge_text": stop_text,
                "comments": comments,
                "phone": phone,
                "email": email,
            }
        )

    ordered_types = ["Dedicated", "Wedge", "Hot Shot", "Other"]
    payload_sections = []
    for trailer_type in ordered_types:
        rows = sections.get(trailer_type) or []
        if not rows:
            continue
        rows.sort(key=lambda row: ((row.get("plant") or "ZZZ"), (row.get("carrier") or "")))
        payload_sections.append({"trailer_type": trailer_type, "rows": rows})
    return {"sections": payload_sections}


def _build_existing_rate_lookup(effective_year):
    lookup = {}
    for row in db.list_rate_matrix():
        try:
            year = int(row.get("effective_year") or 0)
        except (TypeError, ValueError):
            continue
        if year != effective_year:
            continue
        origin = _clean_text(row.get("origin_plant")).upper()
        destination = _clean_text(row.get("destination_state")).upper()
        if not origin or not destination:
            continue
        lookup[(origin, destination)] = float(row.get("rate_per_mile") or 0.0)
    return lookup


def _apply_fls_rates(fls_payload, effective_year, fls_source):
    existing = _build_existing_rate_lookup(effective_year)
    changes = []
    source_note = "Imported from FLS ATL COT Rate Matrix Schedule B (effective 03-01-2026)"
    upsert_count = 0

    for state in fls_payload["states"]:
        for plant in fls_payload["plants"]:
            new_rate = fls_payload["matrix"].get(state, {}).get(plant)
            if new_rate is None:
                continue
            previous_rate = existing.get((plant, state))
            if previous_rate is None or abs(previous_rate - new_rate) > 0.0001:
                changes.append(
                    {
                        "origin_plant": plant,
                        "destination_state": state,
                        "previous_rate": previous_rate,
                        "new_rate": round(new_rate, 2),
                        "change_type": "new" if previous_rate is None else "updated",
                    }
                )
            db.upsert_rate(
                {
                    "origin_plant": plant,
                    "destination_state": state,
                    "rate_per_mile": round(new_rate, 2),
                    "effective_year": effective_year,
                    "notes": source_note,
                }
            )
            upsert_count += 1

    for destination, rows in (fls_payload.get("specials") or {}).items():
        for plant, rate in rows.items():
            if rate is None:
                continue
            db.upsert_rate(
                {
                    "origin_plant": plant,
                    "destination_state": destination,
                    "rate_per_mile": round(rate, 2),
                    "effective_year": effective_year,
                    "notes": source_note,
                }
            )
            upsert_count += 1

    change_payload = {
        "source": fls_source,
        "effective_date": "2026-03-01",
        "effective_year": effective_year,
        "imported_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "changed_cells": changes,
    }
    return upsert_count, change_payload


def _save_nondefault_tables(lst_payload, alternate_payload, lst_source, ssa_source):
    now_value = datetime.now(timezone.utc).isoformat(timespec="seconds")
    lst_setting = {
        "carrier": "LST",
        "source": lst_source,
        "imported_at": now_value,
        "plants": lst_payload.get("plants") or [],
        "states": lst_payload.get("states") or [],
        "state_lane_labels": lst_payload.get("state_lane_labels") or {},
        "matrix": lst_payload.get("matrix") or {},
    }
    alt_setting = {
        "source": ssa_source,
        "imported_at": now_value,
        "sections": alternate_payload.get("sections") or [],
    }
    db.upsert_planning_setting(LST_RATE_TABLE_SETTING_KEY, json.dumps(lst_setting))
    db.upsert_planning_setting(ALTERNATE_TRAILER_RATES_SETTING_KEY, json.dumps(alt_setting))
    db.upsert_planning_setting(
        RATE_TABLE_CONTEXTS_SETTING_KEY,
        json.dumps(
            {
                "default_rate_table_key": "FLS",
                "carrier_dedicated_ryder_rate_table_key": "LST",
                "trailer_hotshot_rate_table_key": "ALTERNATE_TRAILERS",
            }
        ),
    )


def main():
    parser = argparse.ArgumentParser(description="Import FLS/LST/SSA freight tables into app settings.")
    parser.add_argument("--fls", default=str(DEFAULT_FLS_FILE), help="Path to FLS matrix xlsx.")
    parser.add_argument("--lst", default=str(DEFAULT_LST_FILE), help="Path to LST matrix xlsx.")
    parser.add_argument("--ssa", default=str(DEFAULT_SSA_FILE), help="Path to SSA alternate trailer ods.")
    parser.add_argument("--effective-year", type=int, default=2026, help="Effective year for FLS default rates.")
    args = parser.parse_args()

    fls_path = Path(args.fls)
    lst_path = Path(args.lst)
    ssa_path = Path(args.ssa)
    for path in (fls_path, lst_path, ssa_path):
        if not path.exists():
            raise FileNotFoundError(f"Missing source file: {path}")

    fls_payload = _read_matrix_workbook(fls_path)
    lst_payload = _read_matrix_workbook(lst_path)
    alternate_payload = _parse_alternate_trailer_rates(ssa_path)

    upsert_count, change_payload = _apply_fls_rates(fls_payload, args.effective_year, fls_path.name)
    db.upsert_planning_setting(DEFAULT_RATE_CHANGE_METADATA_SETTING_KEY, json.dumps(change_payload))
    _save_nondefault_tables(lst_payload, alternate_payload, lst_path.name, ssa_path.name)

    print(
        json.dumps(
            {
                "fls_upserted_lanes": upsert_count,
                "fls_changed_cells": len(change_payload.get("changed_cells") or []),
                "lst_states": len(lst_payload.get("states") or []),
                "lst_plants": len(lst_payload.get("plants") or []),
                "alternate_trailer_sections": len(alternate_payload.get("sections") or []),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
