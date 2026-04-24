import sqlite3
import os
import re
import csv
import shutil
import tempfile
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[2]


def _is_truthy(value):
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "on"}


def _is_azure_app_service():
    return any(
        os.environ.get(key)
        for key in (
            "WEBSITE_SITE_NAME",
            "WEBSITE_INSTANCE_ID",
            "WEBSITE_HOSTNAME",
            "WEBSITE_OWNER_NAME",
            "WEBSITES_ENABLE_APP_SERVICE_STORAGE",
        )
    )


def _default_prograde_db_path():
    app_db_path_raw = os.environ.get("APP_DB_PATH")
    if app_db_path_raw:
        app_db_path = Path(str(app_db_path_raw))
        if app_db_path.suffix:
            return app_db_path.with_name("prograde.db")
        return app_db_path / "prograde.db"
    if os.environ.get("RENDER") or os.environ.get("RENDER_SERVICE_ID"):
        return Path("/var/data/prograde.db")
    if _is_azure_app_service():
        return Path("/home/site/prograde.db")
    return ROOT_DIR / "data" / "db" / "prograde.db"


DB_PATH = Path(os.environ.get("PROGRADE_DB_PATH", str(_default_prograde_db_path())))
FALLBACK_SEED_DB_PATH = ROOT_DIR / "archive" / "prograde_source_drop_v03" / "prograde.db"
DEFAULT_BT_DATA_WORKBOOK_PATH = Path(
    os.environ.get(
        "PROGRADE_BT_DATA_WORKBOOK_PATH",
        r"c:\Users\gabramowitz\OneDrive - Council Advisors\Bain Capital - ATW - ATW Operations Value Creation\03 - Phase 2\04 - Carry On MFO\PG Freight Tool\BT - Load Sheets\Stacking Guide Master.xlsx",
    )
)
FALLBACK_BT_DATA_WORKBOOK_PATH = ROOT_DIR / "data" / "reference" / "Stacking Guide Master.xlsx"
FALLBACK_BT_TEMP_WORKBOOK_PATH = ROOT_DIR / ".tmp" / "stacking_guide_master.xlsx"
DEFAULT_PJ_DATA_WORKBOOK_PATH = Path(
    os.environ.get(
        "PROGRADE_PJ_DATA_WORKBOOK_PATH",
        r"c:\Users\gabramowitz\OneDrive - Council Advisors\Bain Capital - ATW - ATW Operations Value Creation\03 - Phase 2\04 - Carry On MFO\PG Freight Tool\2024 PJ Product Guide_WORKING (as of 5.28.25).xlsx",
    )
)
FALLBACK_PJ_TEMP_WORKBOOK_PATH = ROOT_DIR / ".tmp" / "pj_product_guide_working.xlsx"

PJ_TOC_MODEL_RE = re.compile(r"\[([A-Za-z0-9]{1,8})\]")
PJ_FT_LENGTH_RE = re.compile(r"(?<!\d)(\d{1,2}(?:\.\d+)?)\s*[\'\u2019]")
PJ_DUMP_CATEGORY_KEYS = {
    "dump_lowside",
    "dump_highside_3ft",
    "dump_highside_4ft",
    "dump_small",
    "dump_gn",
    "dump_variants",
}
PJ_GOOSENECK_CATEGORY_KEYS = {
    "gooseneck",
    "gooseneck_flatdeck",
    "gooseneck_quest",
    "gooseneck_pintle",
    "gooseneck_variants",
    "pintle",
}
PJ_GOOSENECK_MODEL_PREFIXES = {"LD", "LQ", "LS", "LX", "LY", "PL"}
PJ_STACK_HEIGHT_OVERRIDES = {
    "EV": (2.0, 2.5),
    "F8": (2.0, 2.5),
    "H4": (2.0, 2.5),
    "H5": (2.0, 2.5),
    "H6": (2.0, 2.5),
    "H7": (2.0, 2.5),
    "L6": (2.0, 2.5),
    "L7": (2.0, 2.5),
    "P8": (2.9, 2.9),
    "PL": (2.5, 2.5),
}

BIGTEX_CATEGORY_ALIASES = {
    "OL CAR HAULER": "CAR HAULER",
    "OL DUMP": "DUMP",
    "OL SINGLE AXLE": "SINGLE AXLE",
    "OL TANDEM AXLE": "TANDEM AXLE",
    "OL TILT": "TILT DECK",
    "OL EQUIPMENT": "EQUIPMENT HAULER",
    "OL EQUIPMENT HAULER": "EQUIPMENT HAULER",
}

ADVANCED_SCHEMATIC_DEFAULTS = [
    {
        "drawing_key": "utility_profile",
        "drawing_label": "Utility Side Profile",
        "render_mode": "advanced",
        "applies_to_categories": "utility",
        "notes": "Base utility trailer silhouette.",
        "display_order": 10,
    },
    {
        "drawing_key": "car_hauler_profile",
        "drawing_label": "Car Hauler Side Profile",
        "render_mode": "advanced",
        "applies_to_categories": "car_hauler,car_hauler_deckover,tilt",
        "notes": "Shared shape for car hauler and tilt variants.",
        "display_order": 20,
    },
    {
        "drawing_key": "tilt_deckover_profile",
        "drawing_label": "Tilt Deck-Over Side Profile",
        "render_mode": "advanced",
        "applies_to_categories": "tilt_deckover",
        "notes": "Tall deck-over style with hinge details.",
        "display_order": 30,
    },
    {
        "drawing_key": "deck_over_profile",
        "drawing_label": "Deck-Over Side Profile",
        "render_mode": "advanced",
        "applies_to_categories": "deck_over",
        "notes": "Deck-over variants without tilt hinge.",
        "display_order": 40,
    },
    {
        "drawing_key": "dump_profile",
        "drawing_label": "Dump Side Profile",
        "render_mode": "advanced",
        "applies_to_categories": "dump_lowside,dump_highside_3ft,dump_highside_4ft,dump_small,dump_gn,dump_variants",
        "notes": "Dump-family profile with category-specific height behavior.",
        "display_order": 50,
    },
    {
        "drawing_key": "gooseneck_profile",
        "drawing_label": "Gooseneck Side Profile",
        "render_mode": "advanced",
        "applies_to_categories": "gooseneck,gooseneck_flatdeck,gooseneck_quest,gooseneck_pintle,gooseneck_variants,pintle",
        "notes": "High-coupler neck profile used by GN and pintle families.",
        "display_order": 60,
    },
]

PJ_MEASUREMENT_OFFSET_DEFAULTS = [
    {
        "rule_key": "car_hauler_spare_mount_offset",
        "label": "Extra feet for car hauler spare mount",
        "offset_ft": 1.0,
        "notes": "Applied to all car hauler and deck-over categories.",
    },
    {
        "rule_key": "dump_tarp_kit_offset",
        "label": "Extra feet for dump tarp kit",
        "offset_ft": 1.0,
        "notes": "Applied to all dump categories.",
    },
    {
        "rule_key": "dtj_cylinder_extra_offset",
        "label": "Additional feet for DTJ cylinder",
        "offset_ft": 1.0,
        "notes": "Stacks on top of tarp offset; DTJ models only.",
    },
    {
        "rule_key": "gn_in_dump_hidden_ft",
        "label": "Feet of GN tongue hidden inside dump body",
        "offset_ft": 7.0,
        "notes": "Subtracted from GN footprint when nested inside dump.",
    },
    {
        "rule_key": "gn_crisscross_length_save_ft",
        "label": "Length savings for nested GN crisscross pair",
        "offset_ft": 2.0,
        "notes": "Placeholder assumption for opposing GN necks in same stack column.",
    },
    {
        "rule_key": "gn_crisscross_height_save_ft",
        "label": "Height savings for nested GN crisscross pair",
        "offset_ft": 1.0,
        "notes": "Placeholder vertical savings when GN necks interlace.",
    },
    {
        "rule_key": "gn_crisscross_width_save_ft",
        "label": "Width savings for nested GN crisscross pair",
        "offset_ft": 0.6,
        "notes": "Placeholder lateral savings for planner reference only.",
    },
]

REFERENCE_CARRIER_DEFAULTS = [
    {
        "carrier_type": "53_step_deck",
        "brand": "pj",
        "total_length_ft": 53.0,
        "max_height_ft": 13.5,
        "lower_deck_length_ft": 41.5,
        "upper_deck_length_ft": 11.75,
        "lower_deck_ground_height_ft": 3.5,
        "upper_deck_ground_height_ft": 5.0,
        "gn_max_lower_deck_ft": 32.0,
        "notes": "Default PJ step deck carrier profile",
    },
    {
        "carrier_type": "53_flatbed",
        "brand": "bigtex",
        "total_length_ft": 53.0,
        "max_height_ft": 13.5,
        "lower_deck_length_ft": 53.0,
        "upper_deck_length_ft": 0.0,
        "lower_deck_ground_height_ft": 4.0,
        "upper_deck_ground_height_ft": 0.0,
        "gn_max_lower_deck_ft": 0.0,
        "notes": "Default Big Tex flatbed carrier profile",
    },
    {
        "carrier_type": "ground_pull",
        "brand": "bigtex",
        "total_length_ft": 53.0,
        "max_height_ft": 13.5,
        "lower_deck_length_ft": 53.0,
        "upper_deck_length_ft": 0.0,
        "lower_deck_ground_height_ft": 4.0,
        "upper_deck_ground_height_ft": 0.0,
        "gn_max_lower_deck_ft": 0.0,
        "notes": "Ground pull mode: first placed unit defines usable deck length",
    },
]

REFERENCE_PJ_TONGUE_GROUP_DEFAULTS = [
    {
        "group_id": "c_channel",
        "group_label": "C-Channel Hitch",
        "tongue_feet": 4.5,
        "model_codes": "BH, CC, CH, CS, LP, SA, UT",
        "notes": "Standard bumper-pull c-channel tongue",
    },
    {
        "group_id": "deck_over",
        "group_label": "Deck Over",
        "tongue_feet": 5.0,
        "model_codes": "DO, LD, LD2, LD3",
        "notes": "Deck-over bumper-pull tongue",
    },
    {
        "group_id": "dump_small",
        "group_label": "Dump Small",
        "tongue_feet": 5.0,
        "model_codes": "D5, D7",
        "notes": "Small/compact dump tongue",
    },
    {
        "group_id": "dump_std",
        "group_label": "Dump Standard",
        "tongue_feet": 6.0,
        "model_codes": "DL, DM, DT, DV, DW, DX, DTJ, DT1",
        "notes": "Standard dump trailer tongue",
    },
    {
        "group_id": "gooseneck",
        "group_label": "Gooseneck / LDQ",
        "tongue_feet": 9.0,
        "model_codes": "GN, LDG, LDW, LDQ",
        "notes": "Gooseneck draw-bar; tongue hides inside GN neck",
    },
    {
        "group_id": "pintle",
        "group_label": "Pintle Hook",
        "tongue_feet": 4.0,
        "model_codes": "PHT, PT",
        "notes": "Pintle hook hitch",
    },
]

REFERENCE_PJ_HEIGHT_DEFAULTS = [
    {"category": "car_hauler", "label": "Car Hauler", "height_mid_ft": 1.5, "height_top_ft": 2.5, "gn_axle_dropped_ft": None, "notes": "Bottom/mid 1.5'; top 2.5' placeholder from shipping notes"},
    {"category": "car_hauler_deckover", "label": "Car Hauler Deck Over", "height_mid_ft": 2.0, "height_top_ft": 2.5, "gn_axle_dropped_ft": None, "notes": "Deck-over style placeholder: bottom/mid 2.0'; top 2.5'"},
    {"category": "deck_over", "label": "Deck Over", "height_mid_ft": 2.0, "height_top_ft": 2.5, "gn_axle_dropped_ft": None, "notes": "Bottom/mid 2.0'; top 2.5'"},
    {"category": "dump_gn", "label": "Dump - Gooseneck", "height_mid_ft": 3.5, "height_top_ft": 3.5, "gn_axle_dropped_ft": None, "notes": "DDQ-like dump GN placeholder near 3.5'"},
    {"category": "dump_highside_3ft", "label": "Dump - High Side 3'", "height_mid_ft": 5.0, "height_top_ft": 5.0, "gn_axle_dropped_ft": None, "notes": "3' high-side dump placeholder"},
    {"category": "dump_highside_4ft", "label": "Dump - High Side 4'", "height_mid_ft": 6.0, "height_top_ft": 6.0, "gn_axle_dropped_ft": None, "notes": "4' high-side dump placeholder"},
    {"category": "dump_lowside", "label": "Dump - Low Side", "height_mid_ft": 4.0, "height_top_ft": 4.0, "gn_axle_dropped_ft": None, "notes": "Low-side dump placeholder"},
    {"category": "dump_small", "label": "Dump - Small", "height_mid_ft": 3.5, "height_top_ft": 3.5, "gn_axle_dropped_ft": None, "notes": "Compact dump placeholder; D5 often 3.0', D3 often 3.5'"},
    {"category": "dump_variants", "label": "Dump - Variants", "height_mid_ft": 4.0, "height_top_ft": 4.0, "gn_axle_dropped_ft": None, "notes": "Variant dump placeholder (DT1 family trends low-side)"},
    {"category": "gooseneck", "label": "Gooseneck", "height_mid_ft": 2.5, "height_top_ft": 2.5, "gn_axle_dropped_ft": 2.0, "notes": "GN/PLP baseline 2.5'; axle-drop 2.0'; riser envelope handled as 6.0'"},
    {"category": "pintle", "label": "Pintle", "height_mid_ft": 2.5, "height_top_ft": 2.5, "gn_axle_dropped_ft": None, "notes": "PLP-style baseline placeholder 2.5'"},
    {"category": "tilt", "label": "Tilt", "height_mid_ft": 1.5, "height_top_ft": 2.5, "gn_axle_dropped_ft": None, "notes": "Tilt placeholder aligned to car-hauler family"},
    {"category": "tilt_deckover", "label": "Tilt Deck Over", "height_mid_ft": 2.0, "height_top_ft": 2.5, "gn_axle_dropped_ft": None, "notes": "Deck-over tilt placeholder"},
    {"category": "utility", "label": "Utility", "height_mid_ft": 1.3, "height_top_ft": 2.0, "gn_axle_dropped_ft": None, "notes": "Bottom/mid 1.3'; top 2.0'"},
]

REFERENCE_BT_STACK_CONFIG_DEFAULTS = [
    {"config_id": "dump_3stack_stack_1", "label": "3-Stack Dump - Stack 1", "load_type": "dump_3stack", "stack_position": "stack_1", "max_length_ft": 21.0, "max_height_ft": 5.0, "notes": ""},
    {"config_id": "dump_3stack_stack_2", "label": "3-Stack Dump - Stack 2", "load_type": "dump_3stack", "stack_position": "stack_2", "max_length_ft": 16.0, "max_height_ft": 5.0, "notes": ""},
    {"config_id": "dump_3stack_stack_3", "label": "3-Stack Dump - Stack 3", "load_type": "dump_3stack", "stack_position": "stack_3", "max_length_ft": 16.0, "max_height_ft": 5.0, "notes": ""},
    {"config_id": "gooseneck_stack_1", "label": "Gooseneck - Stack 1", "load_type": "gooseneck", "stack_position": "stack_1", "max_length_ft": 53.0, "max_height_ft": 13.5, "notes": "Full deck length; no fixed stack cap for GN loads"},
    {"config_id": "utility_2stack_combined_1_2", "label": "2-Stack Utility - S1+S2 Combined", "load_type": "utility_2stack", "stack_position": "combined_1_2", "max_length_ft": 40.5, "max_height_ft": None, "notes": ""},
    {"config_id": "utility_3stack_combined_1_2", "label": "3-Stack Utility - S1+S2 Combined", "load_type": "utility_3stack", "stack_position": "combined_1_2", "max_length_ft": 40.0, "max_height_ft": None, "notes": "Longest unit in S1 + longest in S2 <= 40'"},
    {"config_id": "utility_3stack_stack_1", "label": "3-Stack Utility - Stack 1", "load_type": "utility_3stack", "stack_position": "stack_1", "max_length_ft": None, "max_height_ft": 5.25, "notes": "Individual cap; combined with S2 <= 40'"},
    {"config_id": "utility_3stack_stack_2", "label": "3-Stack Utility - Stack 2", "load_type": "utility_3stack", "stack_position": "stack_2", "max_length_ft": None, "max_height_ft": 5.25, "notes": "Individual cap; combined with S1 <= 40'"},
    {"config_id": "utility_3stack_stack_3", "label": "3-Stack Utility - Stack 3", "load_type": "utility_3stack", "stack_position": "stack_3", "max_length_ft": 16.0, "max_height_ft": 4.0, "notes": ""},
]


def _ensure_db_file():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists():
        return
    # Use archive bootstrap only when explicitly requested.
    if _is_truthy(os.environ.get("PROGRADE_BOOTSTRAP_FROM_ARCHIVE_DB")) and FALLBACK_SEED_DB_PATH.exists():
        shutil.copyfile(FALLBACK_SEED_DB_PATH, DB_PATH)
        return
    DB_PATH.touch()


def _normalize_profile_name(value):
    return " ".join(str(value or "").strip().split())


def _default_admin_profile_name():
    name = (
        os.environ.get("PROGRADE_DEFAULT_ADMIN_NAME")
        or os.environ.get("USERNAME")
        or "Admin"
    )
    normalized = _normalize_profile_name(name)
    return normalized or "Admin"


def normalize_bigtex_mcat(value):
    label = " ".join(str(value or "").strip().split())
    if not label:
        return ""
    canonical = BIGTEX_CATEGORY_ALIASES.get(label.upper())
    if canonical:
        return canonical
    return label


def _sqlite_busy_timeout_sec():
    timeout_raw = os.environ.get(
        "PROGRADE_SQLITE_BUSY_TIMEOUT_SEC",
        os.environ.get("SQLITE_BUSY_TIMEOUT_SEC", "30"),
    )
    try:
        return max(float(timeout_raw), 1.0)
    except (TypeError, ValueError):
        return 30.0


def get_db():
    _ensure_db_file()
    timeout_sec = _sqlite_busy_timeout_sec()
    timeout_ms = int(timeout_sec * 1000)
    conn = sqlite3.connect(DB_PATH, timeout=timeout_sec)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute(f"PRAGMA busy_timeout={timeout_ms}")
    return conn


def _seed_reference_defaults(cursor):
    now = datetime.utcnow().isoformat()
    cursor.executemany(
        """
        INSERT OR IGNORE INTO carrier_configs
        (
            carrier_type,
            brand,
            total_length_ft,
            max_height_ft,
            lower_deck_length_ft,
            upper_deck_length_ft,
            lower_deck_ground_height_ft,
            upper_deck_ground_height_ft,
            gn_max_lower_deck_ft,
            notes,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["carrier_type"],
                row["brand"],
                row["total_length_ft"],
                row["max_height_ft"],
                row["lower_deck_length_ft"],
                row["upper_deck_length_ft"],
                row["lower_deck_ground_height_ft"],
                row["upper_deck_ground_height_ft"],
                row["gn_max_lower_deck_ft"],
                row["notes"],
                now,
            )
            for row in REFERENCE_CARRIER_DEFAULTS
        ],
    )

    cursor.executemany(
        """
        INSERT OR IGNORE INTO pj_tongue_groups
        (group_id, group_label, tongue_feet, model_codes, notes, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["group_id"],
                row["group_label"],
                row["tongue_feet"],
                row["model_codes"],
                row["notes"],
                now,
            )
            for row in REFERENCE_PJ_TONGUE_GROUP_DEFAULTS
        ],
    )

    cursor.executemany(
        """
        INSERT OR IGNORE INTO pj_height_reference
        (category, label, height_mid_ft, height_top_ft, gn_axle_dropped_ft, notes, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["category"],
                row["label"],
                row["height_mid_ft"],
                row["height_top_ft"],
                row["gn_axle_dropped_ft"],
                row["notes"],
                now,
            )
            for row in REFERENCE_PJ_HEIGHT_DEFAULTS
        ],
    )

    cursor.executemany(
        """
        INSERT OR IGNORE INTO bt_stack_configs
        (config_id, label, load_type, stack_position, max_length_ft, max_height_ft, notes, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["config_id"],
                row["label"],
                row["load_type"],
                row["stack_position"],
                row["max_length_ft"],
                row["max_height_ft"],
                row["notes"],
                now,
            )
            for row in REFERENCE_BT_STACK_CONFIG_DEFAULTS
        ],
    )


def init_db():
    conn = get_db()
    c = conn.cursor()

    # Add new columns to existing tables (safe to run multiple times)
    migrations = [
        "ALTER TABLE load_sessions ADD COLUMN acknowledged_violations TEXT DEFAULT '[]'",
        "ALTER TABLE load_sessions ADD COLUMN is_saved INTEGER DEFAULT 0",
        "ALTER TABLE load_sessions ADD COLUMN created_by_profile_id INTEGER",
        "ALTER TABLE load_sessions ADD COLUMN created_by_name TEXT",
        "ALTER TABLE load_positions ADD COLUMN is_rotated INTEGER DEFAULT 0",
        "ALTER TABLE bt_inventory_upload_log ADD COLUMN source_format TEXT DEFAULT 'workbook'",
        "ALTER TABLE bt_inventory_upload_log ADD COLUMN deduped_rows INTEGER DEFAULT 0",
        "ALTER TABLE bt_inventory_upload_log ADD COLUMN duplicate_rows INTEGER DEFAULT 0",
        "ALTER TABLE bt_inventory_upload_log ADD COLUMN warehouse_count INTEGER DEFAULT 0",
        "ALTER TABLE pj_inventory_upload_log ADD COLUMN source_format TEXT DEFAULT 'csv_inventory'",
        "ALTER TABLE pj_inventory_upload_log ADD COLUMN deduped_rows INTEGER DEFAULT 0",
        "ALTER TABLE pj_inventory_upload_log ADD COLUMN duplicate_rows INTEGER DEFAULT 0",
        "ALTER TABLE pj_inventory_upload_log ADD COLUMN warehouse_count INTEGER DEFAULT 0",
        "ALTER TABLE pj_inventory_upload_log ADD COLUMN matched_rows INTEGER DEFAULT 0",
        "ALTER TABLE pj_inventory_upload_log ADD COLUMN matched_items INTEGER DEFAULT 0",
        "ALTER TABLE pj_inventory_upload_log ADD COLUMN unmatched_items INTEGER DEFAULT 0",
        "ALTER TABLE pj_skus ADD COLUMN height_mid_ft REAL",
        "ALTER TABLE pj_skus ADD COLUMN height_top_ft REAL",
    ]
    for m in migrations:
        try:
            c.execute(m)
        except Exception:
            pass

    c.executescript("""
    CREATE TABLE IF NOT EXISTS carrier_configs (
        carrier_type TEXT PRIMARY KEY,
        brand TEXT,
        total_length_ft REAL,
        max_height_ft REAL,
        lower_deck_length_ft REAL,
        upper_deck_length_ft REAL,
        lower_deck_ground_height_ft REAL,
        upper_deck_ground_height_ft REAL,
        gn_max_lower_deck_ft REAL,
        notes TEXT,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS pj_tongue_groups (
        group_id TEXT PRIMARY KEY,
        group_label TEXT,
        tongue_feet REAL,
        model_codes TEXT,
        notes TEXT,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS pj_height_reference (
        category TEXT PRIMARY KEY,
        label TEXT,
        height_mid_ft REAL,
        height_top_ft REAL,
        gn_axle_dropped_ft REAL,
        notes TEXT,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS pj_measurement_offsets (
        rule_key TEXT PRIMARY KEY,
        label TEXT,
        offset_ft REAL,
        notes TEXT,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS pj_skus (
        item_number TEXT PRIMARY KEY,
        model TEXT,
        pj_category TEXT,
        description TEXT,
        gvwr INTEGER,
        bed_length_stated REAL,
        bed_length_measured REAL,
        tongue_group TEXT,
        tongue_feet REAL,
        height_mid_ft REAL,
        height_top_ft REAL,
        total_footprint REAL,
        dump_side_height_ft REAL,
        can_nest_inside_dump INTEGER DEFAULT 0,
        gn_axle_droppable INTEGER DEFAULT 0,
        tongue_overlap_allowed INTEGER DEFAULT 0,
        pairing_rule TEXT,
        notes TEXT,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS bigtex_skus (
        item_number TEXT PRIMARY KEY,
        mcat TEXT,
        tier INTEGER,
        model TEXT,
        gvwr INTEGER,
        floor_type TEXT,
        bed_length REAL,
        width REAL,
        tongue REAL,
        stack_height REAL,
        total_footprint REAL,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS bt_stack_configs (
        config_id TEXT PRIMARY KEY,
        label TEXT,
        load_type TEXT,
        stack_position TEXT,
        max_length_ft REAL,
        max_height_ft REAL,
        notes TEXT,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS advanced_schematic_links (
        drawing_key TEXT PRIMARY KEY,
        drawing_label TEXT,
        render_mode TEXT DEFAULT 'advanced',
        applies_to_categories TEXT,
        notes TEXT,
        display_order INTEGER DEFAULT 100,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS load_sessions (
        session_id TEXT PRIMARY KEY,
        brand TEXT,
        carrier_type TEXT,
        acknowledged_violations TEXT DEFAULT '[]',
        status TEXT DEFAULT 'draft',
        is_saved INTEGER DEFAULT 0,
        planner_name TEXT,
        created_by_profile_id INTEGER,
        created_by_name TEXT,
        session_label TEXT,
        created_at DATETIME,
        updated_at DATETIME,
        approved_by TEXT,
        approved_at DATETIME,
        notes TEXT
    );

    CREATE TABLE IF NOT EXISTS prograde_access_profiles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE COLLATE NOCASE,
        is_admin INTEGER NOT NULL DEFAULT 0,
        created_at DATETIME,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS load_positions (
        position_id TEXT PRIMARY KEY,
        session_id TEXT,
        brand TEXT,
        item_number TEXT,
        deck_zone TEXT,
        layer INTEGER,
        sequence INTEGER,
        is_nested INTEGER DEFAULT 0,
        nested_inside TEXT,
        gn_axle_dropped INTEGER DEFAULT 0,
        is_rotated INTEGER DEFAULT 0,
        override_reason TEXT,
        added_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS load_patterns (
        pattern_id TEXT PRIMARY KEY,
        brand TEXT,
        pattern_name TEXT,
        load_type TEXT,
        carrier_type TEXT,
        source TEXT,
        confidence INTEGER DEFAULT 3,
        positions_json TEXT,
        unit_count INTEGER,
        notes TEXT,
        created_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS bt_inventory_snapshot (
        item_number TEXT PRIMARY KEY,
        total_count INTEGER DEFAULT 0,
        available_count INTEGER DEFAULT 0,
        assigned_count INTEGER DEFAULT 0,
        built_count INTEGER DEFAULT 0,
        future_build_count INTEGER DEFAULT 0,
        available_built_count INTEGER DEFAULT 0,
        available_future_count INTEGER DEFAULT 0,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS bt_inventory_snapshot_whse (
        item_number TEXT NOT NULL,
        whse_code TEXT NOT NULL,
        total_count INTEGER DEFAULT 0,
        available_count INTEGER DEFAULT 0,
        assigned_count INTEGER DEFAULT 0,
        built_count INTEGER DEFAULT 0,
        future_build_count INTEGER DEFAULT 0,
        available_built_count INTEGER DEFAULT 0,
        available_future_count INTEGER DEFAULT 0,
        updated_at DATETIME,
        PRIMARY KEY (item_number, whse_code)
    );

    CREATE TABLE IF NOT EXISTS bt_inventory_upload_log (
        upload_id TEXT PRIMARY KEY,
        source_filename TEXT,
        sheet_name TEXT,
        source_format TEXT DEFAULT 'workbook',
        processed_rows INTEGER DEFAULT 0,
        valid_rows INTEGER DEFAULT 0,
        distinct_items INTEGER DEFAULT 0,
        deduped_rows INTEGER DEFAULT 0,
        duplicate_rows INTEGER DEFAULT 0,
        warehouse_count INTEGER DEFAULT 0,
        uploaded_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS pj_inventory_snapshot (
        item_number TEXT PRIMARY KEY,
        source_item_number TEXT,
        match_method TEXT,
        normalized_model TEXT,
        normalized_category TEXT,
        footprint_each REAL DEFAULT 0,
        stack_height_each REAL DEFAULT 0,
        total_count INTEGER DEFAULT 0,
        available_count INTEGER DEFAULT 0,
        assigned_count INTEGER DEFAULT 0,
        updated_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS pj_inventory_snapshot_whse (
        item_number TEXT NOT NULL,
        whse_code TEXT NOT NULL,
        source_item_number TEXT,
        match_method TEXT,
        normalized_model TEXT,
        normalized_category TEXT,
        footprint_each REAL DEFAULT 0,
        stack_height_each REAL DEFAULT 0,
        total_count INTEGER DEFAULT 0,
        available_count INTEGER DEFAULT 0,
        assigned_count INTEGER DEFAULT 0,
        updated_at DATETIME,
        PRIMARY KEY (item_number, whse_code)
    );

    CREATE TABLE IF NOT EXISTS pj_inventory_upload_log (
        upload_id TEXT PRIMARY KEY,
        source_filename TEXT,
        source_format TEXT DEFAULT 'csv_inventory',
        processed_rows INTEGER DEFAULT 0,
        valid_rows INTEGER DEFAULT 0,
        deduped_rows INTEGER DEFAULT 0,
        duplicate_rows INTEGER DEFAULT 0,
        distinct_items INTEGER DEFAULT 0,
        warehouse_count INTEGER DEFAULT 0,
        matched_rows INTEGER DEFAULT 0,
        matched_items INTEGER DEFAULT 0,
        unmatched_items INTEGER DEFAULT 0,
        uploaded_at DATETIME
    );

    CREATE TABLE IF NOT EXISTS app_meta (
        meta_key TEXT PRIMARY KEY,
        meta_value TEXT,
        updated_at DATETIME
    );

    CREATE INDEX IF NOT EXISTS idx_bt_inventory_snapshot_available
        ON bt_inventory_snapshot(available_count DESC, item_number);

    CREATE INDEX IF NOT EXISTS idx_bt_inventory_snapshot_whse_lookup
        ON bt_inventory_snapshot_whse(whse_code, available_count DESC, item_number);

    CREATE INDEX IF NOT EXISTS idx_bt_inventory_upload_log_uploaded_at
        ON bt_inventory_upload_log(uploaded_at DESC);

    CREATE INDEX IF NOT EXISTS idx_pj_inventory_snapshot_available
        ON pj_inventory_snapshot(available_count DESC, item_number);

    CREATE INDEX IF NOT EXISTS idx_pj_inventory_snapshot_whse_lookup
        ON pj_inventory_snapshot_whse(whse_code, available_count DESC, item_number);

    CREATE INDEX IF NOT EXISTS idx_pj_inventory_upload_log_uploaded_at
        ON pj_inventory_upload_log(uploaded_at DESC);

    CREATE INDEX IF NOT EXISTS idx_prograde_access_profiles_name
        ON prograde_access_profiles(name COLLATE NOCASE);
    """)

    _seed_reference_defaults(c)

    # One-time migration: preserve all pre-existing sessions as saved so history remains visible.
    backfill_marker = c.execute(
        "SELECT meta_value FROM app_meta WHERE meta_key='is_saved_backfill_v1'"
    ).fetchone()
    if not backfill_marker:
        now = datetime.utcnow().isoformat()
        try:
            c.execute("UPDATE load_sessions SET is_saved=1")
        except Exception:
            pass
        c.execute(
            """
            INSERT OR REPLACE INTO app_meta(meta_key, meta_value, updated_at)
            VALUES ('is_saved_backfill_v1', '1', ?)
            """,
            (now,),
        )

    backfill_marker_v2 = c.execute(
        "SELECT meta_value FROM app_meta WHERE meta_key='is_saved_backfill_v2'"
    ).fetchone()
    if not backfill_marker_v2:
        now = datetime.utcnow().isoformat()
        try:
            c.execute("UPDATE load_sessions SET is_saved=1 WHERE COALESCE(is_saved, 0)=0")
        except Exception:
            pass
        c.execute(
            """
            INSERT OR REPLACE INTO app_meta(meta_key, meta_value, updated_at)
            VALUES ('is_saved_backfill_v2', '1', ?)
            """,
            (now,),
        )

    advanced_link_count = c.execute("SELECT COUNT(*) FROM advanced_schematic_links").fetchone()[0]
    if advanced_link_count == 0:
        now = datetime.utcnow().isoformat()
        c.executemany(
            """
            INSERT INTO advanced_schematic_links
            (drawing_key, drawing_label, render_mode, applies_to_categories, notes, display_order, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["drawing_key"],
                    row["drawing_label"],
                    row["render_mode"],
                    row["applies_to_categories"],
                    row["notes"],
                    row["display_order"],
                    now,
                )
                for row in ADVANCED_SCHEMATIC_DEFAULTS
            ],
        )

    now = datetime.utcnow().isoformat()
    c.executemany(
        """
        INSERT OR IGNORE INTO pj_measurement_offsets
        (rule_key, label, offset_ft, notes, updated_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        [
            (
                row["rule_key"],
                row["label"],
                row["offset_ft"],
                row["notes"],
                now,
            )
            for row in PJ_MEASUREMENT_OFFSET_DEFAULTS
        ],
    )

    c.execute(
        """
        UPDATE load_sessions
        SET created_by_name = COALESCE(NULLIF(created_by_name, ''), planner_name)
        WHERE COALESCE(created_by_name, '') = '' AND COALESCE(planner_name, '') != ''
        """
    )

    admin_name = _default_admin_profile_name()
    now = datetime.utcnow().isoformat()
    admin_row = c.execute(
        "SELECT id, is_admin FROM prograde_access_profiles WHERE lower(name)=lower(?)",
        (admin_name,),
    ).fetchone()
    if admin_row:
        if not int(admin_row["is_admin"] or 0):
            c.execute(
                "UPDATE prograde_access_profiles SET is_admin=1, updated_at=? WHERE id=?",
                (now, int(admin_row["id"])),
            )
    else:
        c.execute(
            """
            INSERT INTO prograde_access_profiles (name, is_admin, created_at, updated_at)
            VALUES (?, 1, ?, ?)
            """,
            (admin_name, now, now),
        )

    conn.commit()

    # Default behavior: keep SKU catalogs synchronized with seed CSV snapshots on startup.
    # Optional preservation mode for future use:
    #   PROGRADE_PRESERVE_SKU_EDITS_ON_START=true
    # In preservation mode we only seed empty tables and do not upsert over existing edits.
    preserve_sku_edits = _is_truthy(os.environ.get("PROGRADE_PRESERVE_SKU_EDITS_ON_START"))
    if preserve_sku_edits:
        _seed_skus_from_csv(conn, mode="bootstrap_if_empty")
    else:
        _seed_skus_from_csv(conn, mode="upsert")

    conn.close()


def _seed_skus_from_csv(conn, mode="bootstrap_if_empty"):
    """Load pj_skus and bigtex_skus from bundled seed CSVs."""
    import csv as _csv

    seed_dir = ROOT_DIR / "data" / "seed"
    c = conn.cursor()

    def _read_seed_rows(seed_filename):
        seed_path = seed_dir / seed_filename
        if not seed_path.exists():
            return []

        with open(seed_path, newline="", encoding="utf-8") as f:
            return list(_csv.DictReader(f))

    def _insert_seed_rows(table_name, rows):
        if not rows:
            return 0
        cols = list(rows[0].keys())
        placeholders = ", ".join("?" * len(cols))
        col_list = ", ".join(cols)
        query = f"INSERT OR IGNORE INTO {table_name} ({col_list}) VALUES ({placeholders})"
        payload = [[row.get(col) or None for col in cols] for row in rows]
        c.executemany(query, payload)
        return len(payload)

    def _upsert_seed_csv(table_name, seed_filename, key_column):
        rows = _read_seed_rows(seed_filename)
        if not rows:
            return 0

        cols = list(rows[0].keys())
        placeholders = ", ".join("?" * len(cols))
        col_list = ", ".join(cols)
        update_cols = [col for col in cols if col != key_column]
        update_clause = ", ".join(f"{col}=excluded.{col}" for col in update_cols)
        query = (
            f"INSERT INTO {table_name} ({col_list}) VALUES ({placeholders}) "
            f"ON CONFLICT({key_column}) DO UPDATE SET {update_clause}"
        )
        payload = [[row.get(col) or None for col in cols] for row in rows]
        c.executemany(query, payload)
        return len(payload)

    def _seed_if_table_empty(table_name, seed_filename):
        existing = c.execute(f"SELECT COUNT(*) AS cnt FROM {table_name}").fetchone()
        if int(existing["cnt"] or 0) > 0:
            return 0
        rows = _read_seed_rows(seed_filename)
        return _insert_seed_rows(table_name, rows)

    normalized_mode = str(mode or "").strip().lower()
    if normalized_mode == "upsert":
        _upsert_seed_csv("pj_skus", "pj_skus.csv", "item_number")
        _upsert_seed_csv("bigtex_skus", "bigtex_skus.csv", "item_number")
    else:
        _seed_if_table_empty("pj_skus", "pj_skus.csv")
        _seed_if_table_empty("bigtex_skus", "bigtex_skus.csv")
    conn.commit()


def has_seed_data():
    with get_db() as conn:
        carrier_count = conn.execute("SELECT COUNT(*) FROM carrier_configs").fetchone()[0]
        pj_count = conn.execute("SELECT COUNT(*) FROM pj_skus").fetchone()[0]
        bt_count = conn.execute("SELECT COUNT(*) FROM bigtex_skus").fetchone()[0]
    return carrier_count > 0 and (pj_count > 0 or bt_count > 0)


def _coerce_float(value, default=None):
    if value is None:
        return default
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return default
        value = raw
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _coerce_int(value, default=None):
    if value is None:
        return default
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return default
        value = raw
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _normalize_model_code(value):
    return "".join(ch for ch in str(value or "").strip().upper() if ch.isalnum())[:2]


def _infer_pj_dump_wall_height_ft(category, model_code=None):
    category_key = str(category or "").strip().lower()
    model = _normalize_model_code(model_code)
    if category_key == "dump_highside_4ft":
        return 4.0
    if category_key == "dump_highside_3ft":
        return 3.0
    if category_key in {"dump_lowside", "dump_small"}:
        return 2.0
    if category_key == "dump_variants":
        if model == "DX":
            return 4.0
        if model == "DV":
            return 3.0
        return 3.0
    if category_key == "dump_gn":
        return 3.0
    return None


def _dump_stacked_height_from_wall_height(dump_side_height_ft):
    wall = _coerce_float(dump_side_height_ft, None)
    if wall is None:
        return None
    if abs(wall - 2.0) <= 0.05:
        return 4.0
    if abs(wall - 3.0) <= 0.05:
        return 5.0
    if abs(wall - 4.0) <= 0.05:
        return 6.0
    return None


def _default_pj_tongue_profile(model_code, category):
    category_key = str(category or "").strip().lower()
    model = _normalize_model_code(model_code)
    if category_key in PJ_GOOSENECK_CATEGORY_KEYS or model in PJ_GOOSENECK_MODEL_PREFIXES:
        return "gooseneck"
    return "standard"


def _default_pj_tongue_feet(model_code, category):
    category_key = str(category or "").strip().lower()
    model = _normalize_model_code(model_code)
    if _default_pj_tongue_profile(model, category_key) == "gooseneck":
        return 9.0
    if category_key in PJ_DUMP_CATEGORY_KEYS:
        return 5.5
    if category_key in {"deck_over", "car_hauler_deckover", "tilt_deckover"}:
        return 6.0
    if model in {"F8", "L6", "L7"}:
        return 6.0
    return 4.0


def _default_pj_stack_heights(model_code, category, dump_side_height_ft=None):
    category_key = str(category or "").strip().lower()
    model = _normalize_model_code(model_code)
    if category_key in PJ_DUMP_CATEGORY_KEYS:
        wall = _coerce_float(dump_side_height_ft, None)
        if wall is None:
            wall = _infer_pj_dump_wall_height_ft(category_key, model)
        stacked = _dump_stacked_height_from_wall_height(wall)
        if stacked is None:
            stacked = 4.0
        return (stacked, stacked)
    if model in PJ_STACK_HEIGHT_OVERRIDES:
        return PJ_STACK_HEIGHT_OVERRIDES[model]
    if model.startswith("T"):
        return (2.0, 2.5)
    if category_key == "utility":
        return (1.25, 2.0)
    if category_key in PJ_GOOSENECK_CATEGORY_KEYS:
        return (2.5, 2.5)
    if category_key in {"car_hauler", "deck_over", "tilt", "car_hauler_deckover", "tilt_deckover"}:
        return (2.0, 2.5)
    return (2.0, 2.5)


def _pj_short_item_code(model_code, item_number, category=None):
    model_prefix = _normalize_model_code(model_code)
    item = "".join(ch for ch in str(item_number or "").strip().upper() if ch.isalnum())
    if not item:
        return ""
    if not model_prefix:
        model_prefix = item[:2]
    tail = item
    if model_prefix and item.startswith(model_prefix):
        tail = item[len(model_prefix):]
    tail = re.sub(r"^[A-Z]+", "", tail)
    digits = "".join(ch for ch in tail if ch.isdigit())
    if not digits:
        digits = "".join(ch for ch in item if ch.isdigit())
    category_key = str(category or "").strip().lower()
    if (
        len(digits) >= 3
        and digits.startswith("2")
        and (model_prefix in {"C4", "C5"} or model_prefix.startswith("U"))
    ):
        digits = digits[1:]
    if len(digits) < 2:
        return item
    return f"{model_prefix}{digits[:2]}"


def _bed_length_from_pj_item_code(model_code, item_number, category=None, fallback=None):
    short_code = _pj_short_item_code(model_code, item_number, category)
    if len(short_code) >= 4 and short_code[-2:].isdigit():
        return float(short_code[-2:])
    return _coerce_float(fallback, None)


def _normalize_header(value):
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _first_readable_path(candidates):
    for path in candidates:
        try:
            if path and path.exists():
                with open(path, "rb"):
                    pass
                return path
        except OSError:
            continue
    return None


def get_bigtex_workbook_path(preferred_path=None):
    candidates = []
    if preferred_path:
        candidates.append(Path(str(preferred_path)))
    env_path = os.environ.get("PROGRADE_BT_DATA_WORKBOOK_PATH")
    if env_path:
        candidates.append(Path(env_path))
    candidates.append(DEFAULT_BT_DATA_WORKBOOK_PATH)
    candidates.append(FALLBACK_BT_DATA_WORKBOOK_PATH)
    candidates.append(FALLBACK_BT_TEMP_WORKBOOK_PATH)
    return _first_readable_path(candidates)


def get_pj_workbook_path(preferred_path=None):
    candidates = []
    if preferred_path:
        candidates.append(Path(str(preferred_path)))
    env_path = os.environ.get("PROGRADE_PJ_DATA_WORKBOOK_PATH")
    if env_path:
        candidates.append(Path(env_path))
    candidates.append(DEFAULT_PJ_DATA_WORKBOOK_PATH)
    candidates.append(FALLBACK_PJ_TEMP_WORKBOOK_PATH)
    return _first_readable_path(candidates)


def _normalize_pj_toc_heading(label):
    text = str(label or "").strip().lower()
    if not text:
        return "unknown"
    if "utility" in text:
        return "utility"
    if "car hauler" in text or "carhauler" in text or "equipment" in text:
        return "car_hauler"
    if "tilt" in text:
        return "tilt"
    if "dump" in text:
        return "dump_variants"
    if "flatdeck" in text or "deckover" in text:
        return "deck_over"
    return "unknown"


def _infer_pj_category(section_label, model_code):
    code = str(model_code or "").strip().upper()
    if code in {"LDQ", "LDG", "LDW", "GN"}:
        return "gooseneck"
    if code in {"PL", "PT", "PHT"}:
        return "pintle"
    if code == "DL":
        return "dump_lowside"
    if code == "DV":
        return "dump_highside_3ft"
    if code == "DX":
        return "dump_highside_4ft"
    if code in {"D5", "D7"}:
        return "dump_small"
    if code.startswith("D"):
        return "dump_variants"
    normalized = _normalize_pj_toc_heading(section_label)
    if normalized != "unknown":
        return normalized
    return "utility"


def _parse_bed_length_ft(description):
    text = str(description or "").strip()
    if not text:
        return None
    m = PJ_FT_LENGTH_RE.search(text)
    if not m:
        return None
    value = m.group(1).lstrip("0")
    if value.startswith("."):
        value = "0" + value
    if value == "":
        value = "0"
    return _coerce_float(value, None)


def _parse_bed_length_from_item_number(item_number, model_code):
    sku = "".join(ch for ch in str(item_number or "").upper() if ch.isalnum())
    model = "".join(ch for ch in str(model_code or "").upper() if ch.isalnum())
    tail = sku
    if model:
        idx = sku.find(model)
        if idx >= 0:
            tail = sku[idx + len(model):]
    for token in re.findall(r"\d{2}", tail):
        value = _coerce_int(token, None)
        if value is not None and 8 <= value <= 40:
            return float(value)
    return None


def _find_code_description_header_row(sheet, max_scan_rows=40, max_scan_cols=12):
    scan_rows = min(max_scan_rows, sheet.max_row or max_scan_rows)
    scan_cols = min(max_scan_cols, sheet.max_column or max_scan_cols)
    for row_idx in range(1, scan_rows + 1):
        code_col = None
        desc_col = None
        for col_idx in range(1, scan_cols + 1):
            name = _normalize_header(sheet.cell(row=row_idx, column=col_idx).value)
            if name == "code":
                code_col = col_idx
            elif name == "description":
                desc_col = col_idx
        if code_col and desc_col:
            return row_idx, code_col, desc_col
    return None, None, None


def _extract_pj_sheet_code_rows(sheet):
    header_row, code_col, desc_col = _find_code_description_header_row(sheet)
    if not header_row:
        return []

    rows = []
    blank_streak = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        code_val = sheet.cell(row=row_idx, column=code_col).value
        desc_val = sheet.cell(row=row_idx, column=desc_col).value

        code = str(code_val or "").strip()
        desc = str(desc_val or "").strip()
        if not code:
            blank_streak += 1
            if blank_streak >= 3 and rows:
                break
            continue
        blank_streak = 0

        marker = _normalize_header(code)
        if marker in {"standardfeatures", "optionalfeatures"}:
            break
        if code.startswith("•"):
            break
        if marker == "code":
            continue

        rows.append(
            {
                "item_number": code.upper(),
                "description": desc,
                "row_idx": row_idx,
            }
        )
    return rows


def _parse_pj_toc_models(workbook, toc_sheet_name="ToC"):
    sheet_name = toc_sheet_name if toc_sheet_name in workbook.sheetnames else None
    if sheet_name is None:
        for name in workbook.sheetnames:
            if str(name).strip().lower() == str(toc_sheet_name).strip().lower():
                sheet_name = name
                break
    if sheet_name is None:
        raise ValueError(f"PJ workbook is missing '{toc_sheet_name}' sheet.")

    toc_sheet = workbook[sheet_name]
    current_section = None
    models = []
    skipped_codes = []
    for row_idx in range(1, toc_sheet.max_row + 1):
        value = toc_sheet.cell(row=row_idx, column=1).value
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue

        matches = PJ_TOC_MODEL_RE.findall(text)
        cell = toc_sheet.cell(row=row_idx, column=1)
        if not matches:
            if bool(getattr(cell.font, "bold", False)) and text.lower() != "table of contents":
                current_section = text
            continue

        model_code = str(matches[-1]).strip().upper()
        if len(model_code) != 2:
            skipped_codes.append({"row": row_idx, "code": model_code, "title": text})
            continue
        models.append(
            {
                "model": model_code,
                "section": current_section or "Uncategorized",
                "sheet_name": model_code,
                "toc_title": text,
                "row": row_idx,
            }
        )

    deduped = {}
    for entry in models:
        deduped.setdefault(entry["model"], entry)
    return {
        "toc_sheet_name": sheet_name,
        "models": list(deduped.values()),
        "skipped_codes": skipped_codes,
    }


def _infer_tongue_group_for_model(model_code, section_label, direct_model_map, group_meta):
    model = str(model_code or "").strip().upper()
    if model in direct_model_map:
        gid = direct_model_map[model]
        meta = group_meta.get(gid, {})
        return {
            "group_id": gid,
            "tongue_feet": _coerce_float(meta.get("tongue_feet"), None),
            "source": "direct",
        }

    inferred_group = None
    if model in {"D5", "D7"}:
        inferred_group = "dump_small"
    elif model.startswith("D"):
        inferred_group = "dump_std"
    elif model in {"PL", "PT", "PHT"}:
        inferred_group = "pintle"
    elif model in {"LDQ", "LDG", "LDW", "GN"}:
        inferred_group = "gooseneck"
    else:
        section_key = _normalize_pj_toc_heading(section_label)
        if section_key in {"utility", "car_hauler", "tilt"}:
            inferred_group = "c_channel"
        elif section_key == "dump_variants":
            inferred_group = "dump_std"
        elif section_key == "deck_over":
            inferred_group = "deck_over"

    if inferred_group and inferred_group in group_meta:
        meta = group_meta[inferred_group]
        return {
            "group_id": inferred_group,
            "tongue_feet": _coerce_float(meta.get("tongue_feet"), None),
            "source": "heuristic",
        }

    return {"group_id": None, "tongue_feet": None, "source": "missing"}


# ── Carrier configs ──────────────────────────────────────────────────────────

def get_carrier_configs():
    with get_db() as conn:
        return conn.execute("SELECT * FROM carrier_configs ORDER BY brand, carrier_type").fetchall()

def get_carrier_config(carrier_type):
    with get_db() as conn:
        return conn.execute("SELECT * FROM carrier_configs WHERE carrier_type=?", (carrier_type,)).fetchone()

def update_carrier_config(carrier_type, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE carrier_configs SET {field}=?, updated_at=? WHERE carrier_type=?",
            (value, datetime.utcnow().isoformat(), carrier_type)
        )


# ── PJ tongue groups ─────────────────────────────────────────────────────────

def get_advanced_schematic_links():
    with get_db() as conn:
        return conn.execute(
            "SELECT * FROM advanced_schematic_links ORDER BY display_order, drawing_key"
        ).fetchall()


def update_advanced_schematic_link(drawing_key, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE advanced_schematic_links SET {field}=?, updated_at=? WHERE drawing_key=?",
            (value, datetime.utcnow().isoformat(), drawing_key),
        )


def get_pj_tongue_groups():
    with get_db() as conn:
        return conn.execute("SELECT * FROM pj_tongue_groups ORDER BY group_id").fetchall()

def update_pj_tongue_group(group_id, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE pj_tongue_groups SET {field}=?, updated_at=? WHERE group_id=?",
            (value, datetime.utcnow().isoformat(), group_id)
        )


# ── PJ height reference ──────────────────────────────────────────────────────

def get_pj_height_reference():
    with get_db() as conn:
        return conn.execute("SELECT * FROM pj_height_reference ORDER BY category").fetchall()

def update_pj_height_reference(category, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE pj_height_reference SET {field}=?, updated_at=? WHERE category=?",
            (value, datetime.utcnow().isoformat(), category)
        )


# ── PJ measurement offsets ───────────────────────────────────────────────────

def get_pj_measurement_offsets():
    with get_db() as conn:
        return conn.execute("SELECT * FROM pj_measurement_offsets ORDER BY rule_key").fetchall()

def get_pj_offsets_dict():
    rows = get_pj_measurement_offsets()
    return {r["rule_key"]: r["offset_ft"] for r in rows}

def update_pj_measurement_offset(rule_key, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE pj_measurement_offsets SET {field}=?, updated_at=? WHERE rule_key=?",
            (value, datetime.utcnow().isoformat(), rule_key)
        )


def get_pj_height_ref_dict():
    """Return {category: dict} derived from PJ SKU cheat-sheet heights."""
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT
                pj_category AS category,
                MAX(height_mid_ft) AS height_mid_ft,
                MAX(height_top_ft) AS height_top_ft
            FROM pj_skus
            WHERE COALESCE(TRIM(pj_category), '') <> ''
            GROUP BY pj_category
            """
        ).fetchall()
    return {
        str(r["category"]): {
            "category": str(r["category"]),
            "height_mid_ft": _coerce_float(r["height_mid_ft"], None),
            "height_top_ft": _coerce_float(r["height_top_ft"], None),
            "gn_axle_dropped_ft": None,
        }
        for r in rows
        if str(r["category"] or "").strip()
    }


# ── PJ SKUs ──────────────────────────────────────────────────────────────────

def get_pj_skus(search=None, category=None):
    sql = "SELECT * FROM pj_skus"
    params = []
    clauses = []
    if search:
        clauses.append("(model LIKE ? OR item_number LIKE ? OR description LIKE ?)")
        params += [f"%{search}%", f"%{search}%", f"%{search}%"]
    if category:
        clauses.append("pj_category=?")
        params.append(category)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY model, bed_length_stated"
    with get_db() as conn:
        return conn.execute(sql, params).fetchall()

def get_pj_sku(item_number):
    with get_db() as conn:
        return conn.execute("SELECT * FROM pj_skus WHERE item_number=?", (item_number,)).fetchone()

def update_pj_sku_field(item_number, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE pj_skus SET {field}=?, updated_at=? WHERE item_number=?",
            (value, datetime.utcnow().isoformat(), item_number)
        )


# ── Big Tex SKUs ─────────────────────────────────────────────────────────────

def recompute_pj_footprint(item_number):
    row = get_pj_sku(item_number)
    if not row:
        return None
    bed_length_measured = _coerce_float(row["bed_length_measured"], None)
    if bed_length_measured is None:
        bed_length_measured = _coerce_float(row["bed_length_stated"], 0.0) or 0.0
    tongue_feet = _coerce_float(row["tongue_feet"], 0.0) or 0.0
    total_footprint = round(float(bed_length_measured) + float(tongue_feet), 2)
    with get_db() as conn:
        conn.execute(
            "UPDATE pj_skus SET total_footprint=?, updated_at=? WHERE item_number=?",
            (total_footprint, datetime.utcnow().isoformat(), item_number),
        )
    return {
        "item_number": item_number,
        "bed_length_measured": round(float(bed_length_measured), 2),
        "tongue_feet": round(float(tongue_feet), 2),
        "total_footprint": total_footprint,
    }


def get_bigtex_skus(search=None, mcat=None):
    sql = "SELECT * FROM bigtex_skus"
    params = []
    clauses = []
    if search:
        clauses.append("(model LIKE ? OR item_number LIKE ?)")
        params += [f"%{search}%", f"%{search}%"]
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    requested_mcat = normalize_bigtex_mcat(mcat) if mcat else ""
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()

    normalized_rows = []
    for row in rows:
        payload = dict(row)
        payload["mcat"] = normalize_bigtex_mcat(payload.get("mcat"))
        if requested_mcat and payload["mcat"] != requested_mcat:
            continue
        normalized_rows.append(payload)

    normalized_rows.sort(
        key=lambda r: (
            str(r.get("mcat") or ""),
            str(r.get("model") or ""),
            _coerce_float(r.get("bed_length"), 0.0) or 0.0,
            str(r.get("item_number") or ""),
        )
    )
    return normalized_rows


def get_bigtex_sku(item_number):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM bigtex_skus WHERE item_number=?", (item_number,)).fetchone()
    if not row:
        return None
    payload = dict(row)
    payload["mcat"] = normalize_bigtex_mcat(payload.get("mcat"))
    return payload


def update_bigtex_sku_field(item_number, field, value):
    if field == "mcat":
        normalized = normalize_bigtex_mcat(value)
        value = normalized or None
    with get_db() as conn:
        conn.execute(
            f"UPDATE bigtex_skus SET {field}=?, updated_at=? WHERE item_number=?",
            (value, datetime.utcnow().isoformat(), item_number)
        )


def recompute_bigtex_footprint(item_number):
    row = get_bigtex_sku(item_number)
    if not row:
        return None
    bed_length = _coerce_float(row["bed_length"], 0.0) or 0.0
    tongue = _coerce_float(row["tongue"], 0.0) or 0.0
    total_footprint = round(bed_length + tongue, 2)
    with get_db() as conn:
        conn.execute(
            "UPDATE bigtex_skus SET total_footprint=?, updated_at=? WHERE item_number=?",
            (total_footprint, datetime.utcnow().isoformat(), item_number),
        )
    return {
        "item_number": item_number,
        "bed_length": round(bed_length, 2),
        "tongue": round(tongue, 2),
        "total_footprint": total_footprint,
    }


def get_bt_inventory_upload_meta():
    with get_db() as conn:
        return conn.execute(
            """
            SELECT *
            FROM bt_inventory_upload_log
            ORDER BY uploaded_at DESC
            LIMIT 1
            """
        ).fetchone()


def _normalize_bt_whse_code(value):
    code = str(value or "").strip().upper()
    if not code or code == "ALL":
        return ""
    return code


def get_bt_inventory_whse_codes():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT whse_code
            FROM bt_inventory_snapshot_whse
            WHERE TRIM(COALESCE(whse_code, '')) <> ''
            ORDER BY whse_code ASC
            """
        ).fetchall()
    return [str(row["whse_code"]).strip().upper() for row in rows if str(row["whse_code"]).strip()]


def get_bt_inventory_snapshot_rows(limit=300, whse_code=None):
    try:
        row_limit = int(limit)
    except (TypeError, ValueError):
        row_limit = 300
    row_limit = max(1, min(row_limit, 2000))
    whse = _normalize_bt_whse_code(whse_code)
    with get_db() as conn:
        if whse:
            return conn.execute(
                """
                SELECT
                    inv.item_number,
                    inv.whse_code,
                    inv.total_count,
                    inv.available_count,
                    inv.assigned_count,
                    inv.built_count,
                    inv.future_build_count,
                    inv.available_built_count,
                    inv.available_future_count,
                    inv.updated_at,
                    sku.mcat AS sku_mcat,
                    sku.model AS sku_model,
                    sku.total_footprint AS sku_total_footprint
                FROM bt_inventory_snapshot_whse inv
                LEFT JOIN bigtex_skus sku
                    ON sku.item_number = inv.item_number
                WHERE inv.whse_code = ?
                ORDER BY inv.available_count DESC, inv.total_count DESC, inv.item_number ASC
                LIMIT ?
                """,
                (whse, row_limit),
            ).fetchall()
        return conn.execute(
            """
            SELECT
                inv.item_number,
                '' AS whse_code,
                inv.total_count,
                inv.available_count,
                inv.assigned_count,
                inv.built_count,
                inv.future_build_count,
                inv.available_built_count,
                inv.available_future_count,
                inv.updated_at,
                sku.mcat AS sku_mcat,
                sku.model AS sku_model,
                sku.total_footprint AS sku_total_footprint
            FROM bt_inventory_snapshot inv
            LEFT JOIN bigtex_skus sku
                ON sku.item_number = inv.item_number
            ORDER BY inv.available_count DESC, inv.total_count DESC, inv.item_number ASC
            LIMIT ?
            """,
            (row_limit,),
        ).fetchall()


def _clear_bt_inventory_snapshots(conn):
    conn.execute("DELETE FROM bt_inventory_snapshot")
    conn.execute("DELETE FROM bt_inventory_snapshot_whse")


def _insert_bt_inventory_snapshot_rows(conn, rows, *, include_whse=False):
    if include_whse:
        conn.executemany(
            """
            INSERT INTO bt_inventory_snapshot_whse
            (
                item_number,
                whse_code,
                total_count,
                available_count,
                assigned_count,
                built_count,
                future_build_count,
                available_built_count,
                available_future_count,
                updated_at
            )
            VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            rows,
        )
        return
    conn.executemany(
        """
        INSERT INTO bt_inventory_snapshot
        (
            item_number,
            total_count,
            available_count,
            assigned_count,
            built_count,
            future_build_count,
            available_built_count,
            available_future_count,
            updated_at
        )
        VALUES (?,?,?,?,?,?,?,?,?)
        """,
        rows,
    )


def _insert_bt_inventory_upload_log(
    conn,
    *,
    upload_id,
    source_filename,
    sheet_name,
    source_format,
    processed_rows,
    valid_rows,
    distinct_items,
    deduped_rows,
    duplicate_rows,
    warehouse_count,
    uploaded_at,
):
    conn.execute(
        """
        INSERT INTO bt_inventory_upload_log
        (
            upload_id,
            source_filename,
            sheet_name,
            source_format,
            processed_rows,
            valid_rows,
            distinct_items,
            deduped_rows,
            duplicate_rows,
            warehouse_count,
            uploaded_at
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            upload_id,
            source_filename,
            sheet_name,
            source_format,
            int(processed_rows),
            int(valid_rows),
            int(distinct_items),
            int(deduped_rows),
            int(duplicate_rows),
            int(warehouse_count),
            uploaded_at,
        ),
    )


def _bt_inventory_metric_template():
    return {
        "total_count": 0,
        "available_count": 0,
        "assigned_count": 0,
        "built_count": 0,
        "future_build_count": 0,
        "available_built_count": 0,
        "available_future_count": 0,
    }


def _bt_parse_inventory_timestamp(raw_value):
    value = str(raw_value or "").strip()
    if not value:
        return None
    for fmt in (
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%y %H:%M",
        "%m/%d/%y %H:%M:%S",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _coerce_nonnegative_int(value):
    parsed = _coerce_int(value, default=0)
    try:
        return max(int(parsed or 0), 0)
    except (TypeError, ValueError):
        return 0


def _open_csv_with_fallback(path):
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        handle = None
        try:
            handle = open(path, "r", newline="", encoding=encoding)
            handle.read(1024)
            handle.seek(0)
            return handle
        except UnicodeDecodeError as exc:
            last_error = exc
            if handle is not None:
                try:
                    handle.close()
                except OSError:
                    pass
    if last_error:
        raise last_error
    return open(path, "r", newline="", encoding="utf-8-sig")


def _import_bigtex_inventory_orders_workbook(source_path, sheet_name="All.Orders.Quick"):
    from openpyxl import load_workbook

    # Match existing ProGrade import pattern: use a local temp copy for stability.
    suffix = source_path.suffix if source_path.suffix else ".xlsx"
    temp_copy = Path(tempfile.gettempdir()) / f"prograde_bt_orders_import_{uuid.uuid4().hex}{suffix}"
    shutil.copyfile(source_path, temp_copy)
    workbook = None

    try:
        workbook = load_workbook(temp_copy, read_only=True, data_only=True)
        selected_sheet = _resolve_sheet_by_name(workbook, sheet_name)
        if selected_sheet is None:
            raise ValueError(
                f"Workbook sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
            )

        sheet = workbook[selected_sheet]
        item_col = 12   # M
        name_col = 2    # C
        days_old_col = 17  # R
        metrics = defaultdict(_bt_inventory_metric_template)
        processed_rows = 0
        valid_rows = 0

        for row in sheet.iter_rows(min_row=1, values_only=True):
            processed_rows += 1
            if not row:
                continue
            if item_col >= len(row):
                continue

            item_number = str(row[item_col] or "").strip().upper()
            if not item_number:
                continue
            normalized_item = _normalize_header(item_number)
            if normalized_item in {"item", "itemnumber", "itemnum", "itemno"}:
                continue

            has_name = not _is_blank_cell(row[name_col] if name_col < len(row) else None)
            has_days_old = not _is_blank_cell(row[days_old_col] if days_old_col < len(row) else None)

            entry = metrics[item_number]
            entry["total_count"] += 1
            valid_rows += 1

            if has_name:
                entry["assigned_count"] += 1
            else:
                entry["available_count"] += 1

            if has_days_old:
                entry["built_count"] += 1
            else:
                entry["future_build_count"] += 1

            if not has_name and has_days_old:
                entry["available_built_count"] += 1
            if not has_name and not has_days_old:
                entry["available_future_count"] += 1

        if not metrics:
            raise ValueError("No inventory rows parsed from All.Orders.Quick (item # column M).")

        now = datetime.utcnow().isoformat()
        upload_id = str(uuid.uuid4())
        with get_db() as conn:
            _clear_bt_inventory_snapshots(conn)
            _insert_bt_inventory_snapshot_rows(
                conn,
                [
                    (
                        item_number,
                        counts["total_count"],
                        counts["available_count"],
                        counts["assigned_count"],
                        counts["built_count"],
                        counts["future_build_count"],
                        counts["available_built_count"],
                        counts["available_future_count"],
                        now,
                    )
                    for item_number, counts in metrics.items()
                ],
                include_whse=False,
            )
            _insert_bt_inventory_upload_log(
                conn,
                upload_id=upload_id,
                source_filename=source_path.name,
                sheet_name=selected_sheet,
                source_format="workbook",
                processed_rows=processed_rows,
                valid_rows=valid_rows,
                distinct_items=len(metrics),
                deduped_rows=valid_rows,
                duplicate_rows=max(processed_rows - valid_rows, 0),
                warehouse_count=0,
                uploaded_at=now,
            )

        sku_item_numbers = {str(r["item_number"]).strip().upper() for r in get_bigtex_skus()}
        unmatched_items = [item for item in metrics.keys() if item not in sku_item_numbers]
        available_total = sum(int(v["available_count"]) for v in metrics.values())
        built_total = sum(int(v["built_count"]) for v in metrics.values())
        future_total = sum(int(v["future_build_count"]) for v in metrics.values())

        return {
            "source_filename": source_path.name,
            "source_format": "workbook",
            "sheet_name": selected_sheet,
            "processed_rows": int(processed_rows),
            "valid_rows": int(valid_rows),
            "deduped_rows": int(valid_rows),
            "duplicate_rows": int(max(processed_rows - valid_rows, 0)),
            "distinct_items": int(len(metrics)),
            "warehouse_count": 0,
            "available_total": int(available_total),
            "built_total": int(built_total),
            "future_build_total": int(future_total),
            "unmatched_item_count": int(len(unmatched_items)),
            "unmatched_items": sorted(unmatched_items)[:25],
        }
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        try:
            temp_copy.unlink(missing_ok=True)
        except OSError:
            pass


def _import_bigtex_inventory_csv_report(source_path):
    with _open_csv_with_fallback(source_path) as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV report is missing a header row.")
        field_map = {_normalize_header(field): field for field in reader.fieldnames}

        item_key = next(
            (
                field_map[k]
                for k in ("itemnum", "itemnumber", "itemno", "item", "sku")
                if k in field_map
            ),
            None,
        )
        whse_key = next((field_map[k] for k in ("whse", "warehouse", "warehousecode") if k in field_map), None)
        serid_key = next((field_map[k] for k in ("serid", "serialid", "serialnumber", "serial") if k in field_map), None)
        onhand_key = next((field_map[k] for k in ("onhand", "qtyonhand", "onhandqty") if k in field_map), None)
        committed_key = next(
            (
                field_map[k]
                for k in ("committed", "committedqty", "qtycommitted", "committedunits")
                if k in field_map
            ),
            None,
        )
        ts_key = next((field_map[k] for k in ("tslastupdated", "lastupdated", "updatedat") if k in field_map), None)

        if not item_key or not whse_key or not serid_key or not onhand_key or not committed_key:
            raise ValueError(
                "CSV report is missing required columns. Expected itemnum, whse, serid, onhand, committed_."
            )

        serial_rows = {}
        processed_rows = 0
        valid_rows = 0
        duplicate_rows = 0
        conflict_serial_rows = 0

        for row in reader:
            processed_rows += 1
            item_number = str(row.get(item_key) or "").strip().upper()
            whse_code = _normalize_bt_whse_code(row.get(whse_key))
            serial_id = str(row.get(serid_key) or "").strip()
            if not item_number or not whse_code or not serial_id:
                continue

            onhand = _coerce_nonnegative_int(row.get(onhand_key))
            committed = _coerce_nonnegative_int(row.get(committed_key))
            ts_raw = row.get(ts_key) if ts_key else None
            ts_value = _bt_parse_inventory_timestamp(ts_raw)
            valid_rows += 1

            incoming = {
                "item_number": item_number,
                "whse_code": whse_code,
                "onhand": onhand,
                "committed": committed,
                "ts_raw": str(ts_raw or "").strip(),
                "ts_value": ts_value,
            }
            existing = serial_rows.get(serial_id)
            if existing is None:
                serial_rows[serial_id] = incoming
                continue

            duplicate_rows += 1
            if (
                existing["item_number"] == incoming["item_number"]
                and existing["whse_code"] == incoming["whse_code"]
                and existing["onhand"] == incoming["onhand"]
                and existing["committed"] == incoming["committed"]
            ):
                continue

            conflict_serial_rows += 1
            existing_ts = existing.get("ts_value")
            incoming_ts = incoming.get("ts_value")
            if incoming_ts and (not existing_ts or incoming_ts >= existing_ts):
                serial_rows[serial_id] = incoming

        if not serial_rows:
            raise ValueError("No inventory rows parsed from CSV report.")

        metrics_by_item = defaultdict(_bt_inventory_metric_template)
        metrics_by_item_whse = defaultdict(_bt_inventory_metric_template)

        for record in serial_rows.values():
            item_number = record["item_number"]
            whse_code = record["whse_code"]
            onhand = int(record["onhand"])
            committed = int(record["committed"])
            available = onhand - committed

            aggregate = metrics_by_item[item_number]
            aggregate["total_count"] += onhand
            aggregate["available_count"] += available
            aggregate["assigned_count"] += committed

            whse_aggregate = metrics_by_item_whse[(item_number, whse_code)]
            whse_aggregate["total_count"] += onhand
            whse_aggregate["available_count"] += available
            whse_aggregate["assigned_count"] += committed

        now = datetime.utcnow().isoformat()
        upload_id = str(uuid.uuid4())
        with get_db() as conn:
            _clear_bt_inventory_snapshots(conn)
            _insert_bt_inventory_snapshot_rows(
                conn,
                [
                    (
                        item_number,
                        counts["total_count"],
                        counts["available_count"],
                        counts["assigned_count"],
                        0,
                        0,
                        0,
                        0,
                        now,
                    )
                    for item_number, counts in metrics_by_item.items()
                ],
                include_whse=False,
            )
            _insert_bt_inventory_snapshot_rows(
                conn,
                [
                    (
                        item_number,
                        whse_code,
                        counts["total_count"],
                        counts["available_count"],
                        counts["assigned_count"],
                        0,
                        0,
                        0,
                        0,
                        now,
                    )
                    for (item_number, whse_code), counts in metrics_by_item_whse.items()
                ],
                include_whse=True,
            )
            _insert_bt_inventory_upload_log(
                conn,
                upload_id=upload_id,
                source_filename=source_path.name,
                sheet_name="",
                source_format="csv_inventory",
                processed_rows=processed_rows,
                valid_rows=valid_rows,
                distinct_items=len(metrics_by_item),
                deduped_rows=len(serial_rows),
                duplicate_rows=duplicate_rows,
                warehouse_count=len({whse for (_, whse) in metrics_by_item_whse.keys()}),
                uploaded_at=now,
            )

        sku_item_numbers = {str(r["item_number"]).strip().upper() for r in get_bigtex_skus()}
        unmatched_items = [item for item in metrics_by_item.keys() if item not in sku_item_numbers]
        available_total = sum(int(v["available_count"]) for v in metrics_by_item.values())

        return {
            "source_filename": source_path.name,
            "source_format": "csv_inventory",
            "sheet_name": "",
            "processed_rows": int(processed_rows),
            "valid_rows": int(valid_rows),
            "deduped_rows": int(len(serial_rows)),
            "duplicate_rows": int(duplicate_rows),
            "conflict_serial_rows": int(conflict_serial_rows),
            "distinct_items": int(len(metrics_by_item)),
            "warehouse_count": int(len({whse for (_, whse) in metrics_by_item_whse.keys()})),
            "available_total": int(available_total),
            "built_total": 0,
            "future_build_total": 0,
            "unmatched_item_count": int(len(unmatched_items)),
            "unmatched_items": sorted(unmatched_items)[:25],
        }


def import_bigtex_inventory_orders_workbook(workbook_path, sheet_name="All.Orders.Quick"):
    source_path = Path(str(workbook_path))
    if not source_path.exists():
        raise FileNotFoundError(f"Inventory source file not found: {source_path}")

    ext = source_path.suffix.lower()
    if ext == ".csv":
        return _import_bigtex_inventory_csv_report(source_path)
    return _import_bigtex_inventory_orders_workbook(source_path, sheet_name=sheet_name)


def get_pj_inventory_upload_meta():
    with get_db() as conn:
        return conn.execute(
            """
            SELECT *
            FROM pj_inventory_upload_log
            ORDER BY uploaded_at DESC
            LIMIT 1
            """
        ).fetchone()


def _normalize_pj_inventory_whse_code(value):
    code = str(value or "").strip().upper()
    if not code or code == "ALL":
        return ""
    return code


def get_pj_inventory_whse_codes():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT whse_code
            FROM pj_inventory_snapshot_whse
            WHERE TRIM(COALESCE(whse_code, '')) <> ''
            ORDER BY whse_code ASC
            """
        ).fetchall()
    return [str(row["whse_code"]).strip().upper() for row in rows if str(row["whse_code"]).strip()]


def get_pj_inventory_snapshot_rows(limit=500, whse_code=None):
    try:
        row_limit = int(limit)
    except (TypeError, ValueError):
        row_limit = 500
    row_limit = max(1, min(row_limit, 3000))
    whse = _normalize_pj_inventory_whse_code(whse_code)
    with get_db() as conn:
        if whse:
            return conn.execute(
                """
                SELECT
                    inv.item_number,
                    inv.source_item_number,
                    inv.match_method,
                    inv.normalized_model,
                    inv.normalized_category,
                    inv.footprint_each,
                    inv.stack_height_each,
                    inv.whse_code,
                    inv.total_count,
                    inv.available_count,
                    inv.assigned_count,
                    inv.updated_at,
                    sku.model AS sku_model,
                    sku.pj_category AS sku_pj_category,
                    sku.total_footprint AS sku_total_footprint,
                    sku.tongue_feet AS sku_tongue_feet,
                    sku.dump_side_height_ft AS sku_dump_side_height_ft
                FROM pj_inventory_snapshot_whse inv
                LEFT JOIN pj_skus sku
                    ON sku.item_number = inv.item_number
                WHERE inv.whse_code = ?
                ORDER BY inv.available_count DESC, inv.total_count DESC, inv.item_number ASC
                LIMIT ?
                """,
                (whse, row_limit),
            ).fetchall()
        return conn.execute(
            """
            SELECT
                inv.item_number,
                inv.source_item_number,
                inv.match_method,
                inv.normalized_model,
                inv.normalized_category,
                inv.footprint_each,
                inv.stack_height_each,
                '' AS whse_code,
                inv.total_count,
                inv.available_count,
                inv.assigned_count,
                inv.updated_at,
                sku.model AS sku_model,
                sku.pj_category AS sku_pj_category,
                sku.total_footprint AS sku_total_footprint,
                sku.tongue_feet AS sku_tongue_feet,
                sku.dump_side_height_ft AS sku_dump_side_height_ft
            FROM pj_inventory_snapshot inv
            LEFT JOIN pj_skus sku
                ON sku.item_number = inv.item_number
            ORDER BY inv.available_count DESC, inv.total_count DESC, inv.item_number ASC
            LIMIT ?
            """,
            (row_limit,),
        ).fetchall()


def _clear_pj_inventory_snapshots(conn):
    conn.execute("DELETE FROM pj_inventory_snapshot")
    conn.execute("DELETE FROM pj_inventory_snapshot_whse")


def _insert_pj_inventory_snapshot_rows(conn, rows, *, include_whse=False):
    if include_whse:
        conn.executemany(
            """
            INSERT INTO pj_inventory_snapshot_whse
            (
                item_number,
                whse_code,
                source_item_number,
                match_method,
                normalized_model,
                normalized_category,
                footprint_each,
                stack_height_each,
                total_count,
                available_count,
                assigned_count,
                updated_at
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            rows,
        )
        return
    conn.executemany(
        """
        INSERT INTO pj_inventory_snapshot
        (
            item_number,
            source_item_number,
            match_method,
            normalized_model,
            normalized_category,
            footprint_each,
            stack_height_each,
            total_count,
            available_count,
            assigned_count,
            updated_at
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """,
        rows,
    )


def _insert_pj_inventory_upload_log(
    conn,
    *,
    upload_id,
    source_filename,
    source_format,
    processed_rows,
    valid_rows,
    deduped_rows,
    duplicate_rows,
    distinct_items,
    warehouse_count,
    matched_rows,
    matched_items,
    unmatched_items,
    uploaded_at,
):
    conn.execute(
        """
        INSERT INTO pj_inventory_upload_log
        (
            upload_id,
            source_filename,
            source_format,
            processed_rows,
            valid_rows,
            deduped_rows,
            duplicate_rows,
            distinct_items,
            warehouse_count,
            matched_rows,
            matched_items,
            unmatched_items,
            uploaded_at
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            upload_id,
            source_filename,
            source_format,
            int(processed_rows),
            int(valid_rows),
            int(deduped_rows),
            int(duplicate_rows),
            int(distinct_items),
            int(warehouse_count),
            int(matched_rows),
            int(matched_items),
            int(unmatched_items),
            uploaded_at,
        ),
    )


def _alnum_upper(value):
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _pj_inventory_stack_height_for_category(category, dump_side_height_ft=None):
    category_key = str(category or "").strip().lower()
    if category_key == "utility":
        return 2.0
    if category_key in PJ_DUMP_CATEGORY_KEYS:
        dump_h = _coerce_float(dump_side_height_ft, None)
        if dump_h is None:
            dump_h = _infer_pj_dump_wall_height_ft(category_key)
        return _dump_stacked_height_from_wall_height(dump_h) or 4.0
    if category_key in PJ_GOOSENECK_CATEGORY_KEYS:
        return 2.5
    if category_key in {"car_hauler", "car_hauler_deckover", "deck_over", "tilt", "tilt_deckover"}:
        return 2.5
    if category_key:
        return 2.0
    return 2.0


def _pj_inventory_tongue_profile_for_model_category(model_code, category):
    category_key = str(category or "").strip().lower()
    if category_key in PJ_GOOSENECK_CATEGORY_KEYS:
        return "gooseneck"
    model_norm = _alnum_upper(model_code)
    if model_norm[:2] in {"LD", "LQ", "LS", "LX", "LY", "PL"}:
        return "gooseneck"
    return "standard"


def _pj_inventory_default_override_reason(category, tongue_profile, dump_side_height_ft=None):
    tokens = [f"tongue_profile:{tongue_profile}"]
    category_key = str(category or "").strip().lower()
    if category_key in PJ_DUMP_CATEGORY_KEYS:
        dump_h = _coerce_float(dump_side_height_ft, None)
        if dump_h is not None:
            tokens.append(f"dump_height_ft:{dump_h:.1f}")
    return ";".join(tokens)


def _pj_inventory_similarity_score(raw_core, candidate_item_number):
    raw = _alnum_upper(raw_core)
    candidate = _alnum_upper(candidate_item_number)
    prefix = 0
    for raw_ch, candidate_ch in zip(raw, candidate):
        if raw_ch != candidate_ch:
            break
        prefix += 1
    length_penalty = abs(len(raw) - len(candidate))
    return (prefix * 10) - length_penalty


def _pj_inventory_pick_best_candidate(raw_core, candidates):
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    ordered = sorted(
        candidates,
        key=lambda row: (
            -_pj_inventory_similarity_score(raw_core, row.get("item_number")),
            abs(
                _coerce_float(row.get("total_footprint"), 0.0)
                - _coerce_float(_parse_bed_length_from_item_number(raw_core, row.get("model")), 0.0)
            ),
            str(row.get("item_number") or ""),
        ),
    )
    return ordered[0]


def _build_pj_inventory_match_indexes():
    sku_rows = [dict(row) for row in get_pj_skus()]
    exact = {}
    models = set()
    by_model_len = defaultdict(list)
    by_item_code = defaultdict(list)
    by_model = defaultdict(list)
    model_tongues = defaultdict(list)
    model_categories = defaultdict(lambda: defaultdict(int))

    for row in sku_rows:
        item_number = _alnum_upper(row.get("item_number"))
        model = _alnum_upper(row.get("model"))
        category = str(row.get("pj_category") or "").strip().lower()
        if not item_number:
            continue
        exact[item_number] = row
        if model:
            models.add(model)
            by_model[model].append(row)
            model_categories[model][category] += 1
            tongue = _coerce_float(row.get("tongue_feet"), None)
            if tongue is not None and tongue > 0:
                model_tongues[model].append(float(tongue))
        bed_length = _coerce_int(row.get("bed_length_measured"), None)
        if bed_length is None:
            bed_length = _coerce_int(row.get("bed_length_stated"), None)
        if model and bed_length is not None:
            key = (model, int(bed_length))
            by_model_len[key].append(row)
            by_item_code[f"{model}{int(bed_length):02d}"].append(row)

    dominant_category_by_model = {}
    median_tongue_by_model = {}
    for model, counts in model_categories.items():
        dominant_category_by_model[model] = sorted(
            counts.items(),
            key=lambda item: (-int(item[1]), str(item[0] or "")),
        )[0][0]
    for model, values in model_tongues.items():
        ordered = sorted(values)
        if not ordered:
            continue
        idx = len(ordered) // 2
        if len(ordered) % 2:
            median_tongue_by_model[model] = float(ordered[idx])
        else:
            median_tongue_by_model[model] = float((ordered[idx - 1] + ordered[idx]) / 2.0)

    return {
        "exact": exact,
        "models": models,
        "by_model_len": by_model_len,
        "by_item_code": by_item_code,
        "by_model": by_model,
        "dominant_category_by_model": dominant_category_by_model,
        "median_tongue_by_model": median_tongue_by_model,
    }


def _normalize_pj_inventory_model_code(raw_itemid, known_models):
    model = _alnum_upper(raw_itemid)
    if not model:
        return ""
    if model in known_models:
        return model
    trimmed = model
    while len(trimmed) > 1 and trimmed not in known_models and trimmed[-1].isdigit():
        trimmed = trimmed[:-1]
    if trimmed in known_models:
        return trimmed
    if len(model) >= 2 and model[:2] in known_models:
        return model[:2]
    return trimmed


def _infer_pj_inventory_category(raw_sku, raw_itemid, normalized_model):
    model = _alnum_upper(normalized_model or raw_itemid)
    raw = _alnum_upper(raw_sku)
    model_prefix = model[:2]

    if model_prefix in {"LD", "LQ", "LS", "LX", "LY", "PL"} or model.startswith("GN"):
        return "gooseneck"
    if model_prefix in {"D5", "D7"}:
        return "dump_small"
    if model.startswith("D"):
        return "dump_variants"
    if model.startswith("T"):
        return "tilt"
    if model_prefix in {"U2", "U6", "U7", "U8", "UC", "UK", "UL"}:
        return "utility"
    if model_prefix in {"B5", "B6", "B8", "C4", "C5", "C8", "CC", "CD", "CE", "CH"}:
        return "car_hauler"
    if model.startswith("F"):
        return "deck_over"

    if raw.startswith("LQ") or raw.startswith("LS") or raw.startswith("LD"):
        return "gooseneck"
    if raw.startswith("D"):
        return "dump_variants"
    if raw.startswith("T"):
        return "tilt"
    return ""


def _resolve_pj_inventory_item(raw_sku, raw_itemid, indexes):
    raw_text = str(raw_sku or "").strip().upper()
    core_text = raw_text.split("-", 1)[0].strip().upper()
    raw_alnum = _alnum_upper(raw_text)
    core_alnum = _alnum_upper(core_text)
    exact = indexes["exact"]

    if raw_alnum and raw_alnum in exact:
        resolved = dict(exact[raw_alnum])
        return {"sku": resolved, "item_number": _alnum_upper(resolved.get("item_number")), "match_method": "exact"}
    if core_alnum and core_alnum in exact:
        resolved = dict(exact[core_alnum])
        return {"sku": resolved, "item_number": _alnum_upper(resolved.get("item_number")), "match_method": "core"}

    known_models = indexes["models"]
    model_code = _normalize_pj_inventory_model_code(raw_itemid, known_models)
    bed_length = _parse_bed_length_from_item_number(core_alnum or raw_alnum, model_code or raw_itemid)

    if model_code and bed_length is not None:
        length_key = (model_code, int(round(float(bed_length))))
        candidates = list(indexes["by_model_len"].get(length_key) or [])
        if candidates:
            resolved = dict(_pj_inventory_pick_best_candidate(core_alnum, candidates))
            return {"sku": resolved, "item_number": _alnum_upper(resolved.get("item_number")), "match_method": "model_length"}

        item_code_key = f"{model_code}{int(round(float(bed_length))):02d}"
        item_code_candidates = list(indexes["by_item_code"].get(item_code_key) or [])
        if item_code_candidates:
            resolved = dict(_pj_inventory_pick_best_candidate(core_alnum, item_code_candidates))
            return {"sku": resolved, "item_number": _alnum_upper(resolved.get("item_number")), "match_method": "item_code"}

    if model_code:
        model_candidates = list(indexes["by_model"].get(model_code) or [])
        if model_candidates:
            resolved = dict(_pj_inventory_pick_best_candidate(core_alnum, model_candidates))
            return {"sku": resolved, "item_number": _alnum_upper(resolved.get("item_number")), "match_method": "model_similarity"}

    normalized_key = core_alnum or raw_alnum
    if not normalized_key:
        normalized_key = f"UNMAPPED-{uuid.uuid4().hex[:8].upper()}"
    return {"sku": None, "item_number": normalized_key, "match_method": "unmapped"}


def _import_pj_inventory_csv_report(source_path):
    with _open_csv_with_fallback(source_path) as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV report is missing a header row.")
        field_map = {_normalize_header(field): field for field in reader.fieldnames}

        sku_key = next(
            (
                field_map[k]
                for k in (
                    "hstrailerconfiglongitemid",
                    "trailerconfiglongitemid",
                    "itemnum",
                    "itemnumber",
                    "item",
                    "sku",
                )
                if k in field_map
            ),
            None,
        )
        itemid_key = next((field_map[k] for k in ("itemid", "model", "itemprefix") if k in field_map), None)
        whse_key = next((field_map[k] for k in ("inventsiteid", "site", "warehouse", "warehousecode") if k in field_map), None)
        row_id_key = next((field_map[k] for k in ("id", "recid", "serialid", "serialnumber") if k in field_map), None)
        avail_key = next((field_map[k] for k in ("availphysical", "available", "availableqty") if k in field_map), None)
        reserved_key = next((field_map[k] for k in ("reservphysical", "reserved", "reservedqty") if k in field_map), None)
        total_key = next((field_map[k] for k in ("physicalinvent", "onhand", "qtyonhand") if k in field_map), None)

        if not sku_key or not whse_key:
            raise ValueError(
                "CSV report is missing required columns. Expected hstrailerconfiglongitemid and inventsiteid."
            )

        indexes = _build_pj_inventory_match_indexes()
        seen_keys = set()
        processed_rows = 0
        valid_rows = 0
        duplicate_rows = 0
        matched_rows = 0

        metrics_by_item = {}
        metrics_by_item_whse = {}

        def _ensure_entry(store, key, *, resolved_item_number, raw_item_number, match_method, model, category, footprint, stack_height):
            if key in store:
                return store[key]
            store[key] = {
                "item_number": resolved_item_number,
                "source_item_number": raw_item_number,
                "match_method": match_method,
                "normalized_model": model,
                "normalized_category": category,
                "footprint_each": float(footprint or 0.0),
                "stack_height_each": float(stack_height or 0.0),
                "total_count": 0,
                "available_count": 0,
                "assigned_count": 0,
            }
            return store[key]

        for row in reader:
            processed_rows += 1
            raw_sku = str(row.get(sku_key) or "").strip().upper()
            raw_itemid = str(row.get(itemid_key) or "").strip().upper() if itemid_key else ""
            whse_code = _normalize_pj_inventory_whse_code(row.get(whse_key))
            if not raw_sku or not whse_code:
                continue

            dedupe_raw = str(row.get(row_id_key) or "").strip() if row_id_key else ""
            if dedupe_raw:
                dedupe_key = f"ID:{dedupe_raw}"
            else:
                dedupe_key = "|".join(
                    [
                        _alnum_upper(raw_sku),
                        _alnum_upper(raw_itemid),
                        whse_code,
                        str(row.get(field_map.get("inventlocationid", ""), "") or "").strip().upper(),
                        str(row.get(field_map.get("wmslocationid", ""), "") or "").strip().upper(),
                    ]
                )
            if dedupe_key in seen_keys:
                duplicate_rows += 1
                continue
            seen_keys.add(dedupe_key)
            valid_rows += 1

            available_raw = _coerce_int(row.get(avail_key), None) if avail_key else None
            reserved_raw = _coerce_int(row.get(reserved_key), None) if reserved_key else None
            total_raw = _coerce_int(row.get(total_key), None) if total_key else None

            available_count = max(int(available_raw), 0) if available_raw is not None else None
            reserved_count = max(int(reserved_raw), 0) if reserved_raw is not None else 0
            total_count = max(int(total_raw), 0) if total_raw is not None else None
            if available_count is None:
                if total_count is not None:
                    available_count = max(total_count - reserved_count, 0)
                else:
                    available_count = 1
            if total_count is None:
                total_count = max(available_count + reserved_count, available_count)
            assigned_count = max(total_count - available_count, 0)

            resolved = _resolve_pj_inventory_item(raw_sku, raw_itemid, indexes)
            resolved_sku = dict(resolved.get("sku") or {})
            resolved_item_number = str(resolved.get("item_number") or "").strip().upper()
            if not resolved_item_number:
                continue
            match_method = str(resolved.get("match_method") or "unmapped").strip().lower()
            if match_method != "unmapped":
                matched_rows += 1

            normalized_model = _alnum_upper(resolved_sku.get("model"))
            if not normalized_model:
                normalized_model = _normalize_pj_inventory_model_code(raw_itemid, indexes["models"])
            normalized_category = str(resolved_sku.get("pj_category") or "").strip().lower()
            if not normalized_category and normalized_model:
                normalized_category = str(indexes["dominant_category_by_model"].get(normalized_model) or "").strip().lower()
            if not normalized_category:
                normalized_category = _infer_pj_inventory_category(raw_sku, raw_itemid, normalized_model)

            tongue_profile = _pj_inventory_tongue_profile_for_model_category(normalized_model, normalized_category)
            footprint_each = _coerce_float(resolved_sku.get("total_footprint"), None)
            if footprint_each is None:
                raw_core_alnum = _alnum_upper(str(raw_sku).split("-", 1)[0])
                bed_length = _parse_bed_length_from_item_number(raw_core_alnum, normalized_model or raw_itemid)
                model_tongue = _coerce_float(indexes["median_tongue_by_model"].get(normalized_model), None)
                if model_tongue is None:
                    model_tongue = 9.0 if tongue_profile == "gooseneck" else 4.5
                if bed_length is not None:
                    footprint_each = float(max(float(bed_length) + float(model_tongue), 0.0))
            if footprint_each is None:
                footprint_each = 0.0
            stack_height_each = _coerce_float(
                _pj_inventory_stack_height_for_category(
                    normalized_category,
                    resolved_sku.get("dump_side_height_ft"),
                ),
                0.0,
            )

            raw_core = str(raw_sku or "").split("-", 1)[0].strip().upper()
            source_item_number = raw_core or raw_sku

            agg = _ensure_entry(
                metrics_by_item,
                resolved_item_number,
                resolved_item_number=resolved_item_number,
                raw_item_number=source_item_number,
                match_method=match_method,
                model=normalized_model,
                category=normalized_category,
                footprint=footprint_each,
                stack_height=stack_height_each,
            )
            agg["total_count"] += int(total_count)
            agg["available_count"] += int(available_count)
            agg["assigned_count"] += int(assigned_count)

            whse_key_value = (resolved_item_number, whse_code)
            agg_whse = _ensure_entry(
                metrics_by_item_whse,
                whse_key_value,
                resolved_item_number=resolved_item_number,
                raw_item_number=source_item_number,
                match_method=match_method,
                model=normalized_model,
                category=normalized_category,
                footprint=footprint_each,
                stack_height=stack_height_each,
            )
            agg_whse["total_count"] += int(total_count)
            agg_whse["available_count"] += int(available_count)
            agg_whse["assigned_count"] += int(assigned_count)

        if not metrics_by_item:
            raise ValueError("No inventory rows parsed from CSV report.")

        now = datetime.utcnow().isoformat()
        upload_id = str(uuid.uuid4())
        with get_db() as conn:
            _clear_pj_inventory_snapshots(conn)
            _insert_pj_inventory_snapshot_rows(
                conn,
                [
                    (
                        row["item_number"],
                        row["source_item_number"],
                        row["match_method"],
                        row["normalized_model"],
                        row["normalized_category"],
                        round(_coerce_float(row["footprint_each"], 0.0), 2),
                        round(_coerce_float(row["stack_height_each"], 0.0), 2),
                        int(row["total_count"]),
                        int(row["available_count"]),
                        int(row["assigned_count"]),
                        now,
                    )
                    for row in metrics_by_item.values()
                ],
                include_whse=False,
            )
            _insert_pj_inventory_snapshot_rows(
                conn,
                [
                    (
                        row["item_number"],
                        whse_code,
                        row["source_item_number"],
                        row["match_method"],
                        row["normalized_model"],
                        row["normalized_category"],
                        round(_coerce_float(row["footprint_each"], 0.0), 2),
                        round(_coerce_float(row["stack_height_each"], 0.0), 2),
                        int(row["total_count"]),
                        int(row["available_count"]),
                        int(row["assigned_count"]),
                        now,
                    )
                    for (_, whse_code), row in metrics_by_item_whse.items()
                ],
                include_whse=True,
            )
            matched_items = sum(1 for row in metrics_by_item.values() if str(row.get("match_method")) != "unmapped")
            unmatched_items = sum(1 for row in metrics_by_item.values() if str(row.get("match_method")) == "unmapped")
            _insert_pj_inventory_upload_log(
                conn,
                upload_id=upload_id,
                source_filename=source_path.name,
                source_format="pj_csv_inventory",
                processed_rows=processed_rows,
                valid_rows=valid_rows,
                deduped_rows=len(seen_keys),
                duplicate_rows=duplicate_rows,
                distinct_items=len(metrics_by_item),
                warehouse_count=len({whse for (_, whse) in metrics_by_item_whse.keys()}),
                matched_rows=matched_rows,
                matched_items=matched_items,
                unmatched_items=unmatched_items,
                uploaded_at=now,
            )

        available_total = sum(int(row["available_count"]) for row in metrics_by_item.values())
        unmatched_items_all = sorted(
            [row["item_number"] for row in metrics_by_item.values() if str(row.get("match_method")) == "unmapped"]
        )
        unmatched_list = unmatched_items_all[:25]
        matched_item_count = sum(1 for row in metrics_by_item.values() if str(row.get("match_method")) != "unmapped")
        matched_pct = (matched_rows / valid_rows) if valid_rows else 0.0
        return {
            "source_filename": source_path.name,
            "source_format": "pj_csv_inventory",
            "processed_rows": int(processed_rows),
            "valid_rows": int(valid_rows),
            "deduped_rows": int(len(seen_keys)),
            "duplicate_rows": int(duplicate_rows),
            "distinct_items": int(len(metrics_by_item)),
            "warehouse_count": int(len({whse for (_, whse) in metrics_by_item_whse.keys()})),
            "available_total": int(available_total),
            "matched_rows": int(matched_rows),
            "matched_item_count": int(matched_item_count),
            "matched_ratio": round(float(matched_pct), 6),
            "unmatched_item_count": int(len(unmatched_items_all)),
            "unmatched_items": unmatched_list,
        }


def import_pj_inventory_report(workbook_path):
    source_path = Path(str(workbook_path))
    if not source_path.exists():
        raise FileNotFoundError(f"Inventory source file not found: {source_path}")
    if source_path.suffix.lower() != ".csv":
        raise ValueError("PJ inventory upload requires a CSV report.")
    return _import_pj_inventory_csv_report(source_path)


def _resolve_sheet_by_name(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        return sheet_name
    expected = str(sheet_name or "").strip().lower()
    for candidate in workbook.sheetnames:
        if str(candidate).strip().lower() == expected:
            return candidate
    return None


def _is_blank_cell(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False

def import_bigtex_skus_from_workbook(workbook_path=None, sheet_name="Data"):
    source_path = get_bigtex_workbook_path(workbook_path)
    if not source_path:
        raise FileNotFoundError("Big Tex workbook not found. Set PROGRADE_BT_DATA_WORKBOOK_PATH or provide a valid path.")

    # OneDrive files can be reparse points. Read from a local temp copy for reliable workbook access.
    temp_copy = Path(tempfile.gettempdir()) / f"prograde_bigtex_import_{uuid.uuid4().hex}.xlsx"
    shutil.copyfile(source_path, temp_copy)

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(temp_copy, read_only=True, data_only=True)
        selected_sheet = None
        if sheet_name in workbook.sheetnames:
            selected_sheet = sheet_name
        else:
            for name in workbook.sheetnames:
                if str(name).strip().lower() == "data":
                    selected_sheet = name
                    break
        if selected_sheet is None:
            selected_sheet = workbook.sheetnames[0]

        sheet = workbook[selected_sheet]
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise ValueError("Workbook sheet has no header row.")

        header_map = {idx: _normalize_header(value) for idx, value in enumerate(header_row)}
        aliases = {
            "mcat": {"mcat", "category", "loadcategory"},
            "tier": {"tier"},
            "model": {"model"},
            "item_number": {"itemnumber", "item", "itemno", "itemnum", "sku", "itemid"},
            "gvwr": {"gvwr"},
            "floor_type": {"floortype", "floor"},
            "bed_length": {"bedlength", "bed", "bedlen", "decklength"},
            "width": {"width"},
            "tongue": {"tongue", "tonguelength", "tongueft"},
            "stack_height": {"stackheight", "stackht", "height"},
        }

        def _find_col(key):
            valid = aliases[key]
            for idx, name in header_map.items():
                if name in valid:
                    return idx
            return None

        required = {"item_number", "bed_length", "tongue"}
        col_idx = {key: _find_col(key) for key in aliases}
        missing = [key for key in required if col_idx.get(key) is None]
        if missing:
            raise ValueError(f"Data sheet missing required columns: {', '.join(missing)}")

        parsed = {}
        for row in rows_iter:
            if not row:
                continue
            item_raw = row[col_idx["item_number"]] if col_idx["item_number"] is not None and col_idx["item_number"] < len(row) else None
            item_number = str(item_raw or "").strip()
            if not item_number:
                continue
            bed_length_raw = row[col_idx["bed_length"]] if col_idx["bed_length"] is not None and col_idx["bed_length"] < len(row) else None
            tongue_raw = row[col_idx["tongue"]] if col_idx["tongue"] is not None and col_idx["tongue"] < len(row) else None
            bed_length = _coerce_float(bed_length_raw, 0.0) or 0.0
            tongue = _coerce_float(tongue_raw, 0.0) or 0.0
            total_footprint = round(bed_length + tongue, 2)

            mcat_raw = row[col_idx["mcat"]] if col_idx["mcat"] is not None and col_idx["mcat"] < len(row) else None
            tier_raw = row[col_idx["tier"]] if col_idx["tier"] is not None and col_idx["tier"] < len(row) else None
            model_raw = row[col_idx["model"]] if col_idx["model"] is not None and col_idx["model"] < len(row) else None
            gvwr_raw = row[col_idx["gvwr"]] if col_idx["gvwr"] is not None and col_idx["gvwr"] < len(row) else None
            floor_raw = row[col_idx["floor_type"]] if col_idx["floor_type"] is not None and col_idx["floor_type"] < len(row) else None
            width_raw = row[col_idx["width"]] if col_idx["width"] is not None and col_idx["width"] < len(row) else None
            stack_height_raw = row[col_idx["stack_height"]] if col_idx["stack_height"] is not None and col_idx["stack_height"] < len(row) else None

            parsed[item_number] = (
                item_number,
                normalize_bigtex_mcat(mcat_raw) if mcat_raw is not None else None,
                _coerce_int(tier_raw, None),
                str(model_raw).strip() if model_raw is not None else None,
                _coerce_int(gvwr_raw, None),
                str(floor_raw).strip() if floor_raw is not None else None,
                round(bed_length, 2),
                _coerce_float(width_raw, None),
                round(tongue, 2),
                _coerce_float(stack_height_raw, None),
                total_footprint,
                datetime.utcnow().isoformat(),
            )

        if not parsed:
            raise ValueError("No Big Tex rows were parsed from workbook Data tab.")

        with get_db() as conn:
            before_count = int(conn.execute("SELECT COUNT(*) FROM bigtex_skus").fetchone()[0] or 0)
            conn.executemany(
                """
                INSERT INTO bigtex_skus
                (item_number, mcat, tier, model, gvwr, floor_type, bed_length, width, tongue, stack_height, total_footprint, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(item_number) DO UPDATE SET
                    mcat = excluded.mcat,
                    tier = excluded.tier,
                    model = excluded.model,
                    gvwr = excluded.gvwr,
                    floor_type = excluded.floor_type,
                    bed_length = excluded.bed_length,
                    width = excluded.width,
                    tongue = excluded.tongue,
                    stack_height = excluded.stack_height,
                    total_footprint = excluded.total_footprint,
                    updated_at = excluded.updated_at
                """,
                list(parsed.values()),
            )
            after_count = int(conn.execute("SELECT COUNT(*) FROM bigtex_skus").fetchone()[0] or 0)
        created_count = max(after_count - before_count, 0)

        return {
            "source_path": str(source_path),
            "sheet_name": selected_sheet,
            "row_count": len(parsed),
            "created_count": created_count,
            "updated_count": max(len(parsed) - created_count, 0),
            "total_row_count": after_count,
        }
    finally:
        try:
            temp_copy.unlink(missing_ok=True)
        except OSError:
            pass


# ── BT stack configs ─────────────────────────────────────────────────────────

def import_pj_skus_from_workbook(workbook_path=None, toc_sheet_name="ToC"):
    source_path = get_pj_workbook_path(workbook_path)
    if not source_path:
        raise FileNotFoundError("PJ workbook not found. Set PROGRADE_PJ_DATA_WORKBOOK_PATH or provide a valid path.")

    # OneDrive files can be reparse points. Read from a local temp copy for reliable workbook access.
    temp_copy = Path(tempfile.gettempdir()) / f"prograde_pj_import_{uuid.uuid4().hex}.xlsx"
    shutil.copyfile(source_path, temp_copy)

    try:
        from openpyxl import load_workbook
        from .services import pj_measurement

        workbook = load_workbook(temp_copy, read_only=True, data_only=True)
        toc_payload = _parse_pj_toc_models(workbook, toc_sheet_name=toc_sheet_name)
        toc_models = toc_payload["models"]
        skipped_codes = toc_payload["skipped_codes"]
        if not toc_models:
            raise ValueError("No 2-character model codes found in PJ ToC sheet.")

        offsets = get_pj_offsets_dict()
        parsed_rows = {}
        items_missing_bed_length = []
        toc_models_missing_sheet = []
        duplicate_items = []

        for entry in toc_models:
            model_code = entry["model"]
            section_label = entry["section"]
            sheet_name = entry["sheet_name"]

            if sheet_name not in workbook.sheetnames:
                toc_models_missing_sheet.append(
                    {
                        "model": model_code,
                        "section": section_label,
                        "sheet_name": sheet_name,
                    }
                )
                continue

            sheet = workbook[sheet_name]
            rows = _extract_pj_sheet_code_rows(sheet)
            if not rows:
                toc_models_missing_sheet.append(
                    {
                        "model": model_code,
                        "section": section_label,
                        "sheet_name": sheet_name,
                        "reason": "missing Code/Description table",
                    }
                )
                continue

            pj_category = _infer_pj_category(section_label, model_code)
            for row in rows:
                item_number = str(row["item_number"]).strip().upper()
                if not item_number:
                    continue
                description = str(row.get("description") or "").strip()
                bed_stated = _parse_bed_length_ft(description)
                if bed_stated is None:
                    bed_stated = _parse_bed_length_from_item_number(item_number, model_code)
                bed_from_item_code = _bed_length_from_pj_item_code(
                    model_code,
                    item_number,
                    pj_category,
                    fallback=bed_stated,
                )
                if bed_from_item_code is not None:
                    bed_stated = bed_from_item_code
                if bed_stated is None:
                    items_missing_bed_length.append(
                        {
                            "item_number": item_number,
                            "model": model_code,
                            "section": section_label,
                            "description": description,
                        }
                    )
                    bed_stated = 0.0
                tongue_profile = _default_pj_tongue_profile(model_code, pj_category)
                tongue_group = tongue_profile
                tongue_feet = _default_pj_tongue_feet(model_code, pj_category)
                dump_side_height_ft = _infer_pj_dump_wall_height_ft(pj_category, model_code)
                height_mid_ft, height_top_ft = _default_pj_stack_heights(
                    model_code,
                    pj_category,
                    dump_side_height_ft=dump_side_height_ft,
                )

                sku_for_calc = {
                    "model": model_code,
                    "bed_length_stated": round(bed_stated, 2),
                    "pj_category": pj_category,
                    "tongue_feet": round(tongue_feet, 2),
                }
                measured = pj_measurement.recompute_sku(sku_for_calc, offsets)

                if item_number in parsed_rows:
                    duplicate_items.append(
                        {
                            "item_number": item_number,
                            "existing_model": parsed_rows[item_number][1],
                            "new_model": model_code,
                        }
                    )

                parsed_rows[item_number] = (
                    item_number,
                    model_code,
                    pj_category,
                    description,
                    None,
                    round(bed_stated, 2),
                    measured["bed_length_measured"],
                    tongue_group,
                    round(tongue_feet, 2),
                    round(float(height_mid_ft), 2),
                    round(float(height_top_ft), 2),
                    measured["total_footprint"],
                    round(float(dump_side_height_ft), 2) if dump_side_height_ft is not None else None,
                    0,
                    0,
                    0,
                    None,
                    None,
                    datetime.utcnow().isoformat(),
                )

        if not parsed_rows:
            raise ValueError("No PJ rows were parsed from workbook tabs listed in ToC.")

        with get_db() as conn:
            before_count = int(conn.execute("SELECT COUNT(*) FROM pj_skus").fetchone()[0] or 0)
            conn.executemany(
                """
                INSERT INTO pj_skus
                (item_number, model, pj_category, description, gvwr, bed_length_stated, bed_length_measured, tongue_group, tongue_feet, height_mid_ft, height_top_ft, total_footprint, dump_side_height_ft, can_nest_inside_dump, gn_axle_droppable, tongue_overlap_allowed, pairing_rule, notes, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(item_number) DO UPDATE SET
                    model = excluded.model,
                    pj_category = excluded.pj_category,
                    description = excluded.description,
                    gvwr = excluded.gvwr,
                    bed_length_stated = excluded.bed_length_stated,
                    bed_length_measured = excluded.bed_length_measured,
                    tongue_group = excluded.tongue_group,
                    tongue_feet = excluded.tongue_feet,
                    height_mid_ft = excluded.height_mid_ft,
                    height_top_ft = excluded.height_top_ft,
                    total_footprint = excluded.total_footprint,
                    dump_side_height_ft = excluded.dump_side_height_ft,
                    can_nest_inside_dump = excluded.can_nest_inside_dump,
                    gn_axle_droppable = excluded.gn_axle_droppable,
                    tongue_overlap_allowed = excluded.tongue_overlap_allowed,
                    pairing_rule = excluded.pairing_rule,
                    notes = excluded.notes,
                    updated_at = excluded.updated_at
                """,
                list(parsed_rows.values()),
            )
            after_count = int(conn.execute("SELECT COUNT(*) FROM pj_skus").fetchone()[0] or 0)
        created_count = max(after_count - before_count, 0)

        disconnects = {
            "models_without_direct_tongue_group": [],
            "items_missing_bed_length": items_missing_bed_length,
            "toc_models_missing_sheet_or_table": toc_models_missing_sheet,
            "duplicate_item_numbers": duplicate_items,
            "toc_skipped_non_two_char_codes": skipped_codes,
        }
        disconnect_counts = {k: len(v) for k, v in disconnects.items()}

        return {
            "source_path": str(source_path),
            "toc_sheet_name": toc_payload["toc_sheet_name"],
            "row_count": len(parsed_rows),
            "created_count": created_count,
            "updated_count": max(len(parsed_rows) - created_count, 0),
            "total_row_count": after_count,
            "model_count": len(toc_models),
            "disconnect_counts": disconnect_counts,
            "disconnects": disconnects,
        }
    finally:
        try:
            temp_copy.unlink(missing_ok=True)
        except OSError:
            pass


def get_bt_stack_configs():
    with get_db() as conn:
        return conn.execute("SELECT * FROM bt_stack_configs ORDER BY load_type, stack_position").fetchall()

def update_bt_stack_config(config_id, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE bt_stack_configs SET {field}=?, updated_at=? WHERE config_id=?",
            (value, datetime.utcnow().isoformat(), config_id)
        )


# ── Load sessions ─────────────────────────────────────────────────────────────

def list_access_profiles():
    with get_db() as conn:
        return conn.execute(
            """
            SELECT id, name, is_admin, created_at, updated_at
            FROM prograde_access_profiles
            ORDER BY lower(name), id
            """
        ).fetchall()


def get_access_profile(profile_id):
    try:
        normalized = int(profile_id)
    except (TypeError, ValueError):
        return None
    with get_db() as conn:
        return conn.execute(
            """
            SELECT id, name, is_admin, created_at, updated_at
            FROM prograde_access_profiles
            WHERE id=?
            """,
            (normalized,),
        ).fetchone()


def get_access_profile_by_name(name):
    normalized = _normalize_profile_name(name)
    if not normalized:
        return None
    with get_db() as conn:
        return conn.execute(
            """
            SELECT id, name, is_admin, created_at, updated_at
            FROM prograde_access_profiles
            WHERE lower(name)=lower(?)
            LIMIT 1
            """,
            (normalized,),
        ).fetchone()


def create_access_profile(name, is_admin=False):
    normalized = _normalize_profile_name(name)
    if not normalized:
        raise ValueError("Account name is required.")
    now = datetime.utcnow().isoformat()
    try:
        with get_db() as conn:
            cursor = conn.execute(
                """
                INSERT INTO prograde_access_profiles (name, is_admin, created_at, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                (normalized, 1 if is_admin else 0, now, now),
            )
            return int(cursor.lastrowid)
    except sqlite3.IntegrityError as exc:
        raise ValueError("Account name already exists.") from exc


def create_session(
    session_id,
    brand,
    carrier_type,
    planner_name,
    session_label,
    is_saved=False,
    created_by_profile_id=None,
    created_by_name=None,
):
    now = datetime.utcnow().isoformat()
    builder_name = _normalize_profile_name(created_by_name or planner_name)
    try:
        builder_profile_id = int(created_by_profile_id) if created_by_profile_id is not None else None
    except (TypeError, ValueError):
        builder_profile_id = None
    with get_db() as conn:
        conn.execute(
            """INSERT INTO load_sessions
               (session_id, brand, carrier_type, status, is_saved, planner_name, created_by_profile_id, created_by_name, session_label, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (
                session_id,
                brand,
                carrier_type,
                "draft",
                1 if is_saved else 0,
                planner_name,
                builder_profile_id,
                builder_name,
                session_label,
                now,
                now,
            ),
        )

def get_session(session_id):
    with get_db() as conn:
        return conn.execute("SELECT * FROM load_sessions WHERE session_id=?", (session_id,)).fetchone()

def get_all_sessions(brand=None, saved_only=False):
    where = []
    params = []
    if brand:
        where.append("lower(ls.brand)=lower(?)")
        params.append(str(brand))
    if saved_only:
        where.append("COALESCE(ls.is_saved, 1)=1")
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    with get_db() as conn:
        return conn.execute(
            f"""
            SELECT
                ls.*,
                COALESCE(lp.position_count, 0) AS trailer_qty
            FROM load_sessions ls
            LEFT JOIN (
                SELECT session_id, COUNT(*) AS position_count
                FROM load_positions
                GROUP BY session_id
            ) lp ON lp.session_id = ls.session_id
            {where_sql}
            ORDER BY ls.created_at DESC
            """,
            tuple(params),
        ).fetchall()


def get_session_daily_sequence(session_id, brand, created_at):
    """Return 1-based sequence for a brand within the session's created date."""
    if not session_id or not created_at:
        return 1
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT COUNT(*) AS seq
            FROM load_sessions
            WHERE lower(brand)=lower(?)
              AND (COALESCE(is_saved, 1)=1 OR session_id=?)
              AND substr(created_at, 1, 10)=substr(?, 1, 10)
              AND (created_at < ? OR (created_at = ? AND session_id <= ?))
            """,
            (brand or "", session_id, created_at, created_at, created_at, session_id),
        ).fetchone()
    seq = int(row["seq"] or 0) if row else 0
    return seq if seq > 0 else 1

def flag_session_stale(session_id):
    with get_db() as conn:
        conn.execute(
            "UPDATE load_sessions SET status='stale', updated_at=? WHERE session_id=?",
            (datetime.utcnow().isoformat(), session_id)
        )

def flag_all_draft_sessions_stale():
    """Called when settings change — flags all open sessions for revalidation."""
    with get_db() as conn:
        conn.execute(
            "UPDATE load_sessions SET status='stale', updated_at=? WHERE status IN ('draft','review')",
            (datetime.utcnow().isoformat(),)
        )

def mark_session_active(session_id):
    """Clear stale flag after revalidation."""
    with get_db() as conn:
        conn.execute(
            "UPDATE load_sessions SET status='draft', updated_at=? WHERE session_id=? AND status='stale'",
            (datetime.utcnow().isoformat(), session_id)
        )

def save_session(session_id):
    """Persist an explicit user save action and mark the session as logged."""
    ts = datetime.utcnow().isoformat()
    with get_db() as conn:
        conn.execute(
            "UPDATE load_sessions SET is_saved=1, updated_at=? WHERE session_id=?",
            (ts, session_id),
        )
        return conn.execute(
            "SELECT session_id, status, is_saved, updated_at FROM load_sessions WHERE session_id=?",
            (session_id,),
        ).fetchone()


def update_session_carrier_type(session_id, carrier_type):
    ts = datetime.utcnow().isoformat()
    with get_db() as conn:
        conn.execute(
            "UPDATE load_sessions SET carrier_type=?, updated_at=? WHERE session_id=?",
            (carrier_type, ts, session_id),
        )
        return conn.execute(
            "SELECT session_id, carrier_type, updated_at FROM load_sessions WHERE session_id=?",
            (session_id,),
        ).fetchone()


def touch_session(session_id):
    """Backward-compatible alias for explicit save action."""
    return save_session(session_id)


def delete_session(session_id):
    with get_db() as conn:
        conn.execute("DELETE FROM load_positions WHERE session_id=?", (session_id,))
        deleted = conn.execute("DELETE FROM load_sessions WHERE session_id=?", (session_id,))
        return int(deleted.rowcount or 0)

def get_acknowledged_violations(session_id):
    """Return list of acknowledged rule_codes for this session."""
    import json
    with get_db() as conn:
        try:
            row = conn.execute(
                "SELECT acknowledged_violations FROM load_sessions WHERE session_id=?",
                (session_id,)
            ).fetchone()
        except sqlite3.OperationalError as exc:
            if "no such column" in str(exc).lower():
                return []
            raise
    if not row or not row["acknowledged_violations"]:
        return []
    try:
        return json.loads(row["acknowledged_violations"])
    except Exception:
        return []

def set_acknowledged_violations(session_id, ack_list):
    """Persist updated acknowledgment list (list of rule_code strings)."""
    import json
    with get_db() as conn:
        try:
            conn.execute(
                "UPDATE load_sessions SET acknowledged_violations=? WHERE session_id=?",
                (json.dumps(ack_list), session_id)
            )
        except sqlite3.OperationalError as exc:
            if "no such column" not in str(exc).lower():
                raise


# ── Load positions ────────────────────────────────────────────────────────────

def get_positions(session_id):
    with get_db() as conn:
        return conn.execute(
            "SELECT * FROM load_positions WHERE session_id=? ORDER BY deck_zone, sequence, layer",
            (session_id,)
        ).fetchall()


def get_position(position_id):
    with get_db() as conn:
        return conn.execute("SELECT * FROM load_positions WHERE position_id=?", (position_id,)).fetchone()

def add_position(
    position_id,
    session_id,
    brand,
    item_number,
    deck_zone,
    layer,
    sequence,
    override_reason=None,
    is_rotated=0,
):
    with get_db() as conn:
        conn.execute(
            """INSERT INTO load_positions
               (position_id, session_id, brand, item_number, deck_zone, layer, sequence, is_rotated, override_reason, added_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (position_id, session_id, brand, item_number, deck_zone, layer, sequence, int(bool(is_rotated)), override_reason,
             datetime.utcnow().isoformat())
        )

def remove_position(position_id):
    with get_db() as conn:
        conn.execute("DELETE FROM load_positions WHERE position_id=?", (position_id,))

def update_position_field(position_id, field, value):
    with get_db() as conn:
        conn.execute(
            f"UPDATE load_positions SET {field}=? WHERE position_id=?",
            (value, position_id)
        )


def _load_zone_columns(connection, session_id, deck_zone):
    rows = connection.execute(
        """
        SELECT position_id, sequence, layer
        FROM load_positions
        WHERE session_id=? AND deck_zone=?
        ORDER BY sequence ASC, layer ASC
        """,
        (session_id, deck_zone),
    ).fetchall()
    grouped = []
    for row in rows:
        sequence = int(row["sequence"])
        if not grouped or grouped[-1]["sequence"] != sequence:
            grouped.append({"sequence": sequence, "ids": []})
        grouped[-1]["ids"].append(row["position_id"])
    return grouped


def _find_column_index(columns, sequence):
    for idx, col in enumerate(columns):
        if int(col["sequence"]) == int(sequence):
            return idx
    return -1


def _apply_zone_columns_layout(connection, session_id, deck_zone, columns):
    for seq_idx, col in enumerate(columns, start=1):
        for layer_idx, position_id in enumerate(col["ids"], start=1):
            connection.execute(
                """
                UPDATE load_positions
                SET deck_zone=?, sequence=?, layer=?
                WHERE session_id=? AND position_id=?
                """,
                (deck_zone, seq_idx, layer_idx, session_id, position_id),
            )


def move_column(session_id, from_zone, sequence, to_zone, insert_index=None):
    with get_db() as conn:
        from_columns = _load_zone_columns(conn, session_id, from_zone)
        if not from_columns:
            return None

        src_idx = _find_column_index(from_columns, sequence)
        if src_idx < 0:
            return None

        source_column = from_columns.pop(src_idx)

        if from_zone == to_zone:
            target_columns = from_columns
            target_idx = len(target_columns) if insert_index is None else int(insert_index)
            target_idx = max(0, min(target_idx, len(target_columns)))
            target_columns.insert(target_idx, source_column)
            _apply_zone_columns_layout(conn, session_id, from_zone, target_columns)
            return {
                "from_zone": from_zone,
                "to_zone": to_zone,
                "insert_index": target_idx,
                "sequence": target_idx + 1,
                "count": len(source_column["ids"]),
            }

        to_columns = _load_zone_columns(conn, session_id, to_zone)
        target_idx = len(to_columns) if insert_index is None else int(insert_index)
        target_idx = max(0, min(target_idx, len(to_columns)))
        to_columns.insert(target_idx, source_column)

        _apply_zone_columns_layout(conn, session_id, from_zone, from_columns)
        _apply_zone_columns_layout(conn, session_id, to_zone, to_columns)
        return {
            "from_zone": from_zone,
            "to_zone": to_zone,
            "insert_index": target_idx,
            "sequence": target_idx + 1,
            "count": len(source_column["ids"]),
        }


def move_position(session_id, position_id, to_zone, to_sequence=None, insert_index=None, to_layer_index=None):
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT position_id, deck_zone, sequence, layer
            FROM load_positions
            WHERE session_id=? AND position_id=?
            """,
            (session_id, position_id),
        ).fetchone()
        if not row:
            return None

        from_zone = row["deck_zone"]
        from_sequence = int(row["sequence"])
        from_layer = int(row["layer"] or 0)
        from_columns = _load_zone_columns(conn, session_id, from_zone)
        src_idx = _find_column_index(from_columns, from_sequence)
        if src_idx < 0:
            return None

        source_column = from_columns[src_idx]
        if position_id not in source_column["ids"]:
            return None
        source_column["ids"].remove(position_id)
        if not source_column["ids"]:
            from_columns.pop(src_idx)

        if to_zone == from_zone:
            target_columns = from_columns
        else:
            target_columns = _load_zone_columns(conn, session_id, to_zone)

        if to_sequence is not None:
            target_idx = _find_column_index(target_columns, int(to_sequence))
            if target_idx < 0:
                return None
            target_ids = target_columns[target_idx]["ids"]
            if to_layer_index is not None:
                normalized_layer_index = int(to_layer_index)
                if (
                    to_zone == from_zone
                    and int(to_sequence) == from_sequence
                    and normalized_layer_index > from_layer
                ):
                    normalized_layer_index -= 1
                insert_at = max(0, min(normalized_layer_index - 1, len(target_ids)))
                target_ids.insert(insert_at, position_id)
                final_layer = insert_at + 1
            else:
                target_ids.append(position_id)
                final_layer = len(target_ids)
            final_sequence = target_idx + 1
        else:
            target_idx = len(target_columns) if insert_index is None else int(insert_index)
            target_idx = max(0, min(target_idx, len(target_columns)))
            target_columns.insert(target_idx, {"sequence": None, "ids": [position_id]})
            final_sequence = target_idx + 1
            final_layer = 1

        if to_zone == from_zone:
            _apply_zone_columns_layout(conn, session_id, from_zone, target_columns)
        else:
            _apply_zone_columns_layout(conn, session_id, from_zone, from_columns)
            _apply_zone_columns_layout(conn, session_id, to_zone, target_columns)

        return {
            "position_id": position_id,
            "from_zone": from_zone,
            "to_zone": to_zone,
            "sequence": final_sequence,
            "layer": final_layer,
        }


def _next_zone_sequence(connection, session_id, deck_zone):
    row = connection.execute(
        "SELECT COALESCE(MAX(sequence), 0) AS max_seq FROM load_positions WHERE session_id=? AND deck_zone=?",
        (session_id, deck_zone),
    ).fetchone()
    return int(row["max_seq"] or 0) + 1


def duplicate_column(session_id, deck_zone, sequence):
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM load_positions
            WHERE session_id=? AND deck_zone=? AND sequence=?
            ORDER BY layer ASC
            """,
            (session_id, deck_zone, int(sequence)),
        ).fetchall()
        if not rows:
            return None

        new_seq = _next_zone_sequence(conn, session_id, deck_zone)
        id_map = {row["position_id"]: str(uuid.uuid4()) for row in rows}
        now = datetime.utcnow().isoformat()
        for row in rows:
            nested_inside = row["nested_inside"]
            if nested_inside in id_map:
                nested_inside = id_map[nested_inside]
            conn.execute(
                """
                INSERT INTO load_positions
                (position_id, session_id, brand, item_number, deck_zone, layer, sequence, is_nested, nested_inside, gn_axle_dropped, is_rotated, override_reason, added_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    id_map[row["position_id"]],
                    row["session_id"],
                    row["brand"],
                    row["item_number"],
                    row["deck_zone"],
                    row["layer"],
                    new_seq,
                    row["is_nested"],
                    nested_inside,
                    row["gn_axle_dropped"],
                    row["is_rotated"],
                    row["override_reason"],
                    now,
                ),
            )
    return {"new_sequence": new_seq, "count": len(rows)}


def move_column_zone(session_id, from_zone, sequence, to_zone):
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT position_id
            FROM load_positions
            WHERE session_id=? AND deck_zone=? AND sequence=?
            ORDER BY layer ASC
            """,
            (session_id, from_zone, int(sequence)),
        ).fetchall()
        if not rows:
            return None
        new_seq = _next_zone_sequence(conn, session_id, to_zone)
        for row in rows:
            conn.execute(
                """
                UPDATE load_positions
                SET deck_zone=?, sequence=?
                WHERE position_id=?
                """,
                (to_zone, new_seq, row["position_id"]),
            )
    return {"new_sequence": new_seq, "count": len(rows)}


def resequence_column(session_id, deck_zone, sequence, direction):
    step = -1 if str(direction).lower() == "left" else 1
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT sequence
            FROM load_positions
            WHERE session_id=? AND deck_zone=?
            ORDER BY sequence ASC
            """,
            (session_id, deck_zone),
        ).fetchall()
        sequences = [int(row["sequence"]) for row in rows]
        if int(sequence) not in sequences:
            return None
        idx = sequences.index(int(sequence))
        target_idx = idx + step
        if target_idx < 0 or target_idx >= len(sequences):
            return {"moved": False}
        src = sequences[idx]
        dst = sequences[target_idx]
        temp_seq = -999999
        conn.execute(
            "UPDATE load_positions SET sequence=? WHERE session_id=? AND deck_zone=? AND sequence=?",
            (temp_seq, session_id, deck_zone, src),
        )
        conn.execute(
            "UPDATE load_positions SET sequence=? WHERE session_id=? AND deck_zone=? AND sequence=?",
            (src, session_id, deck_zone, dst),
        )
        conn.execute(
            "UPDATE load_positions SET sequence=? WHERE session_id=? AND deck_zone=? AND sequence=?",
            (dst, session_id, deck_zone, temp_seq),
        )
    return {"moved": True, "new_sequence": dst}


def clear_session_positions(session_id):
    with get_db() as conn:
        conn.execute("DELETE FROM load_positions WHERE session_id=?", (session_id,))

def get_pj_skus_for_tongue_group(group_id):
    """Return all PJ SKUs belonging to a tongue group."""
    with get_db() as conn:
        return conn.execute(
            "SELECT * FROM pj_skus WHERE tongue_group=?", (group_id,)
        ).fetchall()
